import ome_types
from pycromanager import Core, Acquisition, multi_d_acquisition_events, Dataset, MagellanAcquisition, Magellan, start_headless, XYTiledAcquisition
import numpy as np
import time
from scipy.optimize import curve_fit
import paho.mqtt.client as mqtt
import matplotlib.pyplot as plt
from skimage import io, measure, filters
import os
import math
from datetime import datetime
from tifffile import imsave, imwrite
from openpyxl import load_workbook, Workbook
from ome_types.model import Instrument, Microscope, Objective, InstrumentRef, Image, Pixels, Plane, Channel
from ome_types.model.simple_types import UnitsLength, PixelType, PixelsID, ImageID, ChannelID
from skspatial.objects import Plane, Points
from skspatial.plotting import plot_3d
from sklearn.datasets import make_regression
from sklearn.linear_model import HuberRegressor, LinearRegression, RANSACRegressor
from sklearn.model_selection import cross_val_score
from sklearn.model_selection import RepeatedKFold
from ome_types import from_xml, OME, from_tiff
import sys
from ctypes import *
sys.path.append(r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF\Python_64_elveflow\DLL64')#add the path of the library here
sys.path.append(r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF\Python_64_elveflow')#add the path of the LoadElveflow.py

from array import array
from Elveflow64 import *



#client = mqtt.Client('autocyplex_server')
#client.connect('10.3.141.1', 1883)

core = Core()
magellan = Magellan()

global level
level = []


class brenner:
    def __init__(self):
        brenner.value = []
        return

class exp_level:
    def __init__(self):
        exp_level.value = []
        return

class cycif:

    def __init__(self):

        return

    def tissue_center(self, mag_surface):
        '''
        take magellan surface and find the xy coordinates of the center of the surface
        :param mag_surface:
        :param magellan:
        :return: x tissue center position and y tissue center position
        :rtype: list[float, float]
        '''
        xy_pos = self.tile_xy_pos(mag_surface)
        x_center = (max(xy_pos[0]) + min(xy_pos[0])) / 2
        y_center = (max(xy_pos[1]) + min(xy_pos[1])) / 2
        return x_center, y_center

    def num_surfaces_count(self):
        '''
        Looks at magellan surfaces that start with New Surface in its name, ie. 'New Surface 1' as that is the default generated prefix.

        :param object magellan: magellan object from magellan = Magellan() in pycromanager
        :return: surface_count
        :rtype: int
        '''
        x = 1
        while magellan.get_surface("New Surface " + str(x)) != None:
            x += 1
        surface_count = x - 1
        time.sleep(1)

        return surface_count

    def surface_exist_check(self, surface_name):
        '''
        Checks name of surface to see if exists. If it does, returns 1, if not returns 0

        :param object magellan: magellan object from magellan = Magellan() in pycromanager
        :param str surface_name: name of surface to check if exists

        :return: status
        :rtype: int
        '''

        status = 0
        if magellan.get_surface(surface_name) != None:
            status += 1

        return status

    ####################################################################
    ############ All in section are functions for the autofocus function
    ####################################################################

    def focus_score(self, image, derivative_jump):
        '''
        Calculates focus score on image with Brenners algorithm on downsampled image.


        :param numpy image: single image from hooked from acquistion

        :return: focus score for image
        :rtype: float
        '''
        # Note: Uniform background is a bit mandatory

        # do Brenner score
        a = image[derivative_jump:, :]
        a = a.astype('float64')
        b = image[:-derivative_jump, :]
        b = b.astype('float64')
        c = (a - b)
        c = c * c/10
        f_score_shadow = c.sum(dtype=np.float64) + 0.00001

        return 1 / f_score_shadow

    def sp_array(self, experiment_directory, channel):
        '''
        Generate super pixel array as defined in powerpoint autofocus network
        :param string experiment_directory:
        :param string channel:
        :param y_tiles:
        :param x_tiles:
        :return:
        '''

        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)

        fm_array = np.load('fm_array.npy', allow_pickle=False)

        y_tiles = np.shape(fm_array[0])[0]
        x_tiles = np.shape(fm_array[0])[1]

        x_pixel_count = int((x_tiles + 0.1) * 32)
        y_pixel_count = int((y_tiles + 0.1) * 24)
        sp_array = np.random.rand(5, y_pixel_count, x_pixel_count, 2).astype('float64')

        filename = channel + '_sp_array.npy'
        np.save(filename, sp_array)

    def tile_subsampler(self, experiment_directory):
        '''
        Outs grid of dimensions [y_tiles, x_tiles] of 1 or 0s that indicate with 1 if the tile is chosen to be sampled. Currently, it samples all tiles.
        :return:
        '''

        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)

        full_array = np.load('fm_array.npy', allow_pickle=False)
        numpy_x = full_array[0]
        x_tile_count = np.shape(numpy_x)[1]
        y_tile_count = np.shape(numpy_x)[0]

        subsample_grid = np.ones((y_tile_count, x_tile_count))

        return subsample_grid

    def image_sub_divider(self, whole_image, y_sections, x_sections):
        '''
        takes in image and breaks into subsections of size y_sections by x_sections.
        :param whole_image:
        :param y_sections:
        :param x_sections:
        :return:
        '''

        y_pixels = np.shape(whole_image)[0]
        x_pixels = np.shape(whole_image)[1]
        sub_divided_image = np.random.rand(y_sections, x_sections, int(y_pixels / y_sections), int(x_pixels / x_sections)).astype('uint16')
        for y in range(0, y_sections):
            for x in range(0, x_sections):
                # define y and x start and ends subsection of rebuilt image

                y_start = int(y * (y_pixels / y_sections))
                y_end = y_start + int(y_pixels / y_sections)
                x_start = int(x * (x_pixels / x_sections))
                x_end = x_start + int(x_pixels / x_sections)

                sub_divided_image[y][x] = whole_image[y_start:y_end, x_start:x_end]

        return sub_divided_image

    def sub_divided_2_brenner_sp(self, experiment_directory, sub_divided_image, channel, point_number, z_position, x_tile_number, y_tile_number):
        '''
        Takes in subdivided image and calculated brenner score at each subsection and properly places into sp_array
        :param experiment_directory:
        :param sub_divided_image:
        :param channel:
        :param point_number:
        :param x_tile_number:
        :param y_tile_number:
        :return:
        '''

        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = channel + '_sp_array.npy'
        sp_array = np.load(file_name, allow_pickle=False)
        sp_array_slice = sp_array[point_number]

        derivative_jump = 15

        y_subdivisions = 2
        x_subdivisions = 2

        y_offset = int(y_subdivisions * y_tile_number)
        x_offset = int(x_subdivisions * x_tile_number)

        for y in range(y_offset, y_subdivisions + y_offset):
            for x in range(x_offset, x_subdivisions + x_offset):

                score = self.focus_score(sub_divided_image[y - y_offset][x - x_offset], derivative_jump)
                sp_array_slice[y][x][0] = score
                sp_array_slice[y][x][1] = z_position

        np.save(file_name, sp_array)

    def sp_array_focus_solver(self, experiment_directory, channel):
        '''
        takes fully populated sp point array and applied 3 point brenner solver method to populate predicted focus

        :param experiment_directory:
        :param channel:
        :return:
        '''

        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = channel + '_sp_array.npy'
        sp_array = np.load(file_name, allow_pickle=False)

        y_sp_pixels = np.shape(sp_array[0])[0]
        x_sp_pixels = np.shape(sp_array[0])[1]

        for y in range(0, y_sp_pixels):
            for x in range(0, x_sp_pixels):

                scores = sp_array[0:3, y, x, 0]
                scores = scores / np.min(scores)
                positions = sp_array[0:3, y, x, 1]
                three_point_array = np.stack((scores, positions), axis=1)


                a, b, c, predicted_focus = self.gauss_jordan_solver(three_point_array)
                sp_array[3][y][x][0] = predicted_focus
                print(predicted_focus)

            np.save(file_name, sp_array)

    def sp_array_filter(self, experiment_directory, channel):
        '''
        takes sp array for a channel and generates mask that filters out nonsense answers.
        These answers are solutions that exist outside of the scan range and ones that have a low well depth
        :param experiment_directory:
        :param channel:
        :return:
        '''

        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = channel + '_sp_array.npy'
        sp_array = np.load(file_name, allow_pickle=False)

        y_dim = np.shape(sp_array)[1]
        x_dim = np.shape(sp_array)[2]

        well_depth = np.random.rand(y_dim, x_dim)

        # execute otsu threshold value calc.
        for y in range(0, y_dim):
            for x in range(0, x_dim):
                score_array = sp_array[0:3, y, x, 0]
                score_array = score_array / np.min(score_array)
                depth = np.max(score_array)
                well_depth[y][x] = depth

            threshold = filters.threshold_otsu(well_depth)  # sets filter threshold. If dont want otsu, just chang eto number
            # threshold = 1.1
            for y in range(0, y_dim):
                for x in range(0, x_dim):

                    position_array = sp_array[0:3, y, x, 1]
                    highest_pos = np.max(position_array)
                    lowest_pos = np.min(position_array)

                    if well_depth[y][x] > threshold and lowest_pos < sp_array[3][y][x][0] < highest_pos:
                        sp_array[4][y][x][0] = 1
                    else:
                        sp_array[4][y][x][0] = 0

            np.save(file_name, sp_array)

    def plane_2_z(self, coefficents, xy_point):

        a = coefficents[0]
        b = coefficents[1]
        c = coefficents[2]

        z = a * xy_point[0] + b * xy_point[1] + c

        return z

    def sp_array_surface_2_fm(self, experiment_directory, channel):
        '''
        Takes fully constructed sp array with mask and predicted focus position and fits plane to points.
        :param experiment_directory:
        :param channel:
        :return:
        '''

        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        fm_array = np.load('fm_array.npy', allow_pickle=False)

        y_tiles = np.shape(fm_array[0])[0]
        x_tiles = np.shape(fm_array[0])[1]

        depth_of_focus = 2 #in microns

        point_list = self.map_2_points(experiment_directory, channel)
        X = point_list[:, 0:2]
        y = point_list[:, 2]

        model = HuberRegressor()
        model.fit(X, y)

        a = (model.predict([[1000, 2500]]) - model.predict([[0, 2500]])) / 1000
        b = (model.predict([[1000, 2500]]) - model.predict([[1000, 1500]])) / 1000
        c = model.predict([[2000, 2000]]) - a * 2000 - b * 2000

        coefficents = [a, b, c]

        if channel == 'DAPI':
            channel_index = 2
        if channel == 'A488':
            channel_index = 4
        if channel == 'A555':
            channel_index = 6
        if channel == 'A647':
            channel_index = 8

        #calc number of slices needed

        high_z = self.plane_2_z(coefficents, [0, 0])
        low_z = self.plane_2_z(coefficents, [5056, 2960])
        corner_corner_difference = math.fabs(high_z - low_z)
        number_planes = int(corner_corner_difference/depth_of_focus) + 1

        for y in range(0, y_tiles):
            for x in range(0, x_tiles):
                #I believe the system to count from upper left hand corner starting at 0, 0

                x_point = 5056 * x + 2528
                y_point = 2060 * y + 1480
                focus_z = self.plane_2_z(coefficents, [x_point, y_point])
                fm_array[channel_index][y][x] = focus_z + (number_planes - 1)/2 * depth_of_focus
                fm_array[channel_index + 1][y][x] = number_planes

        np.save('fm_array.npy', fm_array)


    def map_2_points(self, experiment_directory, channel):

        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = channel + '_sp_array.npy'
        sp_array = np.load(file_name, allow_pickle=False)

        y_fm_section = np.shape(sp_array)[1]
        x_fm_section = np.shape(sp_array)[2]

        point_list = np.array([])
        point_list = np.expand_dims(point_list, axis=0)

        first_point_counter = 0

        z_array = sp_array[4, :, :, 0] * sp_array[3, :, :, 0]

        for y in range(0, y_fm_section):
            for x in range(0, x_fm_section):

                x_coord = x * 2528
                y_coord = y * 1480
                z_coord = z_array[y][x]
                single_point = np.array([x_coord, y_coord, z_coord])
                single_point = np.expand_dims(single_point, axis=0)
                if first_point_counter == 0 and z_coord != 0:
                    point_list = np.append(point_list, single_point, axis=1)
                    first_point_counter = 1
                elif first_point_counter == 1 and z_coord != 0:
                    point_list = np.append(point_list, single_point, axis=0)

        point_list = Points(point_list)

        return point_list



    def gauss_jordan_solver(self, three_point_array):
        '''
        Takes 3 points and solves quadratic equation in a generic fashion and returns constants and
        solves for x in its derivative=0 equation

        :param numpy[float, float] three_point_array: numpy array that contains pairs of [focus_score, z]
        :results: z coordinate for in focus plane
        :rtype: float
        '''

        x1 = three_point_array[0][1]
        x2 = three_point_array[1][1]
        x3 = three_point_array[2][1]
        score_0 = three_point_array[0][0]
        score_1 = three_point_array[1][0]
        score_2 = three_point_array[2][0]

        aug_matrix = np.array([[x1 * x1, x1, 1, score_0], [x2 * x2, x2, 1, score_1], [x3 * x3, x3, 1, score_2]])

        aug_matrix[0] = aug_matrix[0] / (aug_matrix[0, 0] + 0.000001)
        aug_matrix[1] = -(aug_matrix[1, 0]) * aug_matrix[0] + aug_matrix[1]
        aug_matrix[2] = -(aug_matrix[2, 0]) * aug_matrix[0] + aug_matrix[2]

        aug_matrix[1] = -(aug_matrix[1, 1] - 1) / (aug_matrix[2, 1] + 0.0000001) * aug_matrix[2] + aug_matrix[1]
        aug_matrix[0] = -aug_matrix[0, 1] * aug_matrix[1] + aug_matrix[0]
        aug_matrix[2] = -aug_matrix[2, 1] * aug_matrix[1] + aug_matrix[2]

        aug_matrix[2] = aug_matrix[2] / (aug_matrix[2, 2] + 0.000001)
        aug_matrix[0] = -aug_matrix[0, 2] * aug_matrix[2] + aug_matrix[0]
        aug_matrix[1] = -aug_matrix[1, 2] * aug_matrix[2] + aug_matrix[1]

        a = aug_matrix[0, 3]
        b = aug_matrix[1, 3]
        c = aug_matrix[2, 3]

        derivative = -b / (2 * (a + 0.00001))

        return a,b,c,derivative

    #########################################################
    # Setup fm_array and sp_array alongside auto focus updates and flat values
    #########################################################

    def establish_fm_array(self, experiment_directory, desired_cycle_count, z_slices, off_array, x_crop_percentage = 0, autofocus = 0):
        # non autofocus. Need to build in auto focus ability
        self.file_structure(experiment_directory, desired_cycle_count)
        xy_points = self.tile_xy_pos('New Surface 1')
        xyz_points = self.nonfocus_tile_DAPI(xy_points)
        self.tile_pattern(xyz_points)
        self.x_overlap_adjuster(x_crop_percentagfe, experiment_directory)

        if x_crop_percentage != 0:
            self.fm_channel_initial(experiment_directory, array, off_array, z_slices)
        else:
            pass

        if autofocus == 1:
            self.fm_array_update_autofocus_autoexpose(experiment_directory)
        else:
            pass

    def fm_array_update_autofocus_autoexpose(self, experiment_directory):

        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)

        y_tiles = np.shape(fm_array[0])[0]
        x_tiles = np.shape(fm_array[0])[1]
        scan_range = 20
        sample_span = [sample_mid_z - scan_range / 2, sample_mid_z, sample_mid_z + scan_range / 2]

        for x in range(0, x_tiles):
            for y in range(0, y_tiles):
                for points in range(0, 3):
                    z_slice = int(sample_span[points])
                    im = self.image_capture(experiment_directory, 'DAPI', 25, y, x, z_slice)
                    # cycif.auto_exposure_calculation(im, 0.99, 'DAPI', points, z_slice, x, y)
                    div_im = self.image_sub_divider(im, 2, 2)
                    self.sub_divided_2_brenner_sp(experiment_directory, div_im, 'DAPI', points, z_slice, x, y)

        # self.calc_array_solver(experiment_directory, 'DAPI')
        # self.calc_array_2_exp_array(experiment_directory, 'DAPI', 0.5)
        self.sp_array_focus_solver(experiment_directory, 'DAPI')
        #self.sp_array_filter(experiment_directory, 'DAPI')
        self.sp_array_surface_2_fm(experiment_directory, 'DAPI')

        np.save('fm_array.npy', fm_array)

    ###########################################################
    #This section is the for the exposure functions.
    ###########################################################

    def auto_exposure_calculation(self, experiment_directory,  image, cut_off_threshold, channel, point_number, z_position, x_tile_number,
                                  y_tile_number):
        '''
        Takes in subdivided image and calculated brenner score at each subsection and properly places into sp_array
        :param experiment_directory:
        :param sub_divided_image:
        :param channel:
        :param point_number:
        :param x_tile_number:
        :param y_tile_number:
        :return:
        '''

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = channel + '_exp_calc_array.npy'
        calc_array = np.load(file_name, allow_pickle=False)
        calc_array_slice = calc_array[point_number]

        intensity = cycif.image_percentile_level(image, cut_off_threshold)
        calc_array_slice[y_tile_number][x_tile_number][0] = intensity
        calc_array_slice[y_tile_number][x_tile_number][1] = z_position

        np.save(file_name, calc_array)

    def calc_array_solver(self, experiment_directory, channel):

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = channel + '_exp_calc_array.npy'
        calc_array = np.load(file_name, allow_pickle=False)

        y_tiles = np.shape(calc_array)[1]
        x_tiles = np.shape(calc_array)[2]

        for x in range(0, x_tiles):
            for y in range(0, y_tiles):
                scores = calc_array[0:3, y, x, 0]
                # scores = 1/scores
                # scores = scores/np.min(scores)
                positions = calc_array[0:3, y, x, 1]
                three_point_array = np.stack((scores, positions), axis=1)

                a, b, c, predicted_focus = cycif.gauss_jordan_solver(three_point_array)
                peak_int = (-(b * b) / (4 * a) + c)
                calc_array[3][y][x][0] = peak_int
                calc_array[3][y][x][1] = predicted_focus
        calc_array[3, :, :, 0] = calc_array[3, :, :, 0]

        np.save(file_name, calc_array)

    def calc_array_2_exp_array(self, experiment_directory, channel, fraction_dynamic_range):
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = channel + '_exp_calc_array.npy'
        calc_array = np.load(file_name, allow_pickle=False)
        exp_array = np.load('exp_array.npy', allow_pickle=False)

        predicted_intensity_image = calc_array[3, :, :, 0]
        threshold = filters.threshold_otsu(predicted_intensity_image)
        thresholded_intensity_array = predicted_intensity_image[predicted_intensity_image > threshold]
        in_focus_int_prediction = np.median(thresholded_intensity_array)
        desired_top_intensity = fraction_dynamic_range * 65535
        scale_up_factor = desired_top_intensity / in_focus_int_prediction

        if channel == 'DAPI':
            index = 0
        if channel == 'A488':
            index = 1
        if channel == 'A555':
            index = 2
        if channel == 'A647':
            index = 3

        predicted_exp_time = exp_array[index] * scale_up_factor
        exp_array[index] = int(predicted_exp_time)

        np.save('exp_array.npy', exp_array)



    ##############################################

    def tile_xy_pos(self, surface_name):
        '''
        imports previously generated micro-magellan surface with name surface_name and outputs
        the coordinates of the center of each tile from it.

        :param str surface_name: name of micro-magellan surface
        :param object magellan: object created via = bridge.get_magellan()

        :return: XY coordinates of the center of each tile from micro-magellan surface
        :rtype: dictionary {{x:(float)}, {y:(float)}}
        '''
        surface = magellan.get_surface(surface_name)
        num = surface.get_num_positions()
        xy = surface.get_xy_positions()
        tile_points_xy = {}
        x_temp = []
        y_temp = []

        for q in range(0, num):
            pos = xy.get(q)
            pos = pos.get_center()
            x_temp.append(pos.x)
            y_temp.append(pos.y)

        tile_points_xy['x'] = x_temp  ## put all points in dictionary to ease use
        tile_points_xy['y'] = y_temp
        x_temp = np.array(tile_points_xy['x'])
        y_temp = np.array(tile_points_xy['y'])

        xy = np.append(x_temp, y_temp)
        xy = np.reshape(xy, (2, int(xy.size/2)))

        return xy


    def tile_pattern(self, numpy_array):
        '''
        Takes numpy array with N rows and known tile pattern and casts into new array that follows
        south-north, west-east snake pattern.


        :param numpy_array: dimensions [N, x_tiles*y_tiles]
        :param x_tiles: number x tiles in pattern
        :param y_tiles: number y tiles in pattern
        :return: numpy array with dimensions [N,x_tiles,y_tiles] with above snake pattern
        '''
        y_tiles = np.unique(numpy_array[1]).size
        x_tiles = np.unique(numpy_array[0]).size
        layers = np.shape(numpy_array)[0]
        numpy_array = numpy_array.reshape(layers, x_tiles, y_tiles)
        dummy = numpy_array.reshape(layers, y_tiles, x_tiles)
        new_numpy = np.empty_like(dummy)
        for x in range(0, layers):
            new_numpy[x] = numpy_array[x].transpose()
            new_numpy[x, ::, 1:y_tiles:2] = np.flipud(new_numpy[x, ::, 1:y_tiles:2])

        return new_numpy

    def fm_channel_initial(self, experiment_directory, off_array, z_slices):


        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)

        a488_channel_offset = off_array[0] #determine if each of these are good and repeatable offsets
        a555_channel_offset = off_array[1]
        a647_channel_offset = off_array[2]

        slice_gap = 2 # space between z slices in microns

        dummy_channel = np.empty_like(fm_array[0])
        dummy_channel = np.expand_dims(dummy_channel, axis=0)
        channel_count = np.shape(fm_array)[0]

        while channel_count < 10:
            full_array = np.append(fm_array, dummy_channel, axis = 0)
            channel_count = np.shape(fm_array)[0]

        fm_array[4] = fm_array[2] + a488_channel_offset #index for a488 = 3
        fm_array[6] = fm_array[2] + a555_channel_offset
        fm_array[8] = fm_array[2] + a647_channel_offset

        y_tiles = int(np.shape(fm_array[0])[0])
        x_tiles = int(np.shape(fm_array[0])[1])
        z_slice_array = np.ones((y_tiles, x_tiles))
        z_slice_array = z_slice_array * z_slices

        fm_array[3] = z_slice_array
        fm_array[5] = z_slice_array
        fm_array[7] = z_slice_array
        fm_array[9] = z_slice_array


        full_array[2] = full_array[2] + int((z_slice_array[0][0] * slice_gap)/2)
        full_array[4] = full_array[4] + int((z_slice_array[0][0] * slice_gap)/2)
        full_array[6] = full_array[6] + int((z_slice_array[0][0] * slice_gap)/2)
        full_array[8] = full_array[8] + int((z_slice_array[0][0] * slice_gap)/2)

        np.save(file_name, full_array)

        return full_array

    def x_overlap_adjuster(self, crop_each_side_percentage, experiment_directory):

        # load in fm array
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)

        # Find pixels in each dimensions
        x_tiles = np.shape(fm_array[0])[1]
        y_tiles = np.shape(fm_array[0])[0]

        x_pixels = ((x_tiles - 1) * (0.9) + 1) * 5056

        # Find number tiles in adjusted grid

        new_frame_x = int(1 - 2 * crop_each_side_percentage * 5056)
        new_x_tiles = (x_pixels / new_frame_x - 1) / 0.9 + 1
        new_x_tiles = math.ceil(new_x_tiles)

        # generate new blank fm_array numpy array

        new_fm_array = np.random.rand(10, y_tiles, new_x_tiles).astype('int16')

        # Find border where x starts on the left (not center point, but x value for left most edge of left most tile

        left_x_center = fm_array[0][0][0]
        left_most_x = left_x_center - 2528 * 0.197

        # Find center point in new image that makes edge of image align with left_most_x
        # Also find x to x + i spacing and populate rest of x values in new_fm_array

        x_col_0 = left_most_x + new_frame_x / 2 * 0.197
        x_spacing = (0.9) * new_frame_x

        # Populate new_fm_array with row 0 x values

        for x in range(0, x_tiles):
            new_fm_array[0, 0:y_tiles, x] = x_col_0 + x * x_spacing

        # populate new_fm_array with y values

        for y in range(0, y_tiles):
            new_fm_array[0, y, 0:x_tiles] = fm_array[1][y][0]

        # populate new_fm_array with dapi z values

        new_fm_array[0, 0:y_tiles, 0:x_tiles] = fm_array[2][0][0]

        np.save('fm_array.npy', new_fm_array)

    def nonfocus_tile_DAPI(self, full_array_no_pattern):
         '''
         Makes micromagellen z the focus position at each XY point for DAPI
         auto_focus and outputs the paired in focus z coordinate

         :param dictionary tile_points_xy: dictionary containing all XY coordinates. In the form: {{x:(int)}, {y:(int)}}

         :return: XYZ points where XY are stage coords and Z is in focus coordinate. {{x:(int)}, {y:(int)}, {z:(float)}}
         :rtype: dictionary
         '''

         numpy_path = experiment_directory + '/' + 'np_arrays'
         os.chdir(numpy_path)
         file_name = 'fm_array.npy'

         z_pos = magellan.get_surface('New Surface 1').get_points().get(0).z
         num = np.shape(full_array_no_pattern)[1]
         z_temp = []
         for q in range(0, num):
             z_temp.append(z_pos)
         z_temp = np.expand_dims(z_temp, axis = 0)
         xyz = np.append(full_array_no_pattern, z_temp, axis =0)

         np.save(file_name, xyz)


############################################
#Using core snap and not pycromanager acquire
############################################


    def image_capture(self, experiment_directory, channel, exp_time, x,y,z):

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        fm_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = fm_array[0]
        numpy_y = fm_array[1]
        numpy_z = fm_array[2]

        core.set_config("Color", channel)
        core.set_exposure(exp_time)

        core.set_xy_position(numpy_x[y][x], numpy_y[y][x])
        time.sleep(.5)
        core.set_position(numpy_z[y][x])

        core.snap_image()
        tagged_image = core.get_tagged_image()
        pixels = np.reshape(tagged_image.pix, newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])

        return pixels


    def core_tile_acquire(self, experiment_directory, channel = 'DAPI'):
        '''
        Makes numpy files that contain all tiles and z slices. Order is z, tiles.

        :param self:
        :param channels:
        :param z_slices:
        :return:
        '''

        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)
        exp_time_array = np.load('exp_array.npy', allow_pickle=False)

        height_pixels = 2960
        width_pixels = 5056

        numpy_x = full_array[0]
        numpy_y = full_array[1]
        x_tile_count = np.unique(numpy_x).size
        y_tile_count = np.unique(numpy_y).size
        total_tile_count = x_tile_count * y_tile_count
        z_slices = full_array[3][0][0]
        slice_gap = 2

        core.set_xy_position(numpy_x[0][0], numpy_y[0][0])
        time.sleep(1)

        tif_stack = np.random.rand(int(z_slices), total_tile_count, height_pixels, width_pixels).astype('float16')


        if channel == 'DAPI':
            channel_index = 2
            tif_stack_c_index = 0
        if channel == 'A488':
            channel_index = 4
            tif_stack_c_index = 1
        if channel == 'A555':
            channel_index = 6
            tif_stack_c_index = 2
        if channel == 'A647':
            channel_index = 8
            tif_stack_c_index = 3

        numpy_z = full_array[channel_index]
        exp_time = int(exp_time_array[tif_stack_c_index])
        core.set_config("Color", channel)
        core.set_exposure(exp_time)
        tile_counter = 0


        for y in range(0, y_tile_count):
            if y % 2 != 0:
                for x in range(x_tile_count - 1, -1, -1):

                    z_end = int(numpy_z[y][x])
                    z_start = int(z_end - z_slices * slice_gap)
                    core.set_xy_position(numpy_x[y][x], numpy_y[y][x])
                    time.sleep(.5)

                    z_counter = 0

                    for z in range(z_start, z_end, slice_gap):
                        core.set_position(z)
                        time.sleep(0.5)
                        core.snap_image()
                        tagged_image = core.get_tagged_image()
                        pixels = np.reshape(tagged_image.pix,newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
                        tif_stack[z_counter][tile_counter] = pixels

                        z_counter += 1

                    tile_counter += 1


            elif y % 2 == 0:
                for x in range(0, x_tile_count):

                    z_end = int(numpy_z[y][x])
                    z_start = int(z_end - z_slices * slice_gap)
                    core.set_xy_position(numpy_x[y][x], numpy_y[y][x])
                    time.sleep(.5)

                    z_counter = 0

                    for z in range(z_start, z_end, slice_gap):
                        core.set_position(z)
                        time.sleep(0.5)
                        core.snap_image()
                        tagged_image = core.get_tagged_image()
                        pixels = np.reshape(tagged_image.pix,
                                            newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])

                        tif_stack[z_counter][tile_counter] = pixels


                        z_counter += 1

                    tile_counter += 1

        return tif_stack
    
    def quick_tile_placement(self, z_tile_stack, overlap = 10):

        numpy_path = 'E:/folder_structure' +'/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]

        x_tile_count = np.unique(numpy_x).size
        y_tile_count = np.unique(numpy_y).size

        height = z_tile_stack[0].shape[0]
        width = z_tile_stack[0].shape[1]
        overlapped_height = int(height * (1 - overlap / 100))
        overlapped_width = int(width * (1 - overlap / 100))

        pna_height = int(y_tile_count * height - int((y_tile_count) * overlap / 100 * height))
        pna_width = int(x_tile_count * width - int((x_tile_count) * overlap / 100 * width))

        pna = np.random.rand(pna_height, pna_width).astype('float16')
        tile_counter = 0

        for y in range(0, y_tile_count):
            if y % 2 != 0:
                for x in range(x_tile_count - 1, -1, -1):
                    pna[y * overlapped_height:(y + 1) * overlapped_height,
                    x * overlapped_width:(x + 1) * overlapped_width] = z_tile_stack[tile_counter][0:overlapped_height,
                                                                       0:overlapped_width]

                    tile_counter += 1


            elif y % 2 == 0:
                for x in range(0, x_tile_count):
                    pna[y * overlapped_height:(y + 1) * overlapped_height,
                    x * overlapped_width:(x + 1) * overlapped_width] = z_tile_stack[tile_counter][0:overlapped_height,
                                                                       0:overlapped_width]

                    tile_counter += 1

        return pna

    def quick_tile_optimal_z(self, z_tile_stack):

        z_slice_count = z_tile_stack.shape[0]
        tile_count = z_tile_stack[0].shape[0]

        height = z_tile_stack[0].shape[1]
        width = z_tile_stack[0].shape[2]

        optimal_stack = np.random.rand(tile_count, height, width).astype('float16')
        score_array = np.random.rand(z_slice_count, 1).astype('float32')


        for tile in range(0, tile_count):

            for z in range(0, z_slice_count):
                score_array[z] = cycif.focus_bin_generator(z_tile_stack[z][tile])

            min_score = np.min(score_array)
            optimal_index = np.where(score_array == min_score)[0][0]
            optimal_stack[tile] = z_tile_stack[optimal_index][tile]

        return optimal_stack

    def optimal_quick_preview_qt(self, z_tile_stack, channel, cycle, experiment_directory,  overlap = 10):

        optimal_stack = self.quick_tile_optimal_z(z_tile_stack)
        optimal_qt = self.quick_tile_placement(optimal_stack, overlap)
        optimal_qt_binned = optimal_qt[0:-1:4, 0:-1:4]
        self.save_optimal_quick_tile(optimal_qt_binned, channel, cycle, experiment_directory)



    def quick_tile_all_z_save(self, z_tile_stack, channel, cycle, experiment_directory, stain_bleach,  overlap = 0):


        z_slice_count = z_tile_stack.shape[0]
        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]

        x_tile_count = int(np.unique(numpy_x).size)
        y_tile_count = int(np.unique(numpy_y).size)

        height = int(z_tile_stack.shape[2])
        width = int(z_tile_stack.shape[3])

        pna_height = int(y_tile_count * height - int((y_tile_count) * overlap / 100 * height))
        pna_width = int(x_tile_count * width - int((x_tile_count) * overlap / 100 * width))

        pna_stack = np.random.rand(z_slice_count, pna_height, pna_width).astype('float16')

        for z in range(0, z_slice_count):
            pna = self.quick_tile_placement(z_tile_stack[z], overlap)
            pna_stack[z] = pna

        self.save_quick_tile(pna_stack, channel, cycle, experiment_directory, stain_bleach)


    def image_cycle_acquire(self, cycle_number, experiment_directory, z_slices, stain_bleach, exp_time_array, offset_array, channels = ['DAPI', 'A488', 'A555', 'A647']):

        self.establish_fm_array(experiment_directory, cycle_number, z_slices, offset_array)

        exp_time = exp_time_array
        np.save('exp_array.npy', exp_time)

        for channel in channels:
            z_tile_stack = self.core_tile_acquire(experiment_directory, channel)
            self.save_files(z_tile_stack, channel, cycle_number, experiment_directory, stain_bleach)

        #self.marker_excel_file_generation(experiment_directory, cycle_number)

    #def full_cycle(self, experiment_directory, cycle_number, offset_array, stain_valve, heater_state = 0):

     #   z_slices = 7

      #  pump.liquid_action('Stain', stain_valve, heater_state)  # nuc is valve=7, pbs valve=8, bleach valve=1 (action, stain_valve, heater state (off = 0, on = 1))
       # microscope.cycle_acquire(cycle_number, experiment_directory, z_slices, 'Stain', offset_array)
        #pump.liquid_action('Bleach')  # nuc is valve=7, pbs valve=8, bleach valve=1 (action, stain_valve, heater state (off = 0, on = 1))
        #microscope.cycle_acquire(cycle_number, experiment_directory, z_slices, 'Bleach', offset_array)

######Folder System Generation########################################################


    def marker_excel_file_generation(self, experiment_directory, cycle_number):

        folder_path = experiment_directory + '/mcmicro'
        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        exp_array = np.load('exp_array.npy', allow_pickle=False)
        exp_array = np.flip(exp_array)
        
        filter_sets = ['A647', 'A555', 'A488', 'DAPI']
        emission_wavelengths = ['675', '565', '525', '450']
        exciation_wavlengths = ['647', '555', '488', '405']

        try:
            os.chdir(folder_path)
            wb = load_workbook('markers.xlsx')
        except:
            wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = 'channel_number'
        ws.cell(row=1, column=2).value = 'cycle_number'
        ws.cell(row=1, column=3).value = 'marker_name'
        ws.cell(row=1, column=4).value = 'Filter'
        ws.cell(row=1, column=5).value = 'excitation_wavlength'
        ws.cell(row=1, column=6).value = 'emission_wavlength'
        ws.cell(row=1, column=7).value = 'background'
        ws.cell(row=1, column=8).value = 'exposure'
        ws.cell(row=1, column=9).value = 'remove'
        
        for row_number in range(2, (cycle_number)*4 + 2):

            cycle_number = 4//(row_number - 2) + 1
            intercycle_channel_number = cycle_number * 4 + 1 - row_number

            ws.cell(row=row_number, column=1).value = row_number
            ws.cell(row=row_number, column=2).value = cycle_number
            ws.cell(row=row_number, column=3).value = 'Marker_' + str(row_number)
            #ws.cell(row=row_number, column=4).value = filter_sets[intercycle_channel_number]
            #ws.cell(row=row_number, column=5).value = exciation_wavlengths[intercycle_channel_number]
            #ws.cell(row=row_number, column=6).value = emission_wavelengths[intercycle_channel_number]


        row_start = (cycle_number - 1)*4 + 2
        row_end = row_start + 4

        for row_number in range(row_start, row_end):

            cycle_number = 4//(row_number - 2) + 1
            intercycle_channel_number = cycle_number * 4 + 1 - row_number

            ws.cell(row=row_number, column=8).value = exp_array[row_number-2]

        os.chdir(folder_path)
        wb.save(filename = 'markers.xlsx')


    def folder_addon(self, parent_directory_path, new_folder_names):

        os.chdir(parent_directory_path)

        for name in new_folder_names:

            add_on_folder = str(name)
            full_directory_path = parent_directory_path + '/' + add_on_folder

            try:
                os.mkdir(full_directory_path)
            except:
               pass

    def file_structure(self, experiment_directory, highest_cycle_count):

        channels = ['DAPI', 'A488', 'A555', 'A647']
        cycles = np.linspace(0,highest_cycle_count).astype(int)

        #folder layer one
        os.chdir(experiment_directory)
        self.folder_addon(experiment_directory, ['Quick_Tile'])
        self.folder_addon(experiment_directory, ['np_arrays'])
        self.folder_addon(experiment_directory, ['mcmicro'])
        self.folder_addon(experiment_directory, channels)

        #folder layer two

        for channel in channels:

            experiment_channel_directory = experiment_directory + '/' + channel

            self.folder_addon(experiment_channel_directory, ['Stain'])
            self.folder_addon(experiment_channel_directory, ['Bleach'])

        #folder layers 3 and 4

            for cycle in cycles:

                experiment_channel_stain_directory = experiment_channel_directory + '/' + 'Stain'
                experiment_channel_bleach_directory = experiment_channel_directory + '/' + 'Bleach'


                self.folder_addon(experiment_channel_stain_directory, ['cy_' + str(cycle)])
                experiment_channel_stain_cycle_directory = experiment_channel_stain_directory + '/' + 'cy_' + str(cycle)

                self.folder_addon(experiment_channel_stain_cycle_directory, ['Tiles'])


                self.folder_addon(experiment_channel_bleach_directory, ['cy_' + str(cycle)])
                experiment_channel_bleach_cycle_directory = experiment_channel_bleach_directory + '/' + 'cy_' + str(cycle)

                self.folder_addon(experiment_channel_bleach_cycle_directory, ['Tiles'])

        os.chdir(experiment_directory + '/' + 'np_arrays')

#####################################################################################################
##########Saving/File Generation Methods#############################################################################

    def z_scan_exposure_hook(image, metadata):
        '''
        Method that hooks from autofocus image acquistion calls. It takes image, calculates a focus score for it
        via focus_score method and exports a list that contains both the focus score and the z position it was taken at

        :param numpy image: single image from hooked from acquistion
        :param list[float] metadata: metadata for image

        :return: Nothing
        '''
        z = metadata.pop('ZPosition_um_Intended')  # moves up while taking z stack
        z_intensity_level = cycif.image_percentile_level(image, 0.99)
        intensity.value.append([z_intensity_level, z])

        return

    def auto_initial_expose(self, seed_expose, benchmark_threshold, channel, z_range, surface_name):
        '''
        Scans z levels around surface z center and finds brightest z position via z_scan_exposure method.
        Moves machine to that z plane and executes auto_expose method to determine proper exposure. This is meant for
        an initial exposure value for autofocus pruposes.

        :param: str surface_name: string of name of magellan surface to use
        :param list z_centers: list of z points associated with xy points where the slide tilt was compensated for
        :param: str channels: list that contains strings with channel names, for example 'DAPI'

        :return: exposure time: time for inputted channels exposure to be used for autofocus
        :rtype: int
        '''

        [x_pos, y_pos] = self.tissue_center(surface_name)
        core.set_xy_position(x_pos, y_pos)

        z_brightest = z_range[0] + z_range[2]
        core.set_position(z_brightest)

        new_exp = cycif.auto_expose(seed_expose, benchmark_threshold, z_brightest, [channel])

        return new_exp

    def post_acquisition_processor(self, experiment_directory):

        mcmicro_path = experiment_directory + r'\mcmicro\raw'
        dapi_im_path = experiment_directory + '\DAPI\Stain'
        cycle_start = 1
        cycle_start_search = 0

        os.chdir(mcmicro_path)
        while cycle_start_search == 0:
            file_name = str(experiment_directory.split("\\")[-1]) + '-cycle-0' + str(cycle_start) + '.ome.tif'
            if os.path.isfile(file_name) == 1:
                cycle_start += 1
            else:
                cycle_start_search = 1

        cycle_end = len(os.listdir(dapi_im_path)) + 1

        for cycle_number in range(cycle_start, cycle_end):
            self.infocus(experiment_directory, cycle_number, 10 ,10)
            self.mcmicro_image_stack_generator(cycle_number, experiment_directory)
            self.stage_placement(experiment_directory, cycle_number)

    def mcmicro_image_stack_generator(self, cycle_number, experiment_directory):

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        xml_metadata = cycif.metadata_generator(experiment_directory)

        numpy_x = full_array[0]
        numpy_y = full_array[1]
        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]
        tile_count = int(x_tile_count * y_tile_count)

        dapi_im_path = experiment_directory + '\DAPI\Stain\cy_' + str(cycle_number) + '\Tiles' + '/focused'
        a488_im_path = experiment_directory + '\A488\Stain\cy_' + str(cycle_number) + '\Tiles' + '/focused'
        a555_im_path = experiment_directory + '\A555\Stain\cy_' + str(cycle_number) + '\Tiles' + '/focused'
        a647_im_path = experiment_directory + '\A647\Stain\cy_' + str(cycle_number) + '\Tiles' + '/focused'

        mcmicro_path = experiment_directory + r'\mcmicro\raw'

        mcmicro_stack = np.random.rand(tile_count * 4, 2960, 5056).astype('uint16')

        tile = 0
        for x in range(0, x_tile_count):
            for y in range(0, y_tile_count):
                dapi_file_name = 'x' + str(x) + '_y_' + str(y) + '_c_DAPI.tif'
                a488_file_name = 'x' + str(x) + '_y_' + str(y) + '_c_A488.tif'
                a555_file_name = 'x' + str(x) + '_y_' + str(y) + '_c_A555.tif'
                a647_file_name = 'x' + str(x) + '_y_' + str(y) + '_c_A647.tif'

                base_count_number_stack = tile * 4

                os.chdir(dapi_im_path)
                image = io.imread(dapi_file_name).astype('uint16')
                mcmicro_stack[base_count_number_stack + 0] = image

                os.chdir(a488_im_path)
                image = io.imread(a488_file_name).astype('uint16')
                mcmicro_stack[base_count_number_stack + 1] = image

                os.chdir(a555_im_path)
                image = io.imread(a555_file_name).astype('uint16')
                mcmicro_stack[base_count_number_stack + 2] = image

                os.chdir(a647_im_path)
                image = io.imread(a647_file_name).astype('uint16')
                mcmicro_stack[base_count_number_stack + 3] = image

                tile += 1

        os.chdir(mcmicro_path)
        mcmicro_file_name = str(experiment_directory.split("\\")[-1]) + '-cycle-0' + str(cycle_number) + '.ome.tif'
        tf.imwrite(mcmicro_file_name, mcmicro_stack, photometric='minisblack', description=xml_metadata)

    def metadata_generator(self, experiment_directory):

        new_ome = OME()
        ome = from_xml(r'C:\Users\mike\Documents\GitHub\AutoCIF/image.xml')
        ome = ome.images[0]

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]
        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]
        total_tile_count = x_tile_count * y_tile_count

        y_gap = 532
        col_col_gap = 10
        for r in range(3, -1, -1):
            numpy_y[r][0] = numpy_y[r + 1][0] - y_gap
            numpy_y[r][1] = numpy_y[r + 1][1] - y_gap - col_col_gap

        # sub in needed pixel size and pixel grid changes
        ome.pixels.physical_size_x = 0.2
        ome.pixels.physical_size_y = 0.2
        ome.pixels.size_x = 5056
        ome.pixels.size_y = 2960
        # sub in other optional numbers to make metadata more accurate

        for x in range(0, total_tile_count):
            tile_metadata = copy.deepcopy(ome)
            new_ome.images.append(tile_metadata)

        # sub in stage positional information into each tile. numpy[y][x]
        tile_counter = 0
        for x in range(0, x_tile_count):
            for y in range(0, y_tile_count):

                for p in range(0, 4):
                    new_x = numpy_x[y][x] - 11000
                    new_y = numpy_y[y][x] + 2300
                    new_ome.images[tile_counter].pixels.planes[p].position_y = copy.deepcopy(new_y)
                    new_ome.images[tile_counter].pixels.planes[p].position_x = copy.deepcopy(new_x)
                    new_ome.images[tile_counter].pixels.tiff_data_blocks[p].ifd = (4 * tile_counter) + p
                tile_counter += 1

        xml = to_xml(new_ome)

        return xml

    def stage_placement(self, experiment_directory, cycle_number):
        '''
        Goal to place images via rough stage coords in a larger image. WIll have nonsense borders
        '''

        quick_tile_path = experiment_directory + r'\Quick_Tile'

        # load in numpy matricies

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]

        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]

        fov_x_pixels = 5056
        fov_y_pixels = 2960
        um_pixel = 0.20

        # generate large numpy image with rand numbers. Big enough to hold all images + 10%

        super_y = int(1.02 * (y_tile_count * fov_y_pixels))
        super_x = int(1.02 * (x_tile_count * fov_x_pixels))
        placed_image = np.random.rand(super_y, super_x).astype('uint16')

        # transform numpy x and y coords into new shifted space that starts at zero and is in units of pixels and not um
        numpy_x_pixels = numpy_x / um_pixel
        numpy_y_pixels = numpy_y / um_pixel

        y_displacement_vector = (2100 / um_pixel + fov_y_pixels / 2) * 1.02
        x_displacement_vector = (-11385 / um_pixel + fov_x_pixels / 2) * 1.02

        numpy_x_pixels = numpy_x_pixels + x_displacement_vector
        numpy_x_pixels = np.ceil(numpy_x_pixels)
        numpy_x_pixels = numpy_x_pixels.astype(int)

        numpy_y_pixels = numpy_y_pixels + y_displacement_vector
        numpy_y_pixels = np.ceil(numpy_y_pixels)
        numpy_y_pixels = numpy_y_pixels.astype(int)

        # load images into python

        channels = ['DAPI', 'A488', 'A555', 'A647']

        for channel in channels:

            im_path = experiment_directory + '/' + channel + '\Stain\cy_' + str(cycle_number) + '\Tiles' + '/focused'
            os.chdir(im_path)

            # place images into large array

            for x in range(0, x_tile_count):
                for y in range(0, y_tile_count):
                    filename = 'x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                    image = io.imread(filename)
                    # define subsection of large array that fits dimensions of single FOV
                    x_center = numpy_x_pixels[y][x]
                    y_center = numpy_y_pixels[y][x]
                    x_start = int(x_center - fov_x_pixels / 2)
                    x_end = int(x_center + fov_x_pixels / 2)
                    y_start = int(y_center - fov_y_pixels / 2)
                    y_end = int(y_center + fov_y_pixels / 2)

                    # placed_image[y_start:y_end, x_start:x_end] = placed_image[y_start:y_end, x_start:x_end] + image
                    placed_image[y_start:y_end, x_start:x_end] = image

            # save output image
            os.chdir(quick_tile_path)
            tf.imwrite(channel + '_cy' + str(cycle_number) + '_placed.tif', placed_image)



    def infocus(self, experiment_directory, cycle_number, x_sub_section_count, y_sub_section_count):

        bin_values = [4]
        channels = ['DAPI', 'A488', 'A555', 'A647']

        dapi_im_path = experiment_directory + '/' + 'DAPI' '\Stain\cy_' + str(cycle_number) + '\Tiles'

        # load numpy arrays in
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]

        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]

        # determine number z slices
        z_checker = 0
        z_slice_count = 0
        os.chdir(dapi_im_path)
        while z_checker == 0:
            file_name = 'z_' + str(z_slice_count) + '_x' + str(0) + '_y_' + str(0) + '_c_DAPI.tif'
            if os.path.isfile(file_name) == 1:
                z_slice_count += 1
            else:
                z_checker = 1

        for channel in channels:
            # generate imstack of z slices for tile
            # channel = 'DAPI'
            im_path = experiment_directory + '/' + channel + '\Stain\cy_' + str(cycle_number) + '\Tiles'
            os.chdir(im_path)

            z_stack = np.random.rand(z_slice_count, 2960, 5056).astype('uint16')
            for x in range(0, x_tile_count):
                for y in range(0, y_tile_count):
                    for z in range(0, z_slice_count):
                        im_path = experiment_directory + '/' + channel + '\Stain\cy_' + str(cycle_number) + '\Tiles'
                        os.chdir(im_path)
                        file_name = 'z_' + str(z) + '_x' + str(x) + '_y_' + str(y) + '_c_' + str(channel) + '.tif'
                        image = tf.imread(file_name)
                        z_stack[z] = image

                        # break into sub sections (2x3)

                    number_bins = len(bin_values)
                    brenner_sub_selector = np.random.rand(z_slice_count, number_bins, y_sub_section_count,
                                                          x_sub_section_count).astype('longlong')
                    for z in range(0, z_slice_count):
                        for y_sub in range(0, y_sub_section_count):
                            for x_sub in range(0, x_sub_section_count):

                                y_end = int((y_sub + 1) * (2960 / y_sub_section_count))
                                y_start = int(y_sub * (2960 / y_sub_section_count))
                                x_end = int((x_sub + 1) * (5056 / x_sub_section_count))
                                x_start = int(x_sub * (5056 / x_sub_section_count))
                                sub_image = z_stack[z][y_start:y_end, x_start:x_end]

                                for b in range(0, number_bins):
                                    bin_value = int(bin_values[b])
                                    score = cycif.focus_score(sub_image, bin_value)
                                    brenner_sub_selector[z][b][y_sub][x_sub] = score

                    reconstruct_array = brenner_reconstruct_array(brenner_sub_selector, z_slice_count, number_bins)
                    reconstruct_array = skimage.filters.median(reconstruct_array)
                    image_reconstructor(reconstruct_array, channel, cycle_number, y, x)

    def brenner_reconstruct_array(self, brenner_sub_selector, z_slice_count, number_bins):
        '''
        take in sub selector array and slice to find max values for brenner scores and then find mode for various bin levels
        output array that shows what slice to grab each sub section from
        :param brenner_sub_selector:
        :return:
        '''
        # array to dictate the z slice to grab each sub section from
        y_sections = np.shape(brenner_sub_selector)[2]
        x_sections = np.shape(brenner_sub_selector)[3]

        reconstruct_array = np.random.rand(y_sections, x_sections).astype('uint16')

        temp_bin_max_indicies = np.random.rand(number_bins).astype('uint16')

        for y in range(0, y_sections):
            for x in range(0, x_sections):
                for b in range(0, number_bins):
                    sub_scores = brenner_sub_selector[0:z_slice_count, b, y, x]
                    max_score = np.max(sub_scores)
                    max_index = np.where(sub_scores == max_score)[0][0]
                    temp_bin_max_indicies[b] = max_index
                sub_section_index_mode = stats.mode(temp_bin_max_indicies)[0][0]
                reconstruct_array[y][x] = sub_section_index_mode

        return reconstruct_array

    def image_reconstructor(self, reconstruct_array, channel, cycle_number, y_tile_number, x_tile_number):

        y_sections = np.shape(reconstruct_array)[0]
        x_sections = np.shape(reconstruct_array)[1]

        cycle_types = ['Stain', 'Bleach']

        for cycle_type in cycle_types:

            im_path = experiment_directory + '/' + channel + '/' + cycle_type + '\cy_' + str(cycle_number) + '\Tiles'
            os.chdir(im_path)
            try:
                os.mkdir('focused')
            except:
                t = 5

            # rebuilt image container
            rebuilt_image = np.random.rand(2960, 5056).astype('uint16')

            for y in range(0, y_sections):
                for x in range(0, x_sections):
                    # define y and x start and ends subsection of rebuilt image
                    y_end = int((y + 1) * (2960 / y_sections))
                    y_start = int(y * (2960 / y_sections))
                    x_end = int((x + 1) * (5056 / x_sections))
                    x_start = int(x * (5056 / x_sections))

                    # find z for specific subsection
                    z_slice = reconstruct_array[y][x]
                    # load in image to extract for subsection
                    file_name = 'z_' + str(z_slice) + '_x' + str(x_tile_number) + '_y_' + str(
                        y_tile_number) + '_c_' + str(channel) + '.tif'
                    image = tf.imread(file_name)
                    rebuilt_image[y_start:y_end, x_start:x_end] = image[y_start:y_end, x_start:x_end]

            reconstruct_path = experiment_directory + '/' + channel + '/' + cycle_type + '\cy_' + str(
                cycle_number) + '\Tiles' + '/focused'
            os.chdir(reconstruct_path)
            filename = 'x' + str(x_tile_number) + '_y_' + str(y_tile_number) + '_c_' + str(channel) + '.tif'
            tf.imwrite(filename, rebuilt_image)

    def save_files(self, z_tile_stack, channel, cycle, experiment_directory, Stain_or_Bleach = 'Stain'):

        save_directory = experiment_directory + '/' + str(channel) + '/' + Stain_or_Bleach + '/' + 'cy_' + str(cycle) + '/' + 'Tiles'

        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]

        x_tile_count = np.unique(numpy_x).size
        y_tile_count = np.unique(numpy_y).size

        z_tile_count = z_tile_stack.shape[0]


        for z in range(0, z_tile_count):

            tile_counter = 0

            for y in range(0, y_tile_count):
                if y % 2 != 0:
                    for x in range(x_tile_count - 1, -1, -1):

                        #meta = self.image_metadata_generation(x, y, channel, experiment_directory)
                        file_name = 'z_' + str(z) + '_x' + str(x) + '_y_' + str(y) + '_c_' + str(channel)+ '.tif'
                        image = z_tile_stack[z][tile_counter]
                        os.chdir(save_directory)
                        imwrite(file_name, image, photometric='minisblack')
                        tile_counter += 1

                if y % 2 == 0:
                    for x in range(0, x_tile_count):

                        #meta = self.image_metadata_generation(x, y, channel, experiment_directory)
                        file_name = 'z_' + str(z) + '_x' + str(x) + '_y_' + str(y) + '_c_' + str(channel)+ '.tif'
                        image = z_tile_stack[z][tile_counter]
                        os.chdir(save_directory)
                        imwrite(file_name, image, photometric='minisblack')
                        tile_counter += 1

    def save_tif_stack(self, tif_stack, cycle_number,  directory_name):

        add_on_folder = 'cycle_' + str(cycle_number)
        full_directory_path = directory_name + add_on_folder
        try:
            os.mkdir(full_directory_path)
        except:
           pass
        os.chdir(full_directory_path)
        file_name = 'image_array.tif'

        imwrite(file_name, tif_stack,
                bigtiff=True,
                photometric='minisblack',
                compression = 'zlib',
                compressionargs = {'level': 8} )


############################################
#depreciated or unused
############################################



############################################
##Control arduino based microfluidic system
############################################

class arduino:

    def __init__(self):
        return

    #def mqtt_publish(self, message, subtopic, topic="control", client=client):
        '''
        takes message and publishes message to server defined by client and under topic of topic/subtopic

        :param str subtopic: second tier of topic heirarchy
        :param str topic: first tier of topic heirarchy
        :param object client: client that MQTT server is on. Established in top of module
        :return:
        '''

        #full_topic = topic + "/" + subtopic

        #client.loop_start()
       # client.publish(full_topic, message)
        #client.loop_stop()

    def heater_state(self, state):

        if state == 'on':
            self.mqtt_publish(210, 'dc_pump')
        elif state == 'off':
            self.mqtt_publish(200, 'dc_pump')

    def chamber(self, fill_drain, run_time=27):
        '''
        Aquarium pumps to fill or drain outer chamber with water. Uses dispense function as backbone.

        :param str fill_drain: fill = fills chamber, drain = drains chamber
        :param int time: time in secs to fill chamber and conversely drain it
        :return: nothing
        '''

        if fill_drain == 'drain':
            self.mqtt_publish(110, 'dc_pump')
            time.sleep(run_time + 3)
            self.mqtt_publish(100, 'dc_pump')

        elif fill_drain == 'fill':
            self.mqtt_publish(111, 'dc_pump')
            time.sleep(run_time)
            self.mqtt_publish(101, 'dc_pump')


class fluidics:

    def __init__(self, mux_com_port, ob1_com_port):

        # OB1 initialize
        ob1_path = 'ASRL' + str(ob1_com_port) + '::INSTR'
        Instr_ID = c_int32()
        pump = OB1_Initialization(ob1_path.encode('ascii'), 0, 0, 0, 0, byref(Instr_ID))
        pump = OB1_Add_Sens(Instr_ID, 1, 5, 1, 0, 7, 0) #16bit working range between 0-1000uL/min, also what are CustomSens_Voltage_5_to_25 and can I really choose any digital range?

        Calib = (c_double * 1000)()
        Elveflow_Calibration_Default(byref(Calib), 1000)
        OB1_Start_Remote_Measurement(Instr_ID.value, byref(Calib), 1000)
        self.calibration_array = byref(Calib)

        set_channel_regulator = int(1)  # convert to int
        set_channel_regulator = c_int32(set_channel_regulator)  # convert to c_int32
        set_channel_sensor = int(1)
        set_channel_sensor = c_int32(set_channel_sensor)  # convert to c_int32
        PID_Add_Remote(Instr_ID.value, set_channel_regulator, Instr_ID.value, set_channel_sensor, .9, 0.002, 1)


        # MUX intiialize
        path = 'ASRL' + str(mux_com_port) + '::INSTR'
        mux_Instr_ID = c_int32()
        MUX_DRI_Initialization(path.encode('ascii'), byref(mux_Instr_ID))  # choose the COM port, it can be ASRLXXX::INSTR (where XXX=port number)

        #home
        #answer = (c_char * 40)()
        self.mux_ID = mux_Instr_ID.value
        #MUX_DRI_Send_Command(self.mux_ID, 0, answer, 40)

        self.pump_ID = Instr_ID.value

        return

    def mux_end(self):

        MUX_DRI_Destructor(self.mux_ID)

    def valve_select(self, valve_number):
        '''
        Selects valve in mux unit with associated mux_id to the valve_number declared.
        :param c_int32 mux_id: mux_id given from mux_initialization method
        :param int valve_number: number of desired valve to be selected
        :return: Nothing
        '''

        valve_number = c_int32(valve_number)
        MUX_DRI_Set_Valve(self.mux_ID, valve_number, 0) #0 is shortest path. clockwise and cc are also options
        time.sleep(1)

    def flow(self, flow_rate):

        set_channel=int(1)#convert to int
        set_channel=c_int32(set_channel)#convert to c_int32

        set_target=float(flow_rate) # in uL/min for flow
        set_target=c_double(set_target)#convert to c_double

        OB1_Set_Remote_Target(self.pump_ID, set_channel, set_target)
        #OB1_Stop_Remote_Measurement(self.pump_ID)

    def ob1_end(self):

        set_channel = int(1)  # convert to int
        set_channel = c_int32(set_channel)  # convert to c_int32

        data_sens=c_double()
        data_reg=c_double()

        x = 0
        self.flow(0)

        while x == 0:
            OB1_Get_Remote_Data(self.pump_ID, set_channel, byref(data_reg), byref(data_sens))
            flow_rate = data_sens.value

            if flow_rate < 10:
                x = 1
            if flow_rate > 10:
                x = 0
            time.sleep(1)

        OB1_Stop_Remote_Measurement(self.pump_ID)
        OB1_Destructor(self.pump_ID)

    def measure(self):

        set_channel = int(1)  # convert to int
        set_channel = c_int32(set_channel)  # convert to c_int32

        data_sens=c_double()
        data_reg=c_double()

        OB1_Get_Remote_Data(self.pump_ID, set_channel, byref(data_reg),byref(data_sens) )

        pressure = data_reg.value
        flow_rate = data_sens.value
        time_stamp = time.time()

        return pressure, flow_rate, time_stamp

    def flow_recorder(self, time_step, total_time, file_name = 'none', plot = 1):

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = 'Time'
        ws.cell(row=1, column=2).value = 'Flow Rate'
        ws.cell(row=1, column=3).value = 'Pressure'

        total_steps = int(total_time/time_step)
        pressure_points = np.random.rand(total_steps).astype('float16')
        time_points = np.random.rand(total_steps).astype('float16')
        flow_points = np.random.rand(total_steps).astype('float16')

        for t in range(0, total_steps):

            pressure_point, flow_point, time_point = self.measure()
            pressure_points[t] = pressure_point
            time_points[t] = t * time_step
            flow_points[t] = flow_point

            ws.cell(row= t + 2, column=1).value = t * time_step
            ws.cell(row= t + 2, column=2).value = flow_point
            ws.cell(row= t + 2, column=3).value = pressure_point

            time.sleep(time_step)

        #wb.save(filename = file_name)

        if plot == 1:
            plt.plot(time_points, flow_points, 'o', color='black')
            plt.show()


    def liquid_action(self, action_type, stain_valve = 0, heater_state = 0):

        bleach_valve = 1
        pbs_valve = 8
        bleach_time = 3 #minutes
        stain_flow_time = 45 #seconds
        if heater_state == 0:
            stain_inc_time = 45 #minutes
        if heater_state == 1:
            stain_inc_time = 45  #minutes
        nuc_valve = 7
        nuc_flow_time = 60 #seconds
        nuc_inc_time = 3 #minutes

        if action_type == 'Bleach':

            self.valve_select(bleach_valve)
            self.flow(500)
            time.sleep(60)
            self.flow(0)
            time.sleep(bleach_time*60)

            self.valve_select(pbs_valve)
            self.flow(500)
            time.sleep(60)
            self.flow(0)

        elif action_type == 'Stain':

            if heater_state == 1:
                arduino.heater_state(1)
                arduino.chamber('drain')
            else:
                pass

            self.valve_select(stain_valve)
            self.flow(500)
            time.sleep(stain_flow_time)
            self.flow(0)
            time.sleep(stain_inc_time*60)

            if heater_state == 1:
                arduino.heater_state(0)
                arduino.chamber('fill')
            else:
                pass

            self.valve_select(pbs_valve)
            self.flow(500)
            time.sleep(60)
            self.flow(0)

        elif action_type == "Wash":

            self.valve_select(pbs_valve)
            self.flow(500)
            time.sleep(60)
            self.flow(0)

        elif action_type == 'Nuc_Touchup':

            self.valve_select(nuc_valve)
            self.flow(500)
            time.sleep(nuc_flow_time)
            self.flow(0)
            time.sleep(nuc_inc_time*60)

            self.valve_select(pbs_valve)
            self.flow(500)
            time.sleep(60)
            self.flow(0)











    
    




