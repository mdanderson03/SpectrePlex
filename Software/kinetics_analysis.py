import copy
from skimage import io, filters, morphology, restoration, util, transform
import numpy as np
import matplotlib.pyplot as plt
from skimage.filters import threshold_otsu, butterworth, median
from skimage import io, filters, morphology
import imageio
from skimage.transform import resize
import math
import time
import os
from openpyxl import Workbook
from datetime import datetime
from csbdeep.utils import normalize
from stardist.models import StarDist2D
import cv2


def tissue_filter(image):
    image = image.astype('bool')
    image_2 = morphology.remove_small_objects(image, min_size=80000, connectivity=1)
    image_2 = image_2.astype('int8')

    return image_2
def tissue_binary_generate(experiment_directory):
    '''
    Generates tissue binary maps from star dist binary maps

    :param experiment_directory:
    :return:
    '''


    star_dist_path = experiment_directory + '/Labelled_Nuc'
    tissue_path = experiment_directory

    foot_print = morphology.disk(70, decomposition='sequence')

    os.chdir(star_dist_path)
    star_dist_filename = 'labelled_DAPI.tif'
    star_dist_im = io.imread(star_dist_filename)

    tissue_binary_im = morphology.binary_dilation(star_dist_im, foot_print)
    tissue_binary_im = tissue_binary_im.astype(np.uint8)
    filtered_image = tissue_filter(tissue_binary_im)
    os.chdir(tissue_path)
    tissue_binary_name = 'tissue_mask.tif'
    io.imsave(tissue_binary_name, filtered_image)
def generate_nuc_mask(experiment_directory):

    model = StarDist2D.from_pretrained('2D_versatile_fluo')

    dapi_im_path = experiment_directory + '/auto_fluorescence'
    os.chdir(experiment_directory)
    try:
        os.mkdir('Labelled_Nuc')
    except:
        pass

    labelled_path = experiment_directory + '/Labelled_Nuc'

    os.chdir(dapi_im_path)
    file_name = 'DAPI.tif'
    labelled_file_name = 'labelled_DAPI.tif'
    img = io.imread(file_name)
    labels, _ = model.predict_instances(normalize(img))
    labels[labels > 0] = 1

    os.chdir(labelled_path)
    io.imsave(labelled_file_name, labels)


def kinetics_df_data(experiment_directory, channel):

    autof_folder_path = experiment_directory  + r'\\auto_fluorescence'
    os.chdir(autof_folder_path)
    auto_fluor = io.imread(channel + '.tif').astype('float32')[::, 1048:4008]



    #generate tissue mask

    generate_nuc_mask(experiment_directory)
    tissue_binary_generate(experiment_directory)

    #stain mask image
    post_wash_path = experiment_directory + '\post_wash'
    os.chdir(post_wash_path)
    post_wash_filename = channel + '.tif'
    image = io.imread(post_wash_filename)[::, 1048:4008]
    image = image - auto_fluor


    thresh = threshold_otsu(image)
    stain_image = image > thresh
    stain_image = stain_image/1
    os.chdir(experiment_directory)

    stain_mask = io.imread(channel + '_stain_mask.tif')[::, 1048:4008]

    #image_save_name = channel + '_stain_mask.tif'
    #io.imsave(image_save_name, stain_mask)

    #non antigen mask

    os.chdir(experiment_directory)
    tissue_mask = io.imread('tissue_mask.tif')[::, 1048:4008]



    non_antigen_mask = copy.deepcopy(stain_mask)
    non_antigen_mask += 1
    non_antigen_mask[non_antigen_mask == 2] = 0
    #non_antigen_image = image < thresh
    #non_antigen_image = non_antigen_image * tissue_mask
    non_antigen_mask = non_antigen_mask * tissue_mask
    non_antigen_save_name = channel + '_nonantigen_tissue_mask.tif'
    io.imsave(non_antigen_save_name, non_antigen_mask)


    #non tissue mask
    slide_mask = tissue_mask
    slide_mask = slide_mask + 1
    slide_mask = slide_mask < 2
    slide_mask = slide_mask/1
    slide_mask_save_name = 'slide_mask.tif'
    io.imsave(slide_mask_save_name, slide_mask)

    #trimming are to not be considered
    trim_window = np.ones((2960, 2960))
    trim_window[0:1000, ::] = 0


    #import files

    stack_folder_path = experiment_directory + r'\\' + channel

    os.chdir(stack_folder_path)
    f_channel = io.imread(channel + '_stain_stack.tif')[::,::, 1048:4008] * trim_window
    os.chdir(experiment_directory)
    tissue_mask = io.imread('tissue_mask.tif')[::, 1048:4008] *trim_window
    nonantigen_mask = io.imread(channel + '_nonantigen_tissue_mask.tif') *trim_window
    nontissue_mask = io.imread('slide_mask.tif') * trim_window
    stain_mask = io.imread(channel + '_stain_mask.tif')[::, 1048:4008] * trim_window

    os.chdir(experiment_directory)

    wb = Workbook()
    ws = wb.active

    #name columns

    ws.cell(row=1, column=1).value = 'Time (minutes)'
    ws.cell(row = 1, column = 2).value = 'Signal'
    ws.cell(row = 1, column = 3).value = 'Non Specific Signal'
    ws.cell(row=1, column=4).value = 'Signal/Non Specific'
    ws.cell(row=1, column=5).value = 'Signal_Avg/Std'

    #make x-axis
    x_axis = np.linspace(0, 90, num = 45)

    #make some type of stack

    stain_used_mask = stain_mask

    f_channel = f_channel - auto_fluor
    f_channel[f_channel < 0] = 0
    signal = f_channel * stain_used_mask
    non_specific_signal = f_channel * nonantigen_mask
    #slide_signal = f_channel * nontissue_mask
    tissue_signal = f_channel[0] * stain_used_mask
    tissue_nonantigen = f_channel[0] * nonantigen_mask

    #make y-axis
    y_axis = np.linspace(0, 45, num = 45)
    slide_y_axis = np.linspace(0, 45, num = 45)
    slide_nonantigen_y_axis = np.linspace(0, 45, num=45)
    non_specific_y_axis = np.linspace(0, 45, num = 45)

    #find pixels in masks
    pixels_in_mask = stain_used_mask.sum(dtype=np.int64)
    slide_pixels = nontissue_mask.sum(dtype=np.int64)
    non_specific_pixels = nonantigen_mask.sum(dtype=np.int64)
    tissue_pixels = tissue_mask.sum(dtype=np.int64)

    #save subbed stack
    df_subbed_stack = np.zeros((45, 2960,2960))

    for x in range(0, 45):

        ws.cell(row=2 + x, column=1).value = x*2
        slide_signal = f_channel[x]
        df = dark_frame_generate(slide_signal, 75, 75)
        slide_signal = slide_signal - df
        slide_signal[slide_signal < 0] = 0

        df_subbed_stack[x] = df

        slide_signal = slide_signal * stain_used_mask



        #slide_y_axis[x] = slide_signal[x].sum(dtype=np.float64) / slide_signal[0].sum(dtype=np.float64) * tissue_signal.sum(dtype=np.float64)/pixels_in_mask
        slide_y_axis[x] = tissue_signal.sum(dtype=np.float64) / pixels_in_mask
        slide_nonantigen_y_axis[x] = slide_signal[x].sum(dtype=np.float64) / slide_signal[0].sum(dtype=np.float64) * tissue_nonantigen.sum(dtype=np.float64) / non_specific_pixels

        non_specific_y_axis[x] = non_specific_signal[x].sum(dtype=np.float64) / non_specific_pixels
        non_specific_y_axis[x] = non_specific_y_axis[x] - slide_nonantigen_y_axis[x] + 1

        ws.cell(row=2 + x, column=3).value = non_specific_y_axis[x]

        y_axis[x] = signal[x].sum(dtype=np.float64) / pixels_in_mask
        y_axis[x] = (y_axis[x] - slide_y_axis[x])

        ws.cell(row=2 + x, column=2).value = y_axis[x]
        ws.cell(row=2 + x, column=4).value = y_axis[x]/non_specific_y_axis[x]

        #y_axis[x] = (y_axis[x]/np.std(signal[x]))**2

        ws.cell(row=2 + x, column=5).value = y_axis[x]/np.std(signal[x])
        #y_axis[x] = slide_y_axis[x]
        #y_axis[x] = non_specific_y_axis[x]
        y_axis[x] = y_axis[x]/non_specific_y_axis[x]
        #y_axis[x] = non_specific_y_axis[x]/slide_nonantigen_y_axis[x]
        print(str(x))

    filename = channel + '.xlsx'
    wb.save(filename)
    os.chdir(stack_folder_path)
    io.imsave('df_subbed_stack.tif', df_subbed_stack)
    return y_axis, x_axis
def kinetics_data(experiment_directory, channel):

    autof_folder_path = experiment_directory  + r'\\auto_fluorescence'
    os.chdir(autof_folder_path)
    auto_fluor = io.imread(channel + '.tif').astype('float32')



    #generate tissue mask

    generate_nuc_mask(experiment_directory)
    tissue_binary_generate(experiment_directory)

    #stain mask image
    post_wash_path = experiment_directory + '\post_wash'
    os.chdir(post_wash_path)
    post_wash_filename = channel + '.tif'
    image = io.imread(post_wash_filename)
    image = image - auto_fluor


    thresh = threshold_otsu(image)
    stain_image = image > thresh
    stain_image = stain_image/1
    os.chdir(experiment_directory)

    stain_mask = io.imread(channel + '_stain_mask.tif')

    #image_save_name = channel + '_stain_mask.tif'
    #io.imsave(image_save_name, stain_mask)

    #non antigen mask

    os.chdir(experiment_directory)
    tissue_mask = io.imread('tissue_mask.tif')



    non_antigen_mask = copy.deepcopy(stain_mask)
    non_antigen_mask += 1
    non_antigen_mask[non_antigen_mask == 2] = 0
    #non_antigen_image = image < thresh
    #non_antigen_image = non_antigen_image * tissue_mask
    non_antigen_mask = non_antigen_mask * tissue_mask
    non_antigen_save_name = channel + '_nonantigen_tissue_mask.tif'
    io.imsave(non_antigen_save_name, non_antigen_mask)


    #non tissue mask
    slide_mask = tissue_mask
    slide_mask = slide_mask + 1
    slide_mask = slide_mask < 2
    slide_mask = slide_mask/1
    slide_mask_save_name = 'slide_mask.tif'
    io.imsave(slide_mask_save_name, slide_mask)


    #import files

    stack_folder_path = experiment_directory + r'\\' + channel

    os.chdir(stack_folder_path)
    f_channel = io.imread(channel + '_stain_stack.tif')
    os.chdir(experiment_directory)
    tissue_mask = io.imread('tissue_mask.tif')
    nonantigen_mask = io.imread(channel + '_nonantigen_tissue_mask.tif')
    nontissue_mask = io.imread('slide_mask.tif')
    stain_mask = io.imread(channel + '_stain_mask.tif')

    os.chdir(experiment_directory)

    wb = Workbook()
    ws = wb.active

    #name columns

    ws.cell(row=1, column=1).value = 'Time (minutes)'
    ws.cell(row = 1, column = 2).value = 'Signal'
    ws.cell(row = 1, column = 3).value = 'Non Specific Signal'
    ws.cell(row=1, column=4).value = 'Signal/Non Specific'
    ws.cell(row=1, column=5).value = 'Signal_Avg/Std'

    #make x-axis
    x_axis = np.linspace(0, 90, num = 45)

    #trimming are to not be considered
    trim_window = np.ones((2960, 5056))
    trim_window[0:1000, ::] = 0

    #make some type of stack

    stain_used_mask = stain_mask

    f_channel = f_channel - auto_fluor
    f_channel[f_channel < 0] = 0
    signal = f_channel * stain_used_mask * trim_window
    non_specific_signal = f_channel * nonantigen_mask * trim_window
    slide_signal = f_channel * nontissue_mask * trim_window
    tissue_signal = f_channel[0] * stain_used_mask * trim_window
    tissue_nonantigen = f_channel[0] * nonantigen_mask * trim_window

    #make y-axis
    y_axis = np.linspace(0, 45, num = 45)
    slide_y_axis = np.linspace(0, 45, num = 45)
    slide_nonantigen_y_axis = np.linspace(0, 45, num=45)
    non_specific_y_axis = np.linspace(0, 45, num = 45)

    #find pixels in masks
    pixels_in_mask = stain_used_mask.sum(dtype=np.int64)
    slide_pixels = nontissue_mask.sum(dtype=np.int64)
    non_specific_pixels = nonantigen_mask.sum(dtype=np.int64)
    tissue_pixels = tissue_mask.sum(dtype=np.int64)

    for x in range(0, 45):

        ws.cell(row=2 + x, column=1).value = x*2


        slide_y_axis[x] = slide_signal[x].sum(dtype=np.float64) / slide_signal[0].sum(dtype=np.float64) * tissue_signal.sum(dtype=np.float64)/pixels_in_mask

        slide_nonantigen_y_axis[x] = slide_signal[x].sum(dtype=np.float64) / slide_signal[0].sum(dtype=np.float64) * tissue_nonantigen.sum(dtype=np.float64) / non_specific_pixels

        non_specific_y_axis[x] = non_specific_signal[x].sum(dtype=np.float64) / non_specific_pixels
        non_specific_y_axis[x] = non_specific_y_axis[x] - slide_nonantigen_y_axis[x] + 1

        ws.cell(row=2 + x, column=3).value = non_specific_y_axis[x]

        y_axis[x] = signal[x].sum(dtype=np.float64) / pixels_in_mask
        y_axis[x] = (y_axis[x] - slide_y_axis[x])

        ws.cell(row=2 + x, column=2).value = y_axis[x]
        ws.cell(row=2 + x, column=4).value = y_axis[x]/non_specific_y_axis[x]

        #y_axis[x] = (y_axis[x]/np.std(signal[x]))**2

        ws.cell(row=2 + x, column=5).value = y_axis[x]/np.std(signal[x])
        #y_axis[x] = slide_y_axis[x]
        #y_axis[x] = non_specific_y_axis[x]
        y_axis[x] = y_axis[x]/non_specific_y_axis[x]
        #y_axis[x] = non_specific_y_axis[x]/slide_nonantigen_y_axis[x]
        print(str(x))

    filename = channel + '.xlsx'
    wb.save(filename)
    return y_axis, x_axis


def block_proc_min(array, block_y_pixels, block_x_pixels):
    '''
    breaks image up into blocks of size (block_y_pixels x block_x_pixels
    and then returns array that is the minimum of each block and is
    effectively down sampled by the block size

    :param array:
    :param block_y_pixels:
    :param block_x_pixles:
    :return:
    '''

    y_pixels = np.shape(array)[0]
    x_pixels = np.shape(array)[1]

    y_num_blocks = int(y_pixels / block_y_pixels)
    x_num_blocks = int(x_pixels / block_x_pixels)

    im = array

    blocked_array = np.lib.stride_tricks.as_strided(im, shape=(x_num_blocks, y_num_blocks, block_x_pixels, block_y_pixels), strides=(im.strides[0] * block_x_pixels, im.strides[1] * block_y_pixels, im.strides[0], im.strides[1]))
    min_im = np.min(blocked_array, axis=2)
    min_im = np.min(min_im, axis=2)

    return min_im

def block_proc_reshaper(array, block_y_pixels, block_x_pixels):
    '''
    takes in array and evenly clips off rows and columns to to their counts be integers
    when divided by block size pixel counts. Does no padding, only shrinks.
    :param array:
    :param block_y_pixels:
    :param block_x_pixels:
    :return:
    '''

    y_pixels = np.shape(array)[0]
    x_pixels = np.shape(array)[1]

    y_num_blocks = math.floor(y_pixels / block_y_pixels)
    x_num_blocks = math.floor(x_pixels / block_x_pixels)

    adjusted_y_pixels = block_y_pixels * y_num_blocks
    adjusted_x_pixels = block_x_pixels * x_num_blocks

    clipped_y_pixel_num = y_pixels - adjusted_y_pixels
    clipped_x_pixel_num = x_pixels - adjusted_x_pixels

    # deal with even and odd cases for y axis
    if clipped_y_pixel_num % 2 == 0:
        top_rows_2_clip = int(clipped_y_pixel_num / 2)
        bottom_rows_2_clip = int(clipped_y_pixel_num / 2)
        bottom_adjusted_row_num = y_pixels - bottom_rows_2_clip
        array = array[top_rows_2_clip:bottom_adjusted_row_num, ::]
    elif clipped_y_pixel_num % 2 != 0:
        top_rows_2_clip = int(clipped_y_pixel_num / 2)
        bottom_rows_2_clip = int((clipped_y_pixel_num / 2) + 1)
        bottom_adjusted_row_num = y_pixels - bottom_rows_2_clip
        array = array[top_rows_2_clip:bottom_adjusted_row_num, ::]

    # deal with even and odd cases for x axis
    if clipped_x_pixel_num % 2 == 0:
        left_col_2_clip = int(clipped_x_pixel_num / 2)
        right_col_2_clip = int(clipped_x_pixel_num / 2)
        right_adjusted_col_num = int(x_pixels - right_col_2_clip)
        array = array[::, left_col_2_clip:right_adjusted_col_num]
    elif clipped_x_pixel_num % 2 != 0:
        left_col_2_clip = int(clipped_x_pixel_num / 2)
        right_col_2_clip = int((clipped_x_pixel_num / 2) + 1)
        right_adjusted_col_num = int(x_pixels - right_col_2_clip)
        array = array[::, left_col_2_clip:right_adjusted_col_num]

    return array
def dark_frame_generate(array, block_y_pixels, block_x_pixels):
    '''

    :param array:
    :param block_y_pixels:
    :param block_x_pixels:
    :return:
    '''


    original_x_pixels = np.shape(array)[1]
    original_y_pixels = np.shape(array)[0]

    mean_kernal_y = math.ceil(block_y_pixels/10)
    mean_kernal_x = math.ceil(block_x_pixels / 10)

    array_mean = cv2.blur(array, (mean_kernal_y, mean_kernal_x))

    adjusted_array = block_proc_reshaper(array_mean, block_y_pixels, block_x_pixels)
    min_array = block_proc_min(adjusted_array, block_y_pixels, block_x_pixels)
    resized_min_array = transform.resize(min_array, (original_y_pixels, original_x_pixels), preserve_range=True,anti_aliasing=True)
    resized_min_array = filters.butterworth(resized_min_array, cutoff_frequency_ratio=0.005, high_pass=False, order=2, npad=1000)

    return resized_min_array



experiment_directory = r'E:\18-12-24_kinetics_20C_cycle5'
'''
os.chdir(experiment_directory +'\A555')
im = io.imread('a555_stain_stack.tif')
im = im[::,::, 1048:4008]
for x in range(0, 45):
    image = im[x]
    df = dark_frame_generate(image, 75, 75)
    image = image - df
    image[image<0] = 0
    im[x] = image
    print('slice '+str(x)+' done')
io.imsave('subbed.tif', im)
'''


#y_axis_5, x_axis = kinetics_data(experiment_directory, 'A488')
#y_axis_5, x_axis = kinetics_data(experiment_directory, 'A555')
#y_axis_5, x_axis = kinetics_data(experiment_directory, 'A647')
#y_axis_7, x_axis = kinetics_data(experiment_directory, 7, 'A488')

y_axis_5, x_axis = kinetics_df_data(experiment_directory, 'A488')

#display_start = 2

#plt.scatter(x_axis[display_start:45], y_axis_5[display_start:45])
#plt.scatter(x_axis[display_start:45], y_axis_7[display_start:45])
#plt.show()
