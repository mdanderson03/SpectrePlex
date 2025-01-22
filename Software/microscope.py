import copy

import ome_types
from pycromanager import Core, Magellan, Studio
import numpy as np
import time
from skimage import io, filters, morphology, restoration, util, transform
import skimage
import os
import math
from tifffile import imwrite
import tifffile as tf
from openpyxl import load_workbook, Workbook
from skspatial.objects import Plane, Points
from sklearn.linear_model import HuberRegressor
from ome_types import from_xml, OME, to_xml
from copy import deepcopy
from pystackreg import StackReg
from pybasic import shading_correction
from path import Path
from csbdeep.utils import normalize
from stardist.models import StarDist2D
from matplotlib import pyplot as plt
import cv2
from pywt import wavedecn, waverecn
from scipy.ndimage import gaussian_filter
from joblib import Parallel, delayed
import multiprocessing
import itertools
import shutil


magellan = Magellan()
core = Core()


class cycif:

    def __init__(self):

        self.hdr_exp_times = np.array([])
        self.hdr = 0
        return

    ####################################################################
    ############ focus scoring
    ####################################################################

    def focus_score(self, image, derivative_jump, labels):
        '''
        Calculates focus score on image with Brenners algorithm on downsampled image.


        :param numpy image: single image from hooked from acquistion

        :return: focus score for image
        :rtype: float
        '''
        # Note: Uniform background is a bit mandatory

        # do Brenner score


        a = image[derivative_jump:, :]
        a = a.astype('float64')
        posinf_nan = np.median(a)
        a = np.nan_to_num(a, posinf=posinf_nan, nan=posinf_nan)
        b = image[:-derivative_jump, :]
        b = b.astype('float64')
        b = np.nan_to_num(b, posinf=posinf_nan, nan=posinf_nan)
        a = np.log(a)
        b = np.log(b)
        c = a-b
        c = c**2
        #c = (a - b)/((a+b)/2)
        #c = c / 1000 * c / 1000
        labels = labels[derivative_jump:, :]
        c = c * labels
        c = np.nan_to_num(c, posinf=0)
        f_score_shadow = c.sum(dtype=np.float64)

        return f_score_shadow

    def focus_score_post_processing(self, image, derivative_jump):
        '''
        Calculates focus score on image with Brenners algorithm on downsampled image.


        :param numpy image: single image from hooked from acquistion

        :return: focus score for image
        :rtype: float
        '''
        # Note: Uniform background is a bit mandatory

        # do Brenner score

        a = image[derivative_jump:, :]
        a = a.astype('float64')
        posinf_nan = np.median(a)
        a = np.nan_to_num(a, posinf=posinf_nan, nan=posinf_nan)
        b = image[:-derivative_jump, :]
        b = b.astype('float64')
        b = np.nan_to_num(b, posinf=posinf_nan, nan=posinf_nan)
        a = np.log(a)
        b = np.log(b)
        c = a - b
        c = c ** 2
        f_score_shadow = c.sum(dtype=np.float64) + 0.00001

        return f_score_shadow

    #########################################################
    # Setup fm_array and exp_array
    #########################################################

    def establish_fm_array(self, experiment_directory, desired_cycle_count, z_slices, off_array, initialize=0,
                           x_frame_size=5056, fm_array_adjuster = 0, autofocus=0, auto_expose=0, focus_position = 'none'):

        self.auto_f = autofocus
        self.file_structure(experiment_directory, desired_cycle_count)

        if initialize == 1:
            xy_points = self.tile_xy_pos('New Surface 1')
            xyz_points = self.nonfocus_tile_DAPI(xy_points, experiment_directory, focus_position=focus_position)
            self.tile_pattern(xyz_points, experiment_directory)
            self.fm_channel_initial(experiment_directory, off_array, z_slices)
            self.establish_exp_arrays(experiment_directory)
            self.hdr_exp_generator(experiment_directory, threshold_level=10000, max_exp=700, min_exp=20)
            self.establish_exp_arrays(experiment_directory)

            if x_frame_size != 5056:
                self.x_overlap_adjuster(x_frame_size, experiment_directory)
                #self.fm_stage_tilt_compensation(experiment_directory, tilt=3.75) #always positive number for tilt
                self.establish_exp_arrays(experiment_directory)
            else:
                pass

        else:
            pass

        if autofocus == 1 and auto_expose == 1:
            if desired_cycle_count == 0:
                self.recursive_stardist_autofocus(experiment_directory, desired_cycle_count)
            else:
                pass
            self.establish_exp_arrays(experiment_directory)
            self.auto_exposure(experiment_directory, x_frame_size, percentage_cut_off = 0.99, target_percentage = 0.25)
        if autofocus == 1 and auto_expose == 0:
            self.recursive_stardist_autofocus(experiment_directory, desired_cycle_count)
            self.hdr = 0

            #self.fm_channel_initial(experiment_directory, off_array, z_slices, 2)
        if autofocus == 0 and auto_expose == 0:
            self.hdr = 0

            #self.fm_channel_initial(experiment_directory, off_array, z_slices, 2)
        if autofocus == 0 and auto_expose == 1:
            self.establish_exp_arrays(experiment_directory)
            self.auto_exposure(experiment_directory, x_frame_size, percentage_cut_off = 0.99, target_percentage = 0.25)
        if autofocus == 0 and auto_expose == 2:
            self.establish_exp_arrays(experiment_directory)
            self.exp_predetermined(experiment_directory, desired_cycle_count)
        if autofocus == 1 and auto_expose == 2:
            if desired_cycle_count == 0:
                self.recursive_stardist_autofocus(experiment_directory, desired_cycle_count)
            else:
                pass
                self.establish_exp_arrays(experiment_directory)
                self.exp_predetermined(experiment_directory, desired_cycle_count)
        if autofocus == 0 and auto_expose == 3:
            self.hdr = 1
            self.hdr_exp_generator(experiment_directory, threshold_level=10000, max_exp=700, min_exp=20)
        if autofocus == 1 and auto_expose == 3:
            self.recursive_stardist_autofocus(experiment_directory, desired_cycle_count)
            self.hdr = 1
            self.hdr_exp_generator(experiment_directory, threshold_level=10000, max_exp=700, min_exp=20)
        else:
            pass

        if fm_array_adjuster == 1:
            self.fm_grid_readjuster(experiment_directory, x_frame_size)
        else:
            pass

        self.fm_channel_update(experiment_directory, off_array, z_slices)

    def establish_exp_arrays(self, experiment_directory):
        '''
        Make exp_array with default exp times and exp_calc_array to use with auto exposure

        :param experiment_directory:
        :return:
        '''
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        fm_array = np.load('fm_array.npy', allow_pickle=False)

        y_tiles = np.shape(fm_array[0])[0]
        x_tiles = np.shape(fm_array[0])[1]

        exp_calc_array = np.random.rand(4, 3, y_tiles, x_tiles)
        exp_array = [10, 50, 50, 50]
        exp_calc_array[::, 0, ::, ::] = 100

        file_name = 'exp_calc_array.npy'
        np.save(file_name, exp_calc_array)
        np.save('exp_array.npy', exp_array)


    ###########################################################
    # This section is the for the exposure functions.
    ###########################################################
    def auto_exposure(self, experiment_directory, x_frame_size, percentage_cut_off = 0.999, target_percentage = 0.1 ):

        # load in data structures
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        exp_filename = 'exp_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)
        exp_array = np.load(exp_filename, allow_pickle=False)
        channels = ['A488', 'A555', 'A647']
        number_channels = len(channels[0])
        slice_gap = 2

        #reset exp array to default values
        exp_array = [75, 30, 30, 30]

        #find tile counts
        x_tiles = np.shape(fm_array[0])[1]
        y_tiles = np.shape(fm_array[0])[0]
        tissue_fm = fm_array[10]

        #get XY arrays
        numpy_x = fm_array[0]
        numpy_y = fm_array[1]

        #make psuedo stack to hold images in
        #exp_image_stack = np.random.rand(number_channels, y_tiles, x_tiles, 2960, x_frame_size)
        exp_image_stack = np.zeros((number_channels, y_tiles, x_tiles, 2960, x_frame_size))
        exp_tile_logger = np.zeros((number_channels, y_tiles, x_tiles))

        #define pixel range in X dimension
        side_pixel_count = int((5056 - x_frame_size)/2)
        start_frame = side_pixel_count
        end_frame = side_pixel_count + x_frame_size

        start = time.time()

        for x in range(0, x_tiles):
            for y in range(0, y_tiles):
                print('x', x, 'y', y)

                #import tissue binary image
                tissue_path = experiment_directory + '/Tissue_Binary'
                os.chdir(tissue_path)
                tissue_binary_name = 'x' + str(x) + '_y_' + str(y) + '_tissue.tif'
                tissue_bin_im = io.imread(tissue_binary_name)

                if tissue_fm[y][x] > 1:

                    for channel in channels:
                        if channel == 'DAPI':
                            channel_index = 0
                        if channel == 'A488':
                            channel_index = 1
                        if channel == 'A555':
                            channel_index = 2
                        if channel == 'A647':
                            channel_index = 3

                        #exp time to use for snap shot image
                        exp_time = exp_array[channel_index]

                        #Go to XY location (mid stack)
                        core.set_xy_position(numpy_x[y][x], numpy_y[y][x])
                        time.sleep(0.3)

                        #Go to Z location
                        numpy_z = fm_array[channel_index*2 + 2]
                        slice_count = fm_array[channel_index*2 + 3][y][x]
                        z_range = slice_count * slice_gap
                        center_z = int(numpy_z[y][x] - z_range/2)
                        core.set_position(center_z)
                        time.sleep(0.3)

                        #set channel and exp time
                        exp_time = int(exp_array[channel_index])
                        core.set_config("Color", channel)
                        core.set_exposure(exp_time)

                        #take reference image
                        core.set_config("amp", 'high')
                        core.snap_image()
                        tagged_image = core.get_tagged_image()
                        pixels = np.reshape(tagged_image.pix,
                                            newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
                        pixels = pixels[::, side_pixel_count:side_pixel_count + x_frame_size]
                        pixels = pixels
                        pixels = np.nan_to_num(pixels, posinf=65500)

                        #Determine if intensity is in bounds and take diff image if not. Record new exp time
                        pixels, exp_time = self.exp_bound_solver(pixels, exp_time, 0.9999)

                        #eliminate offset
                        pixels = pixels - 300

                        #exp_array[channel_index] = exp_time
                        exp_tile_logger[channel_index][y][x] = exp_time

                        #multiply by tissue binary
                        masked_subbed_image = pixels * tissue_bin_im

                        #place subbed image into stack
                        exp_image_stack[channel_index][y][x] = masked_subbed_image

                if tissue_fm[y][x] == 1:
                    pass

        #save image array for analysis
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        np.save('auto_exp_images.npy', exp_image_stack)


        for channel in channels:
            if channel == 'DAPI':
                channel_index = 0
                frame_count_index = 11
            if channel == 'A488':
                channel_index = 1
                frame_count_index = 12
            if channel == 'A555':
                channel_index = 2
                frame_count_index = 13
            if channel == 'A647':
                channel_index = 3
                frame_count_index = 14

            #Normalize exposure pixels to lowest exp count
            channel_exp_times = exp_tile_logger[channel_index]
            print('channel exp times', channel_exp_times)
            lowest_exp = np.min(channel_exp_times.ravel()[np.flatnonzero(channel_exp_times)])
            #add one to eliminate 0 values
            channel_exp_times = channel_exp_times + 1

            #propogate normalization factor
            for x in range(0, x_tiles):
                for y in range(0, y_tiles):
                    scaled_factor_2_lowest = lowest_exp/exp_tile_logger[channel_index][y][x]
                    exp_image_stack[channel_index][y][x] = exp_image_stack[channel_index][y][x] * scaled_factor_2_lowest

            # use Otsu's (or any other threshold method) to find well stained pixels
            channel_pixels = np.nan_to_num(exp_image_stack[channel_index], posinf=65500)
            non_zero_pixels = channel_pixels.ravel()[np.flatnonzero(channel_pixels)]
            thresh = filters.threshold_otsu(non_zero_pixels)

            #apply threshold and find brightest and dimmest pixels
            non_zero_thresh_pixels = channel_pixels.ravel()[np.flatnonzero(channel_pixels > thresh)]
            high_pixel, low_pixel = self.image_percentile_level(non_zero_thresh_pixels, cut_off_threshold= percentage_cut_off)

            #find new exp factor and frame count to average over
            new_exp_factor = (target_percentage * 65500 / low_pixel) * lowest_exp
            new_max_int_value = (new_exp_factor/lowest_exp) * high_pixel

            print('channel', channel, 'thresh', thresh, 'high', high_pixel, 'low', low_pixel, 'new exp', new_exp_factor, 'normalized exp', lowest_exp)
            ratio_new_int_2_max_int = new_max_int_value / (0.8 * 65500)

            try:
                frame_count = math.ceil(ratio_new_int_2_max_int + 0.05)
            except:
                frame_count = 1

            total_exposure_time = new_exp_factor
            new_exp_factor = new_exp_factor/frame_count

            if total_exposure_time < 100:
                total_exposure_time = 100
            else:
                pass

            #if exposure is really high most likely nothing is really there. Just take it one frame
            if new_exp_factor > 1000:
                new_exp_factor = 1000
                frame_count = 1

            if new_exp_factor < 10:
                new_exp_factor = 10
                frame_count = math.ceil(total_exposure_time/new_exp_factor)

            if frame_count < 1:
                frame_count = 1

            else:
                pass

            #add to exp_array and fm_array
            exp_array[channel_index] = new_exp_factor
            fm_array[frame_count_index] = frame_count

        #save new exp_array anmd fm_array
        finish = time.time()

        print('auto_exp time elapsed', finish - start)

        exp_array[0] = 75
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        print(exp_array)
        np.save('exp_array.npy', exp_array)
        np.save('fm_array.npy', fm_array)

    def hdr_exp_generator(self, experiment_directory, threshold_level, max_exp, min_exp):
        '''
        generates exp list for HDR auto exposure.
        :param experiment_directory:
        :param threshold_level:
        :param max_exp:
        :param min_exp:
        :return:
        '''

        # add on exp offset
        exp_offset = 27.43
        max_exp += exp_offset
        min_exp += exp_offset

        # calculate number of images needed for threshold and real threshold
        M = 65536 / threshold_level
        n = math.ceil(np.log(max_exp / min_exp) / np.log(M) + 1)
        M_real = np.power(max_exp / min_exp, 1 / (n - 1))
        T_real = 65536 / M_real

        # create and populate hdr exp array with min and max values
        hdr_exp_list = np.zeros(n)
        hdr_exp_list[0] = min_exp - exp_offset
        hdr_exp_list[n - 1] = max_exp - exp_offset

        for x in range(1, n - 1):
            hdr_exp_list[x] = int(min_exp * np.power(M_real, x) - exp_offset)

        # import exp_calc_array
        exp_path = experiment_directory + '/' + 'exposure_times'
        os.chdir(exp_path)

        # create or open workbook

        if os.path.isfile('HDR_Exp.xlsx') == False:
            wb = Workbook()
            ws = wb.active

            # populate headers
            ws.cell(row=1, column=1).value = 'image count'
            ws.cell(row=1, column=2).value = 'exposure time 1'
            ws.cell(row=1, column=3).value = 'exposure time 2'
            ws.cell(row=1, column=4).value = 'exposure time 3'
            ws.cell(row=1, column=5).value = 'threshold real'
            ws.cell(row=4, column=1).value = 'Cycle'
            ws.cell(row=4, column=2).value = 'Max Int DAPI'
            ws.cell(row=4, column=3).value = 'Max Int A488'
            ws.cell(row=4, column=4).value = 'Max Int A555'
            ws.cell(row=4, column=5).value = 'Max Int A647'

        if os.path.isfile('HDR_Exp.xlsx') == True:
            wb = load_workbook('HDR_Exp.xlsx')
            ws = wb.active

        # populate columns with times and cycle count
        ws.cell(row=2, column=1).value = n
        ws.cell(row=2, column=2).value = hdr_exp_list[0] + exp_offset
        ws.cell(row=2, column=3).value = hdr_exp_list[1] + exp_offset
        ws.cell(row=2, column=4).value = hdr_exp_list[2] + exp_offset
        ws.cell(row=2, column=5).value = T_real

        wb.save('HDR_Exp.xlsx')

        self.hdr_exp_times = hdr_exp_list

    def hdr_manual_compression(self, experiment_directory, cycle_number):
        '''
        Read off cycle and channel highest scaling numbers from excel sheet and apply to flattened, subbed images
        and scale them from 32bit to 16bit. Lowest bound is lowest pixel value
        :param experiment_directory:
        :return:
        '''

        compression_directory = experiment_directory + r'/compression'
        os.chdir(compression_directory)
        wb = load_workbook('compression.xlsx')
        ws = wb.active

        # load in data structures
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)
        tissue_fm = fm_array[10]
        x_tile_count = np.shape(tissue_fm)[1]
        y_tile_count = np.shape(tissue_fm)[0]

        channels = np.array(['A488', 'A555', 'A647'])

        row_number = int(cycle_number + 1)



        for channel in channels:
            low_value = 1000000 #its set higher than it ever will be after adjustments per tile
            high_col = (np.where(channels == channel)[0][0] + 1) * 2
            print(channel, high_col)
            low_col = high_col + 1
            tile_directory = experiment_directory + '//' + str(channel) + r'\Stain\cy_' + str(cycle_number) + r'\Tiles\focused_subbed_basic_corrected'
            print(tile_directory)
            os.chdir(tile_directory)
            #determine low value
            for x in range(0, x_tile_count):
                for y in range(0, y_tile_count):
                    if tissue_fm[y][x] > 2:

                        tile_filename = 'x' + str(x) + '_y_' + str(y) + '_c_' + str(channel) + '.tif'
                        im = io.imread(tile_filename)
                        tile_low_value = np.min(im).astype('float32')
                        if tile_low_value < 0:
                            tile_low_value = 0
                        else:
                            pass
                        if tile_low_value < low_value:
                            low_value = tile_low_value
                        else:
                            pass

                    else:
                        pass

            for x in range(0, x_tile_count):
                for y in range(0, y_tile_count):
                    if tissue_fm[y][x] > 2:

                        high_value = ws.cell(row=row_number, column=high_col).value
                        ws.cell(row=row_number, column=low_col).value = low_value

                        #no reason for range to be under 16bit range ever so make adjustment if it is lower
                        if (high_value - low_value) < 65500:
                            high_value = low_value + 65500
                            ws.cell(row=row_number, column=high_col).value =  high_value
                        else:
                            pass

                        tile_filename = 'x' + str(x) + '_y_' + str(y) + '_c_' + str(channel) + '.tif'
                        im = io.imread(tile_filename)

                        im = im.astype('float32')
                        im[im < 0] = 0
                        im  = im/high_value
                        im[im > 1] = 1
                        im = skimage.util.img_as_uint(im)
                        io.imsave(tile_filename, im)

                    else:
                        pass

        os.chdir(compression_directory)
        wb.save('compression.xlsx')

    def hdr_compression_2(self, experiment_directory, cycle_number):
        '''
        Takes quick tiled image and clips some high and low pixels to define a smaller range to
        use when doing a linear conversion form 32bit space to 16bit space.

        :param experiment_directory:
        :param cycle_number:
        :return:
        '''

        compression_directory = experiment_directory + r'/compression'
        os.chdir(compression_directory)
        try:
            wb = load_workbook('compression.xlsx')
        except:
            wb = Workbook()
        ws = wb.active

        # load in data structures
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)
        tissue_fm = fm_array[10]
        x_tile_count = np.shape(tissue_fm)[1]
        y_tile_count = np.shape(tissue_fm)[0]

        row_number = int(cycle_number + 1)

        min_bin_fraction = 0.000001
        channels = np.array(['A488', 'A555', 'A647'])

        for channel in channels:

            quicktile_path = experiment_directory + '\Quick_Tile/' + channel
            flattened_path = experiment_directory + '//' + channel + '\Stain\cy_' + str(cycle_number) + r'\Tiles\focused_basic_darkframe'
            high_col = (np.where(channels == channel)[0][0] + 1) * 2

            os.chdir(quicktile_path)
            filename = channel + '_cy_' + str(cycle_number) + '_Stain_placed.tif'
            im = io.imread(filename)
            im -= 1
            im[im < 0] = 0
            non_zero_indicies = np.nonzero(im)
            histogram, bin_edges = np.histogram(im[non_zero_indicies], bins=256)

            total_pixels = np.sum(histogram)
            min_bin_size = math.ceil(min_bin_fraction * total_pixels)
            # print(min_bin_size)
            histogram -= min_bin_size
            histogram[histogram < 0] = 0
            try:
                threshold_bin_number = np.where(histogram == 0)[0][0]
            except:
                threshold_bin_number = 255
            top_int = bin_edges[threshold_bin_number - 1]

            if top_int < 65500:
                top_int = 65500
            else:
                pass

            ws.cell(row=row_number, column=high_col).value = top_int

            # apply compression on flattened images
            os.chdir(flattened_path)

            for x in range(0, x_tile_count):
                for y in range(0, y_tile_count):
                    if tissue_fm[y][x] > 1:
                        image_name = r'x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                        flat_image = io.imread(image_name)

                        flat_image[flat_image < 0] = 0
                        flat_image = flat_image.astype('float32')
                        flat_image = flat_image / top_int
                        flat_image[flat_image > 1] = 1
                        flat_image = util.img_as_uint(flat_image)
                        io.imsave(image_name, flat_image)

                    else:
                        pass

        os.chdir(compression_directory)
        wb.save('compression.xlsx')

    def hdr_compression(self, experiment_directory, cycle_number, apply_2_subbed = 1, apply_2_bleached = 1, apply_2_focused = 1, apply_2_flattened = 1,  channels = ['DAPI', 'A488','A555', 'A647']):
        '''
        Looks through all tiles and compresses 32bit images into 16 bit based on 2^16 = highest pixel in any tile.
        By standard, will apply to raw stained tiles. Can be extended to raw bleached tiles and subbed tiles.
        :param experiment_directory:
        :param cycle_number:
        :param apply_2_subbed: 0 = dont compress subbed images, 1= compress
        :param apply_2_bleached: 0 = dont compress raw bleached tiles, 1= compress
        :return:
        '''

        # import fm_array and establish tissue exists array
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        fm_array = np.load('fm_array.npy', allow_pickle=False)
        tissue_fm = fm_array[10]

        #import HDR Exp.xlsx
        exp_path = experiment_directory + '/' + 'exposure_times'
        os.chdir(exp_path)
        wb = load_workbook('HDR_Exp.xlsx')
        ws = wb.active

        #find tile counts
        y_tile_count = np.shape(fm_array[0])[0]
        x_tile_count = np.shape(fm_array[0])[1]
        z_tiles = int(fm_array[3][0][0])

        #determine all types of data to be applied on
        types = ['Stain']
        if apply_2_bleached == 1:
            types.append('Bleach')
        if apply_2_subbed == 1:
            #make anything beyond Stain and Bleach all lower case
            types.append('focused_subbed')
        if apply_2_focused == 1:
            #make anything beyond Stain and Bleach all lower case
            if apply_2_bleached == 1:
                types.append('bleach_focused')
            else:
                pass
            types.append('focused')
        if apply_2_flattened == 1:
            #make anything beyond Stain and Bleach all lower case
            types.append('focused_subbed_basic_corrected')

        for channel in channels:
            for type in types:
                if type == 'Stain':
                    type_path = experiment_directory + r'//' + channel + r'//' + type + r'//cy_' + str(cycle_number) + r'//Tiles'
                    os.chdir(type_path)
                elif type == 'Bleach':
                    type_path = experiment_directory + r'//' + channel + r'//' + type + r'//cy_' + str(cycle_number - 1) + r'//Tiles'
                    os.chdir(type_path)
                elif type == 'bleach_focused':
                    type_path = experiment_directory + r'//' + channel + r'//' + 'Bleach' + r'//cy_' + str(cycle_number - 1) + r'//Tiles/focused'
                    os.chdir(type_path)
                elif channel == 'DAPI' and type == 'focused_subbed':
                    pass
                elif channel == 'DAPI' and type == 'focused_subbed_basic_corrected':
                    type_path = experiment_directory + r'//' + channel + r'//Stain//cy_' + str(cycle_number) + r'//Tiles/' + 'focused_basic_corrected'
                    os.chdir(type_path)
                else:
                    type_path = experiment_directory + r'//' + channel + r'//Stain//cy_' + str(cycle_number) + r'//Tiles/' + type
                    os.chdir(type_path)

                if type == 'Stain':
                    highest_intensity = 0
                    #Find max intensity
                    for x in range(0, x_tile_count):
                        for y in range(0, y_tile_count):
                            if tissue_fm[y][x] > 1:
                                for z in range(0, z_tiles):
                                    file_name = r'z_' +str(z) + '_x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                                    im = io.imread(file_name)
                                    max_tile_int = np.max(im)
                                    if max_tile_int > highest_intensity:
                                        highest_intensity = max_tile_int
                                    else:
                                        pass

                    if im.dtype == 'uint16':
                        if channel == 'DAPI':
                            channel_col_index = 2
                        if channel == 'A488':
                            channel_col_index = 3
                        if channel == 'A555':
                            channel_col_index = 4
                        if channel == 'A647':
                            channel_col_index = 5

                        row = cycle_number + 4
                        highest_intensity = ws.cell(row=row, column=channel_col_index).value

                    if apply_2_flattened == 1:
                        if channel == 'DAPI':
                            ff_path = experiment_directory + r'//' + channel + r'//Stain//cy_' + str(cycle_number) + r'//Tiles/' + 'focused_basic_corrected'
                            os.chdir(ff_path)
                        else:
                            ff_path = experiment_directory + r'//' + channel + r'//Stain//cy_' + str(cycle_number) + r'//Tiles/' + 'focused_subbed_basic_corrected'
                            os.chdir(ff_path)

                        file_names = os.listdir(ff_path)
                        image_tile = io.imread(file_names[2])
                        if image_tile.dtype == 'uint16':
                            pass
                        else:
                            ff_im = io.imread('flatfield.tif')
                            max_ff_ratio = np.max(ff_im)
                            highest_intensity = highest_intensity * max_ff_ratio
                            os.chdir(type_path)

                    #record highest intensity in exp hdr logbook
                    row = cycle_number + 4
                    ws.cell(row=row, column=1).value = cycle_number
                    if channel == 'DAPI':
                        channel_col_index = 2
                    if channel == 'A488':
                        channel_col_index = 3
                    if channel == 'A555':
                        channel_col_index = 4
                    if channel == 'A647':
                        channel_col_index = 5



                    ws.cell(row=row, column=channel_col_index).value = highest_intensity


                    os.chdir(exp_path)
                    wb.save('HDR_Exp.xlsx')
                    #highest_intensity = ws.cell(row=row, column=channel_col_index).value

                    #resave compressed versions of Stain tiles
                    os.chdir(type_path)
                    for x in range(0, x_tile_count):
                        for y in range(0, y_tile_count):
                            if tissue_fm[y][x] > 1:
                                for z in range(0, z_tiles):
                                    file_name = r'z_' +str(z) + '_x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                                    im = io.imread(file_name)
                                    im = np.nan_to_num(im, nan=0, posinf=0)
                                    if im.dtype == 'unit16':
                                        pass
                                    else:
                                        print('stain', np.max(im), highest_intensity)
                                        im = im/highest_intensity
                                        im = skimage.util.img_as_uint(im)
                                        io.imsave(file_name, im)

                elif channel == 'DAPI' and type == 'focused_subbed':
                    pass
                elif type == 'Bleach':
                    #Max int came from Stain loop above this section
                    os.chdir(type_path)
                    for x in range(0, x_tile_count):
                        for y in range(0, y_tile_count):
                            if tissue_fm[y][x] > 1:
                                for z in range(0, z_tiles):
                                    file_name = r'z_' + str(z) + '_x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                                    im = io.imread(file_name)
                                    im = np.nan_to_num(im, nan=0, posinf=0)
                                    if im.dtype == 'unit16':
                                        pass
                                    else:
                                        print('bleach', np.max(im), highest_intensity)
                                        im = im / highest_intensity
                                        im = skimage.util.img_as_uint(im)
                                        #tf.imwrite(file_name,im, compression='zlib')
                                        io.imsave(file_name, im)
                else:
                    #Max int came from Stain loop above this section
                    os.chdir(type_path)
                    for x in range(0, x_tile_count):
                        for y in range(0, y_tile_count):
                            if tissue_fm[y][x] > 1:
                                file_name = r'x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                                im = io.imread(file_name)
                                print(type, channel, 'x', x, 'y', y, np.max(im), highest_intensity)
                                im = np.nan_to_num(im, nan=0, posinf=0)
                                if im.dtype == 'unit16':
                                    pass

                                else:
                                    im = im/highest_intensity
                                    #im = int(im)
                                    #im = im.astype('uint16')

                                    im = skimage.util.img_as_uint(im)
                                    #tf.imwrite(file_name,im, compression='zlib')
                                    io.imsave(file_name, im)

    def hdr_fuser(self, hdr_images):

        hdr_times = self.hdr_exp_times
        max_hdr_time = np.max(hdr_times)
        hdr_array = hdr_images
        weight_array = deepcopy(hdr_array)

        # subtract linear offset
        hdr_array = hdr_array - 300

        # populate weight array
        for index in range(0, 3):
            exp_offset = 27.43
            mag = (max_hdr_time + exp_offset) / (hdr_times[index] + exp_offset)
            del_offset = 1.2 * (max_hdr_time - hdr_times[index]) / (hdr_times[index] + exp_offset) ** 2

            im = copy.deepcopy(hdr_array[index])

            scaled_im = mag * im
            hdr_array[index] = scaled_im
            im[im > 65234] = 0

            del_I = np.sqrt(im)

            weight_array[index] = del_I / scaled_im + del_offset / mag

        total_weight_array = np.sum(weight_array, axis=0)
        scaled_weight_array = np.divide(weight_array, total_weight_array)

        hdr_im = hdr_array[0] * scaled_weight_array[0]
        for x in range(1, np.shape(hdr_array)[0]):
            hdr_im += hdr_array[x] * scaled_weight_array[x]

        return hdr_im

    def autof_factor_estimator(self, image, autof_image, num_images=2):
        top_range = 10
        x_factor = top_range / num_images

        image = np.nan_to_num(image, posinf= 65500)
        autof_image = np.nan_to_num(autof_image, posinf= 65500)

        image = image.astype('float64')
        autof_image = autof_image.astype('float64')

        x_axis = np.linspace(0, top_range, num_images).astype('float64')
        y_axis = np.linspace(0, top_range, num_images).astype('float64')
        for x in range(0, num_images):
            input_image = (image - x_factor * x * autof_image)
            mean = self.absolute_mean(input_image)
            y_axis[x] = mean
            x_axis[x] = x * x_factor

        projected_min_point = self.min_factor(x_axis, y_axis)

        return projected_min_point

    def absolute_mean(self, image):
        abs = np.absolute(image)
        mean = np.mean(abs)

        return mean

    def min_factor(self, x_axis, y_axis):
        sorted_list = np.sort(y_axis)
        index_1 = np.where(y_axis == sorted_list[0])[0][0]
        index_1 = x_axis[index_1]
        index_2 = np.where(y_axis == sorted_list[1])[0][0]
        index_2 = x_axis[index_2]
        min_value_1 = sorted_list[0]
        min_value_2 = sorted_list[1]

        total_value = min_value_1 + min_value_2
        rel_value_1 = total_value / min_value_1
        rel_value_2 = total_value / min_value_2
        total_rel_value = rel_value_1 + rel_value_2
        normalized_rel_value_1 = rel_value_1 / total_rel_value
        normalized_rel_value_2 = rel_value_2 / total_rel_value
        projected_min_point = normalized_rel_value_1 * index_1 + normalized_rel_value_2 * index_2

        return projected_min_point

    def image_percentile_level(self, image, cut_off_threshold=0.99):
        '''
        Takes in image and cut off threshold and finds pixel value that exists at that threshold point.

        :param numpy array image: numpy array image
        :param float cut_off_threshold: percentile for cut off. For example a 0.99 would disregaurd the top 1% of pixels from calculations
        :return: intensity og pixel that resides at the cut off fraction that was entered in the image
        :rtype: int
        '''

        pixel_values = np.sort(image, axis=None)
        indicies = np.nonzero(pixel_values)[0]
        # thresh = filters.threshold_otsu(pixel_values[indicies])
        # indicies = np.where(pixel_values > thresh)[0]
        pixel_values = pixel_values[indicies]

        pixel_count = int(np.size(pixel_values))
        cut_off_index = int(pixel_count * cut_off_threshold)
        min_cut_off_index = int(pixel_count - cut_off_index)
        tail_intensity = pixel_values[cut_off_index]
        min_intensity = pixel_values[min_cut_off_index]

        return tail_intensity, min_intensity

    def exp_bound_solver(self, image, exp_time, percentage_cutoff):
        '''
        Takes in image and determines if its in bounds. If not, it adjusts exp time to compensate and gives indicator
        if exp time alterations were used, ie was it triggered.

        :param image:
        :param exp_time:
        :return:
        '''

        target_intensity = 65535 * 0.05
        max_time = 100
        min_time = 5
        trigger_state = 0

        x_frame_size = np.shape(image)[1]

        intensity, low_intensity = self.image_percentile_level(image, percentage_cutoff)
        #print(intensity)

        while intensity > 60000 and exp_time >= min_time:
            exp_time = exp_time / 3
            core.set_exposure(exp_time)
            core.set_config("amp", 'high')
            core.snap_image()
            tagged_image = core.get_tagged_image()
            new_image = np.reshape(tagged_image.pix, newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
            intensity, low_intensity = self.image_percentile_level(new_image, cut_off_threshold=percentage_cutoff)
            trigger_state = 1

        if intensity < 1000 and exp_time < max_time:
            trigger_state = 1

        if trigger_state == 1:
            scale_factor = target_intensity / (intensity - 300)
            if scale_factor * exp_time > max_time:
                exp_time = max_time
            if scale_factor < min_time:
                scale_factor = min_time
            else:
                exp_time = int(exp_time * scale_factor)

            if exp_time > max_time:
                exp_time = max_time
            elif max_time < min_time:
                exp_time = min_time

        if trigger_state == 1:

            #gather new image
            core.set_exposure(exp_time)
            core.set_config("amp", 'high')
            core.snap_image()
            tagged_image = core.get_tagged_image()
            image = np.reshape(tagged_image.pix, newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
            #crop image to equal first image
            side_pixel_count = int((5056 - x_frame_size)/2)
            image = image[::, side_pixel_count:side_pixel_count + x_frame_size]
            image[image > 60000] = 0
            image = np.nan_to_num(image, posinf=65500)
        else:
            pass


        return image, exp_time

    def exp_predetermined(self, experiment_directory, cycle_number):
        # load in data structures
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        exp_filename = 'exp_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)
        exp_array = np.load(exp_filename, allow_pickle=False)
        channels = ['A488', 'A555', 'A647']
        number_channels = len(channels[0])

        #load in exp excel sheet
        #exp_path = r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF\Software\predetermined_exposure_times\gutage'
        exp_path = r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF\Software\predetermined_exposure_times\marco'
        os.chdir(exp_path)
        wb = load_workbook('Exp.xlsx')
        ws = wb.active
        row_number = cycle_number + 2

        a488_exp = ws.cell(row=row_number, column=4).value
        a555_exp = _exp = ws.cell(row=row_number, column=6).value
        a647_exp = ws.cell(row=row_number, column=8).value

        a488_frames = ws.cell(row=row_number, column=5).value
        a555_frames = _exp = ws.cell(row=row_number, column=7).value
        a647_frames = ws.cell(row=row_number, column=9).value

        exp_array[1] = a488_exp
        exp_array[2] = a555_exp
        exp_array[3] = a647_exp

        fm_array[12][0][0] = a488_frames
        fm_array[13][0][0] = a555_frames
        fm_array[14][0][0] = a647_frames

        #save exp and fm arrays
        os.chdir(numpy_path)
        np.save(file_name, fm_array)
        np.save(exp_filename, exp_array)

    def exp_logbook(self, experiment_directory, cycle):
        '''
        Opens up numpy object for exp times and records in excel sheet
        :param experiment_directory:
        :param cycle:
        :return:
        '''

        # import exp_calc_array
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        exp_array = np.load('exp_array.npy', allow_pickle=False)
        fm_array = np.load('fm_array.npy', allow_pickle=False)

        exp_path = experiment_directory + '/' + 'exposure_times'
        os.chdir(exp_path)

        # create or open workbook

        if os.path.isfile('Exp.xlsx') == False:
            wb = Workbook()
            ws = wb.active

            # populate headers
            ws.cell(row = 1, column = 1).value = 'Cycle #'
            ws.cell(row=1, column=2).value = 'Exp Time DAPI'
            ws.cell(row=1, column=3).value = 'Average # DAPI'
            ws.cell(row=1, column=4).value = 'Exp Time A488'
            ws.cell(row=1, column=5).value = 'Average # A488'
            ws.cell(row=1, column=6).value = 'Exp Time A555'
            ws.cell(row=1, column=7).value = 'Average # A555'
            ws.cell(row=1, column=8).value = 'Exp Time A647'
            ws.cell(row=1, column=9).value = 'Average # A647'

        if os.path.isfile('Exp.xlsx') == True:
            wb = load_workbook('Exp.xlsx')
            ws = wb.active

        # populate columns with times and cycle count
        ws.cell(row=int(cycle + 2), column=1).value = cycle

        ws.cell(row=int(cycle + 2), column=2).value = exp_array[0]
        ws.cell(row=int(cycle + 2), column=4).value = exp_array[1]
        ws.cell(row=int(cycle + 2), column=6).value = exp_array[2]
        ws.cell(row=int(cycle + 2), column=8).value = exp_array[3]

        ws.cell(row=int(cycle + 2), column=3).value = fm_array[11][0][0]
        ws.cell(row=int(cycle + 2), column=5).value = fm_array[12][0][0]
        ws.cell(row=int(cycle + 2), column=7).value = fm_array[13][0][0]
        ws.cell(row=int(cycle + 2), column=9).value = fm_array[14][0][0]

        wb.save('Exp.xlsx')
    ##########################################################
    #Initial population of fm_array and some mods to it
    ##########################################################

    def tile_xy_pos(self, surface_name):
        '''
        imports previously generated micro-magellan surface with name surface_name and outputs
        the coordinates of the center of each tile from it.

        :param str surface_name: name of micro-magellan surface
        :param object magellan: object created via = bridge.get_magellan()

        :return: XY coordinates of the center of each tile from micro-magellan surface
        :rtype: dictionary {{x:(float)}, {y:(float)}}
        '''
        surface = magellan.get_surface(surface_name)
        num = surface.get_num_positions()
        xy = surface.get_xy_positions()
        tile_points_xy = {}
        x_temp = []
        y_temp = []

        for q in range(0, num):
            pos = xy.get(q)
            pos = pos.get_center()
            x_temp.append(pos.x)
            y_temp.append(pos.y)

        tile_points_xy['x'] = x_temp  ## put all points in dictionary to ease use
        tile_points_xy['y'] = y_temp
        x_temp = np.array(tile_points_xy['x'])
        y_temp = np.array(tile_points_xy['y'])

        xy = np.append(x_temp, y_temp)
        xy = np.reshape(xy, (2, int(xy.size / 2)))

        return xy

    def tile_pattern(self, numpy_array, experiment_directory):
        '''
        Takes numpy array with N rows and known tile pattern and casts into new array that follows
        south-north, west-east snake pattern.


        :param numpy_array: dimensions [N, x_tiles*y_tiles]
        :param x_tiles: number x tiles in pattern
        :param y_tiles: number y tiles in pattern
        :return: numpy array with dimensions [N,x_tiles,y_tiles] with above snake pattern
        '''
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        filename = 'fm_array.npy'

        y_tiles = np.unique(numpy_array[1]).size
        x_tiles = np.unique(numpy_array[0]).size
        layers = np.shape(numpy_array)[0]
        numpy_array = numpy_array.reshape(layers, x_tiles, y_tiles)
        dummy = numpy_array.reshape(layers, y_tiles, x_tiles)
        new_numpy = np.empty_like(dummy)
        for x in range(0, layers):
            new_numpy[x] = numpy_array[x].transpose()
            new_numpy[x, ::, 1:y_tiles:2] = np.flipud(new_numpy[x, ::, 1:y_tiles:2])

        np.save(filename, new_numpy)

        return new_numpy

    def fm_channel_update(self, experiment_directory, off_array, z_slices, slice_gap = 2):

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)

        a488_channel_offset = off_array[1]  # determine if each of these are good and repeatable offsets
        a555_channel_offset = off_array[2]
        a647_channel_offset = off_array[3]

        fm_array[4] = fm_array[2] + a488_channel_offset  # index for a488 = 3
        fm_array[6] = fm_array[2] + a555_channel_offset
        fm_array[8] = fm_array[2] + a647_channel_offset

        np.save(file_name, fm_array)

    def fm_channel_initial(self, experiment_directory, off_array, z_slices, slice_gap=2):

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)

        a488_channel_offset = off_array[1]  # determine if each of these are good and repeatable offsets
        a555_channel_offset = off_array[2]
        a647_channel_offset = off_array[3]

        y_tiles = np.shape(fm_array[0])[0]
        x_tiles = np.shape(fm_array[0])[1]

        dummy_channel = np.empty_like(fm_array[0])
        dummy_channel = np.expand_dims(dummy_channel, axis=0)
        channel_count = np.shape(fm_array)[0]

        while channel_count < 15:
            fm_array = np.append(fm_array, dummy_channel, axis=0)
            channel_count = np.shape(fm_array)[0]

        fm_array[4] = fm_array[2] + a488_channel_offset  # index for a488 = 3
        fm_array[6] = fm_array[2] + a555_channel_offset
        fm_array[8] = fm_array[2] + a647_channel_offset
        y_tiles = int(np.shape(fm_array[0])[0])
        x_tiles = int(np.shape(fm_array[0])[1])
        z_slice_array = np.full((y_tiles, x_tiles), z_slices)
        all_ones_array = np.full((y_tiles, x_tiles), 1)

        fm_array[3] = z_slice_array
        fm_array[5] = z_slice_array
        fm_array[7] = z_slice_array
        fm_array[9] = z_slice_array

        fm_array[10] = np.full((y_tiles, x_tiles), 2)
        fm_array[11] = all_ones_array
        fm_array[12] = all_ones_array
        fm_array[13] = all_ones_array
        fm_array[14] = all_ones_array

        fm_array[2] = fm_array[2] + int(((z_slice_array[0][0] - 1) * slice_gap) / 2)
        fm_array[4] = fm_array[4] + int(((z_slice_array[0][0] - 1) * slice_gap) / 2)
        fm_array[6] = fm_array[6] + int(((z_slice_array[0][0] - 1) * slice_gap) / 2)
        fm_array[8] = fm_array[8] + int(((z_slice_array[0][0] - 1) * slice_gap) / 2)

        np.save(file_name, fm_array)

        return fm_array

    def fm_stage_tilt_compensation(self, experiment_directory, tilt=3.75):
        '''
        takes intial seed value and adds in a tilt to give focus values roughly in line with actual positions
        :param experiment_directory:
        :param tilt: slope of displacement in microns per FOV in the y axis
        :return:
        '''

        # load in fm array
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)

        x_tile_count = np.shape(fm_array[0])[1]
        y_tile_count = np.shape(fm_array[0])[0]
        z_planes = fm_array[2]
        starting_focus = z_planes[0][0]
        for x in range(0, x_tile_count):
            for y in range(0, y_tile_count):
                z_planes[y][x] = starting_focus - y * tilt
        np.save('fm_array.npy', fm_array)

    def fm_map_z_shifter(self, experiment_directory, desired_z_slices_dapi, desired_z_slices_other):
        '''
        Looks at focus map and shifts stack and alters desired slice counts from original
        values to new desired values
        :param experiment_directory:
        :param desired_z_slices_dapi:
        :param desired_z_slices_other:
        :return:
        '''

        # load in fm array
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)

        slice_gap = 2

        original_z_dapi = fm_array[3][0][0]
        original_z_others = fm_array[5][0][0]


        #shift stacks
        fm_array[2] -= (original_z_dapi - desired_z_slices_dapi) * slice_gap #dapi
        fm_array[4] -= (original_z_others - desired_z_slices_other) * slice_gap #a488
        fm_array[6] -= (original_z_others - desired_z_slices_other) * slice_gap #a555
        fm_array[8] -= (original_z_others - desired_z_slices_other) * slice_gap #a647

        #update z slices
        fm_array[3] = desired_z_slices_dapi
        fm_array[5] = desired_z_slices_other
        fm_array[7] = desired_z_slices_other
        fm_array[9] = desired_z_slices_other

        np.save('fm_array.npy', fm_array)

    def x_overlap_adjuster(self, new_x_pixel_count, experiment_directory):
        '''
        Increases overlap in focus map while cropping in the x dimension to preserve effective 10% overlap of cropped images
        :param new_x_pixel_count:
        :param experiment_directory:
        :return:
        '''

        # load in fm array
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)

        # Find pixels in each dimensions
        x_tiles = np.shape(fm_array[0])[1]
        y_tiles = np.shape(fm_array[0])[0]
        fm_layers = np.shape(fm_array)[0]

        # find um/pixel in focus map

        try:
            x1 = fm_array[0][0][0]
            x2 = fm_array[0][0][1]
            diff = x2 - x1
        except:
            y1 = fm_array[0][0][0]
            y2 = fm_array[0][1][0]
            diff = y2 - y1

        #um_per_pixel = diff / 4550  # 4550 = 0.9 * 5056
        um_per_pixel = 0.2005

        # Find number tiles in adjusted grid
        x_range_pixels = (x_tiles - 0.2) * 5056
        number_new_x_dim_tiles = x_range_pixels / new_x_pixel_count
        new_x_tiles = math.ceil(number_new_x_dim_tiles)

        # generate new blank fm_array numpy array

        new_fm_array = np.random.rand(fm_layers, y_tiles, new_x_tiles).astype('float64')

        # Find border where x starts on the left (not center point, but x value for left most edge of left most tile

        left_x_center = fm_array[0][0][0]
        left_most_x = left_x_center - 2528 * um_per_pixel

        # Find center point in new image that makes edge of image align with left_most_x
        # Also find x to x + i spacing and populate rest of x values in new_fm_array

        x_col_0 = left_most_x + new_x_pixel_count / 2 * um_per_pixel
        x_spacing = (0.9) * new_x_pixel_count * um_per_pixel

        # Populate new_fm_array with row 0 x values

        for x in range(0, new_x_tiles):
            new_fm_array[0][0:y_tiles, x] = x_col_0 + x * x_spacing

        # populate new_fm_array with y values

        for y in range(0, y_tiles):
            new_fm_array[1][y, 0:new_x_tiles] = fm_array[1][y][0]

        # populate new_fm_array with dapi z values and everything else in planes 2-10
        for slice in range(2, fm_layers):
            new_fm_array[slice, 0:y_tiles, 0:new_x_tiles] = fm_array[slice][0][0]

        np.save('fm_array.npy', new_fm_array)

    def col_row_nonzero(self, image):

        # determine row and column counts
        row_count = np.shape(image)[0]
        column_count = np.shape(image)[1]

        # define row and column arrays
        row_nonzero = np.zeros((row_count))
        column_nonzero = np.zeros((column_count))

        # find column indicies with non zero values
        for x in range(0, column_count):
            column = image[::, x]
            sum = np.sum(column)
            if sum > 0:
                column_nonzero[x] = 1
            else:
                pass

        # find row indicies with non zero values
        for y in range(0, row_count):
            row = image[y, ::]
            sum = np.sum(row)
            if sum > 0:
                row_nonzero[y] = 1
            else:
                pass

        return row_nonzero, column_nonzero

    def fm_grid_readjuster(self, experiment_directory, x_frame_size):
        numpy_path = experiment_directory + r'\np_arrays'
        tissue_path = experiment_directory + r'\Tissue_Binary'

        os.chdir(numpy_path)
        fm_file_name = 'fm_array.npy'
        fm_array = np.load(fm_file_name, allow_pickle=False)

        # find tile counts
        x_tiles = np.shape(fm_array[0])[1]
        y_tiles = np.shape(fm_array[0])[0]

        # make row and column arrays that can contain all tissue images in row or col respectively
        row_image = np.random.rand(2960, x_frame_size * x_tiles).astype('float16')
        col_image = np.random.rand(2960 * y_tiles, x_frame_size).astype('float16')

        os.chdir(tissue_path)

        found_upper_y = 0
        found_lower_y = 0
        found_right_x = 0
        found_left_x = 0

        # find upper and lower tissue containing tiles and what rows within them where the tissue starts showing up

        # upper row
        for y in range(0, y_tiles):
            if found_upper_y == 0:
                for x in range(0, x_tiles):
                    # populate row image
                    filename = 'x' + str(x) + '_y_' + str(y) + '_tissue.tif'
                    individual_image = io.imread(filename)
                    start = x * x_frame_size
                    end = start + x_frame_size
                    row_image[::, start:end] = individual_image
                row_array, col_array = self.col_row_nonzero(row_image)
                try:
                    row_indicies = np.nonzero(row_array)[0]
                    upper_y_index = row_indicies[0]
                    upper_y_tile = y
                    found_upper_y = 1
                except:
                    pass

            else:
                pass

        # lower row
        for y in range(y_tiles - 1, -1, -1):
            if found_lower_y == 0:
                for x in range(0, x_tiles):
                    # populate row image
                    filename = 'x' + str(x) + '_y_' + str(y) + '_tissue.tif'
                    individual_image = io.imread(filename)
                    start = x * x_frame_size
                    end = start + x_frame_size
                    row_image[::, start:end] = individual_image

                row_image = np.flipud(row_image)
                row_array, col_array = self.col_row_nonzero(row_image)
                try:
                    row_indicies = np.nonzero(row_array)[0]
                    lower_y_index = 2960 - row_indicies[0]
                    lower_y_tile = y
                    found_lower_y = 1
                except:
                    pass

            else:
                pass

        # right column
        for x in range(x_tiles - 1, -1, -1):
            if found_right_x == 0:
                for y in range(0, y_tiles):
                    # populate row image
                    filename = 'x' + str(x) + '_y_' + str(y) + '_tissue.tif'
                    individual_image = io.imread(filename)
                    start = y * x_frame_size
                    end = start + 2960
                    col_image[start:end, ::] = individual_image

                col_image = np.fliplr(col_image)
                row_array, col_array = self.col_row_nonzero(col_image)
                try:
                    col_indicies = np.nonzero(col_array)[0]
                    right_x_index = x_frame_size - col_indicies[0]
                    right_x_tile = x
                    found_right_x = 1
                except:
                    pass

            else:
                pass

        # lower column
        for x in range(0, x_tiles):
            if found_left_x == 0:
                for y in range(0, y_tiles):
                    # populate row image
                    filename = 'x' + str(x) + '_y_' + str(y) + '_tissue.tif'
                    individual_image = io.imread(filename)
                    start = y * x_frame_size
                    end = start + 2960
                    col_image[start:end, ::] = individual_image

                row_array, col_array = self.col_row_nonzero(col_image)
                try:
                    col_indicies = np.nonzero(col_array)[0]
                    left_x_index = col_indicies[0]
                    left_x_tile = x
                    found_left_x = 1
                except:
                    pass

            else:
                pass

        # determine new X and Y grid size (physical displacement in microns)

        x_tile_range = fm_array[0][0][right_x_tile] - fm_array[0][0][left_x_tile] + (
                    x_frame_size / 2 - left_x_index) * 0.204 + (right_x_index - x_frame_size / 2) * 0.204
        y_tile_range = fm_array[1][lower_y_tile][0] - fm_array[1][upper_y_tile][0] + (
                    2960 / 2 - upper_y_index) * 0.204 + (lower_y_index - 2960 / 2) * 0.204

        #add in margins. This will add in new tiles if the tissue is close to the frame edge which makes it
        # so that it still fits if the actual tissue is a tad larger than the segmented version

        x_tile_range = x_tile_range * 1.05
        y_tile_range = y_tile_range * 1.05

        # determine min number tiles to encompass tissue

        x_new_tiles = math.ceil(((x_tile_range / (x_frame_size * 0.204*0.9)) - 1) / 0.9 + 1)
        y_new_tiles = math.ceil(((y_tile_range / (2960 * 0.204*0.9)) - 1) / 0.9 + 1)

        # determine displacement vector for xy grid
        margin_frame_x_2_tissue = (((x_new_tiles - 1) + 1) * x_frame_size * 0.204 - x_tile_range) / 2
        displacement_x = (margin_frame_x_2_tissue - left_x_index * 0.204)

        margin_frame_y_2_tissue = (((y_new_tiles - 1) + 1) * 2960 * 0.204 - y_tile_range) / 2
        displacement_y = (margin_frame_y_2_tissue - upper_y_index * 0.204)

        # Alter fm_array tiles
        fm_array_adjusted = fm_array[::, upper_y_tile:(upper_y_tile + y_new_tiles), left_x_tile:(left_x_tile + x_new_tiles)]

        print(margin_frame_x_2_tissue, margin_frame_y_2_tissue, x_tile_range, y_tile_range)
        print('y tiles', upper_y_tile, y_new_tiles, 'x tiles', left_x_tile, x_new_tiles)

        print('dis', displacement_x, displacement_y)

        # add in display vector
        fm_array_adjusted[0] = fm_array_adjusted[0] - displacement_x
        fm_array_adjusted[1] = fm_array_adjusted[1] - displacement_y
        print(fm_array_adjusted[1])

        # save fm_array
        os.chdir(numpy_path)
        np.save(fm_file_name, fm_array_adjusted)

        print(displacement_x, displacement_y)
        print(fm_array_adjusted[0])

    def nonfocus_tile_DAPI(self, full_array_no_pattern, experiment_directory, focus_position = 'none'):
        '''
        Makes micromagellen z the focus position at each XY point for DAPI
        auto_focus and outputs the paired in focus z coordinate

        :param dictionary tile_points_xy: dictionary containing all XY coordinates. In the form: {{x:(int)}, {y:(int)}}

        :return: XYZ points where XY are stage coords and Z is in focus coordinate. {{x:(int)}, {y:(int)}, {z:(float)}}
        :rtype: dictionary
        '''

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'

        if focus_position == 'none':
            z_pos = magellan.get_surface('New Surface 1').get_points().get(0).z
        else:
            z_pos = focus_position

        num = np.shape(full_array_no_pattern)[1]
        z_temp = []
        for q in range(0, num):
            z_temp.append(z_pos)
        z_temp = np.expand_dims(z_temp, axis=0)
        xyz = np.append(full_array_no_pattern, z_temp, axis=0)

        np.save(file_name, xyz)

        return xyz

    #functions for itdentifying is tissue is present or not###################

    def tissue_region_identifier(self, experiment_directory, x_frame_size = 2960, clusters_retained = 1):
        '''
        Looks at tissue binary images and updates fm object to encode information
        :return:
        '''

        # load in fm array
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)

        # Find pixels in each dimensions
        x_tiles = np.shape(fm_array[0])[1]
        y_tiles = np.shape(fm_array[0])[0]

        #make tissue binary images
        self.tissue_binary_generate(experiment_directory, x_frame_size, clusters_retained)
        tissue_path = experiment_directory + '/Tissue_Binary'
        os.chdir(tissue_path)

        tissue_status = 1

        for x in range(0, x_tiles):
            for y in range(0, y_tiles):

                tissue_binary_name = 'x' + str(x) + '_y_' + str(y) + '_label_tissue.tif'
                im = io.imread(tissue_binary_name)
                unique_numbers = np.unique(im)
                tissue_fm_code_number = 0

                for number in unique_numbers:
                    sci_number = '1e+' + str(number)
                    fl_number = float(sci_number)
                    tissue_fm_code_number += fl_number

                fm_array[10][y][x] = tissue_fm_code_number

        os.chdir(numpy_path)
        np.save(file_name, fm_array)

    def tissue_filter(self, image):

        image = image.astype('bool')
        image_2 = morphology.remove_small_objects(image, min_size=50000, connectivity=1)
        image_2 = image_2.astype('int8')

        return image_2

    def tissue_cluster_filter(self, experiment_directory, x_frame_size, number_clusters_retained = 1, area_threshold = 0.25):
        '''
        Looks at tissue binary images, combines them and determines the largest x clusters in
        the joined image and removes the rest. In addition, this will fill in holes
        :param experiment_directory:
        :param x_frame_size:
        :param number_clusters_retained:
        :return:
        '''

        # load numpy arrays in
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]

        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]

        file_path = experiment_directory + '/Tissue_Binary'
        os.chdir(file_path)


        #make image that can hold individual images
        super_image = np.ones((y_tile_count * 2960, x_tile_count * x_frame_size))

        for y in range(0, y_tile_count):
            for x in range(0, x_tile_count):
                filename = 'x' + str(x) + '_y_' +str(y) + '_tissue.tif'
                tile_image = io.imread(filename)

                start_x = x * x_frame_size
                end_x = start_x + x_frame_size
                start_y = y * 2960
                end_y = start_y + 2960

                super_image[start_y:end_y, start_x:end_x] = tile_image

        labelled_super = skimage.measure.label(super_image)
        props = skimage.measure.regionprops(labelled_super)

        # make array to store cluster area and indicies in
        cluster_area_index = np.zeros((6, np.max(labelled_super)))
        '''
        for index in range(0, np.max(labelled_super)):
            area = props[index]['area']
            print(area, index)
            array_min = np.min(cluster_area_index[0])
            if area > array_min:
                min_index = np.where(cluster_area_index[0] == array_min)[0][0]
                cluster_area_index[0][min_index] = area
                cluster_area_index[1][min_index] = int(index + 1)
            else:
                pass
        '''

        for index in range(0, np.max(labelled_super)):
            area = props[index]['area']
            centroid = props[index]['centroid']
            bbox = props[index]['bbox']
            bbox_y_length = bbox[2] - bbox[0]
            bbox_x_length = bbox[3] - bbox[1]

            cluster_area_index[0][index] = area
            cluster_area_index[1][index] = int(index + 1)
            cluster_area_index[2][index] = centroid[0]
            cluster_area_index[3][index] = centroid[1]
            cluster_area_index[4][index] = bbox_y_length
            cluster_area_index[5][index] = bbox_x_length

        #sort array by size and keep index tracked alongside the size sorting

        sorted_cluster_areas = deepcopy(np.sort(cluster_area_index, kind='stable'))

        x = 0
        while x < np.shape(cluster_area_index)[1]:
            area = sorted_cluster_areas[0][x]
            x_index = np.where(cluster_area_index[0] == area)[0]
            cluster_index = cluster_area_index[1][x_index]
            centroid_y_value = cluster_area_index[2][x_index]
            centroid_x_value = cluster_area_index[3][x_index]
            bbox_y_length = cluster_area_index[4][x_index]
            bbox_x_length = cluster_area_index[5][x_index]
            sorted_cluster_areas[1][x:x + np.shape(x_index)[0]] = cluster_index
            sorted_cluster_areas[2][x:x + np.shape(x_index)[0]] = centroid_y_value
            sorted_cluster_areas[3][x:x + np.shape(x_index)[0]] = centroid_x_value
            sorted_cluster_areas[4][x:x + np.shape(x_index)[0]] = bbox_y_length
            sorted_cluster_areas[5][x:x + np.shape(x_index)[0]] = bbox_x_length
            x += np.shape(x_index)[0]

        sorted_cluster_areas = np.fliplr(sorted_cluster_areas)

        #determine how many clusters will pass with threshold based on smallest demanded retained cluster
        min_area = area_threshold * sorted_cluster_areas[0][number_clusters_retained - 1]
        sorted_cluster_areas[0][sorted_cluster_areas[0] < min_area] = 0

        try:
            index_smallest = np.where(sorted_cluster_areas[0] == 0)[0][0]
        except:
            index_smallest = int(number_clusters_retained)

        print(index_smallest)

        sorted_cluster_areas = sorted_cluster_areas[::, 0:index_smallest]
        sorted_y_centroid = sorted_cluster_areas[5]
        sorted_x_centroid = sorted_cluster_areas[4]

        number_actual_clusters_retained = np.shape(sorted_cluster_areas)[1]
        print(number_actual_clusters_retained)

        # make new labelled image with desired clusters retain and renumbered 1 through x
        new_labelled_image = labelled_super
        #new_labelled_image = labelled_super * super_image
        # scale all clusters to same number
        for x in range(0, number_actual_clusters_retained):
            index_value = sorted_cluster_areas[1][x]
            new_labelled_image[new_labelled_image == index_value] = 65535 - x
            sorted_cluster_areas[1][x] = number_actual_clusters_retained - x
            print(sorted_cluster_areas[0][x], number_actual_clusters_retained - x)
        new_labelled_image[new_labelled_image < (65535 - number_actual_clusters_retained -1)] = 0
        new_labelled_image[np.nonzero(new_labelled_image)] = new_labelled_image[np.nonzero(new_labelled_image)] - (65535 - number_actual_clusters_retained)

        #check to see if labelled image exists in folder. If it does, load in and use
        #if not, remake. Delete image in folder if you want to remake

        labelled_image_path = file_path + '/labelled_tissue_filtered.tif'

        if os.path.isfile(labelled_image_path) == True:
            os.chdir(file_path)
            filename = r'labelled_tissue_filtered.tif'
            new_labelled_image = io.imread(filename)
            new_labelled_image -= np.min(new_labelled_image)
        else:
            new_labelled_image = self.cluster_neighborhood(new_labelled_image, sorted_cluster_areas)
            new_labelled_image =  new_labelled_image.astype('uint16')
            tf.imwrite(r'labelled_tissue_filtered.tif', new_labelled_image, compression='zlib', compressionargs={'level': 10})



        #make new binary image
        new_image = copy.deepcopy(new_labelled_image)
        new_image[new_image > 0] = 1

        # fill small holes
        #new_image = skimage.morphology.remove_small_holes(new_image, area_threshold=5000)

        for y in range(0, y_tile_count):
            for x in range(0, x_tile_count):
                filename = 'x' + str(x) + '_y_' + str(y) + '_tissue.tif'
                label_filename = 'x' + str(x) + '_y_' + str(y) + '_label_tissue.tif'

                start_x = x * x_frame_size
                end_x = start_x + x_frame_size
                start_y = y * 2960
                end_y = start_y + 2960

                tile_image = new_image[start_y:end_y, start_x:end_x]
                tile_image = skimage.util.img_as_uint(tile_image)
                tile_image = tile_image/65535

                tile_label_image = new_labelled_image[start_y:end_y, start_x:end_x]
                tile_label_image = tile_label_image.astype('uint16')

                io.imsave(filename, tile_image)
                io.imsave(label_filename, tile_label_image)

        super_image = skimage.util.img_as_uint(super_image)
        new_image = skimage.util.img_as_uint(new_image)

        io.imsave('whole_tissue.tif', super_image)
        io.imsave('whole_tissue_filtered.tif', new_image)

    def tissue_binary_generate(self, experiment_directory, x_frame_size = 2960, clusters_retained = 1, area_threshold = 0.25):
        '''
        Generates tissue binary maps from star dist binary maps

        :param experiment_directory:
        :return:
        '''

        # load numpy arrays in
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]

        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]

        star_dist_path = experiment_directory + '/Labelled_Nuc'
        tissue_path = experiment_directory + '/Tissue_Binary'

        try:
            os.chdir(tissue_path)
        except:
            os.chdir(experiment_directory)
            os.mkdir('Tissue_Binary')

        foot_print = morphology.disk(100, decomposition='sequence')


        for x in range(0, x_tile_count):
            for y in range(0, y_tile_count):

                os.chdir(star_dist_path)
                star_dist_filename =  'x' + str(x) + '_y_' + str(y) + '_c_DAPI.tif'
                star_dist_im = io.imread(star_dist_filename)

                tissue_binary_im = morphology.binary_dilation(star_dist_im, foot_print)
                #tissue_binary_im = morphology.binary_erosion(tissue_binary_im, foot_print)
                #tissue_binary_im = morphology.binary_erosion(tissue_binary_im, foot_print)
                #tissue_binary_im = morphology.binary_dilation(tissue_binary_im, foot_print)

                tissue_binary_im = tissue_binary_im.astype(np.uint8)
                filtered_image = self.tissue_filter(tissue_binary_im)
                os.chdir(tissue_path)
                tissue_binary_name = 'x' + str(x) + '_y_' + str(y) + '_tissue.tif'
                io.imsave(tissue_binary_name, filtered_image)


        self.tissue_cluster_filter(experiment_directory, x_frame_size, clusters_retained, area_threshold=area_threshold)

    def cluster_neighborhood(self, image, sorted_cluster_areas):
        '''
        Peforms series of erosion events in order to establish how close neighboring clusters are to each other
        :param image:
        :param total_dilation: dilation step size x # steps (rounded to nearest integer)
        :param dilation_step_size:
        :return:
        '''

        threshold = 2000 # in pixels
        sorted_y_centroid = sorted_cluster_areas[2]
        sorted_x_centroid = sorted_cluster_areas[3]
        sorted_index = sorted_cluster_areas[1]
        sort_y_length = sorted_cluster_areas[4]
        sort_x_length = sorted_cluster_areas[5]
        number_clusters = np.max(image)

        neighborhood_matrix = np.zeros((number_clusters + 1,number_clusters + 1))
        neighborhood_matrix[0][1:number_clusters + 1] = np.linspace(1,number_clusters, number_clusters)
        neighborhood_matrix[1:number_clusters + 1, 0] = np.linspace(1,number_clusters, number_clusters)
        combos = list(itertools.combinations(np.linspace(1, number_clusters, number_clusters), 2))

        unit_y_axis_vector = [1,0]
        unique_labels = np.linspace(1, number_clusters, number_clusters)
        new_cluster_count = number_clusters

        for combo in combos:
            first_cluster_index = np.where(sorted_index == int(combo[0]))[0][0]
            second_cluster_index = np.where(sorted_index == int(combo[1]))[0][0]
            y1 = sorted_y_centroid[first_cluster_index]
            y2 = sorted_y_centroid[second_cluster_index]
            x1 = sorted_x_centroid[first_cluster_index]
            x2 = sorted_x_centroid[second_cluster_index]
            y_len1 = sort_y_length[first_cluster_index]
            x_len1= sort_x_length[first_cluster_index]
            y_len2= sort_y_length[second_cluster_index]
            x_len2= sort_x_length[second_cluster_index]

            center_center_magnitude = np.sqrt((y2-y1)**2 + (x2-x1)**2)
            center_center_vector = [(y2-y1), (x2-x1)]
            dot = np.dot(center_center_vector, unit_y_axis_vector)/center_center_magnitude
            dot = np.abs(dot)
            angle = np.arccos(dot)

            max_angle1 = np.arctan(x_len1/y_len1)
            max_angle2 = np.arctan(x_len2/y_len2)


            if angle < max_angle1:
                dist_2_edge1 = (y_len1/2)/np.cos(angle)
            elif angle >= max_angle1:
                dist_2_edge1 = (x_len1/2)/np.sin(angle)

            if angle < max_angle2:
                dist_2_edge2 = (y_len2/2)/np.cos(angle)
            elif angle >= max_angle2:
                dist_2_edge2 = (x_len2/2)/np.sin(angle)


            net_distance = center_center_magnitude - dist_2_edge1 - dist_2_edge2

            neighborhood_matrix[int(combo[0])][int(combo[1])] = net_distance
            neighborhood_matrix[int(combo[1])][int(combo[0])] = net_distance

        for x in range(1, number_clusters + 1):
            for y in range(x + 1, number_clusters + 1):
                net_distance = neighborhood_matrix[y][x]

                if net_distance < threshold:

                    #cluster_reference_index = int(neighborhood_matrix[0][x])
                    #cluster_2_be_merged_index = int(neighborhood_matrix[y][0])
                    #remove unique label from list
                    try:
                        unique_second_cluster_index = np.where(unique_labels == neighborhood_matrix[y][0])[0][0]
                        unique_labels = np.delete(unique_labels, unique_second_cluster_index)
                    except:
                        pass

                    #make label in image match
                    image[image == neighborhood_matrix[y][0]] = neighborhood_matrix[0][x]

                    #alter label number to be that of the starting index on x axis
                    neighborhood_matrix[y][0] = neighborhood_matrix[0][x]
                    neighborhood_matrix[0][y] = neighborhood_matrix[0][x]
                    
                    #remove a cluster count as it merged with another cluster
                    new_cluster_count -= 1

                else:
                    pass


        #renumber clusters to be 1 int steps and continuous starting from 1
        desired_numbers = np.linspace(1, new_cluster_count, new_cluster_count)
        not_used_desired_numbers = np.setdiff1d(desired_numbers, unique_labels)
        labels_to_be_replaced  = np.setdiff1d(unique_labels, desired_numbers)
        if len(not_used_desired_numbers) > 0:
            for x in range(0, len(not_used_desired_numbers)):
                image[image==labels_to_be_replaced[x]] = not_used_desired_numbers[x]
        else:
            pass


        io.imshow(image)
        io.show()


        return image


    #recursize autofocusfunctions#####################################

    def highest_index(self, score_array):


        middle_index = (np.shape(score_array)[0] - 1)/2
        middle_index = int(middle_index)

        max_score = np.max(score_array)
        index = np.where(score_array == max_score)
        try:
            index = index[0][0]
        except:
            index = (np.shape(score_array)[0] - 1) / 2
            index = int(index)

        return index

    def generate_nuc_mask(self, experiment_directory, cycle_number):

        model = StarDist2D.from_pretrained('2D_versatile_fluo')

        # load numpy arrays in
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)
        tissue_fm = full_array[10]

        numpy_x = full_array[0]
        numpy_y = full_array[1]

        z_center_index = math.floor(full_array[3][0][0]/2)

        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]
        if cycle_number == 0:
            dapi_im_path = experiment_directory + '/' + 'DAPI' '\Bleach\cy_' + str(cycle_number) + r'\Tiles'
        elif cycle_number == 1:
            dapi_im_path = experiment_directory + '/' + 'DAPI' '\Bleach\cy_' + str(cycle_number - 1) + r'\Tiles'
        else:
            dapi_im_path = experiment_directory + '/' + 'DAPI' '\Stain\cy_' + str(cycle_number - 1) + r'\Tiles'
        os.chdir(experiment_directory)
        try:
            os.mkdir('Labelled_Nuc')
        except:
            pass

        labelled_path = experiment_directory + '/Labelled_Nuc'

        for x in range(0, x_tile_count):
            for y in range(0, y_tile_count):
                if tissue_fm[y][x] > 1:
                    os.chdir(dapi_im_path)
                    file_name = 'z_' + str(z_center_index) + '_x' + str(x) + '_y_' + str(y) + '_c_DAPI.tif'
                    labelled_file_name = 'x' + str(x) + '_y_' + str(y) + '_c_DAPI.tif'
                    img = io.imread(file_name)
                    img = img.astype('int32')
                    img[img < 0] = 0
                    #img = skimage.util.img_as_uint(img)
                    labels, _ = model.predict_instances(normalize(img))
                    labels[labels > 0] = 1

                    os.chdir(labelled_path)
                    io.imsave(labelled_file_name, labels)
                else:
                    pass

    def recursive_stardist_autofocus(self, experiment_directory, cycle, slice_gap=2, remake_nuc_binary = 0):
        '''
        Finds center of dapi z stack for each tile and updates focus map to center it.
        Cycle 1 uses cycle 0 dapi z stacks. So cycle n-1 informs focus map for cycle n.
        A pretrained network for stardist is used to screen out junk in images and increase
        reliability. Note, must have pregenerated binary nuclei images available to use.
        See generate_nuc_mask function to make them. This automakes if cycle = 0.

        :param experimental_directory:
        :param cycle:
        :return:
        '''
        starting_time = time.time()

        model = StarDist2D.from_pretrained('2D_versatile_fluo')

        labelled_path = experiment_directory + '/Labelled_Nuc'
        if cycle == 0:
            dapi_im_path = experiment_directory + '/' + 'DAPI' '\Bleach\cy_' + str(0) + '\Tiles'
        elif cycle == 1:
            dapi_im_path = experiment_directory + '/' + 'DAPI' '\Bleach\cy_' + str(cycle - 1) + '\Tiles'
        else:
            dapi_im_path = experiment_directory + '/' + 'DAPI' '\Stain\cy_' + str(cycle - 1) + '\Tiles'


        if remake_nuc_binary == 1:
            self.generate_nuc_mask(experiment_directory, cycle)
        else:
            pass

        #if cycle == 0:
            #self.tissue_region_identifier(experiment_directory, x_frame_size=2960, clusters_retained=1)
        #else:
        #    pass


        # load numpy arrays in (focus map)
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        fm_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = fm_array[0]
        numpy_y = fm_array[1]
        tissue_fm = fm_array[10]

        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]
        #y_tile_count = 1
        #x_tile_count = 1
        z_count = int(fm_array[3][0][0])
        z_middle = math.floor(z_count / 2)  # if odd, z_count will round down. Since index counts from 0, it is the middle

        step_size = 17  # brenner score step size
        x_axis = np.linspace(0,z_count, z_count)


        # make numpy array to hold scores in for each tile
        score_array = np.zeros(z_count)

        stain_time = 0

        # iterate through tiles and find index of slice most in focus
        for x in range(0, x_tile_count):
            for y in range(0, y_tile_count):

                if cycle != 0:
                    current_time = time.time()
                    new_stain_time = math.floor((current_time - starting_time)/60)
                    if new_stain_time != stain_time:
                        print('Staining Time Elapsed ', new_stain_time)
                        stain_time = new_stain_time
                    else:
                        pass

                if tissue_fm[y][x] > 1:

                    for z in range(0, z_count):
                        # load in binary image mask
                        #os.chdir(labelled_path)
                        #file_name = 'x' + str(x) + '_y_' + str(y) + '_c_DAPI.tif'
                        #labels = io.imread(file_name)
                        # load in z slice
                        os.chdir(dapi_im_path)
                        file_name = 'z_' + str(z) + '_x' + str(x) + '_y_' + str(y) + '_c_DAPI.tif'
                        img = io.imread(file_name)

                        #apply stardist to image
                        if remake_nuc_binary == 1:
                            labels, _ = model.predict_instances(normalize(img))
                            labels[labels > 0] = 1
                        else:
                            os.chdir(labelled_path)
                            file_name = 'x' + str(x) + '_y_' + str(y) + '_c_DAPI.tif'
                            labels = io.imread(file_name)


                        # apply mask to image and find brenner score
                        score = self.focus_score(img, step_size, labels)
                        score_array[z] = score
                    # find highest score slice index and find shift amount
                    focus_index = self.highest_index(score_array)
                    print('cycle',cycle,'x', x, 'y', y)
                    #plt.scatter(x_axis, score_array)
                    #plt.show()
                    center_focus_index_difference = int(focus_index - z_middle)
                    print(center_focus_index_difference)
                    new_fm_z_position = center_focus_index_difference * slice_gap
                    # update focus map for all channels
                    fm_array[2][y][x] = fm_array[2][y][x] + new_fm_z_position

                if tissue_fm[y][x] == 1:
                    pass

        # save updated focus array
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        np.save('fm_array.npy', fm_array)

        ending_time = time.time()
        total_time_elapsed = ending_time - starting_time
        total_time_elapsed = int(total_time_elapsed)

        return total_time_elapsed

    def reacquire_run_autofocus(self, experiment_directory, cycle, z_slices, offset_array, x_frame_size):
        '''
        Acquires dapi images for a cycle and then runs recursive_stardist_autofocus on them with the previously generated masks
        :param experiment_directory:
        :param cycle:
        :param z_slices:
        :param offset_array:
        :param x_frame_size:
        :return:
        '''

        self.image_cycle_acquire(cycle, experiment_directory, z_slices, 'Stain', offset_array, x_frame_size=x_frame_size,establish_fm_array=0, auto_focus_run=0, auto_expose_run=0, channels=['DAPI'])
        self.recursive_stardist_autofocus(experiment_directory, cycle + 1, remake_nuc_binary = 0)


    ############################################
    # Using core snap and not pycromanager acquire
    ############################################

    def image_capture(self, experiment_directory, channel, exp_time, x, y, z):

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        fm_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = fm_array[0]
        numpy_y = fm_array[1]
        numpy_z = fm_array[2]
        if z < 20:
            z_pos = numpy_z[y][x] - z*2
        else:
            z_pos = numpy_z[y][x]


        core.set_config("Color", channel)
        core.set_xy_position(numpy_x[y][x], numpy_y[y][x])
        core.set_position(z_pos)
        time.sleep(0.5)
        core.set_exposure(int(exp_time))

        core.set_config("amp", 'high')
        core.snap_image()
        core.get_tagged_image()

        core.set_config("amp", 'high')
        core.snap_image()
        tagged_image = core.get_tagged_image()

        pixels = np.reshape(tagged_image.pix, newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
        # time.sleep(1)

        return pixels

    def core_capture(self, experiment_directory, x_frame_size, channel, hdr=1):
        '''
        uses core capture to take images with parameters set by the focus map. Has thr ability to average over frames.
        Will not alter XYZ location, channel or exp time. Only captures and averages frames.

        :param experiment_directory:
        :param channels:
        :return:
        '''

        side_pixel_count = int((5056 - x_frame_size)/2)

        if channel =='DAPI':
            hdr = 0

        if hdr ==0:

            # load in focus map and exp array
            #numpy_path = experiment_directory + '/' + 'np_arrays'
            #os.chdir(numpy_path)
            #fm_array = np.load('fm_array.npy', allow_pickle=False)

            # determine proper frame count to average

            if channel == 'DAPI':
                frame_count_index = 11
            if channel == 'A488':
                frame_count_index = 12
            if channel == 'A555':
                frame_count_index = 13
            if channel == 'A647':
                frame_count_index = 14

            frame_count = 1
            frame_count = int(frame_count)

            # make array to hold images in
            average_array = np.random.rand(frame_count, 2960, x_frame_size).astype('float32')

            # acquire and populate array
            for x in range(0, frame_count):
                time.sleep(0.1)
                core.set_config("amp", 'high')
                core.snap_image()
                tagged_image = core.get_tagged_image()
                pixels = np.reshape(tagged_image.pix,newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
                pixels = np.nan_to_num(pixels, posinf=65500)
                average_array[x] = pixels[::, side_pixel_count:side_pixel_count + x_frame_size]

            #find array average and return averaged image
            averaged_image = np.average(average_array, axis = 0)
        if hdr ==1:

            hdr_times = self.hdr_exp_times
            hdr_frame_count = np.shape(hdr_times)[0]

            # make array to hold images in
            hdr_array = np.random.rand(hdr_frame_count, 2960, x_frame_size).astype('float32')

            # acquire and populate array
            for x in range(0, hdr_frame_count):
                exp_time = int(hdr_times[x])
                core.set_exposure(exp_time)
                core.set_config("amp", 'high')
                core.snap_image()
                tagged_image = core.get_tagged_image()
                pixels = np.reshape(tagged_image.pix,
                                    newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
                pixels = np.nan_to_num(pixels, posinf=65535, nan=65535)
                pixels[pixels > 65535] = 65535
                pixels = pixels.astype('float32')
                hdr_array[x] = pixels[::, side_pixel_count:side_pixel_count + x_frame_size]

            # find array average and return averaged image
            averaged_image = self.hdr_fuser(hdr_array)

        return averaged_image

    def multi_channel_z_stack_capture(self, experiment_directory, cycle_number, Stain_or_Bleach,
                                      x_pixels=5056, slice_gap=2, channels=['DAPI', 'A488', 'A555', 'A647']):
        '''
        Captures and saves all images in XY and Z dimensions. Order of operation is ZC XY(snake). Entire z stack with all
        channels is made into a numpy data structure and saved before going to next tile and being reused. This is done
        to reduce overall memory usage.


        :param experiment_directory: directory that main experiment data is stored in
        :param cycle_number: integer of what cycle number it is. Counts from 0
        :param Stain_or_Bleach: string of Stain or Bleach depending on what action it is
        :param slice_gap: micron spacing in z slices
        :param channels: string list of what channels to acquire
        :return:
        '''

        core = Core()
        hdr_value = self.hdr

        # load in focus map and exp array
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)
        exp_time_array = np.load('exp_array.npy', allow_pickle=False)

        height_pixels = 2960
        width_pixels = x_pixels
        # determine attributes like tile counts,z slices and channel counts
        numpy_x = full_array[0]
        numpy_y = full_array[1]
        tissue_fm = full_array[10]

        side_pixel_count = int((5056 - x_pixels)/2)

        x_tile_count = np.unique(numpy_x).size
        y_tile_count = np.unique(numpy_y).size
        #x_tile_count = 1
        #y_tile_count = 1
        #x_tile_count = 1
        #y_tile_count = 1
        z_slices = full_array[5][0][0]
        # z_slices = 11
        # go to upper left corner to start pattern
        core.set_xy_position(numpy_x[0][0], numpy_y[0][0])
        time.sleep(1)
        # generate numpy data structure
        zc_tif_stack = np.random.rand(4, int(z_slices), height_pixels, width_pixels).astype('float32')

        image_number_counter = 0

        for y in range(0, y_tile_count):
            if y % 2 != 0:
                for x in range(x_tile_count - 1, -1, -1):

                    if tissue_fm[y][x] > 1:

                        core.set_xy_position(numpy_x[y][x], numpy_y[y][x])
                        time.sleep(.5)

                        for channel in channels:

                            # determine the proper indecies to use for focus map z positions and exp array
                            if channel == 'DAPI':
                                channel_index = 2
                                tif_stack_c_index = 0
                                zc_index = 0
                            if channel == 'A488':
                                channel_index = 4
                                tif_stack_c_index = 1
                                zc_index = 1
                            if channel == 'A555':
                                channel_index = 6
                                tif_stack_c_index = 2
                                zc_index = 2
                            if channel == 'A647':
                                channel_index = 8
                                tif_stack_c_index = 3
                                zc_index = 3

                            numpy_z = full_array[channel_index]
                            exp_time = int(exp_time_array[tif_stack_c_index])
                            core.set_config("Color", channel)
                            core.set_exposure(exp_time)

                            #burner image due to defect that makes signal 8% higher int he first one
                            core.set_config("amp", 'high')
                            core.snap_image()
                            core.get_tagged_image()

                            z_end = int(numpy_z[y][x]) + slice_gap
                            z_start = int(z_end - (z_slices) * slice_gap)
                            print(z_start, z_end, slice_gap)


                            z_counter = 0

                            for z in range(z_start, z_end, slice_gap):
                                core.set_position(z)
                                time.sleep(0.05)
                                pixels = self.core_capture(experiment_directory,x_pixels, channel, hdr=hdr_value)
                                zc_tif_stack[zc_index][z_counter] = pixels

                                image_number_counter += 1
                                z_counter += 1

                        # save zc stack
                        self.zc_save(zc_tif_stack, channels, x, y, cycle_number, x_pixels, experiment_directory,
                                     Stain_or_Bleach)

                    if tissue_fm[y][x] == 1:
                        pass


            elif y % 2 == 0:
                for x in range(0, x_tile_count):

                    if tissue_fm[y][x] > 1:

                        #print('x', numpy_x[y][x], 'y', numpy_y[y][x])

                        core.set_xy_position(numpy_x[y][x], numpy_y[y][x])
                        time.sleep(.5)

                        for channel in channels:

                            # determine the proper indecies to use for focus map z positions and exp array
                            if channel == 'DAPI':
                                channel_index = 2
                                tif_stack_c_index = 0
                                zc_index = 0
                            if channel == 'A488':
                                channel_index = 4
                                tif_stack_c_index = 1
                                zc_index = 1
                            if channel == 'A555':
                                channel_index = 6
                                tif_stack_c_index = 2
                                zc_index = 2
                            if channel == 'A647':
                                channel_index = 8
                                tif_stack_c_index = 3
                                zc_index = 3

                            numpy_z = full_array[channel_index]
                            exp_time = int(exp_time_array[tif_stack_c_index])
                            core.set_config("Color", channel)
                            core.set_exposure(exp_time)

                            #burner image due to defect that makes signal 8% higher int he first one
                            core.set_config("amp", 'high')
                            core.snap_image()
                            core.get_tagged_image()

                            z_end = int(numpy_z[y][x]) + slice_gap
                            z_start = int(z_end - (z_slices) * slice_gap)
                            z_counter = 0
                            #print('channel', channel, 'z_range', z_start, z_end)

                            for z in range(z_start, z_end, slice_gap):
                                core.set_position(z)
                                time.sleep(0.05)

                                pixels = self.core_capture(experiment_directory, x_pixels, channel, hdr=hdr_value)
                                zc_tif_stack[zc_index][z_counter] = pixels

                                image_number_counter += 1
                                z_counter += 1

                        self.zc_save(zc_tif_stack, channels, x, y, cycle_number, x_pixels, experiment_directory,
                                     Stain_or_Bleach)

                    if tissue_fm[y][x] == 1:
                        pass

        return

    def multi_channel_z_stack_capture_dapi_focus(self, experiment_directory, cycle_number, Stain_or_Bleach,
                                      x_pixels=5056, slice_gap=2, channels=['DAPI', 'A488', 'A555', 'A647']):
        '''
        Captures and saves all images in XY and Z dimensions. Order of operation is ZC XY(snake). Entire z stack with all
        channels is made into a numpy data structure and saved before going to next tile and being reused. This is done
        to reduce overall memory usage. Uses DAPI images from tile to focus and determine focus for other colors. DAPI is
        a multi stack and focused and then acquires one plane for other channels


        :param experiment_directory: directory that main experiment data is stored in
        :param cycle_number: integer of what cycle number it is. Counts from 0
        :param Stain_or_Bleach: string of Stain or Bleach depending on what action it is
        :param slice_gap: micron spacing in z slices
        :param channels: string list of what channels to acquire
        :return:
        '''

        core = Core()
        hdr_value = self.hdr

        # load in focus map and exp array
        numpy_path = experiment_directory + '/' + 'np_arrays'
        dapi_bin_path = experiment_directory + '/' + 'Labelled_Nuc'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)
        exp_time_array = np.load('exp_array.npy', allow_pickle=False)

        height_pixels = 2960
        width_pixels = x_pixels
        # determine attributes like tile counts,z slices and channel counts
        numpy_x = full_array[0]
        numpy_y = full_array[1]
        tissue_fm = full_array[10]

        side_pixel_count = int((5056 - x_pixels)/2)

        x_tile_count = np.unique(numpy_x).size
        y_tile_count = np.unique(numpy_y).size

        z_slices = full_array[5][0][0]
        z_slices_dapi = full_array[3][0][0]
        # go to upper left corner to start pattern
        core.set_xy_position(numpy_x[0][0], numpy_y[0][0])
        time.sleep(1)
        # generate numpy data structures
        zc_tif_stack = np.random.rand(4, int(z_slices), height_pixels, width_pixels).astype('float32')
        zc_dapi_tif_stack = np.random.rand(int(z_slices_dapi), height_pixels, width_pixels).astype('float32')

        image_number_counter = 0

        for y in range(0, y_tile_count):
            if y % 2 != 0:
                for x in range(x_tile_count - 1, -1, -1):

                    if tissue_fm[y][x] > 1:

                        core.set_xy_position(numpy_x[y][x], numpy_y[y][x])
                        time.sleep(.5)

                        for channel in channels:

                            # determine the proper indecies to use for focus map z positions and exp array
                            if channel == 'DAPI':
                                channel_index = 2
                                tif_stack_c_index = 0
                                zc_index = 0
                            if channel == 'A488':
                                channel_index = 4
                                tif_stack_c_index = 1
                                zc_index = 1
                            if channel == 'A555':
                                channel_index = 6
                                tif_stack_c_index = 2
                                zc_index = 2
                            if channel == 'A647':
                                channel_index = 8
                                tif_stack_c_index = 3
                                zc_index = 3

                            z_slices = int(full_array[channel_index + 1][0][0])
                            print(channel, z_slices)

                            if channel == 'DAPI':

                                os.chdir(dapi_bin_path)
                                tissue_name = 'x' + str(x) + '_y_' + str(y) + '_c_DAPI.tif'
                                try:
                                    tissue_im = io.imread(tissue_name)
                                except:
                                    tissue_im = np.ones(2960, x_pixels)

                                numpy_z = full_array[channel_index]
                                exp_time = int(exp_time_array[tif_stack_c_index])
                                core.set_config("Color", channel)
                                core.set_exposure(exp_time)

                                # burner image due to defect that makes signal 8% higher int he first one
                                core.set_config("amp", 'high')
                                core.snap_image()
                                core.get_tagged_image()

                                z_end = int(numpy_z[y][x]) + slice_gap
                                z_start = int(z_end - (z_slices) * slice_gap)

                                z_counter = 0

                                #populate dapi z stack

                                for z in range(z_start, z_end, slice_gap):
                                    core.set_position(z)
                                    time.sleep(0.05)
                                    pixels = self.core_capture(experiment_directory, x_pixels, channel, hdr=hdr_value)
                                    zc_dapi_tif_stack[z_counter] = pixels

                                    image_number_counter += 1
                                    z_counter += 1

                                #score each z slice
                                scores = []
                                for z in range(0, z_slices):

                                    image_slice = zc_dapi_tif_stack[z]
                                    score = self.focus_score(image_slice, 17, tissue_im)
                                    scores.append(score)

                                focus_index = self.highest_index(scores)
                                z_middle = (z_slices-1)/2 #must be odd

                                center_focus_index_difference = int(focus_index - z_middle)
                                new_fm_z_position = center_focus_index_difference * slice_gap
                                # update focus map for all channels
                                full_array[2][y][x] += new_fm_z_position
                                full_array[4][y][x] += new_fm_z_position
                                full_array[6][y][x] += new_fm_z_position
                                full_array[8][y][x] += new_fm_z_position
                                print('x', x, 'y', y, 'z', z)

                                zc_tif_stack[zc_index][0] = zc_dapi_tif_stack[focus_index]

                            else:

                                numpy_z = full_array[channel_index]
                                exp_time = int(exp_time_array[tif_stack_c_index])
                                core.set_config("Color", channel)
                                core.set_exposure(exp_time)

                                #burner image due to defect that makes signal 8% higher int he first one
                                core.set_config("amp", 'high')
                                core.snap_image()
                                core.get_tagged_image()

                                z_end = int(numpy_z[y][x]) + slice_gap
                                z_start = int(z_end - (z_slices) * slice_gap)



                                z_counter = 0

                                for z in range(z_start, z_end, slice_gap):
                                    print(z)
                                    core.set_position(z)
                                    time.sleep(0.05)
                                    pixels = self.core_capture(experiment_directory,x_pixels, channel, hdr=hdr_value)
                                    zc_tif_stack[zc_index][z_counter] = pixels

                                    image_number_counter += 1
                                    z_counter += 1

                        # save zc stack
                        self.zc_save(zc_tif_stack, channels, x, y, cycle_number, x_pixels, experiment_directory,
                                         Stain_or_Bleach)

                    if tissue_fm[y][x] == 1:
                        pass


            elif y % 2 == 0:
                for x in range(0, x_tile_count):

                    if tissue_fm[y][x] > 1:

                        core.set_xy_position(numpy_x[y][x], numpy_y[y][x])
                        time.sleep(.5)

                        for channel in channels:

                            # determine the proper indecies to use for focus map z positions and exp array
                            if channel == 'DAPI':
                                channel_index = 2
                                tif_stack_c_index = 0
                                zc_index = 0
                            if channel == 'A488':
                                channel_index = 4
                                tif_stack_c_index = 1
                                zc_index = 1
                            if channel == 'A555':
                                channel_index = 6
                                tif_stack_c_index = 2
                                zc_index = 2
                            if channel == 'A647':
                                channel_index = 8
                                tif_stack_c_index = 3
                                zc_index = 3

                            z_slices = int(full_array[channel_index + 1][0][0])

                            if channel == 'DAPI':

                                os.chdir(dapi_bin_path)
                                tissue_name = 'x' + str(x) + '_y_' + str(y) + '_c_DAPI.tif'
                                try:
                                    tissue_im = io.imread(tissue_name)
                                except:
                                    tissue_im = np.ones(2960, x_pixels)

                                numpy_z = full_array[channel_index]
                                exp_time = int(exp_time_array[tif_stack_c_index])
                                core.set_config("Color", channel)
                                core.set_exposure(exp_time)

                                # burner image due to defect that makes signal 8% higher int he first one
                                core.set_config("amp", 'high')
                                core.snap_image()
                                core.get_tagged_image()

                                z_end = int(numpy_z[y][x]) + slice_gap
                                z_start = int(z_end - (z_slices) * slice_gap)

                                z_counter = 0

                                # populate dapi z stack

                                for z in range(z_start, z_end, slice_gap):
                                    core.set_position(z)
                                    time.sleep(0.05)
                                    pixels = self.core_capture(experiment_directory, x_pixels, channel, hdr=hdr_value)
                                    zc_dapi_tif_stack[z_counter] = pixels

                                    image_number_counter += 1
                                    z_counter += 1

                                # score each z slice
                                scores = []
                                for z in range(0, z_slices):
                                    image_slice = zc_dapi_tif_stack[z]
                                    score = self.focus_score(image_slice, 17, tissue_im)
                                    scores.append(score)

                                focus_index = self.highest_index(scores)
                                z_middle = (z_slices - 1) / 2  # must be odd

                                center_focus_index_difference = int(focus_index - z_middle)
                                new_fm_z_position = center_focus_index_difference * slice_gap
                                # update focus map for all channels
                                full_array[2][y][x] += new_fm_z_position
                                full_array[4][y][x] += new_fm_z_position
                                full_array[6][y][x] += new_fm_z_position
                                full_array[8][y][x] += new_fm_z_position

                                print('x', x, 'y', y, 'z', z)

                                zc_tif_stack[zc_index][0] = zc_dapi_tif_stack[focus_index]

                            else:

                                numpy_z = full_array[channel_index]
                                exp_time = int(exp_time_array[tif_stack_c_index])
                                core.set_config("Color", channel)
                                core.set_exposure(exp_time)

                                # burner image due to defect that makes signal 8% higher int he first one
                                core.set_config("amp", 'high')
                                core.snap_image()
                                core.get_tagged_image()

                                z_end = int(numpy_z[y][x]) + slice_gap
                                z_start = int(z_end - (z_slices) * slice_gap)


                                z_counter = 0

                                for z in range(z_start, z_end, slice_gap):
                                    print(z)
                                    core.set_position(z)
                                    time.sleep(0.05)
                                    pixels = self.core_capture(experiment_directory, x_pixels, channel, hdr=hdr_value)
                                    zc_tif_stack[zc_index][z_counter] = pixels

                                    image_number_counter += 1
                                    z_counter += 1

                        # save zc stack
                        self.zc_save(zc_tif_stack, channels, x, y, cycle_number, x_pixels, experiment_directory,
                                     Stain_or_Bleach)

                    if tissue_fm[y][x] == 1:
                        pass

        return

    def image_cycle_acquire(self, cycle_number, experiment_directory, z_slices, stain_bleach, offset_array, x_frame_size=5056, fm_array_adjuster = 0, establish_fm_array=0, auto_focus_run=0, auto_expose_run=0,
                            channels=['DAPI', 'A488', 'A555', 'A647'], focus_position = 'none'):

        self.establish_fm_array(experiment_directory, cycle_number, z_slices, offset_array,
                                initialize=establish_fm_array, x_frame_size=x_frame_size, fm_array_adjuster= fm_array_adjuster, autofocus=auto_focus_run,
                                auto_expose=auto_expose_run, focus_position = focus_position)


        #self.image_capture(experiment_directory, 'DAPI', 50, 0, 0, 0)  # wake up lumencor light engine


        '''
        std_dev = np.std(start_image)
        if std_dev > threshold:
            pass
        else: 
            time.sleep(10)
        '''
        #time.sleep(3)  # wait for it to wake up
        ''''
        exp_time = exp_time_array
        np.save('exp_array.npy', exp_time)

        for channel in channels:
            z_tile_stack = self.core_tile_acquire(experiment_directory, channel)
            self.save_files(z_tile_stack, channel, cycle_number, experiment_directory, stain_bleach)

        '''
        self.fm_map_z_shifter(experiment_directory, 3, 3)
        self.exp_logbook(experiment_directory, cycle_number)
        start = time.time()
        #self.multi_channel_z_stack_capture_dapi_focus(experiment_directory, cycle_number, stain_bleach,x_pixels=x_frame_size, slice_gap=2, channels=channels)
        self.multi_channel_z_stack_capture(experiment_directory, cycle_number, stain_bleach,x_pixels=x_frame_size, slice_gap=2, channels=channels)
        end = time.time()
        print('acquistion time', end - start)
        # self.marker_excel_file_generation(experiment_directory, cycle_number)

    def tilt_determination(self):

        current_z = core.get_position()
        number_slices = 25
        z_gap = 1
        bottom_z = current_z - int(math.floor(9/2)) * z_gap

        frame_x_start = 1048
        frame_x_end = frame_x_start + 2960

        sides_scoring = 370


        images = np.zeros((number_slices,2960, 2960))
        mask = np.ones((sides_scoring,sides_scoring))

        for slice in range(0, number_slices):

            z = bottom_z + slice*z_gap
            core.set_position(z)
            time.sleep(0.2)
            core.set_config("amp", 'high')
            core.snap_image()
            tagged_image = core.get_tagged_image()
            pixels = np.reshape(tagged_image.pix, newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
            pixels = np.nan_to_num(pixels, posinf=65500)
            images[slice] = pixels[::, frame_x_start:frame_x_end]

        core.set_position(current_z)

        os.chdir(r'E:\5-9-24 gutage\tilt')
        io.imsave('tilt.tif', images)


        #find focus scores

        scores = np.zeros((5, number_slices))

        for region in range(0, 5):

            if region == 0:
                y_start = 0
                y_end = y_start + sides_scoring
                x_start =0
                x_end = x_start + sides_scoring
            if region == 1:
                y_start = 0
                y_end = y_start + sides_scoring
                x_end = -1
                x_start = x_end - sides_scoring
            if region == 2:
                y_end = -1
                y_start = y_end - sides_scoring
                x_start = 0
                x_end = x_start + sides_scoring
            if region == 3:
                y_end = -1
                y_start = y_end - sides_scoring
                x_end = -1
                x_start = x_end - sides_scoring
            if region == 4:
                y_start = 1480 - sides_scoring
                y_end = y_start + sides_scoring
                x_start = 1480 - sides_scoring
                x_end = x_start + sides_scoring
            else:
                pass

            for slice in range(0, number_slices):
                image = images[slice]
                sub_image = image[y_start:y_end, x_start:x_end]
                #print(np.shape(sub_image), 'region', region)
                scores[region][slice] = self.focus_score(sub_image, 17, mask)


            highest_value = np.max(scores[region])
            highest_index = np.where(scores[region] == highest_value)[0][0]
            print('region', region, 'index', highest_index + 1)

    def tilt_overlap_determination(self):

        current_z = core.get_position()
        number_slices = 25
        z_gap = 1
        bottom_z = current_z - int(math.floor(9/2)) * z_gap

        starting_y = core.get_xy_stage_position().x
        starting_x = core.get_xy_stage_position().y
        displacement_magnitude = 532

        position_list = np.zeros((5,2))
        position_list[0][0] = starting_x
        position_list[0][1] = starting_y

        position_list[1][0] = starting_x
        position_list[1][1] = starting_y - displacement_magnitude

        position_list[2][0] = starting_x - displacement_magnitude
        position_list[2][1] = starting_y

        position_list[3][0] = starting_x - displacement_magnitude
        position_list[3][1] = starting_y - displacement_magnitude

        position_list[4][0] = starting_x - displacement_magnitude/2
        position_list[4][1] = starting_y- displacement_magnitude/2


        frame_x_start = 1048
        frame_x_end = frame_x_start + 2960

        sides_scoring = 370



        images = np.zeros((5, number_slices,2960, 2960))
        mask = np.ones((sides_scoring,sides_scoring))




        for j in range(0, 5):
            core.set_xy_position(position_list[j][1], position_list[j][0])
            time.sleep(1)
            for slice in range(0, number_slices):

                z = bottom_z + slice*z_gap
                core.set_position(z)
                time.sleep(0.2)
                core.set_config("amp", 'high')
                core.snap_image()
                tagged_image = core.get_tagged_image()
                pixels = np.reshape(tagged_image.pix, newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
                pixels = np.nan_to_num(pixels, posinf=65500)
                images[j][slice] = pixels[::, frame_x_start:frame_x_end]

        core.set_position(current_z)
        core.set_xy_position(position_list[0][1], position_list[0][0])
        os.chdir(r'E:\23-8-24 gutage\tilt')
        io.imsave('tilt_0.tif', images[0])
        io.imsave('tilt_1.tif', images[1])
        io.imsave('tilt_2.tif', images[2])
        io.imsave('tilt_3.tif', images[3])
        io.imsave('tilt_4.tif', images[4])

        #find focus scores

        scores = np.zeros((5, number_slices))

        for region in range(0, 5):

            if region == 0:
                y_start = 0
                y_end = y_start + sides_scoring
                x_start =0
                x_end = x_start + sides_scoring
            if region == 1:
                y_start = 0
                y_end = y_start + sides_scoring
                x_end = -1
                x_start = x_end - sides_scoring
            if region == 2:
                y_end = -1
                y_start = y_end - sides_scoring
                x_start = 0
                x_end = x_start + sides_scoring
            if region == 3:
                y_end = -1
                y_start = y_end - sides_scoring
                x_end = -1
                x_start = x_end - sides_scoring
            if region == 4:
                y_start = 1480 - sides_scoring
                y_end = y_start + sides_scoring
                x_start = 1480 - sides_scoring
                x_end = x_start + sides_scoring
            else:
                pass

            for slice in range(0, number_slices):
                image = images[region][slice]
                sub_image = image[y_start:y_end, x_start:x_end]
                #print(np.shape(sub_image), 'region', region)
                scores[region][slice] = self.focus_score(sub_image, 17, mask)


            highest_value = np.max(scores[region])
            highest_index = np.where(scores[region] == highest_value)[0][0]
            print('region', region, 'index', highest_index + 1)

    def wide_net_auto_focus(self, experiment_directory, x_frame_size, offset_array, z_slice_search_range, focus_position, number_clusters_retained = 6, manual_cluster_update = 0):

        #make parent folder for experiment if it isnt made
        os.chdir(r'E:')
        try:
            os.mkdir(experiment_directory)
        except:
            pass

        if manual_cluster_update == 0:
            z_wide_range = z_slice_search_range

            self.image_cycle_acquire(0, experiment_directory,z_wide_range, 'Bleach', offset_array, x_frame_size=x_frame_size,establish_fm_array=1, auto_focus_run=0, auto_expose_run=0, channels=['DAPI'],focus_position=focus_position)
            self.generate_nuc_mask(experiment_directory, 0)
            self.tissue_region_identifier(experiment_directory, x_frame_size = x_frame_size, clusters_retained=number_clusters_retained)

        if manual_cluster_update == 1:
            self.tissue_region_identifier(experiment_directory, x_frame_size=x_frame_size, clusters_retained=number_clusters_retained)


    def initialize(self, experiment_directory, offset_array, z_slices, x_frame_size=2960, focus_position = 'none'):
        '''initialization section. Takes DAPI images, cluster filters tissue, minimally frames sampling grid, acquires all channels.

        :param experiment_directory:
        :param cycle_number:
        :param offset_array:
        :param z_slices:
        :param x_frame_size:
        :param focus_position:
        :return:
        '''

        #make parent folder for experiment if it isnt made
        os.chdir(r'E:')
        try:
            os.mkdir(experiment_directory)
        except:
            pass

        self.recursive_stardist_autofocus(experiment_directory, cycle=0, remake_nuc_binary=0)
        self.fm_map_z_shifter(experiment_directory, desired_z_slices_dapi=3, desired_z_slices_other=3)
        self.image_cycle_acquire(0, experiment_directory, z_slices, 'Bleach', offset_array, x_frame_size=x_frame_size,establish_fm_array=0, auto_focus_run=0, auto_expose_run=0, channels=['DAPI'],focus_position=focus_position)

        for repeat in range(0,3):
            self.recursive_stardist_autofocus(experiment_directory, cycle=0, remake_nuc_binary=0)
            self.image_cycle_acquire(0, experiment_directory, z_slices, 'Bleach', offset_array, x_frame_size=x_frame_size,establish_fm_array=0, auto_focus_run=0, auto_expose_run=0, channels=['DAPI'],focus_position=focus_position)

        self.recursive_stardist_autofocus(experiment_directory, cycle=0)
        self.image_cycle_acquire(0, experiment_directory, 3, 'Bleach', offset_array, x_frame_size=x_frame_size,establish_fm_array=0, auto_focus_run=0, auto_expose_run=0, channels=['DAPI'],focus_position=focus_position)
        self.image_cycle_acquire(0, experiment_directory, 3, 'Stain', offset_array, x_frame_size=x_frame_size,establish_fm_array=0, auto_focus_run=0, auto_expose_run=3)

    def full_cycle(self, experiment_directory, cycle_number, offset_array, stain_valve, fluidics_object, z_slices, incub_val=45, x_frame_size=2960, focus_position = 'none'):

        pump = fluidics_object

        if cycle_number == 0:
            self.initialize(experiment_directory, offset_array, z_slices, x_frame_size=x_frame_size, focus_position = focus_position)
        else:

            # print(status_str)
            print('cycle', cycle_number)
            pump.liquid_action('Stain', incub_val=incub_val, stain_valve=stain_valve,  microscope_object = self, experiment_directory=experiment_directory)  # nuc is valve=7, pbs valve=8, bleach valve=1 (action, stain_valve, heater state (off = 0, on = 1))
            #self.reacquire_run_autofocus(experiment_directory, cycle_number, z_slices, offset_array, x_frame_size)
            # print(status_str)
            #start low flow to constantly flow fluid while imaging to reduce fluorescence of fluidic over time
            pump.liquid_action('low flow on')
            self.image_cycle_acquire(cycle_number, experiment_directory, z_slices, 'Stain', offset_array,x_frame_size=x_frame_size, establish_fm_array=0, auto_focus_run=0,auto_expose_run=3)
            pump.liquid_action('flow off')
            time.sleep(5)

            # print(status_str)
            pump.liquid_action('Bleach', stain_valve=stain_valve)  # nuc is valve=7, pbs valve=8, bleach valve=1 (action, stain_valve, heater state (off = 0, on = 1))
            time.sleep(5)
            # print(status_str)
            self.image_cycle_acquire(cycle_number, experiment_directory, z_slices, 'Bleach', offset_array, x_frame_size=x_frame_size, establish_fm_array=0, auto_focus_run=0,auto_expose_run=0)
            #self.inter_cycle_processing(experiment_directory, cycle_number=cycle_number, x_frame_size=x_frame_size)
            time.sleep(3)

        # self.post_acquisition_processor(experiment_directory, x_frame_size)

    def tissue_integrity_cycles(self, experiment_directory, cycle_number, offset_array, stain_valve, fluidics_object, z_slices, incub_val=45, x_frame_size=2960, focus_position = 'none', number_clusters = 6):

        pump = fluidics_object

        if cycle_number == 0:
            self.initialize(experiment_directory, offset_array, z_slices, x_frame_size=x_frame_size, focus_position = focus_position, number_clusters=number_clusters)
        else:
            if stain_valve == 10:

                # print(status_str)
                print('cycle', cycle_number)
                pump.liquid_action('Stain', incub_val=incub_val, stain_valve=stain_valve,  microscope_object = self, experiment_directory=experiment_directory, cycle = cycle_number)  # nuc is valve=7, pbs valve=8, bleach valve=1 (action, stain_valve, heater state (off = 0, on = 1))
                #self.reacquire_run_autofocus(experiment_directory, cycle_number, z_slices, offset_array, x_frame_size)
                # print(status_str)
                #start low flow to constantly flow fluid while imaging to reduce fluorescence of fluidic over time
                pump.liquid_action('low flow on')
                self.image_cycle_acquire(cycle_number, experiment_directory, z_slices, 'Stain', offset_array,x_frame_size=x_frame_size, establish_fm_array=0, auto_focus_run=0,auto_expose_run=3)
                pump.liquid_action('flow off')
                time.sleep(5)

                # print(status_str)
                pump.liquid_action('Bleach', stain_valve=stain_valve)  # nuc is valve=7, pbs valve=8, bleach valve=1 (action, stain_valve, heater state (off = 0, on = 1))
                time.sleep(5)
                # print(status_str)
                self.image_cycle_acquire(cycle_number, experiment_directory, z_slices, 'Bleach', offset_array, x_frame_size=x_frame_size, establish_fm_array=0, auto_focus_run=0,auto_expose_run=0)
                #self.inter_cycle_processing(experiment_directory, cycle_number=cycle_number, x_frame_size=x_frame_size)
                time.sleep(3)

            else:
                # print(status_str)
                print('cycle', cycle_number)
                pump.liquid_action('Stain', incub_val=incub_val, stain_valve=stain_valve, microscope_object=self,experiment_directory=experiment_directory, cycle = cycle_number)  # nuc is valve=7, pbs valve=8, bleach valve=1 (action, stain_valve, heater state (off = 0, on = 1))
                # self.reacquire_run_autofocus(experiment_directory, cycle_number, z_slices, offset_array, x_frame_size)
                # print(status_str)
                # start low flow to constantly flow fluid while imaging to reduce fluorescence of fluidic over time
                pump.liquid_action('low flow on')
                self.image_cycle_acquire(cycle_number, experiment_directory, z_slices, 'Stain', offset_array,x_frame_size=x_frame_size, establish_fm_array=0, auto_focus_run=0,auto_expose_run=3)
                pump.liquid_action('flow off')
                time.sleep(5)

                # print(status_str)
                pump.liquid_action('Bleach', stain_valve=stain_valve)  # nuc is valve=7, pbs valve=8, bleach valve=1 (action, stain_valve, heater state (off = 0, on = 1))
                time.sleep(5)


        # self.post_acquisition_processor(experiment_directory, x_frame_size)


    ######Folder System Generation########################################################

    def marker_excel_file_generation(self, experiment_directory, cycle_number):

        folder_path = experiment_directory + '/mcmicro'
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        exp_array = np.load('exp_array.npy', allow_pickle=False)
        exp_array = np.flip(exp_array)

        filter_sets = ['A647', 'A555', 'A488', 'DAPI']
        emission_wavelengths = ['675', '565', '525', '450']
        exciation_wavlengths = ['647', '555', '488', '405']

        try:
            os.chdir(folder_path)
            wb = load_workbook('markers.xlsx')
        except:
            wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = 'channel_number'
        ws.cell(row=1, column=2).value = 'cycle_number'
        ws.cell(row=1, column=3).value = 'marker_name'
        ws.cell(row=1, column=4).value = 'Filter'
        ws.cell(row=1, column=5).value = 'excitation_wavlength'
        ws.cell(row=1, column=6).value = 'emission_wavlength'
        ws.cell(row=1, column=7).value = 'background'
        ws.cell(row=1, column=8).value = 'exposure'
        ws.cell(row=1, column=9).value = 'remove'

        for row_number in range(2, (cycle_number) * 4 + 2):
            cycle_number =  (row_number - 2)//4 + 1
            intercycle_channel_number = cycle_number * 4 + 1 - row_number

            ws.cell(row=row_number, column=1).value = row_number - 1
            ws.cell(row=row_number, column=2).value = cycle_number
            ws.cell(row=row_number, column=3).value = 'Marker_' + str(row_number - 1)
            # ws.cell(row=row_number, column=4).value = filter_sets[intercycle_channel_number]
            # ws.cell(row=row_number, column=5).value = exciation_wavlengths[intercycle_channel_number]
            # ws.cell(row=row_number, column=6).value = emission_wavelengths[intercycle_channel_number]

        os.chdir(folder_path)
        wb.save(filename='markers.xlsx')

    def folder_addon(self, parent_directory_path, new_folder_names):

        os.chdir(parent_directory_path)

        for name in new_folder_names:

            add_on_folder = str(name)
            full_directory_path = parent_directory_path + '/' + add_on_folder

            try:
                os.mkdir(full_directory_path)
            except:
                pass

    def file_structure(self, experiment_directory, highest_cycle_count):

        channels = ['DAPI', 'A488', 'A555', 'A647']
        cycles = np.linspace(0, highest_cycle_count).astype(int)

        # folder layer one
        os.chdir(experiment_directory)
        self.folder_addon(experiment_directory, ['Quick_Tile'])
        self.folder_addon(experiment_directory, ['np_arrays'])
        self.folder_addon(experiment_directory, ['mcmicro'])
        self.folder_addon(experiment_directory, ['exposure_times'])
        self.folder_addon(experiment_directory, ['compression'])
        self.folder_addon(experiment_directory, channels)

        # folder layer two

        mc_micro_directory = experiment_directory + '/mcmicro'
        self.folder_addon(mc_micro_directory, ['raw'])
        quick_tile_directory = experiment_directory + '/Quick_Tile'
        self.folder_addon(quick_tile_directory, ['DAPI'])
        self.folder_addon(quick_tile_directory, ['A488'])
        self.folder_addon(quick_tile_directory, ['A555'])
        self.folder_addon(quick_tile_directory, ['A647'])

        for channel in channels:

            experiment_channel_directory = experiment_directory + '/' + channel

            self.folder_addon(experiment_channel_directory, ['Stain'])
            self.folder_addon(experiment_channel_directory, ['Bleach'])

            # folder layers 3 and 4

            for cycle in cycles:
                experiment_channel_stain_directory = experiment_channel_directory + '/' + 'Stain'
                experiment_channel_bleach_directory = experiment_channel_directory + '/' + 'Bleach'

                self.folder_addon(experiment_channel_stain_directory, ['cy_' + str(cycle)])
                experiment_channel_stain_cycle_directory = experiment_channel_stain_directory + '/' + 'cy_' + str(cycle)

                self.folder_addon(experiment_channel_stain_cycle_directory, ['Tiles'])

                self.folder_addon(experiment_channel_bleach_directory, ['cy_' + str(cycle)])
                experiment_channel_bleach_cycle_directory = experiment_channel_bleach_directory + '/' + 'cy_' + str(
                    cycle)

                self.folder_addon(experiment_channel_bleach_cycle_directory, ['Tiles'])

        os.chdir(experiment_directory + '/' + 'np_arrays')

    #####################################################################################################
    ##########Saving/File Generation Methods#############################################################################

    def post_acquisition_processor(self, experiment_directory, x_pixels, rolling_ball = 0):

        mcmicro_path = experiment_directory + r'\mcmicro\raw'
        cycle_start = 1
        cycle_start_search = 0
        '''
        os.chdir(mcmicro_path)
        while cycle_start_search == 0:
            file_name = str(experiment_directory.split("\\")[-1]) + '-cycle-0' + str(cycle_start) + '.ome.tif'
            if os.path.isfile(file_name) == 1:
                cycle_start += 1
            else:
                cycle_start_search = 1
        '''
        cycle_end = 2
        cycle_start = 1

        self.tissue_exist_array_generate(experiment_directory)

        for cycle_number in range(cycle_start, cycle_end):

            #self.background_sub(experiment_directory, cycle_number, hdr_sub= 1,rolling_ball= 0)
            self.focus_excel_creation(experiment_directory, cycle_number)
            self.in_focus_excel_populate(experiment_directory, cycle_number, x_pixels, hdr_sub=0)
            self.excel_2_focus(experiment_directory, cycle_number, hdr_sub=0)
            self.illumination_flattening(experiment_directory, cycle_number, rolling_ball, hdr_sub=0)
            #self.illumination_flattening_per_tile(experiment_directory, cycle_number, rolling_ball=0, hdr_sub=1)
            self.brightness_uniformer(experiment_directory, cycle_number, hdr_sub = 0)
            self.mcmicro_image_stack_generator(cycle_number, experiment_directory, x_pixels, hdr_sub=0)
            self.stage_placement(experiment_directory, cycle_number, x_pixels, hdr_sub = 0)

    def post_acquisition_processor_experimental(self, experiment_directory, x_pixels, rolling_ball = 1):

        mcmicro_path = experiment_directory + r'\mcmicro\raw'
        cycle_start = 1
        cycle_start_search = 0
        '''
        os.chdir(mcmicro_path)
        while cycle_start_search == 0:
            file_name = str(experiment_directory.split("\\")[-1]) + '-cycle-0' + str(cycle_start) + '.ome.tif'
            if os.path.isfile(file_name) == 1:
                cycle_start += 1
            else:
                cycle_start_search = 1
        '''
        cycle_end = 2
        cycle_start = 1

        #self.tissue_binary_generate(experiment_directory)
        #self.tissue_exist_array_generate(experiment_directory)

        for cycle_number in range(cycle_start, cycle_end):
            #self.infocus(experiment_directory, cycle_number, x_pixels, 1, 1)
            #self.max_projector(experiment_directory, cycle_number, x_pixels)
            #self.illumination_flattening(experiment_directory, cycle_number, rolling_ball)
            self.wavelet_background_sub(experiment_directory, cycle_number, resolution_px=100, noise_lvl=1)
            #self.illumination_flattening_per_tile(experiment_directory, cycle_number, rolling_ball)
            self.mcmicro_image_stack_generator(cycle_number, experiment_directory, x_pixels)
            self.stage_placement(experiment_directory, cycle_number, x_pixels)

    def mcmicro_image_stack_generator(self, cycle_number, experiment_directory, x_frame_size):

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)
        tissue_exist = np.load('tissue_exist.npy', allow_pickle=False)


        xml_metadata = self.metadata_generator(experiment_directory, x_frame_size)

        numpy_x = full_array[0]
        numpy_y = full_array[1]
        #numpy_tissue = full_array[10]
        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]
        tile_count = int(tissue_exist.sum())

        dapi_im_path = experiment_directory + '\DAPI\Stain\cy_' + str(
            cycle_number) + '\Tiles' + '/focused_basic_corrected'
        a488_im_path = experiment_directory + '\A488\Stain\cy_' + str(
            cycle_number) + '\Tiles' + '/focused_subbed_basic_corrected'
        a555_im_path = experiment_directory + '\A555\Stain\cy_' + str(
            cycle_number) + '\Tiles' + '/focused_subbed_basic_corrected'
        a647_im_path = experiment_directory + '\A647\Stain\cy_' + str(
            cycle_number) + '\Tiles' + '/focused_subbed_basic_corrected'


        mcmicro_path = experiment_directory + r'\mcmicro\raw'

        mcmicro_stack = np.zeros((tile_count * 4, 2960, x_frame_size)).astype('uint16')

        tile = 0
        for x in range(0, x_tile_count):
            for y in range(0, y_tile_count):

                if tissue_exist[y][x] == 1:

                    dapi_file_name = 'x' + str(x) + '_y_' + str(y) + '_c_DAPI.tif'
                    a488_file_name = 'x' + str(x) + '_y_' + str(y) + '_c_A488.tif'
                    a555_file_name = 'x' + str(x) + '_y_' + str(y) + '_c_A555.tif'
                    a647_file_name = 'x' + str(x) + '_y_' + str(y) + '_c_A647.tif'

                    base_count_number_stack = tile * 4

                    os.chdir(dapi_im_path)
                    try:
                        image = io.imread(dapi_file_name)
                    except:
                        image = cv2.imread(dapi_file_name)[::, ::, 0]

                    image[image > 65500] = 65500
                    image = np.nan_to_num(image, posinf=65500)
                    mcmicro_stack[base_count_number_stack + 0] = image

                    os.chdir(a488_im_path)
                    try:
                        image = io.imread(a488_file_name)
                    except:
                        image = cv2.imread(a488_file_name)[::, ::, 0]

                    image[image > 65500] = 65500
                    image = np.nan_to_num(image, posinf=65500)
                    mcmicro_stack[base_count_number_stack + 1] = image

                    os.chdir(a555_im_path)
                    try:
                        image = io.imread(a555_file_name)
                    except:
                        image = cv2.imread(a555_file_name)[::, ::, 0]

                    image[image > 65500] = 65500
                    image = np.nan_to_num(image, posinf=65500)
                    mcmicro_stack[base_count_number_stack + 2] = image

                    os.chdir(a647_im_path)
                    try:
                        image = io.imread(a647_file_name)
                    except:
                        image = cv2.imread(a647_file_name)[::, ::, 0]

                    image[image > 65500] = 65500
                    image = np.nan_to_num(image, posinf=65500)
                    mcmicro_stack[base_count_number_stack + 3] = image

                    tile += 1

                else:
                    pass

        os.chdir(mcmicro_path)
        mcmicro_file_name = str(experiment_directory.split("\\")[-1]) + '-cycle-0' + str(cycle_number) + '.ome.tif'
        tf.imwrite(mcmicro_file_name, mcmicro_stack, photometric='minisblack', description=xml_metadata)

    def mcmicro_image_stack_generator_separate_clusters(self, cycle_number, experiment_directory, x_frame_size):

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)
        tissue_exist = np.load('tissue_exist.npy', allow_pickle=False)

        #xml_metadata = self.metadata_generator(experiment_directory, x_frame_size)

        numpy_x = full_array[0]
        numpy_y = full_array[1]
        tissue_fm = full_array[10]
        # numpy_tissue = full_array[10]
        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]
        tile_count = int(tissue_exist.sum())

        dapi_im_path = experiment_directory + '\DAPI\Stain\cy_' + str(
            cycle_number) + '\Tiles' + '/focused_basic_darkframe'
        a488_im_path = experiment_directory + '\A488\Stain\cy_' + str(
            cycle_number) + '\Tiles' + '/focused_basic_darkframe'
        a555_im_path = experiment_directory + '\A555\Stain\cy_' + str(
            cycle_number) + '\Tiles' + '/focused_basic_darkframe'
        a647_im_path = experiment_directory + '\A647\Stain\cy_' + str(
            cycle_number) + '\Tiles' + '/focused_basic_darkframe'

        mcmicro_path = experiment_directory + r'\mcmicro'

        # determine max cluster count
        highest_number = np.max(full_array[10])
        number_clusters = math.floor(math.log10(highest_number))
        tiles_in_cluster = self.number_tiles_each_cluster(experiment_directory)
        most_tiles_in_cluster = np.max(tiles_in_cluster)
        most_tiles_in_cluster = int(most_tiles_in_cluster)

        mcmicro_stack = np.zeros((number_clusters, (most_tiles_in_cluster) * 4, 2960, x_frame_size)).astype('uint16')

        # create sub folders in mcmicro folder
        os.chdir(mcmicro_path)
        for x in range(0, number_clusters):
            os.chdir(mcmicro_path)
            sub_folder_name = 'cluster_' + str(x)
            try:
                os.mkdir(sub_folder_name)
            except:
                pass
            os.chdir(sub_folder_name)
            try:
                os.mkdir('raw')
            except:
                pass

        tile = np.zeros(number_clusters).astype('int8')
        for x in range(0, x_tile_count):
            for y in range(0, y_tile_count):

                if tissue_exist[y][x] == 1:

                    clusters_in_tile = self.tissue_fm_decode(tissue_fm[y][x])

                    dapi_file_name = 'x' + str(x) + '_y_' + str(y) + '_c_DAPI.tif'
                    a488_file_name = 'x' + str(x) + '_y_' + str(y) + '_c_A488.tif'
                    a555_file_name = 'x' + str(x) + '_y_' + str(y) + '_c_A555.tif'
                    a647_file_name = 'x' + str(x) + '_y_' + str(y) + '_c_A647.tif'

                    for cluster in clusters_in_tile:

                        base_count_number_stack = int(tile[cluster - 1] * 4)

                        os.chdir(dapi_im_path)
                        try:
                            image = io.imread(dapi_file_name)
                        except:
                            image = cv2.imread(dapi_file_name)[::, ::, 0]

                        #image[image > 65500] = 65500
                        image = np.nan_to_num(image, posinf=0)
                        mcmicro_stack[cluster - 1][base_count_number_stack + 0] = image

                        os.chdir(a488_im_path)
                        try:
                            image = io.imread(a488_file_name)
                        except:
                            image = cv2.imread(a488_file_name)[::, ::, 0]

                        #image[image > 65500] = 65500
                        image = np.nan_to_num(image, posinf=0)
                        mcmicro_stack[cluster - 1][base_count_number_stack + 1] = image

                        os.chdir(a555_im_path)
                        try:
                            image = io.imread(a555_file_name)
                        except:
                            image = cv2.imread(a555_file_name)[::, ::, 0]

                        #image[image > 65500] = 65500
                        image = np.nan_to_num(image, posinf=0)
                        mcmicro_stack[cluster - 1][base_count_number_stack + 2] = image

                        os.chdir(a647_im_path)
                        try:
                            image = io.imread(a647_file_name)
                        except:
                            image = cv2.imread(a647_file_name)[::, ::, 0]

                        #image[image > 65500] = 65500
                        image = np.nan_to_num(image, posinf=0)
                        mcmicro_stack[cluster - 1][base_count_number_stack + 3] = image

                        tile[cluster - 1] += 1

                    else:
                        pass

        for x in range(0, number_clusters):
            mc_micro_cluster_path = mcmicro_path +'\cluster_' + str(x) +r'\raw'
            os.chdir(mc_micro_cluster_path)
            mcmicro_file_name = str(experiment_directory.split("\\")[-1]) + '-cycle-0' + str(cycle_number) + '.ome.tif'
            image_stack = mcmicro_stack[x][0:(int(tiles_in_cluster[x])*4)]
            xml_metadata = self.metadata_generator_separate_clusters(experiment_directory, x_frame_size, cluster_number= x + 1)
            #image_stack = image_stack.astype('uint16')
            os.chdir(mc_micro_cluster_path)
            tf.imwrite(mcmicro_file_name, image_stack, photometric='minisblack', description=xml_metadata)
            #tf.imwrite(mcmicro_file_name, mcmicro_stack, photometric='minisblack')

    def metadata_generator(self, experiment_directory, x_frame_size):

        new_ome = OME()
        #ome = from_xml(r'C:\Users\mike\Documents\GitHub\AutoCIF/image.xml', parser='lxml')
        ome = from_xml(r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF/image.xml', parser='lxml')
        ome = ome.images[0]

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)
        tissue_exist = np.load('tissue_exist.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]
        #numpy_tissue = full_array[10]
        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]
        total_tile_count = int(tissue_exist.sum())

        y_gap = 532
        col_col_gap = 10
        # for r in range(3, -1, -1):
        #    numpy_y[r][0] = numpy_y[r + 1][0] - y_gap
        #    numpy_y[r][1] = numpy_y[r + 1][1] - y_gap - col_col_gap

        # sub in needed pixel size and pixel grid changes
        ome.pixels.physical_size_x = 0.2
        ome.pixels.physical_size_y = 0.2
        ome.pixels.size_x = x_frame_size
        ome.pixels.size_y = 2960
        # sub in other optional numbers to make metadata more accurate

        for x in range(0, total_tile_count):
            tile_metadata = deepcopy(ome)
            new_ome.images.append(tile_metadata)

        # sub in stage positional information into each tile. numpy[y][x]
        tile_counter = 0
        for x in range(0, x_tile_count):
            for y in range(0, y_tile_count):

                if tissue_exist[y][x] == 1:

                    for p in range(0, 4):
                        new_x = numpy_x[y][x] - 11000
                        new_y = numpy_y[y][x] + 2300
                        new_ome.images[tile_counter].pixels.planes[p].position_y = deepcopy(new_y)
                        new_ome.images[tile_counter].pixels.planes[p].position_x = deepcopy(new_x)
                        new_ome.images[tile_counter].pixels.tiff_data_blocks[p].ifd = (4 * tile_counter) + p
                    tile_counter += 1

                else:
                    pass

        xml = to_xml(new_ome)

        return xml

    def metadata_generator_separate_clusters(self, experiment_directory, x_frame_size, cluster_number):

        new_ome = OME()
        #ome = from_xml(r'C:\Users\mike\Documents\GitHub\AutoCIF/image.xml', parser='lxml')
        ome = from_xml(r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF/image.xml', parser='lxml')
        ome = ome.images[0]

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)
        tissue_exist = np.load('tissue_exist.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]
        tissue_fm = full_array[10]
        #numpy_tissue = full_array[10]
        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]
        total_tile_count = int(self.number_tiles_each_cluster(experiment_directory)[cluster_number - 1])
        #total_tile_count = int(tissue_exist.sum())
        print(total_tile_count)

        y_gap = 532
        col_col_gap = 10
        # for r in range(3, -1, -1):
        #    numpy_y[r][0] = numpy_y[r + 1][0] - y_gap
        #    numpy_y[r][1] = numpy_y[r + 1][1] - y_gap - col_col_gap

        # sub in needed pixel size and pixel grid changes
        ome.pixels.physical_size_x = 0.2004
        ome.pixels.physical_size_y = 0.2004
        ome.pixels.size_x = x_frame_size
        ome.pixels.size_y = 2960
        # sub in other optional numbers to make metadata more accurate

        for x in range(0, total_tile_count):
            tile_metadata = deepcopy(ome)
            new_ome.images.append(tile_metadata)


        # sub in stage positional information into each tile. numpy[y][x]
        tile_counter = 0
        for x in range(0, x_tile_count):
            for y in range(0, y_tile_count):

                clusters_in_tile = self.tissue_fm_decode(tissue_fm[y][x])
                
                boolean = np.isin(clusters_in_tile, cluster_number, assume_unique=True)
                if np.sum(boolean) > 0:

                    for p in range(0, 4):
                        new_x = numpy_x[y][x]
                        new_y = numpy_y[y][x]
                        new_ome.images[tile_counter].pixels.planes[p].position_y = deepcopy(new_y)
                        new_ome.images[tile_counter].pixels.planes[p].position_x = deepcopy(new_x)
                        new_ome.images[tile_counter].pixels.tiff_data_blocks[p].ifd = (4 * tile_counter) + p
                    tile_counter += 1

                else:
                    pass


        xml = to_xml(new_ome)

        return xml

    def star_dist_stage_placement(self, experiment_directory, x_pixels):
        '''
        Goal to place images via rough stage coords in a larger image. WIll have nonsense borders
        '''

        star_dist_path = experiment_directory + r'\Labelled_Nuc'

        # load in numpy matricies

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]

        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]

        fov_x_pixels = x_pixels
        fov_y_pixels = 2960
        um_pixel = 0.20

        # generate large numpy image with rand numbers. Big enough to hold all images + 10%

        super_y = int(1.02 * (y_tile_count * fov_y_pixels))
        super_x = int(1.02 * (x_tile_count * fov_x_pixels))
        placed_image = np.random.rand(super_y, super_x).astype('uint16')

        # transform numpy x and y coords into new shifted space that starts at zero and is in units of pixels and not um
        numpy_x_pixels = numpy_x / um_pixel
        numpy_y_pixels = numpy_y / um_pixel

        y_displacement_vector = (2100 / um_pixel + fov_y_pixels / 2) * 1.02
        x_displacement_vector = (-11385 / um_pixel + fov_x_pixels / 2) * 1.02

        numpy_x_pixels = numpy_x_pixels + x_displacement_vector
        numpy_x_pixels = np.ceil(numpy_x_pixels)
        numpy_x_pixels = numpy_x_pixels.astype(int)

        numpy_y_pixels = numpy_y_pixels + y_displacement_vector
        numpy_y_pixels = np.ceil(numpy_y_pixels)
        numpy_y_pixels = numpy_y_pixels.astype(int)

        # load images into python

        os.chdir(star_dist_path)

        # place images into large array

        for x in range(0, x_tile_count):
            for y in range(0, y_tile_count):
                filename = 'x' + str(x) + '_y_' + str(y) + '_c_DAPI.tif'
                try:
                    image = io.imread(filename)
                except:
                    image = cv2.imread(filename)[::,::,0]

                image = np.nan_to_num(image, posinf=65500)

                # define subsection of large array that fits dimensions of single FOV
                # x_center = numpy_x_pixels[y][x]
                # y_center = numpy_y_pixels[y][x]
                # x_start = int(x_center - fov_x_pixels / 2)
                # x_end = int(x_center + fov_x_pixels / 2)
                # y_start = int(y_center - fov_y_pixels / 2)
                # y_end = int(y_center + fov_y_pixels / 2)

                if x == 0 and y == 0:
                    x_start = 0
                    x_end = x_pixels
                    y_start = 0
                    y_end = 2960
                else:

                    x_start = int(x * fov_x_pixels * 0.9)
                    x_end = x_start + x_pixels
                    y_start = int(y * fov_y_pixels * 0.9)
                    y_end = y_start + 2960

                # placed_image[y_start:y_end, x_start:x_end] = placed_image[y_start:y_end, x_start:x_end] + image
                placed_image[y_start:y_end, x_start:x_end] = image

        # save output image
        os.chdir(star_dist_path)
        tf.imwrite('star_dist_placed.tif', placed_image)

    def stage_placement(self, experiment_directory, cycle_number, x_pixels, down_sample_factor = 1, single_fov = 0):
        '''
        Goal to place images via rough stage coords in a larger image. WIll have nonsense borders
        '''

        quick_tile_path = experiment_directory + r'\Quick_Tile'

        # load in numpy matricies

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)
        tissue_exist = np.load('tissue_exist.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]

        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]

        fov_x_pixels = x_pixels
        fov_y_pixels = 2960
        um_pixel = 0.2004

        # generate large numpy image with rand numbers. Big enough to hold all images + 10%

        super_y = int(1.02 * (y_tile_count * fov_y_pixels))
        super_x = int(1.02 * (x_tile_count * fov_x_pixels))

        # transform numpy x and y coords into new shifted space that starts at zero and is in units of pixels and not um
        numpy_x_pixels = numpy_x / um_pixel
        numpy_y_pixels = numpy_y / um_pixel

        y_displacement_vector = (2100 / um_pixel + fov_y_pixels / 2) * 1.02
        x_displacement_vector = (-11385 / um_pixel + fov_x_pixels / 2) * 1.02

        numpy_x_pixels = numpy_x_pixels + x_displacement_vector
        numpy_x_pixels = np.ceil(numpy_x_pixels)
        numpy_x_pixels = numpy_x_pixels.astype(int)

        numpy_y_pixels = numpy_y_pixels + y_displacement_vector
        numpy_y_pixels = np.ceil(numpy_y_pixels)
        numpy_y_pixels = numpy_y_pixels.astype(int)

        # load images into python

        channels = ['DAPI', 'A488', 'A555', 'A647']
        #types = ['Stain', 'Bleach']
        types = ['Stain']

        for type in types:
            for channel in channels:

                if type == 'Stain':
                    if channel == 'DAPI':
                        im_path = experiment_directory + '/' + channel + "/" + type + '\cy_' + str(
                            cycle_number) + '\Tiles' + r'\focused_basic_darkframe'
                    else:
                        im_path = experiment_directory + '/' + channel + "/" + type + '\cy_' + str(
                            cycle_number) + '\Tiles' + '/focused_basic_darkframe'

                elif type == 'Bleach':
                    if single_fov != 1:
                        im_path = experiment_directory + '/' + channel + "/" + type + '\cy_' + str(cycle_number) + '\Tiles' + '/focused'
                    elif single_fov == 1:
                        im_path = experiment_directory + '/' + channel + "/" + type + '\cy_' + str(cycle_number) + '\Tiles'
                os.chdir(im_path)

                # place images into large array

                placed_image = np.random.rand(super_y, super_x).astype('float32')
                for x in range(0, x_tile_count):
                    for y in range(0, y_tile_count):



                        if tissue_exist[y][x] == 1:

                            filename = 'x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                            try:
                                image = io.imread(filename)
                            except:
                                pass
                                #image = cv2.imread(filename)[::,::,0]

                            # define subsection of large array that fits dimensions of single FOV
                            # x_center = numpy_x_pixels[y][x]
                            # y_center = numpy_y_pixels[y][x]
                            # x_start = int(x_center - fov_x_pixels / 2)
                            # x_end = int(x_center + fov_x_pixels / 2)
                            # y_start = int(y_center - fov_y_pixels / 2)
                            # y_end = int(y_center + fov_y_pixels / 2)

                            if x == 0 and y == 0:
                                x_start = 0
                                x_end = x_pixels
                                y_start = 0
                                y_end = 2960
                            else:

                                x_start = int(x * fov_x_pixels * 0.9)
                                x_end = x_start + x_pixels
                                y_start = int(y * fov_y_pixels * 0.9)
                                y_end = y_start + 2960

                            placed_image[y_start:y_end, x_start:x_end] = image

                        else:
                            pass

                #down sample

                placed_image = placed_image[::down_sample_factor, ::down_sample_factor]

                # save output image
                os.chdir(quick_tile_path + '/' + channel)
                tf.imwrite(channel + '_cy_' + str(cycle_number) + '_' + type + '_placed.tif', placed_image)

    def nan_folder_conversion(self, directory_path):
        '''
        goes through every image in a folder, loads in image and runs nan_to_num() on it.
        :param directory_path:
        :return:
        '''

        os.chdir(directory_path)

        filenames = os.listdir(directory_path)
        for x in range(0, len(filenames)):
            try:
                im2 = io.imread(filenames[x]).astype('float32')
                im2 = np.nan_to_num(im2, posinf=65500)
                io.imsave(filenames[x], im2)
            except:
                pass

    def illumination_flattening(self, experiment_directory, cycle_number, single_fov = 0):

        # load numpy arrays in
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        directory_start = experiment_directory + '//'

        for channel in range(0, 4):

            if channel == 0:
                channel_name = 'DAPI'
            if channel == 1:
                channel_name = 'A488'
            if channel == 2:
                channel_name = 'A555'
            if channel == 3:
                channel_name = 'A647'

            print('channel', channel_name, 'cycle', cycle_number)

            if channel == 0:
                if single_fov != 1:
                    stain_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\focused'
                    training_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\focused'
                    stain_output_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\focused_basic_corrected'
                if single_fov == 1:
                    stain_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles'
                    training_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles'
                    stain_output_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\focused_basic_corrected'

            else:
                if single_fov != 1:
                    stain_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\focused'
                    training_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\focused'
                    stain_output_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\focused_basic_corrected'

                if single_fov == 1:
                    stain_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles'
                    training_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles'
                    stain_output_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\focused_basic_corrected'




            try:
                os.mkdir(stain_output_directory)
            except:
                pass


            #resave images without NaN or infinity values
            self.nan_folder_conversion(stain_directory)
            #ff_directory = r'C:\Users\CyCIF PC\Desktop\new A647 FF'

            epsilon = 1e-06
            optimizer = shading_correction.BaSiC(training_directory)
            #optimizer = shading_correction.BaSiC(ff_directory)
            optimizer.prepare()
            optimizer.run()
            # Save the estimated fields (only if the profiles were estimated)
            directory = Path(stain_output_directory)
            flatfield_name = directory / "flatfield.tif"
            darkfield_name = directory / "darkfield.tif"
            cv2.imwrite(str(flatfield_name), optimizer.flatfield_fullsize.astype(np.float32))
            cv2.imwrite(str(darkfield_name), optimizer.darkfield_fullsize.astype(np.float32))
            #optimizer.write_images(stain_output_directory, epsilon=epsilon)

            #run trained FF on subtracted images
            optimizer.directory = stain_directory
            optimizer._sniff_input()
            optimizer._load_images()
            optimizer.write_images(stain_output_directory, epsilon=epsilon)

    def bottom_int_correction(self, experiment_directory, cycle_number):
        '''
        Subs off smallest value in tile

        :param experiment_directory:
        :param cycle_number:
        :return:
        '''

        # load in data structures
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)
        tissue_fm = fm_array[10]
        x_tile_count = np.shape(tissue_fm)[1]
        y_tile_count = np.shape(tissue_fm)[0]

        channels = np.array(['A488', 'A555', 'A647'])

        for channel in channels:

            flattened_path = experiment_directory + '//' + channel + '\Stain\cy_' + str(cycle_number) + r'\Tiles\focused_basic_corrected'
            os.chdir(flattened_path)

            for x in range(0, x_tile_count):
                for y in range(0, y_tile_count):
                    if tissue_fm[y][x] > 1:
                        image_name = r'x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                        flat_image = io.imread(image_name)
                        low_pixel = np.sort(flat_image, axis=None)[100]
                        #print('x', x, 'y', y, 'low pixel', low_pixel)
                        flat_image -= low_pixel
                        flat_image[flat_image<0] = 0
                        io.imsave(image_name, flat_image)

                    else:
                        pass

    def illumination_flattening_per_tile(self, experiment_directory, cycle_number, rolling_ball = 1, hdr_sub = 1):

        # load numpy arrays in
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        directory_start = experiment_directory + '//'

        for channel in range(0, 4):

            if channel == 0:
                channel_name = 'DAPI'
            if channel == 1:
                channel_name = 'A488'
            if channel == 2:
                channel_name = 'A555'
            if channel == 3:
                channel_name = 'A647'

            print('channel', channel_name, 'cycle', cycle_number)

            if channel == 0:
                if hdr_sub == 0:
                    stain_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\focused'
                    stain_temp_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\focused\temp'
                    bleach_directory = directory_start + channel_name + '\Bleach\cy_' + str(cycle_number) + r'\Tiles\focused'
                if hdr_sub == 1:
                    stain_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\subbed_focused'
                    stain_temp_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\subbed_focused\temp'
            else:
                if rolling_ball != 1:
                    if hdr_sub == 0:
                        stain_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\background_subbed'
                        stain_temp_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\background_subbed\temp'
                        bleach_directory = directory_start + channel_name + '\Bleach\cy_' + str(cycle_number) + r'\Tiles\focused'
                    if hdr_sub == 1:
                        stain_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\subbed_focused'
                        stain_temp_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\subbed_focused\temp'
                if rolling_ball == 1:
                    directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\focused\background_subbed_rolling'
            stain_output_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\subbed_focused_basic_corrected'
            #bleach_output_directory = directory_start + channel_name + '\Bleach\cy_' + str(cycle_number) + r'\Tiles\focused_basic_corrected'

            #stain_temp_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\background_subbed\temp'
            #bleach_temp_directory = directory_start + channel_name + '\Bleach\cy_' + str(cycle_number) + r'\Tiles\background_subbed\temp'

            try:
                os.mkdir(stain_temp_directory)
                #os.mkdir(bleach_temp_directory)
            except:
                pass

            try:
                os.mkdir(stain_output_directory)
                #os.mkdir(bleach_output_directory)
            except:
                pass

            #find file names of folder
            names = os.listdir(stain_directory)
            names.pop(0)
            print(names)

            '''
            # build BaSiC model on two identical images and output into dictated folder
            epsilon = 1e-06
            optimizer = shading_correction.BaSiC(stain_directory)
            optimizer.prepare()
            optimizer.run()
            # Save the estimated fields (only if the profiles were estimated)
            directory = Path(stain_output_directory)
            flatfield_name = directory / "flatfield.tif"
            darkfield_name = directory / "darkfield.tif"
            cv2.imwrite(str(flatfield_name), optimizer.flatfield_fullsize.astype(np.float32))
            cv2.imwrite(str(darkfield_name), optimizer.darkfield_fullsize.astype(np.float32))
            optimizer.write_images(stain_output_directory, epsilon=epsilon)
            '''

            for name in names:
                #transfer image and replica of image into temp folder
                first_name = name
                second_name = '2_' + name
                os.chdir(stain_directory)
                image = io.imread(first_name) + 100
                image = np.nan_to_num(image, posinf=65500)
                os.chdir(stain_temp_directory)
                io.imsave(first_name, image)
                io.imsave(second_name, image)
                #build BaSiC model on two identical images and output into dictated folder
                epsilon = 1e-06
                optimizer = shading_correction.BaSiC(stain_temp_directory)
                optimizer.prepare()
                optimizer.run()
                # Save the estimated fields (only if the profiles were estimated)
                directory = Path(stain_output_directory)
                flatfield_name = directory / "flatfield.tif"
                darkfield_name = directory / "darkfield.tif"
                cv2.imwrite(str(flatfield_name), optimizer.flatfield_fullsize.astype(np.float32))
                cv2.imwrite(str(darkfield_name), optimizer.darkfield_fullsize.astype(np.float32))
                optimizer.write_images(stain_output_directory, epsilon=epsilon)

                #delete files in temp and output
                path_one = os.path.join(stain_temp_directory, first_name)
                path_two = os.path.join(stain_temp_directory, second_name)
                os.remove(path_one)
                #os.remove(path_two)

                output_path_file = os.path.join(stain_temp_directory, second_name)
                os.remove(output_path_file)


            '''
            #Replicate process with bleached version

            # find file names of folder
            names = os.listdir(bleach_directory)
            names.pop(0)

            for name in names:
                # transfer image and replica of image into temp folder
                first_name = name
                second_name = '2_' + name
                os.chdir(bleach_directory)
                image = io.imread(first_name)
                image = np.nan_to_num(image, posinf=65500)
                os.chdir(bleach_temp_directory)
                io.imsave(first_name, image)
                io.imsave(second_name, image)
                # build BaSiC model on two identical images and output into dictated folder
                epsilon = 1e-06
                optimizer = shading_correction.BaSiC(bleach_temp_directory)
                optimizer.prepare()
                optimizer.run()
                # Save the estimated fields (only if the profiles were estimated)
                directory = Path(bleach_output_directory)
                flatfield_name = directory / "flatfield.tif"
                darkfield_name = directory / "darkfield.tif"
                cv2.imwrite(str(flatfield_name), optimizer.flatfield_fullsize.astype(np.float32))
                cv2.imwrite(str(darkfield_name), optimizer.darkfield_fullsize.astype(np.float32))
                optimizer.write_images(bleach_output_directory, epsilon=epsilon)

                # delete files in temp and output
                path_one = os.path.join(bleach_temp_directory, first_name)
                path_two = os.path.join(bleach_temp_directory, second_name)
                os.remove(path_one)
                #os.remove(path_two)

                output_path_file = os.path.join(bleach_temp_directory, second_name)
                os.remove(output_path_file)
            '''

    def single_fov_file_rename(self, experiment_directory, cycle_number):

        # load numpy arrays in
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        fm_array = np.load('fm_array.npy', allow_pickle=False)
        tissue_fm = fm_array[10]

        x_tiles = np.shape(fm_array[0])[1]
        y_tiles = np.shape(fm_array[0])[0]

        directory_start = experiment_directory + '//'
        channels = ['DAPI', 'A488', 'A555', 'A647']
        for x in range(0, x_tiles):
            for y in range(0, y_tiles):
                if tissue_fm[y][x] > 1:
                    for channel in channels:

                        stain_directory = directory_start + channel + '\Stain\cy_' + str(cycle_number) + r'\Tiles'
                        os.chdir(stain_directory)
                        file_name = 'z_0_x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                        new_file_name = 'x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                        im = io.imread(file_name)
                        io.imsave(new_file_name, im)

                        stain_directory = directory_start + channel + '\Bleach\cy_' + str(cycle_number) + r'\Tiles'
                        os.chdir(stain_directory)
                        file_name = 'z_0_x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                        new_file_name = 'x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                        im = io.imread(file_name)
                        io.imsave(new_file_name, im)
                else:
                    pass


    def inter_cycle_processing(self, experiment_directory, cycle_number, x_frame_size):

        '''
        To reduce file size between, some processing is done between cycles. In this case: 1. find focus tiles
        2. Subtract bleached from stained
        3. compress files from 32bit to 16bit in general tile folder and subbed folder.
        :param experiment_directory:
        :param cycle_number:
        :return:
        '''



        start = time.time()


        #make tissue exist array if needed
        if cycle_number == 1:
            self.tissue_exist_array_generate(experiment_directory, x_frame_size=x_frame_size)
        else:
            pass
        end = time.time()
        print('binary create', end - start)

        #determine in focus parts first
        self.focus_excel_creation(experiment_directory, cycle_number)
        self.in_focus_excel_populate(experiment_directory, cycle_number, x_frame_size=x_frame_size)

        self.excel_2_focus(experiment_directory, cycle_number, x_frame_size=x_frame_size)
        #self.single_fov_file_rename(experiment_directory, cycle_number)

        end = time.time()
        print('focus', end - start)

        #flatten image

        self.illumination_flattening(experiment_directory, cycle_number, single_fov=0)
        #self.bottom_int_correction(experiment_directory, cycle_number=cycle_number)

        end = time.time()
        print('flatten', end - start)

        self.darkframe_sub(experiment_directory, cycle_number)
        end = time.time()
        print('dark frame subtraction', end - start)


        #compress to 16bit
        self.stage_placement(experiment_directory, cycle_number, x_pixels=x_frame_size, down_sample_factor=4,single_fov=0)
        self.hdr_compression_2(experiment_directory, cycle_number)


        end = time.time()
        print('compress', end - start)

        self.mcmicro_image_stack_generator_separate_clusters(cycle_number, experiment_directory, x_frame_size)

        end = time.time()
        print('mcmicro', end - start)


        #generate stage placement

        self.stage_placement(experiment_directory, cycle_number, x_pixels = x_frame_size, down_sample_factor=4, single_fov=0)

        #if did DAPI focus then acquire one plane, please do the following
        #self.delete_intermediate_folders(experiment_directory, cycle_number)
        #self.zlib_compress_raw(experiment_directory, cycle_number)

        #end = time.time()
        #print('stage placement', end - start)

    def infocus(self, experiment_directory, cycle_number, x_frame_size, x_sub_section_count = 1, y_sub_section_count = 1):

        print('cycle', cycle_number)
        bin_values = [10]
        channels = ['DAPI', 'A488', 'A555', 'A647']
        #channels = ['A647']

        dapi_im_path = experiment_directory + '/' + 'DAPI' '\Stain\cy_' + str(cycle_number) + '\Tiles'
        tissue_path = experiment_directory + '/Tissue_Binary'

        # load numpy arrays in
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)
        tissue_exist = np.load('tissue_exist.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]

        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]

        # determine number z slices
        z_checker = 0
        z_slice_count = 0
        os.chdir(dapi_im_path)
        while z_checker == 0:
            file_name = 'z_' + str(z_slice_count) + '_x' + str(0) + '_y_' + str(0) + '_c_DAPI.tif'
            if os.path.isfile(file_name) == 1:
                z_slice_count += 1
            else:
                z_checker = 1
        z_slice_count = 7

        # make object to hold all tissue binary maps
        tissue_binary_stack = np.random.rand(y_tile_count, x_tile_count, 2960, x_frame_size).astype('uint16')
        for x in range(0, x_tile_count):
            for y in range(0, y_tile_count):
                if tissue_exist[y][x] == 1:
                    os.chdir(tissue_path)
                    file_name = 'x' + str(x) + '_y_' + str(y) + '_tissue.tif'
                    image = tf.imread(file_name)
                    tissue_binary_stack[y][x] = image
                else:
                    pass

        for channel in channels:
            # generate imstack of z slices for tile
            im_path = experiment_directory + '/' + channel + '\Stain\cy_' + str(cycle_number) + '\Tiles'
            os.chdir(im_path)

            z_stack = np.random.rand(z_slice_count, 2960, x_frame_size).astype('float32')
            for x in range(0, x_tile_count):
                for y in range(0, y_tile_count):
                    if tissue_exist[y][x] == 1:
                        for z in range(0, z_slice_count):
                            im_path = experiment_directory + '/' + channel + '\Stain\cy_' + str(cycle_number) + '\Tiles'
                            os.chdir(im_path)
                            file_name = 'z_' + str(z) + '_x' + str(x) + '_y_' + str(y) + '_c_' + str(channel) + '.tif'
                            image = tf.imread(file_name)
                            z_stack[z] = image

                            # break into sub sections (2x2)

                        number_bins = len(bin_values)
                        brenner_sub_selector = np.random.rand(z_slice_count, number_bins, y_sub_section_count,
                                                              x_sub_section_count)
                        for z in range(0, z_slice_count):
                            for y_sub in range(0, y_sub_section_count):
                                for x_sub in range(0, x_sub_section_count):

                                    y_end = int((y_sub + 1) * (2960 / y_sub_section_count))
                                    y_start = int(y_sub * (2960 / y_sub_section_count))
                                    x_end = int((x_sub + 1) * (x_frame_size / x_sub_section_count))
                                    x_start = int(x_sub * (x_frame_size / x_sub_section_count))
                                    sub_image = z_stack[z][y_start:y_end, x_start:x_end]

                                    sub_tissue_bin = tissue_binary_stack[y][x][y_start:y_end, x_start:x_end]

                                    for b in range(0, number_bins):
                                        bin_value = int(bin_values[b])
                                        score =  self.focus_score(sub_image, bin_value, sub_tissue_bin)
                                        #score = self.focus_score_post_processing(sub_image, bin_value)
                                        #score = 500
                                        brenner_sub_selector[z][b][y_sub][x_sub] = score

                        reconstruct_array = self.brenner_reconstruct_array(brenner_sub_selector, z_slice_count, number_bins)
                        #reconstruct_array = skimage.filters.median(reconstruct_array)
                        self.image_reconstructor(experiment_directory, reconstruct_array, channel, cycle_number,
                                             x_frame_size, y, x)

                    else:
                        pass

    #Find focus functions
    def focus_excel_creation(self, experiment_directory, cycle_number):
        os.chdir(experiment_directory)
        folder_name = 'focus_grid_excel'
        file_name = 'focus_grid.xlsx'
        channels = ['DAPI', 'A488', 'A555', 'A647']

        #make directory and set as active directory
        try:
            os.mkdir(folder_name)
            os.chdir(folder_name)
        except:
            os.chdir(folder_name)

        #Create new file or load it in
        try:
            wb = load_workbook(file_name)
        except:
            wb = Workbook()

        sheet_start = (cycle_number -1) * 4



        channel_index = 0
        for x in range(sheet_start, len(channels) + sheet_start):
            sheet_name = channels[channel_index] + '_cycle' + str(cycle_number)
            wb.create_sheet(sheet_name,x)
            channel_index += 1

        wb.save(file_name)

    def in_focus_excel_populate(self, experiment_directory, cycle_number, x_frame_size, x_sub_section_count = 1, y_sub_section_count = 1):

        print('cycle', cycle_number)
        channels = ['DAPI', 'A488', 'A555', 'A647']

        bin_value = 17

        dapi_im_path = experiment_directory + '/' + 'DAPI' '\Stain\cy_' + str(cycle_number) + '\Tiles'
        nuc_binary_path = experiment_directory + '/Labelled_Nuc'

        excel_folder_name = 'focus_grid_excel'
        excel_file_name = 'focus_grid.xlsx'

        # load numpy arrays in
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)
        tissue_exist = np.load('tissue_exist.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]

        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]

        z_slice_count = int(full_array[3][0][0])

        #load in excel file
        os.chdir(experiment_directory + r'/' + excel_folder_name)
        wb = load_workbook(excel_file_name)

        scores = np.zeros(z_slice_count)

        for x in range(0, x_tile_count):
            for y in range(0, y_tile_count):
                if tissue_exist[y][x] >= 1:
                    os.chdir(nuc_binary_path)
                    file_name = 'x' + str(x) + '_y_' + str(y) + '_c_DAPI.tif'
                    nuc_mask = io.imread(file_name)

                    for z in range(0, z_slice_count):
                        os.chdir(dapi_im_path)
                        file_name = 'z_' + str(z) + '_x' + str(x) + '_y_' + str(y) + '_c_DAPI.tif'
                        image = io.imread(file_name)
                        scores[z] = self.focus_score(image, bin_value, nuc_mask)
                        print('x', x, 'y', y, 'z', z, 'score', scores[z])

                    #find in focus slice
                    highest_index = np.where(scores == np.max(scores))[0][0]

                    for channel in channels:
                        sheet_name = channel + '_cycle' + str(cycle_number)
                        ws = wb[sheet_name]
                        ws.cell(row=(y + 1), column=(x + 1)).value = highest_index

                else:
                    pass


        os.chdir(experiment_directory + '/' + excel_folder_name)
        print('saving')
        wb.save(excel_file_name)

    def excel_2_focus(self, experiment_directory, cycle_number, x_frame_size = 2960, hdr_sub = 1):

        # load numpy arrays in
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)
        tissue_exist = np.load('tissue_exist.npy', allow_pickle=False)
        excel_folder_name = 'focus_grid_excel'
        excel_file_name = 'focus_grid.xlsx'


        numpy_x = full_array[0]
        numpy_y = full_array[1]

        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]

        #load in excel file
        os.chdir(experiment_directory + r'/' + excel_folder_name)
        wb = load_workbook(excel_file_name)

        channels = ['DAPI', 'A488', 'A555', 'A647']

        for channel in channels:
            sheet_name = channel + '_cycle' + str(cycle_number)
            ws = wb[sheet_name]
            for x in range(0, x_tile_count):
                for y in range(0, y_tile_count):
                    if tissue_exist[y][x] == 1:

                        slice_number = ws.cell(row= (y+1), column = (x + 1)).value
                        self.image_reconstructor(experiment_directory, slice_number, channel, cycle_number,x_frame_size, y, x)

                    else:
                        pass

    def brenner_reconstruct_array(self, brenner_sub_selector, z_slice_count, number_bins):
        '''
        take in sub selector array and slice to find max values for brenner scores and then find mode for various bin levels
        output array that shows what slice to grab each sub section from
        :param brenner_sub_selector:
        :return:
        '''
        # array to dictate the z slice to grab each sub section from
        y_sections = np.shape(brenner_sub_selector)[2]
        x_sections = np.shape(brenner_sub_selector)[3]

        reconstruct_array = np.random.rand(y_sections, x_sections).astype('uint16')

        temp_bin_max_indicies = np.random.rand(number_bins).astype('uint16')

        for y in range(0, y_sections):
            for x in range(0, x_sections):
                for b in range(0, number_bins):
                    sub_scores = brenner_sub_selector[::, b, y, x]
                    max_score = np.max(sub_scores)
                    max_index = np.where(sub_scores == max_score)[0][0]
                    #max_index = 3
                    # temp_bin_max_indicies[b] = max_index
                # sub_section_index_mode = stats.mode(temp_bin_max_indicies)[0][0]
                # reconstruct_array[y][x] = sub_section_index_mode
                reconstruct_array[y][x] = max_index

        return reconstruct_array

    def image_reconstructor(self, experiment_directory, reconstruct_array, channel, cycle_number, x_frame_size,
                            y_tile_number, x_tile_number):

        #y_sections = np.shape(reconstruct_array)[0]
        #x_sections = np.shape(reconstruct_array)[1]

        #cycle_types = ['Stain', 'Bleach']
        cycle_types = ['Stain']

        for type in cycle_types:

            if type == 'Stain':

                im_path = experiment_directory + '/' + channel + '/' + type + '\cy_' + str(cycle_number) + '\Tiles'
                os.chdir(im_path)
                saving_path = experiment_directory + '/' + channel + '/' + type + '\cy_' + str(cycle_number) + '\Tiles//focused'
                try:
                    os.mkdir('focused')
                except:
                    pass

            if type == 'Bleach':

                im_path = experiment_directory + '/' + channel + '/' + type + '\cy_' + str(cycle_number ) + '\Tiles'
                os.chdir(im_path)
                saving_path = experiment_directory + '/' + channel + '/' + type + '\cy_' + str(cycle_number) + '\Tiles//focused'
                try:
                    os.mkdir('focused')
                except:
                    pass

            os.chdir(im_path)

            # rebuilt image container
            #rebuilt_image = np.random.rand(2960, x_frame_size).astype('float32')

            #for y in range(0, y_sections):
            #    for x in range(0, x_sections):
            #        # define y and x start and ends subsection of rebuilt image
            #        y_end = int((y + 1) * (2960 / y_sections))
            #        y_start = int(y * (2960 / y_sections))
            #        x_end = int((x + 1) * (x_frame_size / x_sections))
            #        x_start = int(x * ((x_frame_size) / x_sections))

                    # find z for specific subsection
                    #z_slice = reconstruct_array[y][x]
            z_slice = reconstruct_array
                    #z_slice = 5
                    # load in image to extract for subsection
            file_name = 'z_' + str(z_slice) + '_x' + str(x_tile_number) + '_y_' + str(y_tile_number) + '_c_' + str(channel) + '.tif'
            print(file_name)
            image = tf.imread(file_name)
            #rebuilt_image[y_start:y_end, x_start:x_end] = image[y_start:y_end, x_start:x_end]

            filename = 'x' + str(x_tile_number) + '_y_' + str(y_tile_number) + '_c_' + str(channel) + '.tif'
            os.chdir(saving_path)
            tf.imwrite(filename, image)

    def background_sub(self, experiment_directory, cycle, rolling_ball = 0):

        experiment_directory = experiment_directory + '/'

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)
        tissue_exist = np.load('tissue_exist.npy', allow_pickle=False)

        y_tile_count = np.shape(fm_array[0])[0]
        x_tile_count = np.shape(fm_array[0])[1]

        channels = ['A488', 'A555', 'A647']


        for y in range(0, y_tile_count):
            for x in range(0, x_tile_count):

                if tissue_exist[y][x] == 1:
                    #tissue_path = experiment_directory + '/Tissue_Binary'
                    #os.chdir(tissue_path)
                    #filename = 'x' + str(x) + '_y_' + str(y) + '_tissue.tif'
                    #tissue_im = io.imread(filename)

                    if rolling_ball != 1:
                        '''
                        #load in tissue binary
                        tissue_path = experiment_directory + r'\Tissue_Binary'
                        os.chdir(tissue_path)
                        tissue_filename = 'x' + str(x) + '_y_' + str(y) + '_tissue.tif'
                        tissue_im = io.imread(tissue_filename)
                        
                        # load reference and "moved" image
                        ref_path = experiment_directory + 'DAPI\Bleach\cy_' + str(cycle) + '\Tiles/focused'
                        os.chdir(ref_path)
                        ref_name = 'x' + str(x) + '_y_' + str(y) + '_c_DAPI.tif'
                        ref = io.imread(ref_name)
                        ref = np.nan_to_num(ref, posinf=65500)
                        mov_path = experiment_directory + 'DAPI\Stain\cy_' + str(cycle) + '\Tiles/focused'
                        os.chdir(mov_path)
                        mov_name = 'x' + str(x) + '_y_' + str(y) + '_c_DAPI.tif'
                        mov = io.imread(mov_name)
                        mov = np.nan_to_num(mov, posinf=65500)

                        # Translational transformation
                        sr = StackReg(StackReg.TRANSLATION)
                        out_tra = sr.register_transform(ref, mov)
                        '''

                        # apply translation to other color channels

                        for channel in channels:

                            #stain_color_path = experiment_directory + channel + r'/Stain/cy_' + str(cycle) + '\Tiles/focused_basic_corrected'
                            stain_color_path = experiment_directory + channel + r'/Stain/cy_' + str(cycle) + '\Tiles/focused'
                            os.chdir(stain_color_path)
                            filename = 'x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                            color_im = io.imread(filename)
                            color_im = np.nan_to_num(color_im, posinf= 65500)
                            #color_reg = sr.transform(color_im)
                            color_factor = color_im

                            # sub background color channels
                            #bleach_color_path = experiment_directory + channel + r'/Bleach/cy_' + str(cycle) + '\Tiles/focused_basic_corrected'
                            bleach_color_path = experiment_directory + channel + r'/Bleach/cy_' + str(cycle) + '\Tiles/focused'
                            os.chdir(bleach_color_path)
                            color_bleach = io.imread(filename)
                            color_bleach_factor = color_bleach
                            color_bleach = np.nan_to_num(color_bleach, posinf=65500)
                            coefficent = self.autof_factor_estimator(color_factor, color_bleach_factor)
                            #color_subbed = color_im - coefficent * color_bleach
                            #color_subbed = color_reg - color_bleach
                            color_subbed = color_im - color_bleach
                            color_subbed[color_subbed < 0] = 0
                            #color_subbed = np.nan_to_num(color_subbed, posinf= 65500)

                            # save

                            #save_path = experiment_directory + channel + r'/Stain/cy_' + str(cycle) + '\Tiles/focused_flattened_subbed'
                            save_path = experiment_directory + channel + r'/Stain/cy_' + str(cycle) + '\Tiles/focused_subbed'
                            try:
                                os.chdir(save_path)
                            except:
                                os.mkdir(save_path)
                                os.chdir(save_path)

                            subbed_filename = 'x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                            # bleach_filename ='x' + str(x) + '_y_' + str(y) + '_c_' + channel + '_bleach.tif'
                            # reg_filename = 'x' + str(x) + '_y_' + str(y) + '_c_' + channel + '_registered.tif'
                            tf.imwrite(subbed_filename, color_subbed)
                            # io.imsave(bleach_filename, color_bleach)
                            # io.imsave(reg_filename, color_reg)


                    if rolling_ball == 1:

                        for channel in channels:
                            stain_color_path = experiment_directory + channel + r'/Stain/cy_' + str(
                                cycle) + '\Tiles/focused'
                            os.chdir(stain_color_path)
                            filename = 'x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                            color_im = io.imread(filename)
                            color_im = np.nan_to_num(color_im, posinf=65500)
                            color_im = color_im.astype('uint16')

                            # sub background color channels

                            #color_background = restoration.rolling_ball(color_im, radius=10)
                            foot = morphology.square(50)
                            color_subbed = morphology.white_tophat(color_im, foot)
                            #color_subbed = color_im - color_background
                            color_subbed[color_subbed < 0] = 0
                            color_subbed = color_subbed.astype('uint16')
                            color_subbed = np.nan_to_num(color_subbed, posinf=65500)

                            # save

                            save_path = experiment_directory + channel + r'/Stain/cy_' + str(
                                cycle) + '\Tiles/focused/background_subbed_rolling'
                            try:
                                os.chdir(save_path)
                            except:
                                os.mkdir(save_path)
                                os.chdir(save_path)

                            subbed_filename = 'x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                            tf.imwrite(subbed_filename, color_subbed)

                else:
                    pass

    def darkframe_sub(self, experiment_directory, cycle_number):
        '''


        :param experimental_directory:
        :param cycle_number:
        :return:
        '''

        experiment_directory = experiment_directory + '/'


        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)
        tissue_exist = np.load('tissue_exist.npy', allow_pickle=False)


        y_tile_count = np.shape(fm_array[0])[0]
        x_tile_count = np.shape(fm_array[0])[1]

        block_y_pixels = 75
        block_x_pixels = 75

        channels = ['DAPI', 'A488', 'A555', 'A647']

        for y in range(0, y_tile_count):
            for x in range(0, x_tile_count):

                if tissue_exist[y][x] == 1:
                    for channel in channels:
                        stain_color_path = experiment_directory + channel + r'/Stain/cy_' + str(cycle_number) + '\Tiles/focused_basic_corrected'
                        #stain_color_path = experiment_directory + channel + r'/Stain/cy_' + str(cycle_number) + '\Tiles/focused'
                        darkframe_color_path = experiment_directory + channel + r'/Stain/cy_' + str(cycle_number) + '\Tiles/focused_basic_darkframe'

                        os.chdir(stain_color_path)
                        filename = 'x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                        color_im = io.imread(filename)


                        darkframe_im = self.dark_frame_generate(color_im, block_y_pixels, block_x_pixels)
                        darkframe_subbed_im = color_im - darkframe_im
                        darkframe_subbed_im[darkframe_subbed_im < 0] = 0

                        try:
                            os.mkdir(darkframe_color_path)
                            os.chdir(darkframe_color_path)
                        except:
                            os.chdir(darkframe_color_path)

                        io.imsave(filename, darkframe_subbed_im)

                else:
                    pass

    def brightness_uniformer(self, experiment_directory, cycle_number, cluster_number):

        experiment_directory = experiment_directory + '/'
        #import numpy focus map info
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)
        channels = ['DAPI', 'A488', 'A555', 'A647']

        #predetermine step sizes for search range on tile multiplier
        step_count = 5
        multiplier_mod_values = np.linspace(0.8, 1.2, step_count)

        #define desired number iterations through tiles
        iteration_count = 1

        tissue_fm = fm_array[10]
        y_tile_count = np.shape(fm_array[0])[0]
        x_tile_count = np.shape(fm_array[0])[1]

        #make arrays +2 in each dimension. This creates a border region around the original array
        #this makes each tile have 4 'neighbors' even if the neighbor = 0 everywhere
        overlap_images = np.zeros([y_tile_count + 2, x_tile_count + 2, 4, 2, 2960, 296])
        multiplier_array = np.zeros([y_tile_count + 2, x_tile_count + 2])

        #bound_width
        x_width = 296
        y_width = 296

        for channel in channels:

            #paths to needed folders

            dapi_file_path = experiment_directory + '/' + channel + '/Stain/cy_' + str(cycle_number) + '/Tiles/subbed_focused_basic_corrected'

            if channel == 'DAPI':
                channel_file_path = dapi_file_path
            else:
                channel_file_path = experiment_directory + '/' + channel + '/Stain/cy_' + str(cycle_number) + '/Tiles/subbed_focused_basic_corrected'
            tissue_path = experiment_directory + '/Tissue_Binary'
            if channel == 'DAPI':
                channel_output_path = experiment_directory + '/' + channel + '/Stain/cy_' + str(cycle_number) + '/Tiles/subbed_focused_basic_brightness_corrected'
            else:
                channel_output_path = experiment_directory + '/' + channel + '/Stain/cy_' + str(cycle_number) + '/Tiles/subbed_focused_basic_brightness_corrected'


            os.chdir(channel_file_path)
            try:
                os.mkdir(channel_output_path)
            except:
                pass

            #populate overlap image array and multipier arrays
            for x in range(0, x_tiles):
                for y in range(0, y_tiles):
                    tile_list = self.tissue_fm_decode(tissue_fm[y][x])
                    if cluster_number in tile_list:

                        multiplier_array[y][x] = 1
                        #load in image
                        os.chdir(channel_file_path)
                        filename = 'x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                        image = io.imread(filename)

                        os.chdir(dapi_file_path)
                        filename = 'x' + str(x) +'_y_' + str(y) +'_tissue.tif'
                        dapi_im = io.imread(filename)

                        #interlace into overlap array (north=0, east=1, south=2, west=3)
                        #1 is added to each of x and y to take blank border into account

                        overlap_images[y+1][x+1][0][0] = np.transpose(dapi_im[0:y_width, ::])
                        overlap_images[y+1][x+1][1][0] = dapi_im[::, -x_width:-1]
                        overlap_images[y+1][x+1][2][0] = np.transpose(dapi_im[-y_width:-1, ::])
                        overlap_images[y+1][x+1][3][0] = dapi_im[::, 0:x_width]

                        overlap_images[y+1][x+1][0][1] = np.transpose(image[0:y_width, ::])
                        overlap_images[y+1][x+1][1][1] = image[::, -296:-1]
                        overlap_images[y+1][x+1][2][1] = np.transpose(image[-y_width:-1, ::])
                        overlap_images[y+1][x+1][3][1] = image[::, 0:x_width]
                    else:
                        pass


                        #make new folder and save brightness uniformed images



            for i in range(0, iteration_count):
                for x in range(0, x_tiles):
                    for y in range(0, y_tiles):
                        if cluster_number in tile_list:

                            tile_north = np.multiply(overlap_images[y+1][x+1][0][1],multiplier_array[y+1][x+1])
                            tile_east =np.multiply(overlap_images[y+1][x+1][1][1],multiplier_array[y+1][x+1])
                            tile_south = np.multiply(overlap_images[y+1][x+1][2][1],multiplier_array[y+1][x+1])
                            tile_west = np.multiply(overlap_images[y+1][x+1][3][1],multiplier_array[y+1][x+1])

                            adjacent_north = np.multiply(overlap_images[y][x+1][0][1],multiplier_array[y][x+1])
                            adjacent_east = np.multiply(overlap_images[y + 1][x + 2][1][1],multiplier_array[y + 1][x + 2])
                            adjacent_south = np.multiply(overlap_images[y + 2][x + 1][2][1],multiplier_array[y + 2][x + 1])
                            adjacent_west = np.multiply(overlap_images[y][x + 1][3][1],multiplier_array[y][x + 1])

                            scores = np.zeros(step_count).astype('float64')
                            for m in range(0, step_count):

                                modded_multiplier = multiplier_mod_values[m]

                                mod_north = np.multiply(tile_north, modded_multiplier)
                                mod_east = np.multiply(tile_east, modded_multiplier)
                                mod_south = np.multiply(tile_south, modded_multiplier)
                                mod_west = np.multiply(tile_west, modded_multiplier)

                                mod_north_difference = np.mean(mod_north - adjacent_north)
                                mod_east_difference = np.mean(mod_east - adjacent_east)
                                mod_south_difference = np.mean(mod_south - adjacent_south)
                                mod_west_difference = np.mean(mod_west - adjacent_west)

                                score = np.sqrt(mod_north_difference**2 + mod_south_difference**2 + mod_east_difference**2 + mod_west_difference**2)
                                scores[m] = score

                            #find better multiplier mod value and update total multiplier
                            lowest_score = np.min(scores)
                            lowest_index = int(np.where(scores == lowest_score)[0][0])
                            best_mod_multiplier = multiplier_mod_values[lowest_index]
                            multiplier_array[y + 1][x + 1] = multiplier_array[y+1][x+1] * best_mod_multiplier


                        else:
                            pass

    def delete_intermediate_folders(self, experiment_directory, cycle_number):
        '''
        Goes through each channel (DAPI, a488, 555, and 647) and deletes every folder used to store
        processed file. Only retains raw images.
        :param experiment_directory:
        :param cycle_number:
        :return:
        '''

        channels = ['DAPI', 'A488', 'A555', 'A647']
        folder_names = ['/focused_basic_corrected', '/focused_basic_darkframe' ]

        for channel in channels:
            for folder_name in folder_names:
                folder_path = experiment_directory + channel + r'/Stain/cy_' + str(cycle_number) + '\Tiles' + folder_name

                try:
                    shutil.rmtree(folder_path)
                    print(f"Folder '{folder_path}' deleted successfully.")
                except FileNotFoundError:
                    print(f"Folder '{folder_path}' not found.")
                except Exception as e:
                    print(f"An error occurred: {e}")

    def zlib_compress_raw(self, experiment_directory, cycle_number):
        '''
        Uses zlib codec to compress raw images and resave them.
        :param experiment_directory:
        :param cycle_number:
        :return:
        '''

        experiment_directory = experiment_directory + '/'

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)
        tissue_exist = np.load('tissue_exist.npy', allow_pickle=False)

        y_tile_count = np.shape(fm_array[0])[0]
        x_tile_count = np.shape(fm_array[0])[1]


        channels = ['DAPI', 'A488', 'A555', 'A647']
        types_images = ['Stain', 'Bleach']

        for y in range(0, y_tile_count):
            for x in range(0, x_tile_count):

                if tissue_exist[y][x] == 1:
                    for channel in channels:
                        for type in types_images:
                            raw_path = experiment_directory + channel + r'/' + type + '/cy_' + str(cycle_number) + '\Tiles'

                            # load in raw image
                            os.chdir(raw_path)
                            filename = 'x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                            raw_im = io.imread(filename)

                            #zlib compress and resave image
                            tifffile.imwrite('compressed_image.tif', image, compression='zlib',compressionargs={'level': 10}, predictor=True)

    def block_proc_min(self, array, block_y_pixels, block_x_pixels):
        '''
        breaks image up into blocks of size (block_y_pixels x block_x_pixels
        and then returns array that is the minimum of each block and is
        effectively down sampled by the block size

        :param array:
        :param block_y_pixels:
        :param block_x_pixles:
        :return:
        '''

        y_pixels = np.shape(array)[0]
        x_pixels = np.shape(array)[1]

        y_num_blocks = int(y_pixels / block_y_pixels)
        x_num_blocks = int(x_pixels / block_x_pixels)

        im = array

        blocked_array = np.lib.stride_tricks.as_strided(im, shape=(x_num_blocks, y_num_blocks, block_x_pixels, block_y_pixels), strides=(im.strides[0] * block_x_pixels, im.strides[1] * block_y_pixels, im.strides[0], im.strides[1]))
        min = np.min(blocked_array, axis=2)
        min = np.min(min, axis=2)

        return min

    def block_proc_reshaper(self, array, block_y_pixels, block_x_pixels):
        '''
        takes in array and evenly clips off rows and columns to to their counts be integers
        when divided by block size pixel counts. Does no padding, only shrinks.
        :param array:
        :param block_y_pixels:
        :param block_x_pixels:
        :return:
        '''

        y_pixels = np.shape(array)[0]
        x_pixels = np.shape(array)[1]

        y_num_blocks = math.floor(y_pixels / block_y_pixels)
        x_num_blocks = math.floor(x_pixels / block_x_pixels)

        adjusted_y_pixels = block_y_pixels * y_num_blocks
        adjusted_x_pixels = block_x_pixels * x_num_blocks

        clipped_y_pixel_num = y_pixels - adjusted_y_pixels
        clipped_x_pixel_num = x_pixels - adjusted_x_pixels

        # deal with even and odd cases for y axis
        if clipped_y_pixel_num % 2 == 0:
            top_rows_2_clip = int(clipped_y_pixel_num / 2)
            bottom_rows_2_clip = int(clipped_y_pixel_num / 2)
            bottom_adjusted_row_num = y_pixels - bottom_rows_2_clip
            array = array[top_rows_2_clip:bottom_adjusted_row_num, ::]
        elif clipped_y_pixel_num % 2 != 0:
            top_rows_2_clip = int(clipped_y_pixel_num / 2)
            bottom_rows_2_clip = int((clipped_y_pixel_num / 2) + 1)
            bottom_adjusted_row_num = y_pixels - bottom_rows_2_clip
            array = array[top_rows_2_clip:bottom_adjusted_row_num, ::]

        # deal with even and odd cases for x axis
        if clipped_x_pixel_num % 2 == 0:
            left_col_2_clip = int(clipped_x_pixel_num / 2)
            right_col_2_clip = int(clipped_x_pixel_num / 2)
            right_adjusted_col_num = int(x_pixels - right_col_2_clip)
            array = array[::, left_col_2_clip:right_adjusted_col_num]
        elif clipped_x_pixel_num % 2 != 0:
            left_col_2_clip = int(clipped_x_pixel_num / 2)
            right_col_2_clip = int((clipped_x_pixel_num / 2) + 1)
            right_adjusted_col_num = int(x_pixels - right_col_2_clip)
            array = array[::, left_col_2_clip:right_adjusted_col_num]

        return array

    def dark_frame_generate(self, array, block_y_pixels, block_x_pixels):
        '''

        :param array:
        :param block_y_pixels:
        :param block_x_pixels:
        :return:
        '''

        original_y_pixels = np.shape(array)[0]
        original_x_pixels = np.shape(array)[1]

        mean_kernal_y = math.ceil(block_y_pixels/10)
        mean_kernal_x = math.ceil(block_x_pixels / 10)

        array_mean = cv2.blur(array, (mean_kernal_y, mean_kernal_x))

        adjusted_array = self.block_proc_reshaper(array_mean, block_y_pixels, block_x_pixels)
        min_array = self.block_proc_min(adjusted_array, block_y_pixels, block_x_pixels)
        resized_min_array = transform.resize(min_array, (original_y_pixels, original_x_pixels), preserve_range=True,anti_aliasing=True)
        resized_min_array = filters.butterworth(resized_min_array, cutoff_frequency_ratio=0.005, high_pass=False, order=2, npad=1000)

        return resized_min_array

    def max_projector(self, experiment_directory, cycle_number, x_frame_size):

        experiment_directory = experiment_directory + '/'
        # import numpy focus map info
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)
        tissue_exist = np.load('tissue_exist.npy', allow_pickle=False)
        channels = ['DAPI', 'A488', 'A555', 'A647']


        #determine number of z slices and x and y tile counts
        z_slice_count = fm_array[3][0][0].astype('int16')
        y_tile_count = np.shape(fm_array[0])[0]
        x_tile_count = np.shape(fm_array[0])[1]

        for channel in channels:

            for y in range(0, y_tile_count):
                for x in range(0, x_tile_count):
                    if tissue_exist[y][x] ==1:
                        file_path = experiment_directory + '/' + channel + '/Stain' + '/cy_' + str(
                            cycle_number) + '/Tiles'
                        os.chdir(file_path)
                        # make z_stack for channel and tile
                        z_stack = np.ones((z_slice_count, 2960, x_frame_size))
                        for z in range(0, z_slice_count):

                            file_name = 'z_' + str(z) + r'_x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                            slice = io.imread(file_name)
                            z_stack[z] = slice

                        #make max projection
                        max_proj = np.max(z_stack, axis=0)

                        #save projection
                        saving_path = experiment_directory + '/' + channel + '/Stain' + '/cy_' + str(cycle_number) + '/Tiles' + '/max_projection'
                        try:
                            os.mkdir(saving_path)
                            os.chdir(saving_path)
                        except:
                            os.chdir(saving_path)
                        file_name = 'x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                        io.imsave(file_name, max_proj)

                    else:
                        pass

    def wavelet_background_sub(self, experiment_directory, cycle_number, resolution_px, noise_lvl):

        experiment_directory = experiment_directory + '/'
        # import numpy focus map info
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)
        tissue_exist = np.load('tissue_exist.npy', allow_pickle=False)
        channels = ['DAPI', 'A488', 'A555', 'A647']

        y_tile_count = np.shape(fm_array[0])[0]
        x_tile_count = np.shape(fm_array[0])[1]

        for channel in channels:
            for x in range(0, x_tile_count):
                for y in range(0, y_tile_count):

                    if tissue_exist[y][x] == 1:

                        image_path = experiment_directory + '/' + channel + '/Stain' + '/cy_' + str(cycle_number) + '/Tiles' + '/focused'
                        os.chdir(image_path)

                        file_name = 'x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                        image = io.imread(file_name) + 1000
                        image = np.nan_to_num(image, posinf=65000)
                        processed_image = self.run_wavelet(image, resolution_px, noise_lvl)

                        saving_path = experiment_directory + '/' + channel + '/Stain' + '/cy_' + str(cycle_number) + '/Tiles' + '/wavelet_subbed'
                        try:
                            os.mkdir(saving_path)
                            os.chdir(saving_path)
                        except:
                            os.chdir(saving_path)
                        io.imsave(file_name, processed_image)

                    else:
                        pass

    def wavelet_based_BG_subtraction_function(self, image, num_levels, noise_lvl):

        coeffs = wavedecn(image, 'db1', level=num_levels)  # decomposition
        coeffs2 = coeffs.copy()

        for BGlvl in range(1, num_levels):
            coeffs[-BGlvl] = {k: np.zeros_like(v) for k, v in coeffs[-BGlvl].items()}  # set lvl 1 details  to zero

        Background = waverecn(coeffs, 'db1')  # reconstruction
        del coeffs
        BG_unfiltered = Background
        Background = gaussian_filter(Background, sigma=2 ** num_levels)  # gaussian filter sigma = 2^#lvls

        coeffs2[0] = np.ones_like(coeffs2[0])  # set approx to one (constant)
        for lvl in range(1, num_levels - noise_lvl):
            coeffs2[lvl] = {k: np.zeros_like(v) for k, v in coeffs2[lvl].items()}  # keep first detail lvl only
        Noise = waverecn(coeffs2, 'db1')  # reconstruction
        del coeffs2

        return Background, Noise, BG_unfiltered

    def run_wavelet(self, image, resolution_px = 100, noise_lvl = 1):

        img_type = image.dtype
        # number of levels for background estimate
        num_levels = np.uint16(np.ceil(np.log2(resolution_px)))

        image = np.array(image, dtype='float32')

        # image = np.array(io.imread(os.path.join(data_dir, file)),dtype = 'float32')
        if np.ndim(image) == 2:
            shape = np.shape(image)
            image = np.reshape(image, [1, shape[0], shape[1]])
        shape = np.shape(image)
        if shape[1] % 2 != 0:
            image = np.pad(image, ((0, 0), (0, 1), (0, 0)), 'edge')
            pad_1 = True
        else:
            pad_1 = False
        if shape[2] % 2 != 0:
            image = np.pad(image, ((0, 0), (0, 0), (0, 1)), 'edge')
            pad_2 = True
        else:
            pad_2 = False

        # extract background and noise
        num_cores = multiprocessing.cpu_count()  # number of cores on your CPU
        res = Parallel(n_jobs=num_cores, max_nbytes=None)(
            delayed(self.wavelet_based_BG_subtraction_function)(image[slice], num_levels, noise_lvl) for slice in
            range(np.size(image, 0)))
        Background, Noise, BG_unfiltered = zip(*res)

        # convert to float64 numpy array
        Noise = np.asarray(Noise, dtype='float32')
        Background = np.asarray(Background, dtype='float32')
        BG_unfiltered = np.asarray(BG_unfiltered, dtype='float32')

        # undo padding
        if pad_1:
            image = image[:, :-1, :]
            Noise = Noise[:, :-1, :]
            Background = Background[:, :-1, :]
            BG_unfiltered = BG_unfiltered[:, :-1, :]
        if pad_2:
            image = image[:, :, :-1]
            Noise = Noise[:, :, :-1]
            Background = Background[:, :, :-1]
            BG_unfiltered = BG_unfiltered[:, :, :-1]

        BG_unfiltered = np.asarray(BG_unfiltered, dtype=img_type.name)
        Background = np.asarray(Background, dtype=img_type.name)

        result = image - Background
        result[result < 0] = 0  # positivity constrait
        result = np.asarray(result, dtype=img_type.name)
        noisy_sig = result

        # correct noise
        Noise[Noise < 0] = 0  # positivity constraint
        noise_threshold = np.mean(Noise) + 2 * np.std(Noise)
        Noise[Noise > noise_threshold] = noise_threshold  # 2 sigma threshold reduces artifacts

        # subtract Noise
        result = image - Background
        result = result - Noise
        result[result < 0] = 0  # positivity constraint
        result = np.asarray(result, dtype=img_type.name)

        return result

    def zc_save(self, zc_tif_stack, channels, x_tile, y_tile, cycle, x_pixels, experiment_directory, Stain_or_Bleach):

        z_tile_count = np.shape(zc_tif_stack)[1]

        for channel in channels:
            if channel == 'DAPI':
                zc_index = 0
            if channel == 'A488':
                zc_index = 1
            if channel == 'A555':
                zc_index = 2
            if channel == 'A647':
                zc_index = 3

            # establish x range to collect in image

            #side_pixel_count = int(5056 - x_pixels)

            save_directory = experiment_directory + '/' + str(channel) + '/' + Stain_or_Bleach + '/' + 'cy_' + str(
                cycle) + '/' + 'Tiles'
            os.chdir(save_directory)

            for z in range(0, z_tile_count):
                file_name = 'z_' + str(z) + '_x' + str(x_tile) + '_y_' + str(y_tile) + '_c_' + str(channel) + '.tif'
                image = zc_tif_stack[zc_index][z]
                imwrite(file_name, image, photometric='minisblack')

    def tissue_exist_array_generate(self, experiment_directory, x_frame_size):

        tissue_path = experiment_directory + '/Tissue_Binary'

        # load numpy arrays in
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]

        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]

        # make object to hold all tissue binary maps
        tissue_binary_stack = np.random.rand(y_tile_count, x_tile_count, 2960, x_frame_size).astype('uint16')
        for x in range(0, x_tile_count):
            for y in range(0, y_tile_count):
                os.chdir(tissue_path)
                file_name = 'x' + str(x) + '_y_' + str(y) + '_label_tissue.tif'
                image = io.imread(file_name)
                tissue_binary_stack[y][x] = image

        # make object that hold info as to if the image has tissue in it or not
        tissue_or_not = np.random.rand(y_tile_count, x_tile_count).astype('uint16')
        for x in range(0, x_tile_count):
            for y in range(0, y_tile_count):
                if tissue_binary_stack[y][x].sum() > 0:
                    tissue_or_not[y][x] = 1
                if tissue_binary_stack[y][x].sum() == 0:
                    tissue_or_not[y][x] = 0

        os.chdir(numpy_path)
        np.save('tissue_exist.npy', tissue_or_not)

    def tissue_fm_decode(self, tissue_fm_yx):
        '''
        Takes in number from tissue_fm array and gives back
        what unique cluster numbers are contained within it (will not output 0)
        :param tissue_fm_yx: tissue_fm[y][x] value
        :return:
        '''


        #unique numbers

        unique_number_list = []
        while tissue_fm_yx > 1:
            power = math.floor(math.log10(tissue_fm_yx))
            unique_number_list.append(power)
            sci_number_str = '1e+' + str(power)
            sci_number = float(sci_number_str)
            tissue_fm_yx -= sci_number

        return unique_number_list

    def number_tiles_each_cluster(self, experiment_directory):
        '''
        Gives list back that contains the amount of tiles in each cluster
        :param experiment_directory:
        :return:
        '''

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        fm_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = fm_array[0]
        numpy_y = fm_array[1]

        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]

        tissue_fm = fm_array[10]

        # determine max cluster count
        highest_number = np.max(tissue_fm)

        number_clusters = math.floor(math.log10(highest_number))

        cluster_tile_count = np.zeros(number_clusters)

        for x in range(0, x_tile_count):
            for y in range(0, y_tile_count):

                tissue_code_number = tissue_fm[y][x]
                clusters_in_tile = self.tissue_fm_decode(tissue_code_number)
                if len(clusters_in_tile) > 0:
                    for number in clusters_in_tile:
                        number = int(number)
                        cluster_tile_count[number - 1] += 1
                else:
                    pass

        return cluster_tile_count


    ######Kinetics and its functions#####################################################

    def kinetic_autofocus(self, experiment_directory, focus_z_position, planes):
        '''
        Takes planes number of images, split in half above and below focus position. Gives back plane position
        that had highest brenner score with skip 17.
        :param experiment_directory: where the experiment files are contained
        :param focus_z_position: expected focus position in the z axis
        :param planes: number planes to acquire for focusing purposes
        :return:
        '''


        os.chdir(experiment_directory)
        try:
            os.mkdir('nuclei_focus_images')
        except:
            pass

        focus_image_path = experiment_directory + '/nuclei_focus_images'
        os.chdir(focus_image_path)

        # set channel and exposure times in ms
        a488_2_dapi_offset = 8
        core.set_config("Color", 'DAPI')
        core.set_exposure(75)
        core.snap_image()
        tagged_image = core.get_tagged_image()
        time.sleep(10)

        # make image stack object
        dapi_tif_stack = np.random.rand(planes, 2960, 5056).astype('float16')

        # find z range to be scanned
        dapi_focus_z_position = focus_z_position + a488_2_dapi_offset
        slice_step_size = 1 # in microns
        bottom_z = dapi_focus_z_position - int(planes/2) * slice_step_size #remember int() rounds down
        top_z = dapi_focus_z_position + int(planes/2) * slice_step_size #remember int() rounds down
        image_slice_counter = 0

        #capture images for stack
        for z in range(bottom_z, top_z + slice_step_size, slice_step_size):

            core.set_position(z)
            time.sleep(0.3)
            core.snap_image()
            tagged_image = core.get_tagged_image()
            pixels = np.reshape(tagged_image.pix,newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
            dapi_tif_stack[image_slice_counter] = pixels
            image_slice_counter += 1

        #save images in folder
        filename = 'dapi_autofocus.tif'
        io.imsave(filename, dapi_tif_stack)

        scores = np.random.rand(planes).astype('float64')

        for plane in range(0, planes):

            score = self.focus_score_post_processing(dapi_tif_stack[plane], 17)
            scores[plane] = score

        #find highest brenner score index
        max_score = np.max(scores)
        focus_index = np.where(scores == max_score)
        #make range of z values scanned
        dapi_scan_range = np.linspace(bottom_z, top_z, num=planes)
        # find focus value
        dapi_in_focus_z_position = dapi_scan_range[focus_index]
        a488_in_focus_z_position = int(dapi_in_focus_z_position - a488_2_dapi_offset)

        return a488_in_focus_z_position


    def antibody_kinetics(self, experiment_directory, capture_rate_staining, duration_staining, stain_valve, fluidic_object, channels=['DAPI', 'A488', 'A555', 'A647']):

        '''
        Dispenses stain avia fluidic system and images until certain time point. After that, bleach dispenses and capturing continues at a different
        rate until set time point


        :param experiment_directory:
        :param capture_rate_staining: points/min
        :param capture_rate_bleaching: points/min
        :param duration_staining: minutes
        :param duration_bleaching: minutes
        :param channels:
        :return:
        '''

        exp_array = [50, 50, 50, 50]
        A488_to_channels_offset = [8, 0, 0, -1]


        # create folders

        # make parent folder for experiment if it isnt made
        os.chdir(r'E:')
        try:
            os.mkdir(experiment_directory)
        except:
            pass

        os.chdir(experiment_directory)
        try:
            os.mkdir('exposure_times')
        except:
            pass
        try:
            os.mkdir('DAPI')
        except:
            pass
        try:
            os.mkdir('A488')
        except:
            pass
        try:
            os.mkdir('A555')
        except:
            pass
        try:
            os.mkdir('A647')
        except:
            pass
        try:
            os.mkdir('DAPI_hdr')
        except:
            pass
        try:
            os.mkdir('A488_hdr')
        except:
            pass
        try:
            os.mkdir('A555_hdr')
        except:
            pass
        try:
            os.mkdir('A647_hdr')
        except:
            pass
        try:
            os.mkdir('auto_fluorescence')
        except:
            pass
        try:
            os.mkdir('post_wash')
        except:
            pass

        dapi_path = experiment_directory + r'\DAPI'
        a488_path = experiment_directory + r'\A488'
        a555_path = experiment_directory + r'\A555'
        a647_path = experiment_directory + r'\A647'

        DAPI_path = experiment_directory + r'\DAPI_hdr'
        A488_path = experiment_directory + r'\A488_hdr'
        A555_path = experiment_directory + r'\A555_hdr'
        A647_path = experiment_directory + r'\A647_hdr'

        self.hdr_exp_generator(experiment_directory, threshold_level=10000, max_exp=700, min_exp=20)

        # dimensional parameters
        y_pixel_count = 2960
        x_pixel_count = 5056
        channel_count = len(channels)
        time_point_stain_count = int(duration_staining * capture_rate_staining)
        time_gap_staining = 1 / capture_rate_staining * 60
        print('time gap stain', time_gap_staining)

        # find current focus position
        focus_position = core.get_position()
        print('focus position acquired')
        focus_position = int(focus_position)



        # create data structure for staining images
        data_points_stain = np.full((time_point_stain_count, channel_count, y_pixel_count, x_pixel_count), 0)
        #fluidic_object.valve_select(stain_valve)
        #print('valve selected')

        #capture pre stain images (so autofluorescence)



        os.chdir(experiment_directory + r'\auto_fluorescence')
        for channel in channels:

            if channel == 'DAPI':
                channel_index = 0
            if channel == 'A488':
                channel_index = 1
            if channel == 'A555':
                channel_index = 2
            if channel == 'A647':
                channel_index = 3

            exp_time = exp_array[channel_index]
            channel_offset = A488_to_channels_offset[channel_index]
            core.set_position(focus_position + channel_offset)
            time.sleep(0.5)

            core.set_config("Color", channel)
            core.set_exposure(exp_time)

            print('capturing pre image', channel)

            core.snap_image()
            tagged_image = core.get_tagged_image()
            pixels = np.reshape(tagged_image.pix,newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
            io.imsave(channel + '.tif', pixels)

            hdr_im = self.core_capture(experiment_directory, x_pixel_count, channel)
            io.imsave(channel + '_HDR.tif', hdr_im)

        os.chdir(experiment_directory)

        #start flow
        fluidic_object.liquid_action('Stain_flow_on', stain_valve=stain_valve)
        print('flowed stain')
        #time.sleep(45)
        #fluidic_object.flow(-3)
        #print('flow stain ended')
        #fluidic_object.valve_select(12)

        # acquire kinetic time points
        print('total time points', time_point_stain_count)
        for time_point in range(0, time_point_stain_count):
            print('time point', time_point)
            start_time = time.time()

            # run auto focus and set new focus position
            focus_position = self.kinetic_autofocus(experiment_directory, focus_z_position= focus_position, planes = 11)
            print('focus position', focus_position)

            for channel in channels:

                if channel == 'DAPI':
                    channel_index = 0
                    hdr_path = DAPI_path
                if channel == 'A488':
                    channel_index = 1
                    hdr_path = A488_path
                if channel == 'A555':
                    channel_index = 2
                    hdr_path = A555_path
                if channel == 'A647':
                    channel_index = 3
                    hdr_path = A647_path

                exp_time = exp_array[channel_index]

                channel_offset = A488_to_channels_offset[channel_index]
                core.set_position(focus_position + channel_offset)
                time.sleep(0.3)

                core.set_config("Color", channel)
                core.set_exposure(exp_time)

                core.snap_image()
                tagged_image = core.get_tagged_image()
                pixels = np.reshape(tagged_image.pix,newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
                data_points_stain[time_point][channel_index] = pixels

                #hdr_im = self.core_capture(experiment_directory, x_pixel_count, channel)
                os.chdir(hdr_path)
                io.imsave(channel + '_t_' + str(time_point) + '.tif', pixels)

            #adjust for time to acquire to make time between frames = time_gap_staining
            end_time = time.time()
            time_elapsed = end_time - start_time
            time.sleep(time_gap_staining - time_elapsed)

        # save images
        print('saving')
        os.chdir(dapi_path)
        tf.imwrite('dapi_stain_stack', data_points_stain[::, 0, ::, ::])
        os.chdir(a488_path)
        tf.imwrite('a488_stain_stack', data_points_stain[::, 1, ::, ::])
        os.chdir(a555_path)
        tf.imwrite('a555_stain_stack', data_points_stain[::, 2, ::, ::])
        os.chdir(a647_path)
        tf.imwrite('a647_stain_stack', data_points_stain[::, 3, ::, ::])
        print('done saving')

        # start flow of PBS
        print('starting wash')
        fluidic_object.liquid_action('Wash')
        #fluidic_object.valve_select(12)
        #fluidic_object.flow(500)
        #time.sleep(45)
        #fluidic_object.flow(-3)
        print('wash completed')

        # acquire post wash images


        os.chdir(experiment_directory + r'/post_wash')
        for channel in channels:

            if channel == 'DAPI':
                channel_index = 0
            if channel == 'A488':
                channel_index = 1
            if channel == 'A555':
                channel_index = 2
            if channel == 'A647':
                channel_index = 3

            exp_time = exp_array[channel_index]
            channel_offset = A488_to_channels_offset[channel_index]
            core.set_position(focus_position  + channel_offset)
            time.sleep(0.3)

            core.set_config("Color", channel)
            core.set_exposure(exp_time)

            core.snap_image()
            tagged_image = core.get_tagged_image()
            pixels = np.reshape(tagged_image.pix,newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])

            #hdr_im = self.core_capture(experiment_directory, x_pixel_count, channel)
            io.imsave(channel + '.tif', pixels)
            #io.imsave(channel + '_hdr.tif', hdr_im)

