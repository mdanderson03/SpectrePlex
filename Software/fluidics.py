import datetime
import os
import math
import kasa
import numpy as np
import time
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import sys
from ctypes import *

sys.path.append(
    r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF\Python_64_elveflow\DLL64')  # add the path of the library here
sys.path.append(r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF\Python_64_elveflow')  # add the path of the LoadElveflow.py
sys.path.append(r'C:\Users\mike\Documents\GitHub\AutoCIF\Python_64_elveflow\DLL64')  # add the path of the library here
sys.path.append(r'C:\Users\mike\Documents\GitHub\AutoCIF\Python_64_elveflow')  # add the path of the LoadElveflow.py

from array import array
from Elveflow64 import *


class fluidics:

    def __init__(self, experiment_path, mux_com_port, ob1_com_port, flow_control=1):

        # OB1 initialize
        ob1_path = 'ASRL' + str(ob1_com_port) + '::INSTR'
        Instr_ID = c_int32()
        OB1_Initialization(ob1_path.encode('ascii'), 0, 0, 0, 0, byref(Instr_ID))
        OB1_Add_Sens(Instr_ID, 1, 5, 1, 0, 7,
                     0)  # 16bit working range between 0-1000uL/min, also what are CustomSens_Voltage_5_to_25 and can I really choose any digital range?

        Calib_path = r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF\Python_64_elveflow\calibration\1_12_24_cal.txt'
        Calib = (c_double * 1000)()
        # Elveflow_Calibration_Load(Calib_path.encode('ascii'), byref(Calib), 1000)
        Elveflow_Calibration_Default(byref(Calib), 1000)

        if flow_control == 1:

            set_channel_regulator = int(1)  # convert to int
            set_channel_regulator = c_int32(set_channel_regulator)  # convert to c_int32
            set_channel_sensor = int(1)
            set_channel_sensor = c_int32(set_channel_sensor)  # convert to c_int32
            PID_Add_Remote(Instr_ID.value, set_channel_regulator, Instr_ID.value, set_channel_sensor, 0.9/22.5, 0.004/.03846, 1)
        else:
            pass

        OB1_Start_Remote_Measurement(Instr_ID.value, byref(Calib), 1000)
        self.calibration_array = byref(Calib)

        # MUX intiialize
        path = 'ASRL' + str(mux_com_port) + '::INSTR'
        mux_Instr_ID = c_int32()
        MUX_DRI_Initialization(path.encode('ascii'), byref(
            mux_Instr_ID))  # choose the COM port, it can be ASRLXXX::INSTR (where XXX=port number)

        # home
        # answer = (c_char * 40)()
        self.mux_ID = mux_Instr_ID.value
        # MUX_DRI_Send_Command(self.mux_ID, 0, answer, 40)

        self.pump_ID = Instr_ID.value
        self.experiment_path = experiment_path
        self.flow_control = flow_control

        self.pressure_on = 1000
        self.pressure_off = 0
        self.flow_on = 500
        self.flow_off = -3

        return

    def fluidics_logger(self, function_used_string, error_code, value_sr):
        experiment_directory = self.experiment_path
        filename = 'logger.xlsx'
        logger_path = experiment_directory + '/fluidics data logger'
        os.chdir(experiment_directory)
        try:
            os.mkdir('fluidics data logger')
            os.chdir(logger_path)
        except:
            os.chdir(logger_path)

        if os.path.isfile(filename) == True:
            wb = load_workbook(filename)
            ws = wb.active
        elif os.path.isfile(filename) == False:
            wb = Workbook()
            ws = wb.active
            # add headers
            ws.cell(row=1, column=1).value = 'Time Stamp'
            ws.cell(row=1, column=2).value = 'Function Used'
            ws.cell(row=1, column=3).value = 'Error Code'
            ws.cell(row=1, column=4).value = 'Value Sent/Recieved'

        # determine row number (add to next line as a logged event)
        current_max_row = ws.max_row
        row_select = current_max_row + 1

        # add in values
        ws.cell(row=row_select, column=1).value = datetime.datetime.now()
        ws.cell(row=row_select, column=2).value = function_used_string
        ws.cell(row=row_select, column=3).value = error_code
        ws.cell(row=row_select, column=4).value = value_sr

        wb.save(filename)

    def mux_end(self):

        error = MUX_DRI_Destructor(self.mux_ID)
        self.fluidics_logger(str(MUX_DRI_Destructor), error, 0)

    def valve_select(self, valve_number):
        '''
        Selects valve in mux unit with associated mux_id to the valve_number declared.
        :param c_int32 mux_id: mux_id given from mux_initialization method
        :param int valve_number: number of desired valve to be selected
        :return: Nothing
        '''

        desired_valve = valve_number
        valve_number = c_int32(valve_number)
        error = MUX_DRI_Set_Valve(self.mux_ID, valve_number, 0)  # 0 is shortest path. clockwise and cc are also options
        self.fluidics_logger(str(MUX_DRI_Set_Valve), error, desired_valve)

        valve = c_int32(-1)
        error = MUX_DRI_Get_Valve(self.mux_ID, byref(valve))
        current_valve = int(valve.value)
        self.fluidics_logger(str(MUX_DRI_Get_Valve), error, current_valve)

        while current_valve != desired_valve:
            MUX_DRI_Get_Valve(self.mux_ID, byref(valve))
            current_valve = int(valve.value)
            self.fluidics_logger(str(MUX_DRI_Get_Valve), error, current_valve)
            time.sleep(1)

    def flow(self, on_off_state):

        run = 1

        while run != 0:

            # disable doing a rerun
            run = 0

            #set target to achieve
            if self.flow_control == 1 and on_off_state == 'ON':
                set_target = self.flow_on
            if self.flow_control == 1 and on_off_state == 'OFF':
                set_target = self.flow_off
            if self.flow_control == 0 and on_off_state == 'ON':
                set_target = self.pressure_on
            if self.flow_control == 0 and on_off_state == 'OFF':
                set_target = self.pressure_off
            set_target_c_types = c_double(set_target)  # convert to c_double


            set_channel = int(1)  # convert to int
            set_channel = c_int32(set_channel)  # convert to c_int32

            # OB1_Start_Remote_Measurement(self.pump_ID, self.calibration_array, 1000)
            error = OB1_Set_Remote_Target(self.pump_ID, set_channel, set_target_c_types)
            self.fluidics_logger(str(OB1_Set_Remote_Target), error, set_target)

            data_sens = c_double()
            data_reg = c_double()
            set_channel = int(1)  # convert to int
            set_channel = c_int32(set_channel)  # convert to c_int32
            time.sleep(3) # wait 3 seconds to stabilize
            error = OB1_Get_Remote_Data(self.pump_ID, set_channel, byref(data_reg), byref(data_sens))

            if self.flow_control == 1:
                current_flow_rate = data_sens.value
            else:
                current_flow_rate = data_reg.value

            self.fluidics_logger(str(OB1_Get_Remote_Data), error, current_flow_rate)

            error = OB1_Get_Remote_Data(self.pump_ID, set_channel, byref(data_reg), byref(data_sens))
            current_flow_rate = data_sens.value
            self.fluidics_logger(str(OB1_Get_Remote_Data), error, current_flow_rate)

            if self.flow_control == 1:

                if set_target > 400 and current_flow_rate < 0.1 * set_target:
                    self.flow_control = 0

                    set_channel = int(1)
                    set_channel = c_int32(set_channel)  # convert to c_int32
                    error = PID_Set_Running_Remote(self.pump_ID, set_channel, 0) # turn off PID loop
                    self.fluidics_logger(str(PID_Set_Running_Remote), error, 0)

                    run = 1 # restart flow function

                if set_target< 40 and current_flow_rate > 100:
                    self.flow_control = 0

                    set_channel = int(1)
                    set_channel = c_int32(set_channel)  # convert to c_int32
                    error = PID_Set_Running_Remote(self.pump_ID, set_channel, 0) # TURN OFF pid LOOP
                    self.fluidics_logger(str(PID_Set_Running_Remote), error, 0)

                    run = 1 # restart flow function

            else:
                pass

    def flow_meter_freeze_check(self):
        '''
        checks if flow meter value is changing (ie checks 3 times and if all are the same value,
        says that the meter is frozen). If frozen, asks to reboot OB1
        :return:
        '''

        #make ctypes structures
        data_sens = c_double()
        data_reg = c_double()
        set_channel = int(1)  # convert to int
        set_channel = c_int32(set_channel)  # convert to c_int32

        #make triple value array
        flows = np.zeros((3))
        #populate with 3 values with second spacing
        for x in range(0,3):
            error = OB1_Get_Remote_Data(self.pump_ID, set_channel, byref(data_reg), byref(data_sens))
            current_flow_rate = data_sens.value
            self.fluidics_logger('checking if frozen_' + str(OB1_Get_Remote_Data), error, current_flow_rate)
            flows[x] = current_flow_rate

        #determine if frozen, ie see how many unique values there are
        unique_value_count = np.unique(flows)

        if unique_value_count == 1:
            self.fluidics_logger('rebooting ob1', error, current_flow_rate)
            self.ob1_reboot()
        else:
            pass

    def ob1_reboot(self):

        kasa --type strip --host < addr > off --name 'Socket1'  # turns off socket named 'Socket1'
        time.sleep(2)
        kasa --type strip --host < addr > on --name 'Socket1'  # turns off socket named 'Socket1'

    def filter_pump(self, filtering_time):

            kasa --host < addr > off --name 'Socket1'  # turns off socket named 'Socket1'
            time.sleep(filtering_time)
            kasa --host < addr > on --name 'Socket1'  # turns off socket named 'Socket1'

    def ob1_end(self):

        set_channel = int(1)
        set_channel = c_int32(set_channel)  # convert to c_int32
        error = PID_Set_Running_Remote(self.pump_ID, set_channel, 0)
        self.fluidics_logger(str(PID_Set_Running_Remote), error, 0)

        error = OB1_Stop_Remote_Measurement(self.pump_ID)
        self.fluidics_logger(str(OB1_Stop_Remote_Measurement), error, 0)

        OB1_Destructor(self.pump_ID)
        self.fluidics_logger(str(OB1_Destructor), error, 0)

    def ob1_start(self, ob1_com_port, flow_control=1):

        # OB1 initialize
        ob1_path = 'ASRL' + str(ob1_com_port) + '::INSTR'
        Instr_ID = c_int32()
        error = OB1_Initialization(ob1_path.encode('ascii'), 0, 0, 0, 0, byref(Instr_ID))
        self.fluidics_logger(str(OB1_Initialization), error, 1)
        error = OB1_Add_Sens(Instr_ID, 1, 5, 1, 0, 7,
                             0)  # 16bit working range between 0-1000uL/min, also what are CustomSens_Voltage_5_to_25 and can I really choose any digital range?
        self.fluidics_logger(str(OB1_Add_Sens), error, 1)
        Calib_path = r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF\Python_64_elveflow\calibration\1_12_24_cal.txt'
        Calib = (c_double * 1000)()
        error = Elveflow_Calibration_Load(Calib_path.encode('ascii'), byref(Calib), 1000)
        self.fluidics_logger(str(Elveflow_Calibration_Load), error, 1)
        # Elveflow_Calibration_Default(byref(Calib), 1000)

        if flow_control == 1:

            set_channel_regulator = int(1)  # convert to int
            set_channel_regulator = c_int32(set_channel_regulator)  # convert to c_int32
            set_channel_sensor = int(1)
            set_channel_sensor = c_int32(set_channel_sensor)  # convert to c_int32
            error = PID_Add_Remote(Instr_ID.value, set_channel_regulator, Instr_ID.value, set_channel_sensor, 0.9,
                                   0.004, 1)
            self.fluidics_logger(str(PID_Add_Remote), error, 1)
        else:
            pass

        error = OB1_Start_Remote_Measurement(Instr_ID.value, byref(Calib), 1000)
        self.fluidics_logger(str(OB1_Start_Remote_Measurement), error, 1)
        self.calibration_array = byref(Calib)

        self.pump_ID = Instr_ID.value

    def measure(self):

        set_channel = int(1)  # convert to int
        set_channel = c_int32(set_channel)  # convert to c_int32

        data_sens = c_double()
        data_reg = c_double()

        OB1_Get_Remote_Data(self.pump_ID, set_channel, byref(data_reg), byref(data_sens))

        pressure = data_reg.value
        flow_rate = data_sens.value
        time_stamp = time.time()

        return pressure, flow_rate, time_stamp

    def flow_recorder(self, time_step, total_time, file_name='none', plot=1):

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = 'Time'
        ws.cell(row=1, column=2).value = 'Flow Rate'
        ws.cell(row=1, column=3).value = 'Pressure'

        total_steps = int(total_time / time_step)
        pressure_points = np.random.rand(total_steps).astype('float16')
        time_points = np.random.rand(total_steps).astype('float16')
        flow_points = np.random.rand(total_steps).astype('float16')

        for t in range(0, total_steps):
            pressure_point, flow_point, time_point = self.measure()
            pressure_points[t] = pressure_point
            time_points[t] = t * time_step
            flow_points[t] = flow_point

            ws.cell(row=t + 2, column=1).value = t * time_step
            ws.cell(row=t + 2, column=2).value = flow_point
            ws.cell(row=t + 2, column=3).value = pressure_point

            time.sleep(time_step)

        # wb.save(filename = file_name)

        if plot == 1:
            plt.plot(time_points, flow_points, 'o', color='black')
            plt.show()

    def liquid_action(self, action_type, stain_valve=0, incub_val=45, heater_state=0, microscope_object = 0, experiment_directory = 0):

        bleach_valve = 11
        pbs_valve = 12
        bleach_time = 15  # minutes
        stain_flow_time = 45  # seconds
        if heater_state == 0:
            stain_inc_time = incub_val  # minutes
        if heater_state == 1:
            stain_inc_time = 45  # minutes
        nuc_valve = 4
        nuc_flow_time = 45  # seconds
        nuc_inc_time = 3  # minutes

        flow_rate = "ON"
        flow_rate_stop = 'OFF'

        #flow_control = self.flow_control

        #if flow_control != 1:
        #    flow_rate = 1100
        #    flow_rate_stop = 0
        #else:
        #    pass

        if action_type == 'Bleach':

            self.valve_select(bleach_valve)
            self.flow(flow_rate)
            time.sleep(90)
            self.flow(flow_rate_stop)
            # time.sleep(bleach_time*60)
            self.valve_select(pbs_valve)

            for x in range(0, bleach_time):
                time.sleep(60)

            self.flow(flow_rate)
            time.sleep(150)
            self.flow(flow_rate_stop)
            time.sleep(5)

        elif action_type == 'Stain':

            stain_start = 0

            self.valve_select(stain_valve)
            self.flow(flow_rate)
            time.sleep(stain_flow_time)
            self.flow(flow_rate_stop)
            self.valve_select(pbs_valve)

            if microscope_object != 0:
                microscope = microscope_object

                time_elapsed = microscope.recursive_stardist_autofocus(experiment_directory, stain_valve)  # int time in seconds
                whole_minutes_elapsed = math.floor(time_elapsed/60)
                seconds_remaining = time_elapsed % 60 # remainder in seconds of time remaining after dividing by 60
                print(seconds_remaining)
                time.sleep(seconds_remaining)
                stain_start = int(whole_minutes_elapsed + 1)

            for x in range(stain_start, stain_inc_time):


                time.sleep(60)
                print('Staining Time Elapsed ', x)

            # if heater_state == 1:
            #    arduino.heater_state(0)
            #    arduino.chamber('fill')
            # else:
            #    pass



            self.valve_select(pbs_valve)
            time.sleep(1)
            self.flow(flow_rate)
            time.sleep(150)
            self.flow(flow_rate_stop)
            time.sleep(5)



        elif action_type == "Wash":

            self.valve_select(pbs_valve)
            self.flow(flow_rate)
            time.sleep(100)
            self.flow(flow_rate_stop)


        elif action_type == 'Nuc_Touchup':

            self.valve_select(nuc_valve)
            self.flow(500)
            time.sleep(nuc_flow_time)
            self.flow(0)
            time.sleep(nuc_inc_time * 60)

            self.valve_select(pbs_valve)
            self.flow(450)
            time.sleep(70)
            self.flow(0)

        elif action_type == 'PBS_flow_on':

            self.valve_select(pbs_valve)
            self.flow(500)
            time.sleep(10)

        elif action_type == 'PBS_flow_off':

            self.valve_select(pbs_valve)
            self.flow(0)
            time.sleep(10)