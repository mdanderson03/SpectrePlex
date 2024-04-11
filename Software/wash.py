import datetime
import os

import numpy as np
import time
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import sys
from ctypes import *

sys.path.append(
    r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF\Python_64_elveflow\DLL64')  # add the path of the library here
sys.path.append(r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF\Python_64_elveflow')  # add the path of the LoadElveflow.py
sys.path.append(r'C:\Users\mike\Documents\GitHub\AutoCIF\Python_64_elveflow\DLL64')  # add the path of the library here
sys.path.append(r'C:\Users\mike\Documents\GitHub\AutoCIF\Python_64_elveflow')  # add the path of the LoadElveflow.py

from array import array
from Elveflow64 import *




experiment_path = r'E:\8-4-24 celiac'
ob1_com_port = 13
flow_control = 1


# OB1 initialize
ob1_path = 'ASRL' + str(ob1_com_port) + '::INSTR'
Instr_ID = c_int32()
OB1_Initialization(ob1_path.encode('ascii'), 0, 0, 0, 0, byref(Instr_ID))
OB1_Add_Sens(Instr_ID, 1, 5, 1, 0, 7,
             0)  # 16bit working range between 0-1000uL/min, also what are CustomSens_Voltage_5_to_25 and can I really choose any digital range?

Calib_path = r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF\Python_64_elveflow\calibration\1_12_24_cal.txt'
Calib = (c_double * 1000)()
# Elveflow_Calibration_Load(Calib_path.encode('ascii'), byref(Calib), 1000)
Elveflow_Calibration_Default(byref(Calib), 1000)

if flow_control == 1:

    set_channel_regulator = int(1)  # convert to int
    set_channel_regulator = c_int32(set_channel_regulator)  # convert to c_int32
    set_channel_sensor = int(1)
    set_channel_sensor = c_int32(set_channel_sensor)  # convert to c_int32
    PID_Add_Remote(Instr_ID.value, set_channel_regulator, Instr_ID.value, set_channel_sensor, 0.9, 0.004, 1)
else:
    pass

OB1_Start_Remote_Measurement(Instr_ID.value, byref(Calib), 1000)
calibration_array = byref(Calib)


pump_ID = Instr_ID.value

flow_control = flow_control


pressure_on = 1000
pressure_off = 0
flow_on = 500
flow_off = -3


def fluidics_logger(self, function_used_string, error_code, value_sr):
    experiment_directory = self.experiment_path
    filename = 'logger.xlsx'
    logger_path = experiment_directory + '/fluidics data logger'
    os.chdir(experiment_directory)
    try:
        os.mkdir('fluidics data logger')
        os.chdir(logger_path)
    except:
        os.chdir(logger_path)

    if os.path.isfile(filename) == True:
        wb = load_workbook(filename)
        ws = wb.active
    elif os.path.isfile(filename) == False:
        wb = Workbook()
        ws = wb.active
        # add headers
        ws.cell(row=1, column=1).value = 'Time Stamp'
        ws.cell(row=1, column=2).value = 'Function Used'
        ws.cell(row=1, column=3).value = 'Error Code'
        ws.cell(row=1, column=4).value = 'Value Sent/Recieved'

    # determine row number (add to next line as a logged event)
    current_max_row = ws.max_row
    row_select = current_max_row + 1

    # add in values
    ws.cell(row=row_select, column=1).value = datetime.datetime.now()
    ws.cell(row=row_select, column=2).value = function_used_string
    ws.cell(row=row_select, column=3).value = error_code
    ws.cell(row=row_select, column=4).value = value_sr

    wb.save(filename)
def flow(self, on_off_state):

        run = 1

        while run != 0:

            # disable doing a rerun
            run = 0

            #set target to achieve
            if self.flow_control == 1 and on_off_state == 'ON':
                set_target = flow_on
            if self.flow_control == 1 and on_off_state == 'OFF':
                set_target = flow_off
            if self.flow_control == 0 and on_off_state == 'ON':
                set_target = pressure_on
            if self.flow_control == 0 and on_off_state == 'OFF':
                set_target = pressure_off
            set_target_c_types = c_double(set_target)  # convert to c_double


            set_channel = int(1)  # convert to int
            set_channel = c_int32(set_channel)  # convert to c_int32

            # OB1_Start_Remote_Measurement(self.pump_ID, self.calibration_array, 1000)
            error = OB1_Set_Remote_Target(self.pump_ID, set_channel, set_target_c_types)
            self.fluidics_logger(str(OB1_Set_Remote_Target), error, set_target)

            data_sens = c_double()
            data_reg = c_double()
            set_channel = int(1)  # convert to int
            set_channel = c_int32(set_channel)  # convert to c_int32
            time.sleep(3) # wait 3 seconds to stabilize
            error = OB1_Get_Remote_Data(self.pump_ID, set_channel, byref(data_reg), byref(data_sens))

            if self.flow_control == 1:
                current_flow_rate = data_sens.value
            else:
                current_flow_rate = data_reg.value

            self.fluidics_logger(str(OB1_Get_Remote_Data), error, current_flow_rate)

            error = OB1_Get_Remote_Data(self.pump_ID, set_channel, byref(data_reg), byref(data_sens))
            current_flow_rate = data_sens.value
            self.fluidics_logger(str(OB1_Get_Remote_Data), error, current_flow_rate)

            if self.flow_control == 1:

                if set_target > 400 and current_flow_rate < 0.1 * set_target:
                    self.flow_control = 0

                    set_channel = int(1)
                    set_channel = c_int32(set_channel)  # convert to c_int32
                    error = PID_Set_Running_Remote(self.pump_ID, set_channel, 0) # turn off PID loop
                    self.fluidics_logger(str(PID_Set_Running_Remote), error, 0)

                    run = 1 # restart flow function

                if set_target< 40 and current_flow_rate > 100:
                    self.flow_control = 0

                    set_channel = int(1)
                    set_channel = c_int32(set_channel)  # convert to c_int32
                    error = PID_Set_Running_Remote(self.pump_ID, set_channel, 0) # TURN OFF pid LOOP
                    self.fluidics_logger(str(PID_Set_Running_Remote), error, 0)

                    run = 1 # restart flow function

            else:
                pass


flow('ON')
time.sleep(150)
flow('OFF')

set_channel = int(1)
set_channel = c_int32(set_channel)  # convert to c_int32
error = PID_Set_Running_Remote(self.pump_ID, set_channel, 0)
self.fluidics_logger(str(PID_Set_Running_Remote), error, 0)

error = OB1_Stop_Remote_Measurement(self.pump_ID)
self.fluidics_logger(str(OB1_Stop_Remote_Measurement), error, 0)

OB1_Destructor(self.pump_ID)
self.fluidics_logger(str(OB1_Destructor), error, 0)
set_channel = int(1)
set_channel = c_int32(set_channel)  # convert to c_int32
error = PID_Set_Running_Remote(self.pump_ID, set_channel, 0)
self.fluidics_logger(str(PID_Set_Running_Remote), error, 0)

error = OB1_Stop_Remote_Measurement(self.pump_ID)
self.fluidics_logger(str(OB1_Stop_Remote_Measurement), error, 0)

OB1_Destructor(self.pump_ID)
self.fluidics_logger(str(OB1_Destructor), error, 0)


