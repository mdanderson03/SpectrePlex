import datetime
import os
import numpy as np
import time
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import sys
from ctypes import *
from KasaSmartPowerStrip import SmartPowerStrip

sys.path.append(
    r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF\Python_64_elveflow\DLL64')  # add the path of the library here
sys.path.append(r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF\Python_64_elveflow')  # add the path of the LoadElveflow.py
sys.path.append(r'C:\Users\mike\Documents\GitHub\AutoCIF\Python_64_elveflow\DLL64')  # add the path of the library here
sys.path.append(r'C:\Users\mike\Documents\GitHub\AutoCIF\Python_64_elveflow')  # add the path of the LoadElveflow.py

from array import array
from Elveflow64 import *


class fluidics:

    def __init__(self, experiment_path, ob1_com_port, flow_control=1):

        # OB1 initialize
        ob1_path = 'ASRL' + str(ob1_com_port) + '::INSTR'
        Instr_ID = c_int32()
        OB1_Initialization(ob1_path.encode('ascii'), 0, 0, 0, 0, byref(Instr_ID))
        OB1_Add_Sens(Instr_ID, 1, 5, 1, 0, 7, 0)  # 16bit working range between 0-1000uL/min, also what are CustomSens_Voltage_5_to_25 and can I really choose any digital range?

        Calib_path = r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF\Python_64_elveflow\calibration\1_12_24_cal.txt'
        Calib = (c_double * 1000)()
        # Elveflow_Calibration_Load(Calib_path.encode('ascii'), byref(Calib), 1000)
        Elveflow_Calibration_Default(byref(Calib), 1000)

        if flow_control == 1:

            set_channel_regulator = int(1)  # convert to int
            set_channel_regulator = c_int32(set_channel_regulator)  # convert to c_int32
            set_channel_sensor = int(1)
            set_channel_sensor = c_int32(set_channel_sensor)  # convert to c_int32
            PID_Add_Remote(Instr_ID.value, set_channel_regulator, Instr_ID.value, set_channel_sensor, 0.9/22.5, 0.0015/.03846, 1)
            #PID_Add_Remote(Instr_ID.value, set_channel_regulator, Instr_ID.value, set_channel_sensor, 0.035, 0.00000000003, 1)
        else:
            pass

        OB1_Start_Remote_Measurement(Instr_ID.value, byref(Calib), 1000)
        self.calibration_array = byref(Calib)


        self.pump_ID = Instr_ID.value
        self.experiment_directory = experiment_path
        self.experiment_path = experiment_path
        self.flow_control = flow_control

        self.pressure_on = 500
        self.pressure_off = 0
        self.flow_on = 500
        self.flow_off = -3
        self.low_flow_on = 120
        self.high_flow_on = 500

        return

    def fluidics_logger(self, function_used_string, error_code, value_sr):
        experiment_directory = self.experiment_path
        filename = 'logger.xlsx'
        logger_path = experiment_directory + '/fluidics data logger'
        os.chdir(experiment_directory)
        try:
            os.mkdir('fluidics data logger')
            os.chdir(logger_path)
        except:
            os.chdir(logger_path)

        if os.path.isfile(filename) == True:
            wb = load_workbook(filename)
            ws = wb.active
        elif os.path.isfile(filename) == False:
            wb = Workbook()
            ws = wb.active
            # add headers
            ws.cell(row=1, column=1).value = 'Time Stamp'
            ws.cell(row=1, column=2).value = 'Function Used'
            ws.cell(row=1, column=3).value = 'Error Code'
            ws.cell(row=1, column=4).value = 'Value Sent/Recieved'

        # determine row number (add to next line as a logged event)
        current_max_row = ws.max_row
        row_select = current_max_row + 1

        # add in values
        ws.cell(row=row_select, column=1).value = datetime.datetime.now()
        ws.cell(row=row_select, column=2).value = function_used_string
        ws.cell(row=row_select, column=3).value = error_code
        ws.cell(row=row_select, column=4).value = value_sr

        wb.save(filename)

    def flow(self, on_off_state):
        '''
        ON turns flow on as dictated in init function and ON LOW does it for th low flow in the init function

        :param on_off_state: 'ON', 'ON LOW', 'ON HIGH', 'OFF'
        :return:
        '''

        # load in data structures
        numpy_path = self.experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fluid_info_array.npy'
        fluid_array = np.load(file_name, allow_pickle=False)

        # set target to achieve
        if self.flow_control == 1 and on_off_state == 'ON':
            set_target = self.flow_on
        if self.flow_control == 1 and on_off_state == 'ON LOW':
            set_target = self.low_flow_on
        if self.flow_control == 1 and on_off_state == 'ON HIGH':
            set_target = self.high_flow_on
        if self.flow_control == 1 and on_off_state == 'OFF':
            set_target = self.flow_off
        if self.flow_control == 0 and on_off_state == 'ON':
            set_target = self.pressure_on
        if self.flow_control == 0 and on_off_state == 'OFF':
            set_target = self.pressure_off
        set_target_c_types = c_double(set_target)  # convert to c_double

        set_channel = int(1)  # convert to int
        set_channel = c_int32(set_channel)  # convert to c_int32
        data_sens = c_double()
        data_reg = c_double()

        error = OB1_Get_Remote_Data(self.pump_ID, set_channel, byref(data_reg), byref(data_sens))
        starting_flow_rate = data_sens.value
        self.fluidics_logger(str(OB1_Get_Remote_Data), error, starting_flow_rate)

        # OB1_Start_Remote_Measurement(self.pump_ID, self.calibration_array, 1000)
        error = OB1_Set_Remote_Target(self.pump_ID, set_channel, set_target_c_types)
        self.fluidics_logger(str(OB1_Set_Remote_Target), error, set_target)

        time.sleep(4)  # wait 3 seconds to stabilize
        error = OB1_Get_Remote_Data(self.pump_ID, set_channel, byref(data_reg), byref(data_sens))

        if self.flow_control == 1:
            #current_flow_rate = 0
            current_flow_rate = data_sens.value
        else:
            current_flow_rate = data_reg.value


        self.fluidics_logger(str(OB1_Get_Remote_Data), error, current_flow_rate)

        delta_flow_rate = current_flow_rate - starting_flow_rate
        print('start', starting_flow_rate, 'end', current_flow_rate, 'del', delta_flow_rate)

        error = OB1_Get_Remote_Data(self.pump_ID, set_channel, byref(data_reg), byref(data_sens))
        # current_flow_rate = data_sens.value
        # self.fluidics_logger(str(OB1_Get_Remote_Data), error, current_flow_rate)

        if self.flow_control == 1:

            if set_target != self.flow_off  and delta_flow_rate < 0.5 * set_target:
                print('flow failed')
                zero_target = c_double(self.flow_off)

                OB1_Set_Remote_Target(self.pump_ID, set_channel, zero_target)
                # self.flow_control = 0

                # set_channel = int(1)
                # set_channel = c_int32(set_channel)  # convert to c_int32
                # error = PID_Set_Running_Remote(self.pump_ID, set_channel, 0) # turn off PID loop
                # self.fluidics_logger(str(PID_Set_Running_Remote), error, 0)

                # run = 1 # restart flow function
                fluid_array[2] = 1
                self.ob1_reboot()

            if set_target == self.flow_off:
                #time.sleep(0.4)
                print('turning off ob1')
                # self.flow_control = 0

                # set_channel = int(1)
                # set_channel = c_int32(set_channel)  # convert to c_int32
                # error = PID_Set_Running_Remote(self.pump_ID, set_channel, 0) # TURN OFF pid LOOP
                # self.fluidics_logger(str(PID_Set_Running_Remote), error, 0)

                # run = 1 # restart flow function
                self.ob1_reboot()
                #fluid_array[2] = 1

        else:
            pass

        os.chdir(numpy_path)
        np.save(file_name, fluid_array)

    def flow_check(self):

        # make ctypes structures
        data_sens = c_double()
        data_reg = c_double()
        set_channel = int(1)  # convert to int
        set_channel = c_int32(set_channel)  # convert to c_int32

        # make triple value array
        flows = np.zeros((5))
        # populate with 3 values with second spacing
        for x in range(0, 5):
            error = OB1_Get_Remote_Data(self.pump_ID, set_channel, byref(data_reg), byref(data_sens))
            current_flow_rate = data_sens.value
            self.fluidics_logger('checking if frozen_' + str(OB1_Get_Remote_Data), error, current_flow_rate)
            flows[x] = current_flow_rate
            print(current_flow_rate)
            time.sleep(0.5)

        # determine if frozen, ie see how many unique values there are
        unique_value_count = np.shape(np.unique(flows))[0]

        if unique_value_count <= 2:
            print('reboot')
            self.fluidics_logger('rebooting ob1', error, current_flow_rate)
            self.ob1_reboot()
        else:
            pass

    '''
    def ob1_reboot(self):

        power_strip = SmartPowerStrip('10.3.141.157')
        time.sleep(1)
        try:
            power_strip.toggle_plug('off', plug_num=4)
        except:
            power_strip.toggle_plug('off', plug_num=4)
        time.sleep(3)
        try:
            power_strip.toggle_plug('on', plug_num=4)  # turns off socket named 'Socket1'
        except:
            power_strip.toggle_plug('on', plug_num=4)  # turns off socket named 'Socket1'
        time.sleep(3)
    '''

    def ob1_reboot(self):

        self.ob1_off(nretry=0)
        self.ob1_on(nretry=0)

    def ob1_on(self, retry_limit=10, nretry=0):

        if nretry <= retry_limit:
            try:
                power_strip = SmartPowerStrip('10.3.141.157')
                time.sleep(1)
                power_strip.toggle_plug('on', plug_num=4)
            except:
                return self.ob1_on(nretry=nretry + 1)

    def ob1_off(self, retry_limit=10, nretry=0):

        if nretry <= retry_limit:
            try:
                power_strip = SmartPowerStrip('10.3.141.157')
                #time.sleep(0.5)
                power_strip.toggle_plug('off', plug_num=4)
            except:
                return self.ob1_off(nretry=nretry + 1)

    def ob1_end(self):

        error = OB1_Stop_Remote_Measurement(self.pump_ID)
        error = OB1_Destructor(self.pump_ID)
        self.fluidics_logger(str(OB1_Destructor), error, 0)
