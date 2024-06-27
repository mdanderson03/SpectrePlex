import copy

import numpy as np
import os

import skimage.util
from skimage import io, filters, util
import time
from copy import deepcopy
import math
from openpyxl import Workbook, load_workbook
from tifffile import imread


experiment_directory = r'E:\delete'
def hdr_exp_generator(experiment_directory, threshold_level, max_exp, min_exp):
    #add on exp offset
    exp_offset = 27.43
    max_exp += exp_offset
    min_exp += exp_offset

    #calculate number of images needed for threshold and real threshold
    M = 65536/threshold_level
    n = math.ceil(np.log(max_exp/min_exp)/np.log(M) + 1)
    M_real = np.power(max_exp/min_exp, 1/(n-1))
    T_real = 65536/M_real

    #create and populate hdr exp array with min and max values
    hdr_exp_list = np.zeros(n)
    hdr_exp_list[0] = min_exp - exp_offset
    hdr_exp_list[n-1] = max_exp - exp_offset

    for x in range(1,n-1):
        hdr_exp_list[x] = int(min_exp * np.power(M_real, x) - exp_offset)


    # import exp_calc_array
    exp_path = experiment_directory + '/' + 'exposure_times'
    os.chdir(exp_path)

    # create or open workbook

    if os.path.isfile('HDR_Exp.xlsx') == False:
        wb = Workbook()
        ws = wb.active

        # populate headers
        ws.cell(row=1, column=1).value = 'image count'
        ws.cell(row=1, column=2).value = 'exposure time 1'
        ws.cell(row=1, column=3).value = 'exposure time 2'
        ws.cell(row=1, column=4).value = 'exposure time 3'
        ws.cell(row=1, column=5).value = 'threshold real'
        ws.cell(row=4, column=1).value = 'Cycle'
        ws.cell(row=4, column=2).value = 'Max Int A488'
        ws.cell(row=4, column=3).value = 'Max Int A555'
        ws.cell(row=4, column=4).value = 'Max Int A647'

    if os.path.isfile('HDR_Exp.xlsx') == True:
        wb = load_workbook('HDR_Exp.xlsx')
        ws = wb.active

    # populate columns with times and cycle count
    ws.cell(row=2, column=1).value = n
    ws.cell(row=2, column=2).value = hdr_exp_list[0] + exp_offset
    ws.cell(row=2, column=3).value = hdr_exp_list[1] + exp_offset
    ws.cell(row=2, column=4).value = hdr_exp_list[2] + exp_offset
    ws.cell(row=2, column=5).value = T_real

    wb.save('HDR_Exp.xlsx')

    return hdr_exp_list





print(hdr_exp_generator(experiment_directory,10000, 1000, 20))


times = np.array([10, 20, 30, 50, 100, 150, 250, 400, 800, 2000])
hdr_indicies = [0,3,9]

os.chdir(r'C:\Users\CyCIF PC\Desktop\linearity')
array = io.imread('linearity.tif')

hdr_array = array[hdr_indicies]
times = times[hdr_indicies]

start_time = time.time()

max_hdr_time = np.max(times)
hdr_array = hdr_array.astype('float32')
weight_array = deepcopy(hdr_array)

#subtract linear offset
hdr_array = hdr_array - 300

#populate weight array
for index in range(0,3):

    exp_offset = 27.43
    mag = (max_hdr_time + exp_offset)/(times[index] + exp_offset)
    del_offset = 1.2*(max_hdr_time - times[index])/(times[index] + exp_offset)**2

    im = copy.deepcopy(hdr_array[index])

    scaled_im = mag * im
    hdr_array[index] = scaled_im
    im[im > 65234] = 0

    del_I = np.sqrt(im)

    weight_array[index] = del_I/scaled_im + del_offset/mag

total_weight_array = np.sum(weight_array, axis=0)
scaled_weight_array = np.divide(weight_array,total_weight_array)

hdr_im = hdr_array[0]*scaled_weight_array[0]
for x in range(1, np.shape(hdr_array)[0]):
    hdr_im += hdr_array[x]*scaled_weight_array[x]

end_time = time.time()
print('time to run', end_time - start_time)

hdr_im[hdr_im<0] = 0
os.chdir(r'C:\Users\CyCIF PC\Desktop\linearity\hdr')
io.imsave('bit32_hdr.tif', hdr_im)
hdr_im = hdr_im/np.max(hdr_im)
hdr_im = util.img_as_uint(hdr_im)
io.imsave('bit16_hdr.tif', hdr_im)



experiment_directory = r'E:\delete'
x = 1
y = 1

labelled_path = experiment_directory + '/Labelled_Nuc'
dapi_im_path = experiment_directory + '/' + 'DAPI' '\Bleach\cy_' + str(0) + '\Tiles'

os.chdir(dapi_im_path)
file_name = 'z_' + str(4) + '_x' + str(x) + '_y_' + str(y) + '_c_DAPI.tif'
#file_name = r'bit32.tif'
img = io.imread(file_name)
print(img[1600][600])

os.chdir(labelled_path)
file_name = 'x' + str(x) + '_y_' + str(y) + '_c_DAPI.tif'
labels = io.imread(file_name)


def focus_score(image, derivative_jump, labels):
    '''
    Calculates focus score on image with Brenners algorithm on downsampled image.


    :param numpy image: single image from hooked from acquistion

    :return: focus score for image
    :rtype: float
    '''
    # Note: Uniform background is a bit mandatory

    # do Brenner score

    a = image[derivative_jump:, :]
    a = a.astype('float64')
    print('a', a[1600][600])
    a = np.nan_to_num(a, posinf=65500, nan=65550)
    print('a', a[1600][600])
    b = image[:-derivative_jump, :]
    print('b', b[1600][600])
    b = b.astype('float64')

    b = np.nan_to_num(b, posinf=65500, nan=65550)
    a = np.log(a)
    print('a', a[1600][600])
    b = np.log(b)
    print('b', b[1600][600])
    c = a - b
    print('c', c[1600][600])
    c = c ** 2
    # c = (a - b)/((a+b)/2)
    # c = c / 1000 * c / 1000
    labels = labels[derivative_jump:, :]
    c = c * labels
    f_score_shadow = c.sum(dtype=np.float64)

    return f_score_shadow


print(focus_score(img, 17, labels))
#io.imshow(img)
#io.show()
