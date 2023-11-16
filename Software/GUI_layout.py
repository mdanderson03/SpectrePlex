import PySimpleGUI as sg
from datetime import datetime, date
import openpyxl

list_status = []

sg.set_options(font=('Arial Bold', 16))
cycle_options = ["0","1","2","3","4","5","6","7","8","9","10"]
layout_cycle = [[sg.Text("Directory:"), sg.Input(key="-path-"), sg.FolderBrowse(target="-path-")],
                [sg.Text("Add New Folder"),sg.Input(key="-new_path-"), sg.Button("Create",key="-new_folder-")], 
                [sg.Checkbox("Multiple Cycles", enable_events=True, key="-checkbox-",default=True),sg.Text("Start Cycle Number:"), sg.Combo(["0","1","2","3","4","5","6","7","8","9"],default_value="0", key="-cycle-", enable_events=True, readonly=True),sg.Text("End Cycle Number:", key="-txt_end-",visible=True), sg.Combo(cycle_options, default_value="8",key="-cycle_end-", visible=True, readonly=True)],
                # [sg.Checkbox("Change number of z-slices (3 slices)", key="-slice-", enable_events=True), sg.Text("z-slices:", key="-slice_txt-", visible=False), sg.Input(key="-num_z-", size=(5,1),visible=False)],
                [sg.Checkbox("Antibody Incubation Time (45 min)", key="-incub_time-", enable_events=True), sg.Text("Incubation Time:", key="-fc_incub_txt-", visible=False), sg.Input(key="-fc_incub_input-", visible=False, size=(10,1))],
                [sg.Checkbox("Perform post-acquisition processing", key="-post_acq-", default=True)]]

# ----------- Add frames around changeable layouts ----------- #
cycle_list = [sg.Column(layout_cycle)]
frame_cycle = sg.Frame("Full Cycle", [cycle_list],title_color="black", key='-COL_1-')

listbox_list = [sg.Listbox(list_status, key="-listbox-", size=(30,3), expand_x=True, no_scrollbar=True)]
frame_listbox = sg.Frame("Status: idle", [listbox_list], title_color="black", expand_x=True, key="-frame_list-")

# ----------- Main Layout ----------- #
layout = [[frame_cycle],
 [sg.Button("PBS On", button_color=("white","green"),key="-pbs_on-"), 
sg.Button("PBS Off", button_color=("white","red"),key="-pbs_off-")],
 [frame_listbox, sg.Push(),sg.Button("Start Run")],
 [sg.ProgressBar(10, key="-PBAR-", visible=False, expand_x=True, expand_y=True, size=(10,10))]]


# ----------- functions for updating and saving status messages ----------- #
try: 
    filename = 'multiplex_gui_log.xlsx'
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
except:
    filename = 'multiplex_gui_log.xlsx'
    workbook = openpyxl.Workbook()
    workbook.save(filename)
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    sheet.append(("Date","Time","Status Message"))

def excel_save(message):
    current_time = (datetime.now()).strftime("%H:%M:%S")
    current_day = (date.today()).strftime("%b-%d-%Y")
    sheet.append((current_day,current_time, message))
    workbook.save(filename)


def status_update(message, list_status, window):
    if len(list_status) == 3:
        list_status.remove(list_status[0])
        list_status.append(message)
        window["-listbox-"].update(list_status)
        window.refresh()
        excel_save(message)
    else:
        list_status.append(message)
        window["-listbox-"].update(list_status)
        window.refresh()
        excel_save(message)
