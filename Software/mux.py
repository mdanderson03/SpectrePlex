import datetime
import os

import numpy as np
import time
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import sys
from ctypes import *

sys.path.append(
    r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF\Python_64_elveflow\DLL64')  # add the path of the library here
sys.path.append(r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF\Python_64_elveflow')  # add the path of the LoadElveflow.py
sys.path.append(r'C:\Users\mike\Documents\GitHub\AutoCIF\Python_64_elveflow\DLL64')  # add the path of the library here
sys.path.append(r'C:\Users\mike\Documents\GitHub\AutoCIF\Python_64_elveflow')  # add the path of the LoadElveflow.py

from array import array
from Elveflow64 import *


class mux_valve:

    def __init__(self, experiment_path, mux_com_port):

        # MUX intiialize
        path = 'ASRL' + str(mux_com_port) + '::INSTR'
        mux_Instr_ID = c_int32()
        MUX_DRI_Initialization(path.encode('ascii'), byref(
            mux_Instr_ID))  # choose the COM port, it can be ASRLXXX::INSTR (where XXX=port number)

        self.mux_ID = mux_Instr_ID.value
        self.experiment_path = experiment_path

        return

    def fluidics_logger(self, function_used_string, error_code, value_sr):
        experiment_directory = self.experiment_path
        filename = 'logger.xlsx'
        logger_path = experiment_directory + '/fluidics data logger'
        os.chdir(experiment_directory)
        try:
            os.mkdir('fluidics data logger')
            os.chdir(logger_path)
        except:
            os.chdir(logger_path)

        if os.path.isfile(filename) == True:
            wb = load_workbook(filename)
            ws = wb.active
        elif os.path.isfile(filename) == False:
            wb = Workbook()
            ws = wb.active
            # add headers
            ws.cell(row=1, column=1).value = 'Time Stamp'
            ws.cell(row=1, column=2).value = 'Function Used'
            ws.cell(row=1, column=3).value = 'Error Code'
            ws.cell(row=1, column=4).value = 'Value Sent/Recieved'

        # determine row number (add to next line as a logged event)
        current_max_row = ws.max_row
        row_select = current_max_row + 1

        # add in values
        ws.cell(row=row_select, column=1).value = datetime.datetime.now()
        ws.cell(row=row_select, column=2).value = function_used_string
        ws.cell(row=row_select, column=3).value = error_code
        ws.cell(row=row_select, column=4).value = value_sr

        wb.save(filename)

    def mux_end(self):

        error = MUX_DRI_Destructor(self.mux_ID)
        self.fluidics_logger(str(MUX_DRI_Destructor), error, 0)

    def valve_select(self, valve_number):
        '''
        Selects valve in mux unit with associated mux_id to the valve_number declared.
        :param c_int32 mux_id: mux_id given from mux_initialization method
        :param int valve_number: number of desired valve to be selected
        :return: Nothing
        '''

        desired_valve = valve_number
        valve_number = c_int32(valve_number)
        error = MUX_DRI_Set_Valve(self.mux_ID, valve_number, 0)  # 0 is shortest path. clockwise and cc are also options
        self.fluidics_logger(str(MUX_DRI_Set_Valve), error, desired_valve)

        valve = c_int32(-1)
        error = MUX_DRI_Get_Valve(self.mux_ID, byref(valve))
        current_valve = int(valve.value)
        self.fluidics_logger(str(MUX_DRI_Get_Valve), error, current_valve)

        while current_valve != desired_valve:
            MUX_DRI_Get_Valve(self.mux_ID, byref(valve))
            current_valve = int(valve.value)
            self.fluidics_logger(str(MUX_DRI_Get_Valve), error, current_valve)
            time.sleep(1)

