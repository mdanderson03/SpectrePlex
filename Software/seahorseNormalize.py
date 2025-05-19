from csbdeep.utils import normalize
from stardist.models import StarDist2D
import os
from skimage import io
from openpyxl import Workbook
import numpy as np

model = StarDist2D.from_pretrained('2D_versatile_fluo')

def well_cell_counter(experiment_directory):

    os.chdir(experiment_directory)
    #create new labeled image directory
    try:
        os.mkdir('Labelled_Nuc')
    except:
        pass
    labelled_directory = experiment_directory + '/' + 'Labelled_Nuc'

    #define what rows and columns to be iterated through
    columns= ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J' ]
    rows = [1,2,3,4,5,6,7,8]

    #create excel workbook and label columns
    wb = Workbook()
    ws = wb.active
    ws.cell(row = 1, column=1).value = 'Well'
    ws.cell(row=1, column=2).value = 'Cell Count'

    for row in range(0, len(rows)):
        for column in range(0, len(columns)):

            try:
                merged_filepath = experiment_directory + '/' + columns[column]
                merged_filename = str(rows[row]) + '_Merged.tif'
                labelled_filename = columns[column] + '_' + str(rows[row]) + '.tif'

                os.chdir(merged_filepath)
                img = io.imread(merged_filename)

                #use stardist and find cell count

                labels, _ = model.predict_instances(normalize(img))
                cell_count = np.max(labels)
                ws.cell(row=row+2, column=column).value = cell_count

                #save labelled image in folder
                os.chdir(labelled_directory)
                io.imsave(labelled_filename, labels)
            except:
                pass

    os.chdir(experiment_directory)
    wb.save('Well_cell_counts')




well_cell_counter(r'C:\Users\CyCIF PC\Desktop\TileScan2')
