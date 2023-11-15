import ome_types
#from pycromanager import Core, Acquisition, multi_d_acquisition_events, Dataset, MagellanAcquisition, Magellan, start_headless, XYTiledAcquisition, Studio
import numpy as np
import time
from scipy.optimize import curve_fit
import paho.mqtt.client as mqtt
import matplotlib.pyplot as plt
from skimage import io, measure, filters
import skimage
import os
import math
from datetime import datetime
from tifffile import imsave, imwrite
import tifffile as tf
from openpyxl import load_workbook, Workbook
from ome_types.model import Instrument, Microscope, Objective, InstrumentRef, Image, Pixels, Plane, Channel
from ome_types.model.simple_types import UnitsLength, PixelType, PixelsID, ImageID, ChannelID
from skspatial.objects import Plane, Points
from skspatial.plotting import plot_3d
from sklearn.datasets import make_regression
from sklearn.linear_model import HuberRegressor, LinearRegression, RANSACRegressor
from sklearn.model_selection import cross_val_score
from sklearn.model_selection import RepeatedKFold
from ome_types import from_xml, OME, from_tiff, to_xml
from scipy import stats
from copy import copy, deepcopy
import sys
from pybasic import shading_correction
from path import Path
import cv2
from ctypes import *
#from GUI_layout import *
sys.path.append(r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF\Python_64_elveflow\DLL64')#add the path of the library here
sys.path.append(r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF\Python_64_elveflow')#add the path of the LoadElveflow.py
sys.path.append(r'C:\Users\mike\Documents\GitHub\AutoCIF\Python_64_elveflow\DLL64')#add the path of the library here
sys.path.append(r'C:\Users\mike\Documents\GitHub\AutoCIF\Python_64_elveflow')#add the path of the LoadElveflow.py

from array import array
#from Elveflow64 import *


#mm_app_path = 'C:\Program Files\Micro-Manager-2.0'
#config_file = r'C:\Users\CyCIF PC\Desktop\lumencor_auto_cycif.cfg'

#start_headless(mm_app_path, config_file, buffer_size_mb=50000)
#client = mqtt.Client('autocyplex_server')
#client.connect('10.3.141.1', 1883)

#core = Core()
#magellan = Magellan()

global level
level = []


class brenner:
    def __init__(self):
        brenner.value = []
        return

class exp_level:
    def __init__(self):
        exp_level.value = []
        return

class cycif:

    def __init__(self):

        return

    def tissue_center(self, mag_surface):
        '''
        take magellan surface and find the xy coordinates of the center of the surface
        :param mag_surface:
        :param magellan:
        :return: x tissue center position and y tissue center position
        :rtype: list[float, float]
        '''
        xy_pos = self.tile_xy_pos(mag_surface)
        x_center = (max(xy_pos[0]) + min(xy_pos[0])) / 2
        y_center = (max(xy_pos[1]) + min(xy_pos[1])) / 2
        return x_center, y_center

    def num_surfaces_count(self):
        '''
        Looks at magellan surfaces that start with New Surface in its name, ie. 'New Surface 1' as that is the default generated prefix.

        :param object magellan: magellan object from magellan = Magellan() in pycromanager
        :return: surface_count
        :rtype: int
        '''
        x = 1
        while magellan.get_surface("New Surface " + str(x)) != None:
            x += 1
        surface_count = x - 1
        time.sleep(1)

        return surface_count

    def surface_exist_check(self, surface_name):
        '''
        Checks name of surface to see if exists. If it does, returns 1, if not returns 0

        :param object magellan: magellan object from magellan = Magellan() in pycromanager
        :param str surface_name: name of surface to check if exists

        :return: status
        :rtype: int
        '''

        status = 0
        if magellan.get_surface(surface_name) != None:
            status += 1

        return status

    ####################################################################
    ############ All in section are functions for the autofocus function
    ####################################################################

    def focus_score(self, image, derivative_jump):
        '''
        Calculates focus score on image with Brenners algorithm on downsampled image.


        :param numpy image: single image from hooked from acquistion

        :return: focus score for image
        :rtype: float
        '''
        # Note: Uniform background is a bit mandatory

        # do Brenner score
        a = image[derivative_jump:, :]
        a = a.astype('float64')
        b = image[:-derivative_jump, :]
        b = b.astype('float64')
        c = (a - b)
        c = c/10000 * c/10000
        f_score_shadow = c.sum(dtype=np.float64) + 0.00001

        return  f_score_shadow

    def sp_array(self, experiment_directory):
        '''
        Generate super pixel array as defined in powerpoint autofocus network
        :param string experiment_directory:
        :param string channel:
        :param y_tiles:
        :param x_tiles:
        :return:
        '''

        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)

        fm_array = np.load('fm_array.npy', allow_pickle=False)

        channels = ['DAPI', 'A488', 'A555', 'A647']

        y_tiles = np.shape(fm_array[0])[0]
        x_tiles = np.shape(fm_array[0])[1]

        x_pixel_count = int((x_tiles) * 32) #32
        y_pixel_count = int((y_tiles) * 24) #24
        sp_array = np.random.rand(5, y_pixel_count, x_pixel_count, 2).astype('float64')
        print(np.shape(sp_array))
        for channel in channels:
            filename = channel + '_sp_array.npy'
            np.save(filename, sp_array)

    def tile_subsampler(self, experiment_directory):
        '''
        Outs grid of dimensions [y_tiles, x_tiles] of 1 or 0s that indicate with 1 if the tile is chosen to be sampled. Currently, it samples all tiles.
        :return:
        '''

        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)

        full_array = np.load('fm_array.npy', allow_pickle=False)
        numpy_x = full_array[0]
        x_tile_count = np.shape(numpy_x)[1]
        y_tile_count = np.shape(numpy_x)[0]

        subsample_grid = np.ones((y_tile_count, x_tile_count))

        return subsample_grid

    def image_sub_divider(self, whole_image, y_sections, x_sections):
        '''
        takes in image and breaks into subsections of size y_sections by x_sections.
        :param whole_image:
        :param y_sections:
        :param x_sections:
        :return:
        '''

        y_pixels = np.shape(whole_image)[0]
        x_pixels = np.shape(whole_image)[1]
        sub_divided_image = np.random.rand(y_sections, x_sections, int(y_pixels / y_sections), int(x_pixels / x_sections)).astype('uint16')
        for y in range(0, y_sections):
            for x in range(0, x_sections):
                # define y and x start and ends subsection of rebuilt image

                y_start = int(y * (y_pixels / y_sections))
                y_end = y_start + int(y_pixels / y_sections)
                x_start = int(x * (x_pixels / x_sections))
                x_end = x_start + int(x_pixels / x_sections)

                sub_divided_image[y][x] = whole_image[y_start:y_end, x_start:x_end]

        return sub_divided_image

    def sub_divided_2_brenner_sp(self, experiment_directory, sub_divided_image, channel, point_number, z_position, x_tile_number, y_tile_number):
        '''
        Takes in subdivided image and calculated brenner score at each subsection and properly places into sp_array
        :param experiment_directory:
        :param sub_divided_image:
        :param channel:
        :param point_number:
        :param x_tile_number:
        :param y_tile_number:
        :return:
        '''

        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = channel + '_sp_array.npy'
        sp_array = np.load(file_name, allow_pickle=False)
        sp_array_slice = sp_array[point_number]

        derivative_jump = 10

        y_subdivisions = 24
        x_subdivisions = 32

        y_offset = int(y_subdivisions * y_tile_number)
        x_offset = int(x_subdivisions * x_tile_number)

        for y in range(y_offset, y_subdivisions + y_offset):
            for x in range(x_offset, x_subdivisions + x_offset):


                score = self.focus_score(sub_divided_image[y - y_offset][x - x_offset], derivative_jump)
                sp_array_slice[y][x][0] = score
                sp_array_slice[y][x][1] = z_position

        np.save(file_name, sp_array)

    def sp_array_focus_solver(self, experiment_directory, channel):
        '''
        takes fully populated sp point array and applied 3 point brenner solver method to populate predicted focus

        :param experiment_directory:
        :param channel:
        :return:
        '''

        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = channel + '_sp_array.npy'
        sp_array = np.load(file_name, allow_pickle=False)

        y_sp_pixels = np.shape(sp_array[0])[0]
        x_sp_pixels = np.shape(sp_array[0])[1]

        for y in range(0, y_sp_pixels):
            for x in range(0, x_sp_pixels):

                scores = sp_array[0:3, y, x, 0]
                scores = scores / np.min(scores)
                positions = sp_array[0:3, y, x, 1]
                three_point_array = np.stack((scores, positions), axis=1)


                a, b, c, predicted_focus = self.gauss_jordan_solver(three_point_array)
                sp_array[3][y][x][0] = predicted_focus

            np.save(file_name, sp_array)

    def sp_array_filter(self, experiment_directory, channel):
        '''
        takes sp array for a channel and generates mask that filters out nonsense answers.
        These answers are solutions that exist outside of the scan range and ones that have a low well depth
        :param experiment_directory:
        :param channel:
        :return:
        '''

        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = channel + '_sp_array.npy'
        sp_array = np.load(file_name, allow_pickle=False)

        y_dim = np.shape(sp_array)[1]
        x_dim = np.shape(sp_array)[2]

        well_depth = np.random.rand(y_dim, x_dim)

        # execute otsu threshold value calc.
        for y in range(0, y_dim):
            for x in range(0, x_dim):
                score_array = sp_array[0:3, y, x, 0]
                score_array = score_array / np.min(score_array)
                depth = np.max(score_array)
                well_depth[y][x] = depth

            threshold = filters.threshold_otsu(well_depth)  # sets filter threshold. If dont want otsu, just chang eto number
            # threshold = 1.1
            for y in range(0, y_dim):
                for x in range(0, x_dim):

                    position_array = sp_array[0:3, y, x, 1]
                    highest_pos = np.max(position_array)
                    lowest_pos = np.min(position_array)

                    if well_depth[y][x] > threshold and lowest_pos < sp_array[3][y][x][0] < highest_pos:
                        sp_array[4][y][x][0] = 1
                    else:
                        sp_array[4][y][x][0] = 0

            np.save(file_name, sp_array)

    def plane_2_z(self, coefficents, xy_point):

        a = coefficents[0]
        b = coefficents[1]
        c = coefficents[2]

        z = a * xy_point[0] + b * xy_point[1] + c

        return z

    def sp_array_surface_2_fm(self, experiment_directory, channel):
        '''
        Takes fully constructed sp array with mask and predicted focus position and fits plane to points.
        :param experiment_directory:
        :param channel:
        :return:
        '''

        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        fm_array = np.load('fm_array.npy', allow_pickle=False)

        y_tiles = np.shape(fm_array[0])[0]
        x_tiles = np.shape(fm_array[0])[1]

        depth_of_focus = 3 #in microns

        point_list = self.map_2_points(experiment_directory, channel)
        X = point_list[:, 0:2]
        y = point_list[:, 2]

        model = HuberRegressor()
        model.fit(X, y)

        a = (model.predict([[1000, 2500]]) - model.predict([[0, 2500]])) / 1000
        b = (model.predict([[1000, 2500]]) - model.predict([[1000, 1500]])) / 1000
        c = model.predict([[2000, 2000]]) - a * 2000 - b * 2000

        coefficents = [a, b, c]
        print('coefficents', coefficents)

        if channel == 'DAPI':
            channel_index = 2
        if channel == 'A488':
            channel_index = 4
        if channel == 'A555':
            channel_index = 6
        if channel == 'A647':
            channel_index = 8

        #calc number of slices needed

        high_z = self.plane_2_z(coefficents, [0, 0])
        low_z = self.plane_2_z(coefficents, [5056, 2960])
        corner_corner_difference = math.fabs(high_z - low_z)
        #number_planes = int(corner_corner_difference/depth_of_focus) + 1
        number_planes = 9

        for y in range(0, y_tiles):
            for x in range(0, x_tiles):
                #I believe the system to count from upper left hand corner starting at 0, 0

                x_point = 5056 * x + 2528
                y_point = 2060 * y + 1480
                focus_z = self.plane_2_z(coefficents, [x_point, y_point])
                fm_array[channel_index][y][x] = focus_z + (number_planes - 1)/2 * depth_of_focus
                fm_array[channel_index][y][x] = focus_z
                fm_array[channel_index + 1][y][x] = number_planes

        np.save('fm_array.npy', fm_array)


    def map_2_points(self, experiment_directory, channel):

        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = channel + '_sp_array.npy'
        sp_array = np.load(file_name, allow_pickle=False)
        fm_array = np.load('fm_array.npy', allow_pickle=False)

        y_fm_section = np.shape(sp_array)[1]
        x_fm_section = np.shape(sp_array)[2]

        y_tiles = np.shape(fm_array[0])[0]
        x_tiles = np.shape(fm_array[0])[1]



        y_pixels_per_section = int((y_tiles * 2960)/y_fm_section)
        x_pixels_per_section = int((x_tiles * 2960)/x_fm_section)

        point_list = np.array([])
        point_list = np.expand_dims(point_list, axis=0)

        first_point_counter = 0

        z_array = sp_array[4, :, :, 0] * sp_array[3, :, :, 0]

        for y in range(0, y_fm_section):
            for x in range(0, x_fm_section):

                x_coord = x * x_pixels_per_section
                y_coord = y * y_pixels_per_section
                z_coord = z_array[y][x]
                single_point = np.array([x_coord, y_coord, z_coord])
                single_point = np.expand_dims(single_point, axis=0)
                if first_point_counter == 0 and z_coord != 0:
                    point_list = np.append(point_list, single_point, axis=1)
                    first_point_counter = 1
                elif first_point_counter == 1 and z_coord != 0:
                    point_list = np.append(point_list, single_point, axis=0)

        point_list = Points(point_list)

        return point_list


    def gauss_jordan_solver(self, three_point_array):
        '''
        Takes 3 points and solves quadratic equation in a generic fashion and returns constants and
        solves for x in its derivative=0 equation

        :param numpy[float, float] three_point_array: numpy array that contains pairs of [focus_score, z]
        :results: z coordinate for in focus plane
        :rtype: float
        '''

        x1 = three_point_array[0][1]
        x2 = three_point_array[1][1]
        x3 = three_point_array[2][1]
        score_0 = three_point_array[0][0]
        score_1 = three_point_array[1][0]
        score_2 = three_point_array[2][0]

        aug_matrix = np.array([[x1 * x1, x1, 1, score_0], [x2 * x2, x2, 1, score_1], [x3 * x3, x3, 1, score_2]])

        aug_matrix[0] = aug_matrix[0] / (aug_matrix[0, 0] + 0.000001)
        aug_matrix[1] = -(aug_matrix[1, 0]) * aug_matrix[0] + aug_matrix[1]
        aug_matrix[2] = -(aug_matrix[2, 0]) * aug_matrix[0] + aug_matrix[2]

        aug_matrix[1] = -(aug_matrix[1, 1] - 1) / (aug_matrix[2, 1] + 0.0000001) * aug_matrix[2] + aug_matrix[1]
        aug_matrix[0] = -aug_matrix[0, 1] * aug_matrix[1] + aug_matrix[0]
        aug_matrix[2] = -aug_matrix[2, 1] * aug_matrix[1] + aug_matrix[2]

        aug_matrix[2] = aug_matrix[2] / (aug_matrix[2, 2] + 0.000001)
        aug_matrix[0] = -aug_matrix[0, 2] * aug_matrix[2] + aug_matrix[0]
        aug_matrix[1] = -aug_matrix[1, 2] * aug_matrix[2] + aug_matrix[1]

        a = aug_matrix[0, 3]
        b = aug_matrix[1, 3]
        c = aug_matrix[2, 3]

        derivative = -b / (2 * (a + 0.00001))

        return a,b,c,derivative

    #########################################################
    # Setup fm_array and sp_array alongside auto focus updates and flat values
    #########################################################

    def establish_fm_array(self, experiment_directory, desired_cycle_count, z_slices, off_array, initialize = 0, x_frame_size = 5056, autofocus = 0, auto_expose = 0):

        self.file_structure(experiment_directory, desired_cycle_count)


        if initialize == 1:
            xy_points = self.tile_xy_pos('New Surface 1')
            xyz_points = self.nonfocus_tile_DAPI(xy_points, experiment_directory)
            self.tile_pattern(xyz_points, experiment_directory)
            self.fm_channel_initial(experiment_directory, off_array, z_slices)
            self.establish_exp_arrays(experiment_directory)

            if x_frame_size != 5056:
                self.x_overlap_adjuster(x_crop_percentage, experiment_directory)
            else:
                pass

        else:
            pass


        if autofocus == 1 and auto_expose == 1:
            self.DAPI_surface_autofocus(experiment_directory, 30, 2)
            self.fm_channel_initial(experiment_directory, off_array, z_slices, 2)
            self.fm_array_update_autofocus_autoexpose(experiment_directory, exp=1)
        if autofocus == 1 and auto_expose == 0:
            self.DAPI_surface_autofocus(experiment_directory, 30, 2)
            self.fm_channel_initial(experiment_directory, off_array, z_slices, 2)
        if autofocus == 0 and auto_expose == 1:
            self.fm_array_update_autofocus_autoexpose(experiment_directory, exp = 1)
        else:
            pass


    def highest_brenner_index_solver(self, image_stack):

        slice_count = np.shape(image_stack)[0]
        scores = np.random.rand(slice_count)

        for x in range(0, scores):

            image = image_stack[x]
            score = self.focus_score(image, 17)
            scores[x] = score

        highest_score = np.max(scores)
        index = np.where(scores == highest_score)[0][0]

        return index


    def DAPI_surface_autofocus(self, experiment_directory, z_slices, z_slice_gap, x_frame_size):
        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)

        numpy_x = fm_array[0]
        numpy_y = fm_array[1]

        y_tile_count = numpy_y[0]
        x_tile_count = numpy_y[1]

        center_z = magellan.get_surface('New Surface 1').get_points().get(0).z
        bottom_z = int(center_z - z_slices/2 * z_slice_gap)
        top_z = int(center_z + z_slices/2 * z_slice_gap)

        #find crop range for x dimension

        side_pixels = int(5056 - x_frame_size)


        core.set_config("Color", 'DAPI')
        core.set_exposure(50)

        self.image_capture(experiment_directory, 'DAPI', 50, 0, 0, 0) #wake up lumencor light engine
        print('wait 10 seconds')

        core.set_xy_position(numpy_x[0][0], numpy_y[0][0])
        time.sleep(1)

        for x in range(0, x_tile_count):
            for y in range(0, y_tile_count):
                core.set_xy_position(numpy_x[y][x], numpy_y[y][x])
                z_stack = np.random.rand(z_slices, 2960, x_frame_size)
                time.sleep(1)
                stack_index = 0

                for z in range(bottom_z, top_z, z_slice_gap):

                    core.set_position(z)
                    time.sleep(0.5)

                    core.snap_image()
                    tagged_image = core.get_tagged_image()
                    pixels = np.reshape(tagged_image.pix, newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
                    z_stack[stack_index] = pixels[::, side_pixels:x_frame_size + side_pixels]

                    stack_index += 1

                z_index = self.highest_brenner_index_solver(z_stack)
                focus_z_position = bottom_z + z_index * z_slice_gap
                fm_array[2][y][x] = focus_z_position

        np.save('fm_array.npy', fm_array)


    def fm_array_update_autofocus_autoexpose(self, experiment_directory, exp = 0, focus = 0):
        '''
        Executes auto focus and exposure algorithms.

        :param experiment_directory:
        :return:
        '''
        #make super pixel array
        self.sp_array(experiment_directory)

        # load in data structures
        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        exp_filename = 'exp_array.npy'
        exp_calc_filename = 'exp_calc_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)
        exp_array = np.load(exp_filename, allow_pickle=False)
        exp_calc_array = np.load(exp_calc_filename, allow_pickle=False)
        channels = ['DAPI', 'A488', 'A555', 'A647']

        # break data structures into more usable components
        numpy_y = fm_array[1]
        numpy_x = fm_array[0]
        numpy_z = fm_array[2]
        number_slices = fm_array[3][0][0]
        y_tiles = np.shape(fm_array[0])[0]
        x_tiles = np.shape(fm_array[0])[1]

        # Go to upper left corner to start
        core.set_xy_position(numpy_x[0][0], numpy_y[0][0])

        #define range to scan through. 3 equally distributed points
        images = np.random.rand(3, 2960, 5056).astype('uint16')
        #order is x,y channel and z points

        self.image_capture(experiment_directory, 'DAPI', 50, 0, 0, 0) #wake up lumencor light engine
        time.sleep(10) #wait for it to wake up


        for x in range(0, x_tiles):
            for y in range(0, y_tiles):

                scan_range = 25
                sample_mid_z = numpy_z[0][0] - 2 * number_slices/2
                sample_span = [sample_mid_z - scan_range / 2, sample_mid_z, sample_mid_z + scan_range / 2]
                print(sample_span)

                for channel_index in range(0, 4):

                    if exp == 1 and focus == 0:
                        point = 1
                        end_point_number = 1
                    if exp == 1 and focus == 1 and channel_index == 0:
                        point = 0
                        end_point_number = 2
                    if exp == 1 and focus == 1 and channel_index != 0:
                        point = 1
                        end_point_number = 1
                    if exp == 0 and focus == 1 and channel_index == 0:
                        point = 0
                        end_point_number = 2
                    if exp == 0 and focus == 1 and channel_index != 0:
                        point = 2
                        end_point_number = 1

                    exp_time = int(exp_array[channel_index])
                    exp_calc_array_channel_xy = exp_calc_array[channel_index][y][x]

                    while point <= end_point_number:
                        #take image at XYZ position and determine 99 percentile intensity
                        z_slice = int(sample_span[point])
                        im = self.image_capture(experiment_directory, channels[channel_index], exp_time, x, y, z_slice)
                        # determine if intensity is too low or too high. Adjust exp time to compensate
                        exp_time, trigger_state = self.exp_bound_solver(im, exp_time)

                        exp_array[channel_index] = exp_time #allow 'memory' to happen. Effectively a markov model
                        time.sleep(0.5)


                        if trigger_state == 1: # restart z point aquistions
                            point = 0
                        if trigger_state == 0: # input points into data structure and move to next point
                            if exp == 1:
                                intensity = self.image_percentile_level(im, 0.99)  # 99th percentile intensity
                                exp_calc_array_channel_xy[0][0] = intensity - 300  # 300 is camera offset
                                exp_calc_array_channel_xy[0][1] = z_slice
                                exp_calc_array_channel_xy[0][2] = core.get_exposure()
                                exp_calc_array_channel_xy[1][0] = intensity - 300  # 300 is camera offset
                                exp_calc_array_channel_xy[1][1] = z_slice
                                exp_calc_array_channel_xy[1][2] = core.get_exposure()
                                exp_calc_array_channel_xy[2][0] = intensity - 300  # 300 is camera offset
                                exp_calc_array_channel_xy[2][1] = z_slice
                                exp_calc_array_channel_xy[2][2] = core.get_exposure()
                            else:
                                pass

                            #auto focus
                            if focus == 1 and channel_index == 0:
                                sub_im = self.image_sub_divider(im, 24, 32)
                                self.sub_divided_2_brenner_sp(experiment_directory, sub_im, channels[channel_index], point, z_slice, x, y)
                                images[point] = im
                            else:
                                pass


                            point += 1


        # solve and populate exp_array
        if exp == 1:
            np.save(exp_calc_filename, exp_calc_array)
            #self.one_slice_calc_array_solver(experiment_directory)
            self.calc_array_solver(experiment_directory)
            self.calc_array_2_exp_array(experiment_directory, 0.1)  # 0.2 =20% dynamic range used
        else:
            pass

        # solve sp array and populate fm array
        if focus == 1:
            self.sp_array_focus_solver(experiment_directory, 'DAPI')
            self.sp_array_filter(experiment_directory, 'DAPI')
            self.sp_array_surface_2_fm(experiment_directory, 'DAPI')
        else:
            pass


    ###########################################################
    #This section is the for the exposure functions.
    ###########################################################

    def establish_exp_arrays(self, experiment_directory):
        '''
        Make exp_array with default exp times and exp_calc_array to use with auto exposure

        :param experiment_directory:
        :return:
        '''
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        fm_array = np.load('fm_array.npy', allow_pickle=False)

        y_tiles = np.shape(fm_array[0])[0]
        x_tiles = np.shape(fm_array[0])[1]

        exp_calc_array = np.random.rand(4, y_tiles, x_tiles, 4, 3)
        exp_array = np.array([100,100,100,100])

        file_name = 'exp_calc_array.npy'
        np.save(file_name, exp_calc_array)
        np.save('exp_array.npy', exp_array)


    def exp_bound_solver(self, image, exp_time):
        '''
        Takes in image and determines if its in bounds. If not, it adjusts exp time to compensate and gives indicator
        if exp time alterations were used, ie was it triggered.

        :param image:
        :param exp_time:
        :return:
        '''

        target_intensity = 65535 * 0.05
        max_time = 100
        trigger_state = 0

        intensity = self.image_percentile_level(image, 0.99)

        while intensity > 65000:

            exp_time = exp_time/10
            core.set_exposure(exp_time)
            core.snap_image()
            tagged_image = core.get_tagged_image()
            new_image = np.reshape(tagged_image.pix, newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
            intensity = self.image_percentile_level(new_image, 0.99)
            trigger_state = 1

        if intensity < 1000 and exp_time < max_time:
            trigger_state = 1

        if trigger_state == 1:
            scale_factor = target_intensity/intensity
            if scale_factor * exp_time > max_time:
                exp_time = max_time
            else:
                exp_time = int(exp_time * scale_factor)

        return exp_time, trigger_state


    def image_percentile_level(self, image, cut_off_threshold = 0.9):
        '''
        Takes in image and cut off threshold and finds pixel value that exists at that threshold point.

        :param numpy array image: numpy array image
        :param float cut_off_threshold: percentile for cut off. For example a 0.99 would disregaurd the top 1% of pixels from calculations
        :return: intensity og pixel that resides at the cut off fraction that was entered in the image
        :rtype: int
        '''
        #cut_off_threshold = 0.9
        threshy_image = image / 10
        thresh = filters.threshold_otsu(threshy_image)
        index = np.where(threshy_image > thresh)
        pixel_values = np.sort(image[index], axis=None)
        pixel_count = int(np.size(pixel_values))
        cut_off_index = int(pixel_count * cut_off_threshold)
        tail_intensity = pixel_values[cut_off_index]


        return tail_intensity

    def one_slice_calc_array_solver(self, experiment_directory):

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'exp_calc_array.npy'
        calc_array = np.load(file_name, allow_pickle=False)
        file_name = 'exp_array.npy'
        exp_array = np.load(file_name, allow_pickle=False)

        y_tiles = np.shape(calc_array)[1]
        x_tiles = np.shape(calc_array)[2]

        goal_int = 65500 * 0.2



        for channel_index in range(0, 4):
            exp_time_list = np.ones([y_tiles, x_tiles])
            int_time_list = np.ones([y_tiles, x_tiles])
            for x in range(0, x_tiles):
                for y in range(0, y_tiles):
                    intensity = calc_array[channel_index, y, x, 0, 0]
                    exp_time_used = calc_array[channel_index, y, x, 0, 2]
                    exp_time_list[y][x] = exp_time_used
                    int_time_list[y][x] = intensity

            scaled_int_time_list = np.max(int_time_list)/int_time_list
            scaled_exp_time_list = exp_time_list * scaled_int_time_list
            brightest = np.max(scaled_exp_time_list)
            y_index = np.where(scaled_exp_time_list == brightest)[0][0]
            x_index = np.where(scaled_exp_time_list == brightest)[1][0]
            brightest_int = int_time_list[y_index][x_index]
            brightest_exp = exp_time_list[y_index][x_index]

            scale_factor = goal_int/brightest_int
            new_exp = scale_factor * brightest_exp

            print(channel_index, new_exp)

            exp_array[channel_index] = new_exp

        np.save('exp_array.npy', exp_array)



    def calc_array_solver(self, experiment_directory):
        '''
        Uses calc array and 3 point gauss jordan reduction method to solve for projected intensity in focal plane

        :param experiment_directory:
        :return:
        '''

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'exp_calc_array.npy'
        calc_array = np.load(file_name, allow_pickle=False)

        y_tiles = np.shape(calc_array)[1]
        x_tiles = np.shape(calc_array)[2]

        for channel_index in range(0, 4):
            for x in range(0, x_tiles):
                for y in range(0, y_tiles):
                    scores = calc_array[channel_index, y, x, 0:3, 0]
                    positions = calc_array[channel_index, y, x, 0:3, 1]
                    #three_point_array = np.stack((scores, positions), axis=1)

                    #a, b, c, predicted_focus = self.gauss_jordan_solver(three_point_array)
                    #peak_int = (-(b * b) / (4 * a) + c)
                    #calc_array[channel_index][y][x][3][0] = peak_int
                    #calc_array[channel_index][y][x][3][1] = predicted_focus
                    calc_array[channel_index][y][x][3][0] = calc_array[channel_index][y][x][0][0]

        np.save(file_name, calc_array)

    def calc_array_2_exp_array(self, experiment_directory, fraction_dynamic_range):
        '''
        Takes calc array and determines time to place into exp_array for use. In short, it scales intensities and
        find highest scaled intensity and then scales it from there again to get to the desired dynamic range occupied.
        Employs max time cut off as well

        :param experiment_directory:
        :param fraction_dynamic_range:
        :return:
        '''

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        calc_array = np.load('exp_calc_array.npy', allow_pickle=False)
        exp_array = np.load('exp_array.npy', allow_pickle=False)

        y_tiles = np.shape(calc_array)[1]
        x_tiles = np.shape(calc_array)[2]

        max_time = 2000

        desired_top_intensity = fraction_dynamic_range * 65535

        for channel_index in range(0, 4):
            predicted_int_list = np.ones([y_tiles, x_tiles])
            exp_time_list = np.ones([y_tiles, x_tiles])
            for x in range(0, x_tiles):
                for y in range(0, y_tiles):
                    exp_time_list[y][x] = calc_array[channel_index][y][x][0][2]
                    predicted_int_list[y][x]= calc_array[channel_index][y][x][3][0]

            lowest_exp_time = np.min(exp_time_list)
            scaled_exp_list = lowest_exp_time/exp_time_list
            scaled_int_list = predicted_int_list * scaled_exp_list
            highest_intensity = np.max(scaled_int_list)
            index = np.where(scaled_int_list == highest_intensity)
            dimensions = np.shape(index)[0]
            if dimensions == 1:
                highest_intensity = predicted_int_list[index[0][0]]
                exp_time_for_highest_int = exp_time_list[index[0][0]]
            if dimensions == 2:
                highest_intensity = predicted_int_list[index[0][0]][index[1][0]]
                exp_time_for_highest_int = exp_time_list[index[0][0]][index[1][0]]

            scale_up_factor = desired_top_intensity / highest_intensity
            new_exp_time = int(exp_time_for_highest_int * scale_up_factor)

            if new_exp_time > max_time:
                new_exp_time = max_time
            else:
                pass

            if new_exp_time < 50:
                new_exp_time = 50
            else:
                pass

            exp_array[channel_index] = new_exp_time
            print(channel_index, new_exp_time)

        np.save('exp_array.npy', exp_array)


    ##############################################

    def tile_xy_pos(self, surface_name):
        '''
        imports previously generated micro-magellan surface with name surface_name and outputs
        the coordinates of the center of each tile from it.

        :param str surface_name: name of micro-magellan surface
        :param object magellan: object created via = bridge.get_magellan()

        :return: XY coordinates of the center of each tile from micro-magellan surface
        :rtype: dictionary {{x:(float)}, {y:(float)}}
        '''
        surface = magellan.get_surface(surface_name)
        num = surface.get_num_positions()
        xy = surface.get_xy_positions()
        tile_points_xy = {}
        x_temp = []
        y_temp = []

        for q in range(0, num):
            pos = xy.get(q)
            pos = pos.get_center()
            x_temp.append(pos.x)
            y_temp.append(pos.y)

        tile_points_xy['x'] = x_temp  ## put all points in dictionary to ease use
        tile_points_xy['y'] = y_temp
        x_temp = np.array(tile_points_xy['x'])
        y_temp = np.array(tile_points_xy['y'])

        xy = np.append(x_temp, y_temp)
        xy = np.reshape(xy, (2, int(xy.size/2)))

        return xy


    def tile_pattern(self, numpy_array, experiment_directory):
        '''
        Takes numpy array with N rows and known tile pattern and casts into new array that follows
        south-north, west-east snake pattern.


        :param numpy_array: dimensions [N, x_tiles*y_tiles]
        :param x_tiles: number x tiles in pattern
        :param y_tiles: number y tiles in pattern
        :return: numpy array with dimensions [N,x_tiles,y_tiles] with above snake pattern
        '''
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        filename = 'fm_array.npy'

        y_tiles = np.unique(numpy_array[1]).size
        x_tiles = np.unique(numpy_array[0]).size
        layers = np.shape(numpy_array)[0]
        numpy_array = numpy_array.reshape(layers, x_tiles, y_tiles)
        dummy = numpy_array.reshape(layers, y_tiles, x_tiles)
        new_numpy = np.empty_like(dummy)
        for x in range(0, layers):
            new_numpy[x] = numpy_array[x].transpose()
            new_numpy[x, ::, 1:y_tiles:2] = np.flipud(new_numpy[x, ::, 1:y_tiles:2])


        np.save(filename, new_numpy)

        return new_numpy

    def fm_channel_initial(self, experiment_directory, off_array, z_slices, slice_gap = 2):


        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)

        a488_channel_offset = off_array[1] #determine if each of these are good and repeatable offsets
        a555_channel_offset = off_array[2]
        a647_channel_offset = off_array[3]

        y_tiles = np.shape(fm_array[0])[0]
        x_tiles = np.shape(fm_array[0])[1]

        dummy_channel = np.empty_like(fm_array[0])
        dummy_channel = np.expand_dims(dummy_channel, axis=0)
        channel_count = np.shape(fm_array)[0]

        while channel_count < 10:
            fm_array = np.append(fm_array, dummy_channel, axis = 0)
            channel_count = np.shape(fm_array)[0]

        fm_array[4] = fm_array[2] + a488_channel_offset #index for a488 = 3
        fm_array[6] = fm_array[2] + a555_channel_offset
        fm_array[8] = fm_array[2] + a647_channel_offset
        y_tiles = int(np.shape(fm_array[0])[0])
        x_tiles = int(np.shape(fm_array[0])[1])
        z_slice_array = np.full((y_tiles, x_tiles), z_slices)

        fm_array[3] = z_slice_array
        fm_array[5] = z_slice_array
        fm_array[7] = z_slice_array
        fm_array[9] = z_slice_array


        fm_array[2] = fm_array[2] + int((z_slice_array[0][0] * slice_gap)/2) - 1
        fm_array[4] = fm_array[4] + int((z_slice_array[0][0] * slice_gap)/2) - 1
        fm_array[6] = fm_array[6] + int((z_slice_array[0][0] * slice_gap)/2) - 1
        fm_array[8] = fm_array[8] + int((z_slice_array[0][0] * slice_gap)/2) - 1

        np.save(file_name, fm_array)

        return fm_array

    def x_overlap_adjuster(self, new_x_pixel_count, experiment_directory):
        '''
        Increases overlap in focus map while cropping in the x dimension to preserve effective 10% overlap of cropped images
        :param new_x_pixel_count:
        :param experiment_directory:
        :return:
        '''


        # load in fm array
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        file_name = 'fm_array.npy'
        fm_array = np.load(file_name, allow_pickle=False)

        # Find pixels in each dimensions
        x_tiles = np.shape(fm_array[0])[1]
        y_tiles = np.shape(fm_array[0])[0]

        # find um/pixel in focus map

        try:
            x1 = fm_array[0][0][0]
            x2 = fm_array[0][0][1]
            diff = x2 - x1
        except:
            y1 = fm_array[0][0][0]
            y2 = fm_array[0][1][0]
            diff = y2 - y1

        um_per_pixel = diff/4550 # 4550 = 0.9 * 5056

        # Find number tiles in adjusted grid
        x_range_pixels = (x_tiles - 0.2) * 5056
        number_new_x_dim_tiles = x_range_pixels/new_x_pixel_count
        new_x_tiles = math.ceil(number_new_x_dim_tiles)

        # generate new blank fm_array numpy array

        new_fm_array = np.random.rand(10, y_tiles, new_x_tiles).astype('int16')

        # Find border where x starts on the left (not center point, but x value for left most edge of left most tile

        left_x_center = fm_array[0][0][0]
        left_most_x = left_x_center - 2528 * um_per_pixel

        # Find center point in new image that makes edge of image align with left_most_x
        # Also find x to x + i spacing and populate rest of x values in new_fm_array

        x_col_0 = left_most_x + new_x_pixel_count / 2 * um_per_pixel
        x_spacing = (0.9) * new_x_pixel_count * um_per_pixel

        # Populate new_fm_array with row 0 x values

        for x in range(0, x_tiles):
            new_fm_array[0, 0:y_tiles, x] = x_col_0 + x * x_spacing

        # populate new_fm_array with y values

        for y in range(0, y_tiles):
            new_fm_array[0, y, 0:x_tiles] = fm_array[1][y][0]

        # populate new_fm_array with dapi z values and everything else in planes 2-9
        for slice in range(2,10):
            new_fm_array[slice, 0:y_tiles, 0:x_tiles] = fm_array[slice][0][0]

        np.save('fm_array.npy', new_fm_array)

    def nonfocus_tile_DAPI(self, full_array_no_pattern, experiment_directory):
         '''
         Makes micromagellen z the focus position at each XY point for DAPI
         auto_focus and outputs the paired in focus z coordinate

         :param dictionary tile_points_xy: dictionary containing all XY coordinates. In the form: {{x:(int)}, {y:(int)}}

         :return: XYZ points where XY are stage coords and Z is in focus coordinate. {{x:(int)}, {y:(int)}, {z:(float)}}
         :rtype: dictionary
         '''

         numpy_path = experiment_directory + '/' + 'np_arrays'
         os.chdir(numpy_path)
         file_name = 'fm_array.npy'

         z_pos = magellan.get_surface('New Surface 1').get_points().get(0).z
         num = np.shape(full_array_no_pattern)[1]
         z_temp = []
         for q in range(0, num):
             z_temp.append(z_pos)
         z_temp = np.expand_dims(z_temp, axis = 0)
         xyz = np.append(full_array_no_pattern, z_temp, axis =0)

         np.save(file_name, xyz)

         return xyz


############################################
#Using core snap and not pycromanager acquire
############################################

    def position_verify(self, z_position):

        difference_range = 1
        current_z = core.get_position()
        difference = abs(z_position - current_z)
        while difference > difference_range:
            core.set_position(z_position)
            current_z = core.get_position()
            difference = abs(z_position - current_z)
            time.sleep(0.5)

    def image_capture(self, experiment_directory, channel, exp_time, x,y,z):

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        fm_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = fm_array[0]
        numpy_y = fm_array[1]
        numpy_z = fm_array[2]

        core.set_config("Color", channel)
        core.set_exposure(exp_time)
        core.set_xy_position(numpy_x[y][x], numpy_y[y][x])
        #time.sleep(1)
        core.set_position(z)
        time.sleep(0.5)

        core.snap_image()
        tagged_image = core.get_tagged_image()
        pixels = np.reshape(tagged_image.pix, newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
        #time.sleep(1)

        return pixels


    def multi_channel_z_stack_capture(self, experiment_directory, cycle_number, Stain_or_Bleach, list_status, window, x_pixels = 5056, slice_gap = 2, channels = ['DAPI', 'A488', 'A555', 'A647']):
        '''
        Captures and saves all images in XY and Z dimensions. Order of operation is ZC XY(snake). Entire z stack with all
        channels is made into a numpy data structure and saved before going to next tile and being reused. This is done
        to reduce overall memory usage.


        :param experiment_directory: directory that main experiment data is stored in
        :param cycle_number: integer of what cycle number it is. Counts from 0
        :param Stain_or_Bleach: string of Stain or Bleach depending on what action it is
        :param slice_gap: micron spacing in z slices
        :param channels: string list of what channels to acquire
        :return:
        '''

        core = Core()


        #load in focus map and exp array
        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)
        exp_time_array = np.load('exp_array.npy', allow_pickle=False)

        height_pixels = 2960
        width_pixels = 5056
        #determine attributes like tile counts,z slices and channel counts
        numpy_x = full_array[0]
        numpy_y = full_array[1]

        x_tile_count = np.unique(numpy_x).size
        y_tile_count = np.unique(numpy_y).size
        z_slices = full_array[5][0][0]
        #z_slices = 11
        #go to upper left corner to start pattern
        core.set_xy_position(numpy_x[0][0], numpy_y[0][0])
        time.sleep(1)
        #generate numpy data structure
        zc_tif_stack = np.random.rand(4, int(z_slices), height_pixels, width_pixels).astype('float16')

        image_number_counter = 0

        for y in range(0, y_tile_count):
            if y % 2 != 0:
                for x in range(x_tile_count - 1, -1, -1):

                    core.set_xy_position(numpy_x[y][x], numpy_y[y][x])
                    time.sleep(1)

                    for channel in channels:

                        #determine the proper indecies to use for focus map z positions and exp array
                        if channel == 'DAPI':
                            channel_index = 2
                            tif_stack_c_index = 0
                            zc_index = 0
                        if channel == 'A488':
                            channel_index = 4
                            tif_stack_c_index = 1
                            zc_index = 1
                        if channel == 'A555':
                            channel_index = 6
                            tif_stack_c_index = 2
                            zc_index = 2
                        if channel == 'A647':
                            channel_index = 8
                            tif_stack_c_index = 3
                            zc_index = 3

                        numpy_z = full_array[channel_index]
                        exp_time = int(exp_time_array[tif_stack_c_index])
                        core.set_config("Color", channel)
                        core.set_exposure(exp_time)

                        z_end = int(numpy_z[y][x])
                        z_start = int(z_end - z_slices * slice_gap)

                        z_counter = 0

                        for z in range(z_start, z_end, slice_gap):
                            core.set_position(z)
                            time.sleep(0.3)
                            core.snap_image()
                            tagged_image = core.get_tagged_image()
                            pixels = np.reshape(tagged_image.pix,newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
                            zc_tif_stack[zc_index][z_counter] = pixels

                            #core.pop_next_tagged_image()
                            image_number_counter += 1
                            z_counter += 1

                    #save zc stack
                    status_str = f'Cycle {cycle_number}: {x} {y} start saving z stack'
                    print(status_str)
                    status_update(status_str, list_status, window)
                    self.zc_save(zc_tif_stack, channels, x, y, cycle_number, experiment_directory, Stain_or_Bleach)
                    status_str = f'Cycle {cycle_number}: {x} {y} finished saving z stack'
                    print(status_str)
                    status_update(status_str, list_status, window)


            elif y % 2 == 0:
                for x in range(0, x_tile_count):

                    core.set_xy_position(numpy_x[y][x], numpy_y[y][x])
                    time.sleep(1)

                    for channel in channels:

                        # determine the proper indecies to use for focus map z positions and exp array
                        if channel == 'DAPI':
                            channel_index = 2
                            tif_stack_c_index = 0
                            zc_index = 0
                        if channel == 'A488':
                            channel_index = 4
                            tif_stack_c_index = 1
                            zc_index = 1
                        if channel == 'A555':
                            channel_index = 6
                            tif_stack_c_index = 2
                            zc_index = 2
                        if channel == 'A647':
                            channel_index = 8
                            tif_stack_c_index = 3
                            zc_index = 3

                        numpy_z = full_array[channel_index]
                        exp_time = int(exp_time_array[tif_stack_c_index])
                        core.set_config("Color", channel)
                        core.set_exposure(exp_time)

                        z_end = int(numpy_z[y][x])
                        z_start = int(z_end - z_slices * slice_gap)
                        z_counter = 0

                        for z in range(z_start, z_end, slice_gap):
                            core.set_position(z)
                            time.sleep(0.3)
                            core.snap_image()
                            tagged_image = core.get_tagged_image()
                            pixels = np.reshape(tagged_image.pix,
                                                newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
                            zc_tif_stack[zc_index][z_counter] = pixels

                            #core.pop_next_tagged_image()
                            image_number_counter += 1
                            z_counter += 1

                    status_str = f'Cycle {cycle_number}: {x} {y} start saving z stack'
                    print(status_str)
                    status_update(status_str, list_status, window)
                    self.zc_save(zc_tif_stack, channels, x, y, cycle_number, x_pixels, experiment_directory, Stain_or_Bleach)
                    status_str = f'Cycle {cycle_number}: {x} {y} finished saving z stack'
                    print(status_str)
                    status_update(status_str, list_status, window)

        print('all finished acquire events for this cycle')
        status_str = f'Cycle {cycle_number}: finished acquiring'
        status_update(status_str, list_status, window)
        return


    def image_cycle_acquire(self, cycle_number, experiment_directory, z_slices, stain_bleach, offset_array, list_status, window, x_frame_size = 5056, establish_fm_array = 0, auto_focus_run = 0, auto_expose_run = 0, channels = ['DAPI', 'A488', 'A555', 'A647']):

        status_str = f'Cycle {cycle_number}: establish fm array'
        print(status_str)
        status_update(status_str, list_status, window)
        self.establish_fm_array(experiment_directory, cycle_number, z_slices, offset_array, initialize= establish_fm_array, x_frame_size = x_frame_size, autofocus=auto_focus_run, auto_expose=auto_expose_run)
        status_str = f'Cycle {cycle_number}: image capture to wake up engine'
        print(status_str)
        status_update(status_str, list_status, window)
        self.image_capture(experiment_directory, 'DAPI', 50, 0, 0, 0) #wake up lumencor light engine
        status_str = f'Cycle {cycle_number}: wait 10 seconds'
        print(status_str)
        status_update(status_str, list_status, window)

        '''
        std_dev = np.std(start_image)
        if std_dev > threshold:
            pass
        else: 
            time.sleep(10)
        '''
        time.sleep(10) #wait for it to wake up
        ''''
        exp_time = exp_time_array
        np.save('exp_array.npy', exp_time)

        for channel in channels:
            z_tile_stack = self.core_tile_acquire(experiment_directory, channel)
            self.save_files(z_tile_stack, channel, cycle_number, experiment_directory, stain_bleach)
            
        '''
        status_str = f'Cycle {cycle_number}:acquire all images'
        print(status_str)
        status_update(status_str, list_status, window)
        self.multi_channel_z_stack_capture(experiment_directory, cycle_number, stain_bleach, list_status, window, x_frame_size = x_frame_size,  slice_gap=2, channels = channels)
        #self.marker_excel_file_generation(experiment_directory, cycle_number)


    def full_cycle(self, experiment_directory, cycle_number, offset_array, stain_valve, fluidics_object, z_slices, window, list_status, x_frame_size = 5056, incub_val=45):

        pump = fluidics_object
        # z_slices = 9

        if cycle_number == 0:
            status_str = f'Cycle {cycle_number}: baseline bleach image acquiring'
            status_update(status_str, list_status, window)
            # print(status_str)
            self.image_cycle_acquire(cycle_number, experiment_directory, z_slices, 'Bleach', offset_array, list_status, window, x_frame_size, establish_fm_array = 1, auto_focus_run=0, auto_expose_run=0)
        else:
            status_str = f'Cycle {cycle_number}: Stain in progress'
            status_update(status_str, list_status, window)
            # print(status_str)
            pump.liquid_action('Stain', incub_val, stain_valve, window, list_status)  # nuc is valve=7, pbs valve=8, bleach valve=1 (action, stain_valve, heater state (off = 0, on = 1))
            #print('washing')
            time.sleep(5)
            status_str = f'Cycle {cycle_number}: washing in progress'
            status_update(status_str, list_status, window)
            pump.liquid_action('Wash', stain_valve, window, list_status)  # nuc is valve=7, pbs valve=8, bleach valve=1 (action, stain_valve, heater state (off = 0, on = 1))
            #pump.liquid_action('PBS flow off')  # nuc is valve=7, pbs valve=8, bleach valve=1 (action, stain_valve, heater state (off = 0, on = 1))
            #time.sleep(5)
            status_str = f'Cycle {cycle_number}: stain image acquistion in progress'
            status_update(status_str, list_status, window)
            # print(status_str)
            self.image_cycle_acquire(cycle_number, experiment_directory, z_slices, 'Stain', offset_array, list_status, window, x_frame_size = x_frame_size, establish_fm_array=0, auto_focus_run=0, auto_expose_run = 1)
            time.sleep(5)
            status_str = f'Cycle {cycle_number}: bleaching in progress'
            status_update(status_str, list_status, window)
            # print(status_str)
            pump.liquid_action('Bleach', stain_valve, list_status, window)  # nuc is valve=7, pbs valve=8, bleach valve=1 (action, stain_valve, heater state (off = 0, on = 1))
            #print('washing')
            #time.sleep(5)
            status_str = f'Cycle {cycle_number}: washing in progress'
            status_update(status_str, list_status, window)
            pump.liquid_action('Wash', stain_valve, window, list_status)  # nuc is valve=7, pbs valve=8, bleach valve=1 (action, stain_valve, heater state (off = 0, on = 1))
            #pump.liquid_action('PBS flow off')  # nuc is valve=7, pbs valve=8, bleach valve=1 (action, stain_valve, heater state (off = 0, on = 1))
            time.sleep(5)
            status_str = 'bleach images acquiring'
            status_update(status_str, list_status, window)
            # print(status_str)
            self.image_cycle_acquire(cycle_number, experiment_directory, z_slices, 'Bleach', offset_array, list_status, window, x_frame_size = x_frame_size, establish_fm_array=0, auto_focus_run=0, auto_expose_run = 0)
            time.sleep(10)


    def antibody_kinetics(self, experiment_directory, capture_rate_staining, capture_rate_bleaching, duration_staining, duration_bleaching, stain_valve,  fluidic_object, channels = ['DAPI', 'A488', 'A555', 'A647']):

        '''
        Dispenses stain avia fluidic system and images until certain time point. After that, bleach dispenses and capturing continues at a different
        rate until set time point


        :param experiment_directory:
        :param capture_rate_staining: points/min
        :param capture_rate_bleaching: points/min
        :param duration_staining: minutes
        :param duration_bleaching: minutes
        :param channels:
        :return:
        '''

        exp_array = [50,5,5,5]

        #create folders

        os.chdir(experiment_directory)
        os.mkdir('DAPI')
        os.mkdir('A488')
        os.mkdir('A555')
        os.mkdir('A647')

        dapi_path = experiment_directory + r'\DAPI'
        a488_path = experiment_directory + r'\A488'
        a555_path = experiment_directory + r'\A555'
        a647_path = experiment_directory + r'\A647'

        #dimensional parameters
        y_pixel_count = 2960
        x_pixel_count = 5056
        channel_count = len(channels)
        time_point_stain_count = int(duration_staining * capture_rate_staining)
        time_point_bleach_count = int(duration_bleaching * capture_rate_bleaching)
        time_gap_staining = 1/capture_rate_staining * 60
        time_gap_bleach = 1/capture_rate_bleaching * 60

        # create data structure for staining images
        data_points_stain = np.full((time_point_stain_count, channel_count, y_pixel_count, x_pixel_count), 0)
        data_points_bleach = np.full((time_point_bleach_count, channel_count, y_pixel_count, x_pixel_count), 0)

        fluidic_object.valve_select(stain_valve)
        fluidic_object.flow(500)
        time.sleep(45)
        fluidic_object.flow(0)
        fluidic_object.valve_select(12)

        for time_point in range(0, time_point_stain_count):
            for channel in channels:

                if channel == 'DAPI':
                    channel_index = 0
                if channel == 'A488':
                    channel_index = 1
                if channel == 'A555':
                    channel_index = 2
                if channel == 'A647':
                    channel_index = 3

                exp_time = exp_array[channel_index]

                core.set_config("Color", channel)
                core.set_exposure(exp_time)

                core.snap_image()
                tagged_image = core.get_tagged_image()
                pixels = np.reshape(tagged_image.pix,newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
                data_points_stain[time_point][channel_index] = pixels

            time.sleep(time_gap_staining)

            os.chdir(dapi_path)
            io.imsave('dapi_stain_stack', data_points_stain[::, 0, ::, ::])
            os.chdir(a488_path)
            io.imsave('a488_stain_stack', data_points_stain[::, 1, ::, ::])
            os.chdir(a555_path)
            io.imsave('a555_stain_stack', data_points_stain[::, 2, ::, ::])
            os.chdir(a647_path)
            io.imsave('a647_stain_stack', data_points_stain[::, 3, ::, ::])

            fluidic_object.valve_select(11)
            fluidic_object.flow(500)
            time.sleep(45)
            fluidic_object.flow(0)

            for time_point in range(0, time_point_bleach_count):
                for channel in channels:

                    if channel == 'DAPI':
                        channel_index = 0
                    if channel == 'A488':
                        channel_index = 1
                    if channel == 'A555':
                        channel_index = 2
                    if channel == 'A647':
                        channel_index = 3

                    exp_time = exp_array[channel_index]

                    core.set_config("Color", channel)
                    core.set_exposure(exp_time)

                    core.snap_image()
                    tagged_image = core.get_tagged_image()
                    pixels = np.reshape(tagged_image.pix,
                                        newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
                    data_points_bleach[time_point][channel_index] = pixels

                time.sleep(time_gap_bleach)

            os.chdir(dapi_path)
            io.imsave('dapi_stain_stack', data_points_bleach[::, 0, ::, ::])
            os.chdir(a488_path)
            io.imsave('a488_stain_stack', data_points_bleach[::, 1, ::, ::])
            os.chdir(a555_path)
            io.imsave('a555_stain_stack', data_points_bleach[::, 2, ::, ::])
            os.chdir(a647_path)
            io.imsave('a647_stain_stack', data_points_bleach[::, 3, ::, ::])

            fluidic_object.valve_select(12)
            fluidic_object.flow(500)
            time.sleep(70)
            fluidic_object.flow(0)



    ######Folder System Generation########################################################


    def marker_excel_file_generation(self, experiment_directory, cycle_number):

        folder_path = experiment_directory + '/mcmicro'
        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        exp_array = np.load('exp_array.npy', allow_pickle=False)
        exp_array = np.flip(exp_array)
        
        filter_sets = ['A647', 'A555', 'A488', 'DAPI']
        emission_wavelengths = ['675', '565', '525', '450']
        exciation_wavlengths = ['647', '555', '488', '405']

        try:
            os.chdir(folder_path)
            wb = load_workbook('markers.xlsx')
        except:
            wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = 'channel_number'
        ws.cell(row=1, column=2).value = 'cycle_number'
        ws.cell(row=1, column=3).value = 'marker_name'
        ws.cell(row=1, column=4).value = 'Filter'
        ws.cell(row=1, column=5).value = 'excitation_wavlength'
        ws.cell(row=1, column=6).value = 'emission_wavlength'
        ws.cell(row=1, column=7).value = 'background'
        ws.cell(row=1, column=8).value = 'exposure'
        ws.cell(row=1, column=9).value = 'remove'
        
        for row_number in range(2, (cycle_number)*4 + 2):

            cycle_number = 4//(row_number - 2) + 1
            intercycle_channel_number = cycle_number * 4 + 1 - row_number

            ws.cell(row=row_number, column=1).value = row_number
            ws.cell(row=row_number, column=2).value = cycle_number
            ws.cell(row=row_number, column=3).value = 'Marker_' + str(row_number)
            #ws.cell(row=row_number, column=4).value = filter_sets[intercycle_channel_number]
            #ws.cell(row=row_number, column=5).value = exciation_wavlengths[intercycle_channel_number]
            #ws.cell(row=row_number, column=6).value = emission_wavelengths[intercycle_channel_number]


        row_start = (cycle_number - 1)*4 + 2
        row_end = row_start + 4

        for row_number in range(row_start, row_end):

            cycle_number = 4//(row_number - 2) + 1
            intercycle_channel_number = cycle_number * 4 + 1 - row_number

            ws.cell(row=row_number, column=8).value = exp_array[row_number-2]

        os.chdir(folder_path)
        wb.save(filename = 'markers.xlsx')


    def folder_addon(self, parent_directory_path, new_folder_names):

        os.chdir(parent_directory_path)

        for name in new_folder_names:

            add_on_folder = str(name)
            full_directory_path = parent_directory_path + '/' + add_on_folder

            try:
                os.mkdir(full_directory_path)
            except:
               pass

    def file_structure(self, experiment_directory, highest_cycle_count):

        channels = ['DAPI', 'A488', 'A555', 'A647']
        cycles = np.linspace(0,highest_cycle_count).astype(int)

        #folder layer one
        os.chdir(experiment_directory)
        self.folder_addon(experiment_directory, ['Quick_Tile'])
        self.folder_addon(experiment_directory, ['np_arrays'])
        self.folder_addon(experiment_directory, ['mcmicro'])
        self.folder_addon(experiment_directory, channels)

        #folder layer two

        mc_micro_directory = experiment_directory + '/mcmicro'
        self.folder_addon(mc_micro_directory, ['raw'])
        quick_tile_directory = experiment_directory + '/Quick_Tile'
        self.folder_addon(quick_tile_directory, ['DAPI'])
        self.folder_addon(quick_tile_directory, ['A488'])
        self.folder_addon(quick_tile_directory, ['A555'])
        self.folder_addon(quick_tile_directory, ['A647'])

        for channel in channels:

            experiment_channel_directory = experiment_directory + '/' + channel

            self.folder_addon(experiment_channel_directory, ['Stain'])
            self.folder_addon(experiment_channel_directory, ['Bleach'])

        #folder layers 3 and 4

            for cycle in cycles:

                experiment_channel_stain_directory = experiment_channel_directory + '/' + 'Stain'
                experiment_channel_bleach_directory = experiment_channel_directory + '/' + 'Bleach'


                self.folder_addon(experiment_channel_stain_directory, ['cy_' + str(cycle)])
                experiment_channel_stain_cycle_directory = experiment_channel_stain_directory + '/' + 'cy_' + str(cycle)

                self.folder_addon(experiment_channel_stain_cycle_directory, ['Tiles'])


                self.folder_addon(experiment_channel_bleach_directory, ['cy_' + str(cycle)])
                experiment_channel_bleach_cycle_directory = experiment_channel_bleach_directory + '/' + 'cy_' + str(cycle)

                self.folder_addon(experiment_channel_bleach_cycle_directory, ['Tiles'])

        os.chdir(experiment_directory + '/' + 'np_arrays')

#####################################################################################################
##########Saving/File Generation Methods#############################################################################

    def z_scan_exposure_hook(image, metadata):
        '''
        Method that hooks from autofocus image acquistion calls. It takes image, calculates a focus score for it
        via focus_score method and exports a list that contains both the focus score and the z position it was taken at

        :param numpy image: single image from hooked from acquistion
        :param list[float] metadata: metadata for image

        :return: Nothing
        '''
        z = metadata.pop('ZPosition_um_Intended')  # moves up while taking z stack
        z_intensity_level = cycif.image_percentile_level(image, 0.99)
        intensity.value.append([z_intensity_level, z])

        return

    def auto_initial_expose(self, seed_expose, benchmark_threshold, channel, z_range, surface_name):
        '''
        Scans z levels around surface z center and finds brightest z position via z_scan_exposure method.
        Moves machine to that z plane and executes auto_expose method to determine proper exposure. This is meant for
        an initial exposure value for autofocus pruposes.

        :param: str surface_name: string of name of magellan surface to use
        :param list z_centers: list of z points associated with xy points where the slide tilt was compensated for
        :param: str channels: list that contains strings with channel names, for example 'DAPI'

        :return: exposure time: time for inputted channels exposure to be used for autofocus
        :rtype: int
        '''

        [x_pos, y_pos] = self.tissue_center(surface_name)
        core.set_xy_position(x_pos, y_pos)

        z_brightest = z_range[0] + z_range[2]
        core.set_position(z_brightest)

        new_exp = cycif.auto_expose(seed_expose, benchmark_threshold, z_brightest, [channel])

        return new_exp

    def post_acquisition_processor(self, experiment_directory, x_pixels):

        mcmicro_path = experiment_directory + r'\mcmicro\raw'
        cycle_start = 1
        cycle_start_search = 0

        os.chdir(mcmicro_path)
        while cycle_start_search == 0:
            file_name = str(experiment_directory.split("\\")[-1]) + '-cycle-0' + str(cycle_start) + '.ome.tif'
            if os.path.isfile(file_name) == 1:
                cycle_start += 1
            else:
                cycle_start_search = 1

        #cycle_end = len(os.listdir(dapi_im_path)) + 1
        cycle_end = 9
        cycle_start = 1

        for cycle_number in range(cycle_start, cycle_end):
            self.infocus(experiment_directory, cycle_number, 3 ,4)
            self.illumination_flattening(experiment_directory, cycle_number)
            self.mcmicro_image_stack_generator(cycle_number, experiment_directory)
            self.stage_placement(experiment_directory, cycle_number, x_pixels)

    def mcmicro_image_stack_generator(self, cycle_number, experiment_directory):

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        xml_metadata = self.metadata_generator(experiment_directory)

        numpy_x = full_array[0]
        numpy_y = full_array[1]
        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]
        tile_count = int(x_tile_count * y_tile_count)

        dapi_im_path = experiment_directory + '\DAPI\Stain\cy_' + str(cycle_number) + '\Tiles' + '/focused_basic_corrected'
        a488_im_path = experiment_directory + '\A488\Stain\cy_' + str(cycle_number) + '\Tiles' + '/focused_basic_corrected'
        a555_im_path = experiment_directory + '\A555\Stain\cy_' + str(cycle_number) + '\Tiles' + '/focused_basic_corrected'
        a647_im_path = experiment_directory + '\A647\Stain\cy_' + str(cycle_number) + '\Tiles' + '/focused_basic_corrected'

        mcmicro_path = experiment_directory + r'\mcmicro\raw'

        mcmicro_stack = np.random.rand(tile_count * 4, 2960, 5056).astype('uint16')

        tile = 0
        for x in range(0, x_tile_count):
            for y in range(0, y_tile_count):
                dapi_file_name = 'x' + str(x) + '_y_' + str(y) + '_c_DAPI.tif'
                a488_file_name = 'x' + str(x) + '_y_' + str(y) + '_c_A488.tif'
                a555_file_name = 'x' + str(x) + '_y_' + str(y) + '_c_A555.tif'
                a647_file_name = 'x' + str(x) + '_y_' + str(y) + '_c_A647.tif'

                base_count_number_stack = tile * 4

                os.chdir(dapi_im_path)
                image = io.imread(dapi_file_name).astype('uint16')
                mcmicro_stack[base_count_number_stack + 0] = image

                os.chdir(a488_im_path)
                image = io.imread(a488_file_name).astype('uint16')
                mcmicro_stack[base_count_number_stack + 1] = image

                os.chdir(a555_im_path)
                image = io.imread(a555_file_name).astype('uint16')
                mcmicro_stack[base_count_number_stack + 2] = image

                os.chdir(a647_im_path)
                image = io.imread(a647_file_name).astype('uint16')
                mcmicro_stack[base_count_number_stack + 3] = image

                tile += 1

        os.chdir(mcmicro_path)
        mcmicro_file_name = str(experiment_directory.split("\\")[-1]) + '-cycle-0' + str(cycle_number) + '.ome.tif'
        tf.imwrite(mcmicro_file_name, mcmicro_stack, photometric='minisblack', description=xml_metadata)

    def metadata_generator(self, experiment_directory):

        new_ome = OME()
        ome = from_xml(r'C:\Users\mike\Documents\GitHub\AutoCIF/image.xml', parser='lxml')
        #ome = from_xml(r'C:\Users\CyCIF PC\Documents\GitHub\AutoCIF/image.xml', parser='lxml')
        ome = ome.images[0]

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]
        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]
        total_tile_count = x_tile_count * y_tile_count

        y_gap = 532
        col_col_gap = 10
        #for r in range(3, -1, -1):
        #    numpy_y[r][0] = numpy_y[r + 1][0] - y_gap
        #    numpy_y[r][1] = numpy_y[r + 1][1] - y_gap - col_col_gap

        # sub in needed pixel size and pixel grid changes
        ome.pixels.physical_size_x = 0.2
        ome.pixels.physical_size_y = 0.2
        ome.pixels.size_x = 5056
        ome.pixels.size_y = 2960
        # sub in other optional numbers to make metadata more accurate

        for x in range(0, total_tile_count):
            tile_metadata = deepcopy(ome)
            new_ome.images.append(tile_metadata)

        # sub in stage positional information into each tile. numpy[y][x]
        tile_counter = 0
        for x in range(0, x_tile_count):
            for y in range(0, y_tile_count):

                for p in range(0, 4):

                    new_x = numpy_x[y][x] - 11000
                    new_y = numpy_y[y][x] + 2300
                    new_ome.images[tile_counter].pixels.planes[p].position_y = deepcopy(new_y)
                    new_ome.images[tile_counter].pixels.planes[p].position_x = deepcopy(new_x)
                    new_ome.images[tile_counter].pixels.tiff_data_blocks[p].ifd = (4 * tile_counter) + p
                tile_counter += 1

        xml = to_xml(new_ome)

        return xml

    def stage_placement(self, experiment_directory, cycle_number, x_pixels):
        '''
        Goal to place images via rough stage coords in a larger image. WIll have nonsense borders
        '''

        quick_tile_path = experiment_directory + r'\Quick_Tile'

        # load in numpy matricies

        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]

        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]

        fov_x_pixels = x_pixels
        fov_y_pixels = 2960
        um_pixel = 0.20

        # generate large numpy image with rand numbers. Big enough to hold all images + 10%

        super_y = int(1.02 * (y_tile_count * fov_y_pixels))
        super_x = int(1.02 * (x_tile_count * fov_x_pixels))
        placed_image = np.random.rand(super_y, super_x).astype('uint16')


        # transform numpy x and y coords into new shifted space that starts at zero and is in units of pixels and not um
        numpy_x_pixels = numpy_x / um_pixel
        numpy_y_pixels = numpy_y / um_pixel

        y_displacement_vector = (2100 / um_pixel + fov_y_pixels / 2) * 1.02
        x_displacement_vector = (-11385 / um_pixel + fov_x_pixels / 2) * 1.02

        numpy_x_pixels = numpy_x_pixels + x_displacement_vector
        numpy_x_pixels = np.ceil(numpy_x_pixels)
        numpy_x_pixels = numpy_x_pixels.astype(int)

        numpy_y_pixels = numpy_y_pixels + y_displacement_vector
        numpy_y_pixels = np.ceil(numpy_y_pixels)
        numpy_y_pixels = numpy_y_pixels.astype(int)

        # load images into python

        channels = ['DAPI', 'A488', 'A555', 'A647']
        types = ['\Stain', '\Bleach']

        for type in types:
            for channel in channels:

                im_path = experiment_directory + '/' + channel + type + '\cy_' + str(cycle_number) + '\Tiles' + '/focused_basic_corrected'
                os.chdir(im_path)

                # place images into large array

                for x in range(0, x_tile_count):
                    for y in range(0, y_tile_count):
                        filename = 'x' + str(x) + '_y_' + str(y) + '_c_' + channel + '.tif'
                        image = io.imread(filename)
                        # define subsection of large array that fits dimensions of single FOV
                        #x_center = numpy_x_pixels[y][x]
                        #y_center = numpy_y_pixels[y][x]
                        #x_start = int(x_center - fov_x_pixels / 2)
                        #x_end = int(x_center + fov_x_pixels / 2)
                        #y_start = int(y_center - fov_y_pixels / 2)
                        #y_end = int(y_center + fov_y_pixels / 2)

                        if x == 0 and y ==0:
                            x_start = 0
                            x_end = x_pixels
                            y_start =0
                            y_end = 2960
                        else:

                            x_start = int(x * fov_x_pixels * 0.87)
                            x_end = x_start + x_pixels
                            y_start = int(y * fov_y_pixels * 0.9)
                            y_end = y_start + 2960


                    # placed_image[y_start:y_end, x_start:x_end] = placed_image[y_start:y_end, x_start:x_end] + image
                    placed_image[y_start:y_end, x_start:x_end] = image

                # save output image
                os.chdir(quick_tile_path + '/' + channel)
                tf.imwrite(channel + '_cy_' + str(cycle_number) + '_' + type + '_placed.tif', placed_image)


    def illumination_flattening(self, experiment_directory, cycle_number):

        directory_start = experiment_directory + '//'

        for channel in range(0, 4):

            if channel == 0:
                channel_name = 'DAPI'
            if channel == 1:
                channel_name = 'A488'
            if channel == 2:
                channel_name = 'A555'
            if channel == 3:
                channel_name = 'A647'

            directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\focused'
            output_directory = directory_start + channel_name + '\Stain\cy_' + str(cycle_number) + r'\Tiles\focused_basic_corrected'

            try:
                os.mkdir(output_directory)
            except:
                pass

            epsilon = 1e-06
            optimizer = shading_correction.BaSiC(directory)
            optimizer.prepare()
            optimizer.run()
            # Save the estimated fields (only if the profiles were estimated)
            directory = Path(output_directory)
            flatfield_name = directory / "flatfield.tif"
            darkfield_name = directory / "darkfield.tif"
            cv2.imwrite(str(flatfield_name), optimizer.flatfield_fullsize.astype(np.float32))
            cv2.imwrite(str(darkfield_name), optimizer.darkfield_fullsize.astype(np.float32))
            optimizer.write_images(output_directory, epsilon=epsilon)

    def infocus(self, experiment_directory, cycle_number, x_sub_section_count, y_sub_section_count):

        bin_values = [4]
        channels = ['DAPI', 'A488', 'A555', 'A647']

        dapi_im_path = experiment_directory + '/' + 'DAPI' '\Stain\cy_' + str(cycle_number) + '\Tiles'

        # load numpy arrays in
        numpy_path = experiment_directory + '/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]

        y_tile_count = numpy_x.shape[0]
        x_tile_count = numpy_y.shape[1]

        # determine number z slices
        z_checker = 0
        z_slice_count = 0
        os.chdir(dapi_im_path)
        while z_checker == 0:
            file_name = 'z_' + str(z_slice_count) + '_x' + str(0) + '_y_' + str(0) + '_c_DAPI.tif'
            if os.path.isfile(file_name) == 1:
                z_slice_count += 1
            else:
                z_checker = 1

        for channel in channels:
            # generate imstack of z slices for tile
            # channel = 'DAPI'
            im_path = experiment_directory + '/' + channel + '\Stain\cy_' + str(cycle_number) + '\Tiles'
            os.chdir(im_path)

            z_stack = np.random.rand(z_slice_count, 2960, 5056).astype('uint16')
            for x in range(0, x_tile_count):
                for y in range(0, y_tile_count):
                    for z in range(0, z_slice_count):
                        im_path = experiment_directory + '/' + channel + '\Stain\cy_' + str(cycle_number) + '\Tiles'
                        os.chdir(im_path)
                        file_name = 'z_' + str(z) + '_x' + str(x) + '_y_' + str(y) + '_c_' + str(channel) + '.tif'
                        image = tf.imread(file_name)
                        z_stack[z] = image

                        # break into sub sections (2x3)

                    number_bins = len(bin_values)
                    brenner_sub_selector = np.random.rand(z_slice_count, number_bins, y_sub_section_count,
                                                          x_sub_section_count).astype('longlong')
                    for z in range(0, z_slice_count):
                        for y_sub in range(0, y_sub_section_count):
                            for x_sub in range(0, x_sub_section_count):

                                y_end = int((y_sub + 1) * (2960 / y_sub_section_count))
                                y_start = int(y_sub * (2960 / y_sub_section_count))
                                x_end = int((x_sub + 1) * (5056 / x_sub_section_count))
                                x_start = int(x_sub * (5056 / x_sub_section_count))
                                sub_image = z_stack[z][y_start:y_end, x_start:x_end]

                                for b in range(0, number_bins):
                                    bin_value = int(bin_values[b])
                                    score = self.focus_score(sub_image, bin_value)
                                    brenner_sub_selector[z][b][y_sub][x_sub] = score

                    reconstruct_array = self.brenner_reconstruct_array(brenner_sub_selector, z_slice_count, number_bins)
                    reconstruct_array = skimage.filters.median(reconstruct_array)
                    self.image_reconstructor(experiment_directory, reconstruct_array, channel, cycle_number, y, x)

    def brenner_reconstruct_array(self, brenner_sub_selector, z_slice_count, number_bins):
        '''
        take in sub selector array and slice to find max values for brenner scores and then find mode for various bin levels
        output array that shows what slice to grab each sub section from
        :param brenner_sub_selector:
        :return:
        '''
        # array to dictate the z slice to grab each sub section from
        y_sections = np.shape(brenner_sub_selector)[2]
        x_sections = np.shape(brenner_sub_selector)[3]

        reconstruct_array = np.random.rand(y_sections, x_sections).astype('uint16')

        temp_bin_max_indicies = np.random.rand(number_bins).astype('uint16')

        for y in range(0, y_sections):
            for x in range(0, x_sections):
                for b in range(0, number_bins):
                    sub_scores = brenner_sub_selector[0:z_slice_count, b, y, x]
                    max_score = np.max(sub_scores)
                    max_index = np.where(sub_scores == max_score)[0][0]
                    #temp_bin_max_indicies[b] = max_index
                #sub_section_index_mode = stats.mode(temp_bin_max_indicies)[0][0]
                #reconstruct_array[y][x] = sub_section_index_mode
                reconstruct_array[y][x] = max_index

        return reconstruct_array

    def image_reconstructor(self, experiment_directory, reconstruct_array, channel, cycle_number, y_tile_number, x_tile_number):

        y_sections = np.shape(reconstruct_array)[0]
        x_sections = np.shape(reconstruct_array)[1]

        cycle_types = ['Stain', 'Bleach']

        for cycle_type in cycle_types:

            im_path = experiment_directory + '/' + channel + '/' + cycle_type + '\cy_' + str(cycle_number) + '\Tiles'
            os.chdir(im_path)
            try:
                os.mkdir('focused')
            except:
                t = 5

            # rebuilt image container
            rebuilt_image = np.random.rand(2960, 5056).astype('uint16')

            for y in range(0, y_sections):
                for x in range(0, x_sections):
                    # define y and x start and ends subsection of rebuilt image
                    y_end = int((y + 1) * (2960 / y_sections))
                    y_start = int(y * (2960 / y_sections))
                    x_end = int((x + 1) * (5056 / x_sections))
                    x_start = int(x * (5056 / x_sections))

                    # find z for specific subsection
                    z_slice = reconstruct_array[y][x]
                    # load in image to extract for subsection
                    file_name = 'z_' + str(z_slice) + '_x' + str(x_tile_number) + '_y_' + str(
                        y_tile_number) + '_c_' + str(channel) + '.tif'
                    image = tf.imread(file_name)
                    rebuilt_image[y_start:y_end, x_start:x_end] = image[y_start:y_end, x_start:x_end]

            reconstruct_path = experiment_directory + '/' + channel + '/' + cycle_type + '\cy_' + str(
                cycle_number) + '\Tiles' + '/focused'
            os.chdir(reconstruct_path)
            filename = 'x' + str(x_tile_number) + '_y_' + str(y_tile_number) + '_c_' + str(channel) + '.tif'
            tf.imwrite(filename, rebuilt_image)


    def zc_save(self, zc_tif_stack, channels, x_tile, y_tile, cycle, x_pixels, experiment_directory, Stain_or_Bleach):

        z_tile_count = np.shape(zc_tif_stack)[1]

        for channel in channels:
            if channel == 'DAPI':
                zc_index = 0
            if channel == 'A488':
                zc_index = 1
            if channel == 'A555':
                zc_index = 2
            if channel == 'A647':
                zc_index = 3

            # establish x range to collect in image

            side_pixel_count = int(5056 - x_pixels)

            save_directory = experiment_directory + '/' + str(channel) + '/' + Stain_or_Bleach + '/' + 'cy_' + str(cycle) + '/' + 'Tiles'
            os.chdir(save_directory)

            for z in range(0, z_tile_count):

                        file_name = 'z_' + str(z) + '_x' + str(x_tile) + '_y_' + str(y_tile) + '_c_' + str(channel)+ '.tif'
                        image = zc_tif_stack[zc_index][z][::, side_pixel_count:side_pixel_count + x_pixels]
                        imwrite(file_name, image, photometric='minisblack')



    def save_files(self, z_tile_stack, channel, cycle, experiment_directory, Stain_or_Bleach = 'Stain'):

        save_directory = experiment_directory + '/' + str(channel) + '/' + Stain_or_Bleach + '/' + 'cy_' + str(cycle) + '/' + 'Tiles'

        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]

        x_tile_count = np.unique(numpy_x).size
        y_tile_count = np.unique(numpy_y).size

        z_tile_count = z_tile_stack.shape[0]


        for z in range(0, z_tile_count):

            tile_counter = 0

            for y in range(0, y_tile_count):
                if y % 2 != 0:
                    for x in range(x_tile_count - 1, -1, -1):

                        #meta = self.image_metadata_generation(x, y, channel, experiment_directory)
                        file_name = 'z_' + str(z) + '_x' + str(x) + '_y_' + str(y) + '_c_' + str(channel)+ '.tif'
                        image = z_tile_stack[z][tile_counter]
                        os.chdir(save_directory)
                        imwrite(file_name, image, photometric='minisblack')
                        tile_counter += 1

                if y % 2 == 0:
                    for x in range(0, x_tile_count):

                        #meta = self.image_metadata_generation(x, y, channel, experiment_directory)
                        file_name = 'z_' + str(z) + '_x' + str(x) + '_y_' + str(y) + '_c_' + str(channel)+ '.tif'
                        image = z_tile_stack[z][tile_counter]
                        os.chdir(save_directory)
                        imwrite(file_name, image, photometric='minisblack')
                        tile_counter += 1

    def save_tif_stack(self, tif_stack, cycle_number,  directory_name):

        add_on_folder = 'cycle_' + str(cycle_number)
        full_directory_path = directory_name + add_on_folder
        try:
            os.mkdir(full_directory_path)
        except:
           pass
        os.chdir(full_directory_path)
        file_name = 'image_array.tif'

        imwrite(file_name, tif_stack,
                bigtiff=True,
                photometric='minisblack',
                compression = 'zlib',
                compressionargs = {'level': 8} )


############################################
#depreciated or unused
############################################
    def quick_tile_placement(self, z_tile_stack, overlap = 10):

        numpy_path = 'E:/folder_structure' +'/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]

        x_tile_count = np.unique(numpy_x).size
        y_tile_count = np.unique(numpy_y).size

        height = z_tile_stack[0].shape[0]
        width = z_tile_stack[0].shape[1]
        overlapped_height = int(height * (1 - overlap / 100))
        overlapped_width = int(width * (1 - overlap / 100))

        pna_height = int(y_tile_count * height - int((y_tile_count) * overlap / 100 * height))
        pna_width = int(x_tile_count * width - int((x_tile_count) * overlap / 100 * width))

        pna = np.random.rand(pna_height, pna_width).astype('float16')
        tile_counter = 0

        for y in range(0, y_tile_count):
            if y % 2 != 0:
                for x in range(x_tile_count - 1, -1, -1):
                    pna[y * overlapped_height:(y + 1) * overlapped_height,
                    x * overlapped_width:(x + 1) * overlapped_width] = z_tile_stack[tile_counter][0:overlapped_height,
                                                                       0:overlapped_width]

                    tile_counter += 1


            elif y % 2 == 0:
                for x in range(0, x_tile_count):
                    pna[y * overlapped_height:(y + 1) * overlapped_height,
                    x * overlapped_width:(x + 1) * overlapped_width] = z_tile_stack[tile_counter][0:overlapped_height,
                                                                       0:overlapped_width]

                    tile_counter += 1

        return pna

    def quick_tile_optimal_z(self, z_tile_stack):

        z_slice_count = z_tile_stack.shape[0]
        tile_count = z_tile_stack[0].shape[0]

        height = z_tile_stack[0].shape[1]
        width = z_tile_stack[0].shape[2]

        optimal_stack = np.random.rand(tile_count, height, width).astype('float16')
        score_array = np.random.rand(z_slice_count, 1).astype('float32')


        for tile in range(0, tile_count):

            for z in range(0, z_slice_count):
                score_array[z] = cycif.focus_bin_generator(z_tile_stack[z][tile])

            min_score = np.min(score_array)
            optimal_index = np.where(score_array == min_score)[0][0]
            optimal_stack[tile] = z_tile_stack[optimal_index][tile]

        return optimal_stack

    def optimal_quick_preview_qt(self, z_tile_stack, channel, cycle, experiment_directory,  overlap = 10):

        optimal_stack = self.quick_tile_optimal_z(z_tile_stack)
        optimal_qt = self.quick_tile_placement(optimal_stack, overlap)
        optimal_qt_binned = optimal_qt[0:-1:4, 0:-1:4]
        self.save_optimal_quick_tile(optimal_qt_binned, channel, cycle, experiment_directory)

    def core_tile_acquire(self, experiment_directory, channel = 'DAPI'):
        '''
        Makes numpy files that contain all tiles and z slices. Order is z, tiles.

        :param self:
        :param channels:
        :param z_slices:
        :return:
        '''

        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)
        exp_time_array = np.load('exp_array.npy', allow_pickle=False)

        height_pixels = 2960
        width_pixels = 5056

        numpy_x = full_array[0]
        numpy_y = full_array[1]
        x_tile_count = np.unique(numpy_x).size
        y_tile_count = np.unique(numpy_y).size
        total_tile_count = x_tile_count * y_tile_count
        z_slices = full_array[3][0][0]
        slice_gap = 2

        core.set_xy_position(numpy_x[0][0], numpy_y[0][0])
        time.sleep(1)

        tif_stack = np.random.rand(int(z_slices), total_tile_count, height_pixels, width_pixels).astype('float16')


        if channel == 'DAPI':
            channel_index = 2
            tif_stack_c_index = 0
        if channel == 'A488':
            channel_index = 4
            tif_stack_c_index = 1
        if channel == 'A555':
            channel_index = 6
            tif_stack_c_index = 2
        if channel == 'A647':
            channel_index = 8
            tif_stack_c_index = 3

        numpy_z = full_array[channel_index]
        exp_time = int(exp_time_array[tif_stack_c_index])
        core.set_config("Color", channel)
        core.set_exposure(exp_time)
        tile_counter = 0


        for y in range(0, y_tile_count):
            if y % 2 != 0:
                for x in range(x_tile_count - 1, -1, -1):

                    z_end = int(numpy_z[y][x])
                    z_start = int(z_end - z_slices * slice_gap)
                    core.set_xy_position(numpy_x[y][x], numpy_y[y][x])
                    time.sleep(.5)

                    z_counter = 0

                    for z in range(z_start, z_end, slice_gap):
                        core.set_position(z)
                        time.sleep(0.5)
                        core.snap_image()
                        tagged_image = core.get_tagged_image()
                        pixels = np.reshape(tagged_image.pix,newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])
                        tif_stack[z_counter][tile_counter] = pixels
                        print(core.getRemainingImageCount())

                        z_counter += 1

                    tile_counter += 1


            elif y % 2 == 0:
                for x in range(0, x_tile_count):

                    z_end = int(numpy_z[y][x])
                    z_start = int(z_end - z_slices * slice_gap)
                    core.set_xy_position(numpy_x[y][x], numpy_y[y][x])
                    time.sleep(.5)

                    z_counter = 0

                    for z in range(z_start, z_end, slice_gap):
                        core.set_position(z)
                        time.sleep(0.5)
                        core.snap_image()
                        tagged_image = core.get_tagged_image()
                        pixels = np.reshape(tagged_image.pix,
                                            newshape=[tagged_image.tags["Height"], tagged_image.tags["Width"]])

                        tif_stack[z_counter][tile_counter] = pixels


                        z_counter += 1

                    tile_counter += 1

        return tif_stack

    def quick_tile_all_z_save(self, z_tile_stack, channel, cycle, experiment_directory, stain_bleach,  overlap = 0):


        z_slice_count = z_tile_stack.shape[0]
        numpy_path = experiment_directory +'/' + 'np_arrays'
        os.chdir(numpy_path)
        full_array = np.load('fm_array.npy', allow_pickle=False)

        numpy_x = full_array[0]
        numpy_y = full_array[1]

        x_tile_count = int(np.unique(numpy_x).size)
        y_tile_count = int(np.unique(numpy_y).size)

        height = int(z_tile_stack.shape[2])
        width = int(z_tile_stack.shape[3])

        pna_height = int(y_tile_count * height - int((y_tile_count) * overlap / 100 * height))
        pna_width = int(x_tile_count * width - int((x_tile_count) * overlap / 100 * width))

        pna_stack = np.random.rand(z_slice_count, pna_height, pna_width).astype('float16')

        for z in range(0, z_slice_count):
            pna = self.quick_tile_placement(z_tile_stack[z], overlap)
            pna_stack[z] = pna

        self.save_quick_tile(pna_stack, channel, cycle, experiment_directory, stain_bleach)


############################################
##Control arduino based microfluidic system
############################################

class arduino:

    def __init__(self):
        return

    #def mqtt_publish(self, message, subtopic, topic="control", client=client):
        '''
        takes message and publishes message to server defined by client and under topic of topic/subtopic

        :param str subtopic: second tier of topic heirarchy
        :param str topic: first tier of topic heirarchy
        :param object client: client that MQTT server is on. Established in top of module
        :return:
        '''

        #full_topic = topic + "/" + subtopic

        #client.loop_start()
       # client.publish(full_topic, message)
        #client.loop_stop()

    def heater_state(self, state):

        if state == 'on':
            self.mqtt_publish(210, 'dc_pump')
        elif state == 'off':
            self.mqtt_publish(200, 'dc_pump')

    def chamber(self, fill_drain, run_time=27):
        '''
        Aquarium pumps to fill or drain outer chamber with water. Uses dispense function as backbone.

        :param str fill_drain: fill = fills chamber, drain = drains chamber
        :param int time: time in secs to fill chamber and conversely drain it
        :return: nothing
        '''

        if fill_drain == 'drain':
            self.mqtt_publish(110, 'dc_pump')
            time.sleep(run_time + 3)
            self.mqtt_publish(100, 'dc_pump')

        elif fill_drain == 'fill':
            self.mqtt_publish(111, 'dc_pump')
            time.sleep(run_time)
            self.mqtt_publish(101, 'dc_pump')


class fluidics:

    def __init__(self, mux_com_port, ob1_com_port):

        # OB1 initialize
        ob1_path = 'ASRL' + str(ob1_com_port) + '::INSTR'
        Instr_ID = c_int32()
        pump = OB1_Initialization(ob1_path.encode('ascii'), 0, 0, 0, 0, byref(Instr_ID))
        pump = OB1_Add_Sens(Instr_ID, 1, 5, 1, 0, 7, 0) #16bit working range between 0-1000uL/min, also what are CustomSens_Voltage_5_to_25 and can I really choose any digital range?

        Calib = (c_double * 1000)()
        Elveflow_Calibration_Default(byref(Calib), 1000)
        OB1_Start_Remote_Measurement(Instr_ID.value, byref(Calib), 1000)
        self.calibration_array = byref(Calib)

        set_channel_regulator = int(1)  # convert to int
        set_channel_regulator = c_int32(set_channel_regulator)  # convert to c_int32
        set_channel_sensor = int(1)
        set_channel_sensor = c_int32(set_channel_sensor)  # convert to c_int32
        PID_Add_Remote(Instr_ID.value, set_channel_regulator, Instr_ID.value, set_channel_sensor, 0.9, 0.004, 1)


        # MUX intiialize
        path = 'ASRL' + str(mux_com_port) + '::INSTR'
        mux_Instr_ID = c_int32()
        MUX_DRI_Initialization(path.encode('ascii'), byref(mux_Instr_ID))  # choose the COM port, it can be ASRLXXX::INSTR (where XXX=port number)

        #home
        #answer = (c_char * 40)()
        self.mux_ID = mux_Instr_ID.value
        #MUX_DRI_Send_Command(self.mux_ID, 0, answer, 40)

        self.pump_ID = Instr_ID.value

        return

    def mux_end(self):

        MUX_DRI_Destructor(self.mux_ID)

    def valve_select(self, valve_number):
        '''
        Selects valve in mux unit with associated mux_id to the valve_number declared.
        :param c_int32 mux_id: mux_id given from mux_initialization method
        :param int valve_number: number of desired valve to be selected
        :return: Nothing
        '''

        desired_valve =  valve_number
        valve_number = c_int32(valve_number)
        MUX_DRI_Set_Valve(self.mux_ID, valve_number, 0) #0 is shortest path. clockwise and cc are also options

        valve = c_int32(-1)
        MUX_DRI_Get_Valve(self.mux_ID, byref(valve))
        current_valve = int(valve.value)

        while current_valve != desired_valve:
            MUX_DRI_Get_Valve(self.mux_ID, byref(valve))
            current_valve = int(valve.value)
            #print('valve', current_valve, 'deired valve', desired_valve)
            time.sleep(1)


    def flow(self, flow_rate):

        set_channel=int(1)#convert to int
        set_channel=c_int32(set_channel)#convert to c_int32

        set_target=float(flow_rate) # in uL/min for flow
        set_target=c_double(set_target)#convert to c_double

        #OB1_Start_Remote_Measurement(self.pump_ID, self.calibration_array, 1000)
        OB1_Set_Remote_Target(self.pump_ID, set_channel, set_target)


        data_sens=c_double()
        data_reg=c_double()
        set_channel=int(1)#convert to int
        set_channel=c_int32(set_channel)#convert to c_int32
        OB1_Get_Remote_Data(self.pump_ID,set_channel, byref(data_reg),byref(data_sens))
        current_flow_rate = data_sens.value
        current_pressure = int(data_reg.value)
        #print('current flow rate', int(current_flow_rate))
        time.sleep(3)

        time_log = 0
        '''

        while current_flow_rate < flow_rate * 0.95 or current_flow_rate > flow_rate * 1.05 and current_pressure > 0:

            data_sens = c_double()
            data_reg = c_double()
            set_channel = int(1)  # convert to int
            set_channel = c_int32(set_channel)  # convert to c_int32
            OB1_Get_Remote_Data(self.pump_ID, set_channel, byref(data_reg), byref(data_sens))
            current_flow_rate = int(data_sens.value)
            current_pressure = data_reg.value
            #print('current flow rate', current_flow_rate)
            time.sleep(1)
            time_log += 1
            if time_log > 10:
                OB1_Set_Remote_Target(self.pump_ID, set_channel, set_target)
                time_log = 0
            else:
                pass
        
        '''

        #OB1_Stop_Remote_Measurement(self.pump_ID)

    def ob1_end(self):

        set_channel = int(1)  # convert to int
        set_channel = c_int32(set_channel)  # convert to c_int32

        data_sens=c_double()
        data_reg=c_double()

        x = 0
        self.flow(0)

        while x == 0:
            OB1_Get_Remote_Data(self.pump_ID, set_channel, byref(data_reg), byref(data_sens))
            flow_rate = data_sens.value

            if flow_rate < 10:
                x = 1
            if flow_rate > 10:
                x = 0
            time.sleep(1)

        OB1_Stop_Remote_Measurement(self.pump_ID)
        OB1_Destructor(self.pump_ID)

    def measure(self):

        set_channel = int(1)  # convert to int
        set_channel = c_int32(set_channel)  # convert to c_int32

        data_sens=c_double()
        data_reg=c_double()

        OB1_Get_Remote_Data(self.pump_ID, set_channel, byref(data_reg),byref(data_sens) )

        pressure = data_reg.value
        flow_rate = data_sens.value
        time_stamp = time.time()

        return pressure, flow_rate, time_stamp

    def flow_recorder(self, time_step, total_time, file_name = 'none', plot = 1):

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = 'Time'
        ws.cell(row=1, column=2).value = 'Flow Rate'
        ws.cell(row=1, column=3).value = 'Pressure'

        total_steps = int(total_time/time_step)
        pressure_points = np.random.rand(total_steps).astype('float16')
        time_points = np.random.rand(total_steps).astype('float16')
        flow_points = np.random.rand(total_steps).astype('float16')

        for t in range(0, total_steps):

            pressure_point, flow_point, time_point = self.measure()
            pressure_points[t] = pressure_point
            time_points[t] = t * time_step
            flow_points[t] = flow_point

            ws.cell(row= t + 2, column=1).value = t * time_step
            ws.cell(row= t + 2, column=2).value = flow_point
            ws.cell(row= t + 2, column=3).value = pressure_point

            time.sleep(time_step)

        #wb.save(filename = file_name)

        if plot == 1:
            plt.plot(time_points, flow_points, 'o', color='black')
            plt.show()


    def liquid_action(self, action_type, incub_val=0, stain_valve = 0, window = 0, list_status =0, heater_state = 0):

        bleach_valve = 11
        pbs_valve = 12
        bleach_time = 3 #minutes
        stain_flow_time = 45 #seconds
        if heater_state == 0:
            stain_inc_time = incub_val #minutes
        if heater_state == 1:
            stain_inc_time = 45  #minutes
        nuc_valve = 4
        nuc_flow_time = 45 #seconds
        nuc_inc_time = 3 #minutes

        if action_type == 'Bleach':

            self.valve_select(bleach_valve)
            self.flow(500)
            time.sleep(70)
            self.flow(0)
            # time.sleep(bleach_time*60)
            self.valve_select(pbs_valve)

            for x in range(0, bleach_time):
                status_str = f'Cycle {stain_valve}: bleaching time elapsed {x}'
                status_update(status_str, list_status, window)
                time.sleep(60)
            
            self.flow(500)
            time.sleep(70)
            self.flow(0)
            time.sleep(5)

        elif action_type == 'Stain':

            if heater_state == 1:
                arduino.heater_state(1)
                arduino.chamber('drain')
            else:
                pass

            time.sleep(4)
            self.valve_select(stain_valve)
            self.flow(500)
            time.sleep(stain_flow_time)
            self.flow(0)
            self.valve_select(pbs_valve)

            for x in range(0, stain_inc_time):
                status_str = f'Cycle {stain_valve}: staining time elapsed {x}'
                status_update(status_str, list_status, window)
                time.sleep(60)
                # print('Staining Time Elapsed ', x)

            #if heater_state == 1:
            #    arduino.heater_state(0)
            #    arduino.chamber('fill')
            #else:
            #    pass

            self.flow(500)
            time.sleep(70)
            self.flow(0)


        elif action_type == "Wash":

            status_str = f'Cycle {stain_valve}: start PBS wash'
            status_update(status_str, list_status, window)
            self.valve_select(pbs_valve)
            self.flow(500)
            time.sleep(70)
            self.flow(0)
            status_str = f'Cycle {stain_valve}: PBS wash finished'
            status_update(status_str, list_status, window)


        elif action_type == 'Nuc_Touchup':

            self.valve_select(nuc_valve)
            self.flow(500)
            time.sleep(nuc_flow_time)
            self.flow(0)
            time.sleep(nuc_inc_time*60)

            self.valve_select(pbs_valve)
            self.flow(450)
            time.sleep(70)
            self.flow(0)

        elif action_type == 'PBS_flow_on':

            self.valve_select(pbs_valve)
            self.flow(500)
            time.sleep(10)

        elif action_type == 'PBS_flow_off':

            self.valve_select(pbs_valve)
            self.flow(0)
            time.sleep(10)











    
    




